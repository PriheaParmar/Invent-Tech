from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from calendar import monthcalendar
from datetime import timedelta, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
import json
import logging
from django.urls import reverse
from zoneinfo import ZoneInfo
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from django.http import HttpResponse
from django import forms
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import PermissionDenied, ValidationError
from django.core.paginator import Paginator
from django.core.validators import validate_email
from django.db import transaction
from django.db.models import Count, Prefetch, Q, Sum
from django.http import HttpResponse, HttpResponseForbidden, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.template import TemplateDoesNotExist
from django.template.loader import get_template
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.decorators.http import require_GET, require_POST, require_http_methods

from .forms import (
    AccessoryForm,
    BOMAccessoryItemFormSet,
    BOMForm,
    BOMImageFormSet,
    BOMMaterialItemFormSet,
    BOMJobberTypeProcessFormSet,
    BOMJobberDetailFormSet,
    BOMExpenseItemFormSet,
    BrandForm,
    CatalogueForm,ReadyPOReviewForm,
    CategoryForm,
    ClientForm,
    DashboardProfileForm,
    DyeingMaterialLinkDetailFormSet,
    DyeingMaterialLinkForm,
    DyeingOtherChargeForm,
    DyeingPOInwardForm,MaintenanceExpenseItemFormSet,
    DyeingPurchaseOrderForm,
    DyeingPurchaseOrderItemFormSet,
    ExpenseForm,
    FirmForm,
    FirmAddressFormSet,
    GreigePOInwardForm,    QRCodeRecordForm,
    QualityCheckForm,
    QualityCheckParameterFormSet,
    QualityCheckDefectFormSet,
    GreigePurchaseOrderItemFormSet,
    GreigePOReviewForm,
    GreigePurchaseOrderForm,
    InwardTypeForm,
    JobberForm,
    JobberTypeForm,
    LocationForm,
    MainCategoryForm,
    MaterialForm,
    MaterialShadeForm,
    MaterialSubTypeForm,
    MaterialTypeForm,
    MaterialUnitForm,
    PartyForm,
    PatternTypeForm,
    ProgramForm,
    ProgramJobberDetailFormSet,
    ProgramStartForm,
    ProgramStartFabricFormSet,
    ProgramStartSizeFormSet,
    ProgramStartJobberFormSet,
    ProgramJobberChallanForm,
    ProgramJobberChallanSizeFormSet,
    ProgramJobberChallanApprovalForm,
    validate_program_jobber_challan_size_formset,
    ProgramInvoiceForm,
    MaintenanceRecordForm,CostingSnapshotForm,
    ReadyPOInwardForm,
    ReadyPurchaseOrderForm,
    ReadyPurchaseOrderItemFormSet,
    SubCategoryForm,
    TermsConditionForm,
    VendorForm,
    YarnPOInwardForm,
    YarnPOReviewForm,
    YarnPurchaseOrderForm,
    YarnPurchaseOrderItemFormSet,
)

DyeingPOReviewForm = GreigePOReviewForm

from .forms_permissions import ERPCompanyForm, ERPRoleForm, TeamUserForm
from .permissions import PERMISSION_GROUPS, has_erp_permission, get_actor, get_company, is_company_admin

try:
    from .forms import DispatchChallanForm
except ImportError:
    class DispatchChallanForm(forms.Form):
        pass

from .models import (
    Accessory,
    BOM,
    BOMAccessoryItem,
    BOMImage,
    BOMJobberTypeProcess,MaintenanceExpenseItem,
    BOMJobberDetail,
    BOMMaterialItem,
    Program,
    ProgramSizeDetail,CostingSnapshot,
    ProgramJobberDetail,
    ProgramStart,
    ProgramStartFabric,
    ProgramStartSize,
    ProgramStartJobber,
    ProgramJobberChallan,
    ProgramJobberChallanSize,
    ProgramInvoice,
    ProgramInvoiceItem,
    MaintenanceRecord,
    Brand,
    Catalogue,
    Category,
    Client,
    DyeingMaterialLink,
    DyeingMaterialLinkDetail,
    DyeingOtherCharge,
    ERPNotification,
    ERPCompany,
    ERPUserProfile,
    ERPRole,
    DyeingPOInward,    InventoryLot,
    InventoryMovement,
    
    FirmAddress,
    QRCodeRecord,
    QualityCheck,
    DyeingPOInwardItem,
    DyeingPurchaseOrder,
    DyeingPurchaseOrderItem,
    Expense,
    Firm,
    GreigePOInward,
    GreigePOInwardItem,
    GreigePurchaseOrder,
    GreigePurchaseOrderItem,
    InwardType,
    Jobber,
    JobberType,
    Location,
    MainCategory,
    Material,
    MaterialShade,
    MaterialSubType,
    MaterialType,
    MaterialUnit,
    Party,
    PatternType,
    ReadyPOInward,
    ReadyPOInwardItem,
    ReadyPurchaseOrder,
    ReadyPurchaseOrderItem,
    SubCategory,
    TermsCondition,
    UserExtra,
    Vendor,
    YarnPOInward,
    YarnPOInwardItem,
    YarnPurchaseOrder,
    YarnPurchaseOrderItem,
        InventoryLot,
    InventoryMovement,
    QRCodeRecord,
    QualityCheck,
    next_quality_check_number,
    next_qr_code_number,
)
from .navigation import get_utility_groups

try:
    from .models import DispatchChallan
except ImportError:
    DispatchChallan = None


logger = logging.getLogger(__name__)





# ============================================================
# PLATFORM + COMPANY ROLE PERMISSIONS
# ============================================================


def _require_platform_admin(request):
    actor = get_actor(request)
    if not actor or not actor.is_authenticated or not actor.is_superuser:
        raise PermissionDenied("Only platform super admin can access this page.")
    return actor


def _require_company_admin(request):
    actor = get_actor(request)
    if not actor or not actor.is_authenticated:
        raise PermissionDenied("Login required.")
    if actor.is_superuser:
        raise PermissionDenied("Platform super admin cannot manage company roles/users directly. Login as the company admin.")
    if not is_company_admin(request):
        raise PermissionDenied("Only company admin can access this page.")
    return actor


@login_required
def platform_company_list(request):
    _require_platform_admin(request)
    companies = (
        ERPCompany.objects.select_related("admin_user")
        .prefetch_related("user_profiles")
        .order_by("name")
    )
    return render(
        request,
        "accounts/platform/company_list.html",
        {
            "companies": companies,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def platform_company_create(request):
    _require_platform_admin(request)
    form = ERPCompanyForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        company = form.save()
        messages.success(request, f"{company.name} company created with admin login.")
        return redirect("accounts:platform_company_list")

    return render(
        request,
        "accounts/platform/company_form.html",
        {
            "form": form,
            "mode": "add",
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def platform_company_update(request, pk):
    _require_platform_admin(request)
    company = get_object_or_404(ERPCompany.objects.select_related("admin_user"), pk=pk)
    form = ERPCompanyForm(request.POST or None, instance=company)

    if request.method == "POST" and form.is_valid():
        company = form.save()
        messages.success(request, f"{company.name} company updated.")
        return redirect("accounts:platform_company_list")

    return render(
        request,
        "accounts/platform/company_form.html",
        {
            "form": form,
            "mode": "edit",
            "company": company,
        },
    )


@login_required
@require_POST
def platform_company_toggle(request, pk):
    _require_platform_admin(request)
    company = get_object_or_404(ERPCompany, pk=pk)
    company.status = ERPCompany.STATUS_INACTIVE if company.is_active_company else ERPCompany.STATUS_ACTIVE
    company.save(update_fields=["status", "updated_at"])
    messages.success(request, f"{company.name} status updated.")
    return redirect("accounts:platform_company_list")


def _company_for_settings(request):
    company = get_company(request)
    if company:
        return company

    actor = get_actor(request)
    if actor and actor.is_superuser:
        company_id = request.GET.get("company") or request.POST.get("company")
        if company_id:
            return get_object_or_404(ERPCompany, pk=company_id)
        return None

    try:
        return request.user.erp_company_admin
    except Exception:
        return None


@login_required
def role_list(request):
    _require_company_admin(request)
    company = _company_for_settings(request)
    if company is None:
        messages.warning(request, "Create an ERP company first.")
        return redirect("accounts:platform_company_list") if get_actor(request).is_superuser else redirect("accounts:dashboard")

    roles = ERPRole.objects.filter(company=company).order_by("name")
    users = (
        ERPUserProfile.objects.filter(company=company, user_type=ERPUserProfile.USER_TYPE_STAFF)
        .select_related("user", "role")
        .prefetch_related("allowed_firms")
        .order_by("user__username")
    )
    firms = Firm.objects.filter(owner=company.admin_user).order_by("firm_name")

    return render(
        request,
        "accounts/permissions/role_list.html",
        {
            "company": company,
            "roles": roles,
            "team_users": users,
            "firms": firms,
            "permission_groups": PERMISSION_GROUPS,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def role_create(request):
    _require_company_admin(request)
    company = _company_for_settings(request)
    if company is None:
        messages.warning(request, "Create an ERP company first.")
        return redirect("accounts:dashboard")

    form = ERPRoleForm(request.POST or None, company=company)

    if request.method == "POST" and form.is_valid():
        role = form.save()
        messages.success(request, f"{role.name} role created.")
        return redirect("accounts:role_list")

    return render(
        request,
        "accounts/permissions/role_form.html",
        {
            "form": form,
            "company": company,
            "mode": "add",
            "permission_groups": PERMISSION_GROUPS,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def role_update(request, pk):
    _require_company_admin(request)
    company = _company_for_settings(request)
    role = get_object_or_404(ERPRole, pk=pk, company=company)
    form = ERPRoleForm(request.POST or None, instance=role, company=company)

    if request.method == "POST" and form.is_valid():
        role = form.save()
        messages.success(request, f"{role.name} role updated.")
        return redirect("accounts:role_list")

    return render(
        request,
        "accounts/permissions/role_form.html",
        {
            "form": form,
            "company": company,
            "role": role,
            "mode": "edit",
            "permission_groups": PERMISSION_GROUPS,
        },
    )


@login_required
@require_POST
def role_delete(request, pk):
    _require_company_admin(request)
    company = _company_for_settings(request)
    role = get_object_or_404(ERPRole, pk=pk, company=company)
    role.delete()
    messages.success(request, "Role deleted.")
    return redirect("accounts:role_list")


@login_required
@require_http_methods(["GET", "POST"])
def team_user_create(request):
    _require_company_admin(request)
    company = _company_for_settings(request)
    if company is None:
        messages.warning(request, "Create an ERP company first.")
        return redirect("accounts:dashboard")

    form = TeamUserForm(request.POST or None, company=company)

    if request.method == "POST" and form.is_valid():
        profile = form.save()
        messages.success(request, f"{profile.user.username} user created.")
        return redirect("accounts:role_list")

    return render(
        request,
        "accounts/permissions/team_user_form.html",
        {
            "form": form,
            "company": company,
            "mode": "add",
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def team_user_update(request, pk):
    _require_company_admin(request)
    company = _company_for_settings(request)
    profile = get_object_or_404(
        ERPUserProfile.objects.select_related("user", "role"),
        pk=pk,
        company=company,
        user_type=ERPUserProfile.USER_TYPE_STAFF,
    )
    form = TeamUserForm(
        request.POST or None,
        instance=profile,
        company=company,
        user_instance=profile.user,
    )

    if request.method == "POST" and form.is_valid():
        profile = form.save()
        messages.success(request, f"{profile.user.username} user updated.")
        return redirect("accounts:role_list")

    return render(
        request,
        "accounts/permissions/team_user_form.html",
        {
            "form": form,
            "company": company,
            "profile": profile,
            "mode": "edit",
        },
    )


@login_required
@require_POST
def team_user_toggle(request, pk):
    _require_company_admin(request)
    company = _company_for_settings(request)
    profile = get_object_or_404(
        ERPUserProfile,
        pk=pk,
        company=company,
        user_type=ERPUserProfile.USER_TYPE_STAFF,
    )
    profile.is_active = not profile.is_active
    profile.user.is_active = profile.is_active
    profile.user.save(update_fields=["is_active"])
    profile.save(update_fields=["is_active", "updated_at"])
    messages.success(request, f"{profile.user.username} status updated.")
    return redirect("accounts:role_list")


@login_required
@require_POST
def notifications_mark_all_read(request):
    ERPNotification.objects.filter(owner=request.user, is_read=False).update(
        is_read=True,
        read_at=timezone.now(),
    )
    next_url = request.POST.get("next") or request.META.get("HTTP_REFERER") or reverse("accounts:dashboard")
    if not url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
        next_url = reverse("accounts:dashboard")
    return redirect(next_url)

def _first_form_error(form):
    for field_name, errors in form.errors.items():
        if errors:
            label = field_name
            try:
                if field_name != "__all__":
                    label = form.fields[field_name].label or field_name.replace("_", " ").title()
            except Exception:
                label = field_name.replace("_", " ").title()

            return {
                "field": field_name,
                "label": label,
                "message": str(errors[0]),
            }

    return {
        "field": "",
        "label": "",
        "message": "Please check the form.",
    }

def _is_embed(request) -> bool:
    return (
        request.GET.get("embed") == "1"
        or request.POST.get("embed") == "1"
        or request.headers.get("X-Requested-With") == "XMLHttpRequest"
    )


def _utility_page_size(request, default=25):
    allowed = (10, 25, 50, 100)
    raw = request.GET.get("per_page") or request.GET.get("page_size") or default
    try:
        value = int(raw)
    except (TypeError, ValueError):
        value = default
    return value if value in allowed else default


def _paginate_utility_queryset(request, queryset, *, default_page_size=25):
    page_size = _utility_page_size(request, default_page_size)
    paginator = Paginator(queryset, page_size)
    page_obj = paginator.get_page(request.GET.get("page") or 1)

    page_params = request.GET.copy()
    page_params.pop("page", None)

    size_params = request.GET.copy()
    size_params.pop("page", None)
    size_params.pop("per_page", None)
    size_params.pop("page_size", None)

    return {
        "object_list": page_obj.object_list,
        "page_obj": page_obj,
        "paginator": paginator,
        "page_size": page_size,
        "page_sizes": (10, 25, 50, 100),
        "page_querystring": page_params.urlencode(),
        "page_size_querystring": size_params.urlencode(),
        "total_results": paginator.count,
    }



def _template_exists(template_name: str) -> bool:
    try:
        get_template(template_name)
        return True
    except TemplateDoesNotExist:
        return False


def _pick_template(*template_names: str) -> str:
    for template_name in template_names:
        if template_name and _template_exists(template_name):
            return template_name
    return template_names[-1]


def _model_has_fields(model, *field_names: str) -> bool:
    try:
        existing_fields = {field.name for field in model._meta.get_fields()}
    except Exception:
        return False
    return all(field_name in existing_fields for field_name in field_names)

def _flatten_form_errors(form):
    items = []
    for field_name, errors in form.errors.items():
        if field_name == "__all__":
            label = "Form"
        else:
            try:
                label = form.fields[field_name].label or field_name.replace("_", " ").title()
            except Exception:
                label = field_name.replace("_", " ").title()

        for error in errors:
            items.append(
                {
                    "section": "Main BOM Form",
                    "row": "",
                    "field": label,
                    "message": str(error),
                }
            )
    return items


def _flatten_formset_errors(section_name, formset):
    items = []

    for error in formset.non_form_errors():
        items.append(
            {
                "section": section_name,
                "row": "",
                "field": "Formset",
                "message": str(error),
            }
        )

    for index, form in enumerate(formset.forms, start=1):
        if not form.errors:
            continue

        for field_name, errors in form.errors.items():
            if field_name == "__all__":
                label = "Row"
            else:
                try:
                    label = form.fields[field_name].label or field_name.replace("_", " ").title()
                except Exception:
                    label = field_name.replace("_", " ").title()

            for error in errors:
                items.append(
                    {
                        "section": section_name,
                        "row": index,
                        "field": label,
                        "message": str(error),
                    }
                )

    return items


def _collect_bom_debug_errors(
    form,
    material_formset,
    accessory_formset,
    image_formset,
    jobber_process_formset,
    jobber_detail_formset,
    expense_formset,
):
    errors = []
    errors.extend(_flatten_form_errors(form))
    errors.extend(_flatten_formset_errors("Materials", material_formset))
    errors.extend(_flatten_formset_errors("Accessories", accessory_formset))
    errors.extend(_flatten_formset_errors("Images", image_formset))
    errors.extend(_flatten_formset_errors("Jobber Processes", jobber_process_formset))
    errors.extend(_flatten_formset_errors("Jobber Details", jobber_detail_formset))
    errors.extend(_flatten_formset_errors("Expenses", expense_formset))
    return errors

def _client_list_url(request):
    url = reverse("accounts:client_list")
    if _is_embed(request):
        url += "?embed=1"
    return url



def _client_usage_rows(client):
    rows = []

    dispatch_count = 0
    invoice_count = 0

    if hasattr(client, "dispatch_challans"):
        try:
            dispatch_count = client.dispatch_challans.count()
        except Exception:
            dispatch_count = 0

    if hasattr(client, "program_invoices"):
        try:
            invoice_count = client.program_invoices.count()
        except Exception:
            invoice_count = 0

    if dispatch_count:
        rows.append({"label": "Dispatch Challans", "count": dispatch_count})

    if invoice_count:
        rows.append({"label": "Program Invoices", "count": invoice_count})

    return rows



def _client_list_context(request):
    q = (request.GET.get("q") or "").strip()
    status = (request.GET.get("status") or "").strip()

    qs = Client.objects.filter(owner=request.user)

    if q:
        qs = qs.filter(
            Q(name__icontains=q)
            | Q(contact_person__icontains=q)
            | Q(phone__icontains=q)
            | Q(email__icontains=q)
            | Q(gst_number__icontains=q)
            | Q(pan_number__icontains=q)
            | Q(city__icontains=q)
        )

    if status == "active":
        qs = qs.filter(is_active=True)
    elif status == "inactive":
        qs = qs.filter(is_active=False)

    all_clients = Client.objects.filter(owner=request.user)

    return {
        "clients": qs.order_by("name"),
        "q": q,
        "status": status,
        "stats": {
            "total": all_clients.count(),
            "active": all_clients.filter(is_active=True).count(),
            "inactive": all_clients.filter(is_active=False).count(),
        },
    }


@login_required
def client_list(request):
    tpl = "accounts/clients/list_embed.html" if _is_embed(request) else "accounts/clients/list.html"
    return render(request, tpl, _client_list_context(request))


@login_required
@require_http_methods(["GET", "POST"])
def client_create(request):
    form = ClientForm(request.POST or None, user=request.user)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.owner = request.user
        obj.save()

        url = _client_list_url(request)
        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})
        return redirect(url)

    tpl = "accounts/clients/form_embed.html" if _is_embed(request) else "accounts/clients/form.html"
    return render(request, tpl, {"form": form, "mode": "add"})


@login_required
@require_http_methods(["GET", "POST"])
def client_update(request, pk: int):
    client = get_object_or_404(Client, pk=pk, owner=request.user)
    form = ClientForm(request.POST or None, instance=client, user=request.user)

    if request.method == "POST" and form.is_valid():
        form.save()

        url = _client_list_url(request)
        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})
        return redirect(url)

    tpl = "accounts/clients/form_embed.html" if _is_embed(request) else "accounts/clients/form.html"
    return render(request, tpl, {"form": form, "mode": "edit", "client": client})


@login_required
@require_POST
def client_delete(request, pk: int):
    client = get_object_or_404(Client, pk=pk, owner=request.user)
    usage_rows = _client_usage_rows(client)

    if usage_rows:
        error_message = (
            f'Cannot delete "{client.name}" because it is already used in linked records. '
            f'Mark it inactive instead.'
        )

        if _is_embed(request):
            context = _client_list_context(request)
            context["delete_error"] = error_message
            context["delete_label"] = client.name
            context["delete_usage"] = usage_rows
            return render(request, "accounts/clients/list_embed.html", context)

        messages.error(request, error_message)
        return redirect("accounts:client_list")

    client.delete()

    url = _client_list_url(request)
    if _is_embed(request):
        return JsonResponse({"ok": True, "url": url})
    return redirect(url)


def _next_yarn_po_number() -> str:
    last = YarnPurchaseOrder.objects.order_by("-id").first()
    next_id = (last.id + 1) if last else 1
    return f"YPO-{next_id:04d}"

def _next_greige_po_number() -> str:
    last = GreigePurchaseOrder.objects.order_by("-id").first()
    next_id = (last.id + 1) if last else 1
    return f"GPO-{next_id:04d}"


def _next_greige_inward_number() -> str:
    last = GreigePOInward.objects.order_by("-id").first()
    next_id = (last.id + 1) if last else 1
    return f"GIN-{next_id:04d}"


def _next_dyeing_po_number() -> str:
    last = DyeingPurchaseOrder.objects.order_by("-id").first()
    next_id = (last.id + 1) if last else 1
    return f"DPO-{next_id:04d}"

def _next_ready_po_number() -> str:
    last = ReadyPurchaseOrder.objects.order_by("-id").first()
    next_id = (last.id + 1) if last else 1
    return f"RPO-{next_id:04d}"

def _compact_spaces(value):
    """Normalize user-entered/material names for safe ID/name fallback matching."""
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def _firm_address(firm):
    if not firm:
        return ""
    default_address = getattr(firm, "default_shipping_address", None)
    if default_address:
        return default_address.full_address
    parts = [firm.address_line, firm.city, firm.state, firm.pincode]
    return ", ".join([p for p in parts if p])


def _recalculate_yarn_po(po: YarnPurchaseOrder):
    subtotal = Decimal("0")
    total_weight = Decimal("0")

    for item in po.items.all():
        qty = item.quantity or Decimal("0")
        rate = item.rate or Decimal("0")
        calculated_amount = qty * rate
        if item.final_amount != calculated_amount or item.material_type_id != getattr(item.material, "material_type_id", None):
            item.final_amount = calculated_amount
            item.material_type = item.material.material_type if item.material else None
            item.save(update_fields=["final_amount", "material_type"])
        subtotal += item.final_amount or Decimal("0")
        total_weight += qty

    discount_percent = po.discount_percent or Decimal("0")
    others = po.others or Decimal("0")
    cgst_percent = po.cgst_percent or Decimal("0")
    sgst_percent = po.sgst_percent or Decimal("0")

    after_discount_value = subtotal - (subtotal * discount_percent / Decimal("100"))
    tax_value = after_discount_value * (cgst_percent + sgst_percent) / Decimal("100")
    grand_total = after_discount_value + others + tax_value

    po.total_weight = total_weight
    po.subtotal = subtotal
    po.after_discount_value = after_discount_value
    po.grand_total = grand_total
    if not po.system_number:
        po.system_number = _next_yarn_po_number()
    po.save(
        update_fields=[
            "system_number",
            "total_weight",
            "subtotal",
            "after_discount_value",
            "grand_total",
            "updated_at",
        ]
    )


def _can_review_yarn_po(request_or_user):
    """
    Review permission must be checked against the real ERP actor, not only
    request.user. Staff requests are tenant-bridged so request.user may be the
    company owner, while request.erp_actor is the person actually logged in.
    """
    if hasattr(request_or_user, "erp_actor") or hasattr(request_or_user, "erp_is_company_admin"):
        return has_erp_permission(request_or_user, "yarn_po.review")

    user = request_or_user
    if not user or not getattr(user, "is_authenticated", False):
        return False

    if getattr(user, "is_superuser", False) or getattr(user, "is_staff", False) or user.username.lower() == "admin":
        return True

    try:
        profile = user.erp_profile
    except Exception:
        profile = None

    if not profile or not getattr(profile, "is_active", False):
        return False
    if not getattr(user, "is_active", True):
        return False
    if getattr(profile, "company", None) and not getattr(profile.company, "is_active_company", True):
        return False
    if getattr(profile, "is_company_admin", False):
        return True

    role = getattr(profile, "role", None)
    return bool(role and getattr(role, "is_active", False) and "yarn_po.review" in (role.permissions or []))

def _greige_po_lock_reason(po: GreigePurchaseOrder):
    if getattr(po, "approval_status", "") == "approved":
        return "This Greige PO is locked because it has already been approved."
    if po.inwards.exists():
        return "This Greige PO is locked because inward has already started against it."
    if po.dyeing_pos.exists() or po.inwards.filter(generated_dyeing_pos__isnull=False).exists():
        return "This Greige PO is locked because it is already linked to Dyeing PO generation."
    return ""


def _greige_po_delete_lock_reason(po: GreigePurchaseOrder):
    if getattr(po, "approval_status", "") == "approved":
        return "Approved Greige PO cannot be deleted."
    if po.inwards.exists():
        return "Greige PO with inward history cannot be deleted."
    if po.dyeing_pos.exists() or po.inwards.filter(generated_dyeing_pos__isnull=False).exists():
        return "Greige PO linked to Dyeing PO cannot be deleted."
    return ""


def _recalculate_greige_po(po: GreigePurchaseOrder):
    total_qty = Decimal("0")

    for item in po.items.select_related("material").all():
        qty = item.quantity or Decimal("0")
        rate = item.rate or Decimal("0")
        calculated_amount = qty * rate

        update_fields = []

        if item.final_amount != calculated_amount:
            item.final_amount = calculated_amount
            update_fields.append("final_amount")

        desired_fabric_name = item.material.name if item.material else (item.fabric_name or "")
        if desired_fabric_name != (item.fabric_name or ""):
            item.fabric_name = desired_fabric_name
            update_fields.append("fabric_name")

        if update_fields:
            item.save(update_fields=update_fields)

        total_qty += qty

    po.available_qty = total_qty
    po.save(update_fields=["available_qty", "updated_at"])


def _apply_greige_source_links(po: GreigePurchaseOrder):
    source_rows = []

    if po.source_yarn_inward_id:
        source_rows = [
            inward_item.po_item
            for inward_item in po.source_yarn_inward.items.select_related(
                "po_item__material",
                "po_item__material_type",
            ).all()
            if inward_item.po_item_id
        ]
    elif po.source_yarn_po_id:
        source_rows = list(
            po.source_yarn_po.items.select_related("material", "material_type").all()
        )

    greige_items = list(
        po.items.select_related("material", "source_yarn_po_item").all()
    )

    for index, item in enumerate(greige_items):
        source_item = source_rows[index] if index < len(source_rows) else None

        desired_fabric_name = item.material.name if item.material else (item.fabric_name or "")
        desired_yarn_name = item.yarn_name or ""

        if source_item is not None:
            desired_yarn_name = (
                source_item.material.name
                if source_item.material
                else (source_item.material_type.name if source_item.material_type else "")
            )

        update_fields = []

        if source_item is not None and item.source_yarn_po_item_id != source_item.id:
            item.source_yarn_po_item = source_item
            update_fields.append("source_yarn_po_item")

        if item.fabric_name != desired_fabric_name:
            item.fabric_name = desired_fabric_name
            update_fields.append("fabric_name")

        if item.yarn_name != desired_yarn_name:
            item.yarn_name = desired_yarn_name
            update_fields.append("yarn_name")

        if update_fields:
            item.save(update_fields=update_fields)

def _yarn_po_lock_reason(po: YarnPurchaseOrder):
    if po.approval_status == "approved":
        return "This Yarn PO is locked because it has already been approved."
    if po.inwards.exists():
        return "This Yarn PO is locked because inward has already started against it."
    if po.greige_pos.exists() or po.inwards.filter(generated_greige_pos__isnull=False).exists():
        return "This Yarn PO is locked because it is already linked to Greige PO generation."
    return ""


def _yarn_po_delete_lock_reason(po: YarnPurchaseOrder):
    if po.approval_status == "approved":
        return "Approved Yarn PO cannot be deleted."
    if po.inwards.exists():
        return "Yarn PO with inward history cannot be deleted."
    if po.greige_pos.exists() or po.inwards.filter(generated_greige_pos__isnull=False).exists():
        return "Yarn PO linked to Greige PO cannot be deleted."
    return ""

@require_http_methods(["GET", "POST"])
def signup_view(request):
    """Public signup is disabled for this ERP.

    Company admins must be created by the platform super admin from
    Platform → Companies, and company staff must be created by the company
    admin from Settings → Roles & Users. This prevents profile-less users from
    entering company ERP data without a tenant/profile scope.
    """
    if request.user.is_authenticated:
        return redirect("accounts:dashboard")

    return render(
        request,
        "accounts/signup.html",
        {
            "error": "Public signup is disabled. Ask the platform admin to create your company/admin account.",
            "form_data": {
                "username": "",
                "email": "",
                "password": "",
                "password2": "",
            },
            "signup_disabled": True,
        },
        status=403 if request.method == "POST" else 200,
    )

@require_http_methods(["GET", "POST"])
def login_view(request):
    if request.user.is_authenticated:
        return redirect("accounts:dashboard")

    error = None

    if request.method == "POST":
        identifier = request.POST.get("username", "").strip()
        password = request.POST.get("password", "")
        remember_me = request.POST.get("remember_me")

        username = identifier
        if "@" in identifier:
            user_obj = User.objects.filter(email__iexact=identifier).first()
            username = user_obj.username if user_obj else None

        user = authenticate(request, username=username, password=password) if username else None

        if user is not None:
            login(request, user)

            if remember_me:
                request.session.set_expiry(60 * 60 * 24 * 14)
            else:
                request.session.set_expiry(0)

            next_url = request.GET.get("next") or request.POST.get("next")
            if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
                return redirect(next_url)

            return redirect("accounts:dashboard")

        error = "Invalid username/email or password."

    return render(request, "accounts/login.html", {"error": error})


@require_http_methods(["POST"])
def logout_view(request):
    logout(request)
    return redirect("accounts:login")


@login_required
def dashboard_view(request):
    dashboard_tz = ZoneInfo("Asia/Kolkata")
    now_local = timezone.now().astimezone(dashboard_tz)
    hour = now_local.hour

    if 5 <= hour < 12:
        greeting = "Good morning"
    elif 12 <= hour < 17:
        greeting = "Good afternoon"
    elif 17 <= hour < 21:
        greeting = "Good evening"
    else:
        greeting = "Good night"

    month_weeks = monthcalendar(now_local.year, now_local.month)
    calendar_weeks = []

    for week in month_weeks:
        calendar_weeks.append([
            {
                "day": day,
                "is_current_day": day == now_local.day,
                "is_empty": day == 0,
            }
            for day in week
        ])

    today_local = now_local.date()
    owner = request.user

    def inward_summary(model):
        return model.objects.filter(owner=owner).aggregate(
            total=Count("id"),
            today=Count("id", filter=Q(inward_date=today_local)),
        )

    yarn_summary = inward_summary(YarnPOInward)
    greige_summary = inward_summary(GreigePOInward)
    dyeing_summary = inward_summary(DyeingPOInward)
    ready_summary = inward_summary(ReadyPOInward)

    yarn_inward_count = yarn_summary["total"] or 0
    greige_inward_count = greige_summary["total"] or 0
    dyeing_inward_count = dyeing_summary["total"] or 0
    ready_inward_count = ready_summary["total"] or 0

    fabric_inward_count = greige_inward_count + dyeing_inward_count + ready_inward_count
    total_inward_count = yarn_inward_count + greige_inward_count + dyeing_inward_count + ready_inward_count
    today_inward_count = sum([
        yarn_summary["today"] or 0,
        greige_summary["today"] or 0,
        dyeing_summary["today"] or 0,
        ready_summary["today"] or 0,
    ])

    # -----------------------------
    # PROGRAM / WIP
    # -----------------------------
    total_programs = Program.objects.filter(owner=owner).count()
    open_programs = Program.objects.filter(owner=owner, status="open").count()
    closed_programs = Program.objects.filter(owner=owner, status="closed").count()
    started_programs = ProgramStart.objects.filter(owner=owner, is_started=True).count()
    programs_today = Program.objects.filter(owner=owner, program_date=today_local).count()

    recent_programs = (
        Program.objects.filter(owner=owner)
        .select_related("bom", "firm")
        .order_by("-id")[:6]
    )

    recent_program_rows = []
    for program in recent_programs:
        start_record = getattr(program, "start_record", None)

        challans = ProgramJobberChallan.objects.filter(
            owner=owner,
            program=program,
        )

        total_issued = challans.aggregate(total=Sum("total_issued_qty")).get("total") or Decimal("0")
        total_inward = challans.aggregate(total=Sum("inward_qty")).get("total") or Decimal("0")
        pending_qty = total_issued - total_inward
        if pending_qty < Decimal("0"):
            pending_qty = Decimal("0")

        if not start_record or not getattr(start_record, "is_started", False):
            production_status = "Not Started"
        elif total_issued > 0 and pending_qty == Decimal("0"):
            production_status = "Completed"
        else:
            production_status = "In Progress"

        recent_program_rows.append({
            "program_no": program.program_no,
            "sku": getattr(program.bom, "sku", "") or "-",
            "product_name": getattr(program.bom, "product_name", "") or "-",
            "qty": program.total_qty or 0,
            "is_started": bool(start_record and start_record.is_started),
            "status": program.status,
            "total_issued": total_issued,
            "total_inward": total_inward,
            "pending_qty": pending_qty,
            "production_status": production_status,
        })

    # -----------------------------
    # DISPATCH / INVOICE
    # -----------------------------
    total_dispatch = DispatchChallan.objects.filter(owner=owner).count()
    dispatch_today = DispatchChallan.objects.filter(owner=owner, challan_date=today_local).count()

    total_invoices = ProgramInvoice.objects.filter(owner=owner).count()
    invoices_today = ProgramInvoice.objects.filter(owner=owner, invoice_date=today_local).count()

    # -----------------------------
    # LOT / QC SNAPSHOT
    # -----------------------------
    total_lots = InventoryLot.objects.filter(owner=owner).count()
    pending_lots = InventoryLot.objects.filter(owner=owner, qc_status="pending").count()
    approved_lots = InventoryLot.objects.filter(owner=owner, qc_status="approved").count()
    hold_lots = InventoryLot.objects.filter(owner=owner, qc_status="hold").count()
    rejected_lots = InventoryLot.objects.filter(owner=owner, qc_status="rejected").count()

    def safe_percent(value, total):
        if not total:
            return 0
        return round((value / total) * 100)

    pending_lots_pct = safe_percent(pending_lots, total_lots)
    approved_lots_pct = safe_percent(approved_lots, total_lots)
    hold_lots_pct = safe_percent(hold_lots, total_lots)

    lot_watch_qs = (
        InventoryLot.objects.filter(owner=owner)
        .select_related("material")
        .order_by("qc_status", "available_qty", "-id")[:8]
    )

    lot_watch_rows = []
    for lot in lot_watch_qs:
        needs_attention = (
            lot.qc_status in {"pending", "hold", "rejected"}
            or (lot.available_qty or 0) <= 0
        )
        if not needs_attention:
            continue

        lot_watch_rows.append({
            "lot_code": lot.lot_code,
            "material_name": getattr(lot.material, "name", "") or "-",
            "available_qty": lot.available_qty or 0,
            "qc_status": lot.qc_status,
        })

    # -----------------------------
    # 14-DAY ACTIVITY TREND
    # -----------------------------
    last_14_days = [today_local - timedelta(days=offset) for offset in range(13, -1, -1)]

    trend_map = {
        day: {
            "label": day.strftime("%d %b"),
            "programs": 0,
            "inwards": 0,
            "dispatch": 0,
            "invoices": 0,
            "total": 0,
        }
        for day in last_14_days
    }

    for obj in Program.objects.filter(owner=owner, program_date__gte=last_14_days[0], program_date__lte=today_local):
        if obj.program_date in trend_map:
            trend_map[obj.program_date]["programs"] += 1

    for model in [YarnPOInward, GreigePOInward, DyeingPOInward, ReadyPOInward]:
        for obj in model.objects.filter(owner=owner, inward_date__gte=last_14_days[0], inward_date__lte=today_local):
            if obj.inward_date in trend_map:
                trend_map[obj.inward_date]["inwards"] += 1

    for obj in DispatchChallan.objects.filter(owner=owner, challan_date__gte=last_14_days[0], challan_date__lte=today_local):
        if obj.challan_date in trend_map:
            trend_map[obj.challan_date]["dispatch"] += 1

    for obj in ProgramInvoice.objects.filter(owner=owner, invoice_date__gte=last_14_days[0], invoice_date__lte=today_local):
        if obj.invoice_date in trend_map:
            trend_map[obj.invoice_date]["invoices"] += 1

    trend_points = []
    max_total = 0
    for day in last_14_days:
        row = trend_map[day]
        row["total"] = row["programs"] + row["inwards"] + row["dispatch"] + row["invoices"]
        max_total = max(max_total, row["total"])
        trend_points.append(row)

    if max_total <= 0:
        max_total = 1

    for row in trend_points:
        row["bar_height"] = max(10, round((row["total"] / max_total) * 100)) if row["total"] else 6

    # -----------------------------
    # RECENT TIMELINE
    # -----------------------------
    timeline_items = []

    for obj in Program.objects.filter(owner=owner).select_related("bom").order_by("-created_at")[:5]:
        timeline_items.append({
            "when": timezone.localtime(obj.created_at, dashboard_tz),
            "time": timezone.localtime(obj.created_at, dashboard_tz).strftime("%I:%M %p"),
            "title": f"Program created - {obj.program_no}",
            "sub": f"{getattr(obj.bom, 'sku', '') or '-'} • Qty {obj.total_qty or 0}",
        })

    for model, label in [
        (YarnPOInward, "Yarn inward"),
        (GreigePOInward, "Greige inward"),
        (DyeingPOInward, "Dyeing inward"),
        (ReadyPOInward, "Ready inward"),
    ]:
        for obj in model.objects.filter(owner=owner).order_by("-created_at")[:3]:
            timeline_items.append({
                "when": timezone.localtime(obj.created_at, dashboard_tz),
                "time": timezone.localtime(obj.created_at, dashboard_tz).strftime("%I:%M %p"),
                "title": label,
                "sub": getattr(obj, "inward_no", "") or getattr(obj, "grn_no", "") or f"Entry #{obj.id}",
            })

    for obj in DispatchChallan.objects.filter(owner=owner).order_by("-created_at")[:4]:
        timeline_items.append({
            "when": timezone.localtime(obj.created_at, dashboard_tz),
            "time": timezone.localtime(obj.created_at, dashboard_tz).strftime("%I:%M %p"),
            "title": f"Dispatch challan - {obj.challan_no}",
            "sub": getattr(obj.program, "program_no", "-"),
        })

    for obj in ProgramInvoice.objects.filter(owner=owner).order_by("-created_at")[:4]:
        timeline_items.append({
            "when": timezone.localtime(obj.created_at, dashboard_tz),
            "time": timezone.localtime(obj.created_at, dashboard_tz).strftime("%I:%M %p"),
            "title": f"Invoice - {obj.invoice_no}",
            "sub": getattr(obj.program, "program_no", "-"),
        })

    timeline_items = sorted(timeline_items, key=lambda x: x["when"], reverse=True)[:6]

    return render(
        request,
        "accounts/dashboard.html",
        {
            "greeting": greeting,
            "current_month_label": now_local.strftime("%B %Y"),
            "current_date_label": now_local.strftime("%d %b %Y"),
            "current_time_label": now_local.strftime("%I:%M %p"),
            "calendar_weeks": calendar_weeks,

            "total_inward_count": total_inward_count,
            "today_inward_count": today_inward_count,
            "fabric_inward_count": fabric_inward_count,
            "yarn_inward_count": yarn_inward_count,
            "greige_inward_count": greige_inward_count,
            "dyeing_inward_count": dyeing_inward_count,
            "ready_inward_count": ready_inward_count,

            "total_programs": total_programs,
            "open_programs": open_programs,
            "closed_programs": closed_programs,
            "started_programs": started_programs,
            "programs_today": programs_today,
            "recent_program_rows": recent_program_rows,

            "total_dispatch": total_dispatch,
            "dispatch_today": dispatch_today,
            "total_invoices": total_invoices,
            "invoices_today": invoices_today,

            "total_lots": total_lots,
            "pending_lots": pending_lots,
            "approved_lots": approved_lots,
            "hold_lots": hold_lots,
            "rejected_lots": rejected_lots,
            "pending_lots_pct": pending_lots_pct,
            "approved_lots_pct": approved_lots_pct,
            "hold_lots_pct": hold_lots_pct,
            "lot_watch_rows": lot_watch_rows,

            "trend_points": trend_points,
            "timeline_items": timeline_items,
        },
    )


@login_required
def utilities_view(request):
    return render(
        request,
        "accounts/utilities.html",
        {
            "utility_groups": get_utility_groups(request),
        },
    )


@login_required
def developer_stats_view(request):
    now = timezone.now()

    total_users = User.objects.count()

    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    new_today = User.objects.filter(date_joined__gte=today_start).count()

    week_start = now - timedelta(days=7)
    new_7_days = User.objects.filter(date_joined__gte=week_start).count()

    active_24h = User.objects.filter(last_login__gte=now - timedelta(hours=24)).count()
    active_7d = User.objects.filter(last_login__gte=now - timedelta(days=7)).count()

    return render(
        request,
        "accounts/developer_stats.html",
        {
            "total_users": total_users,
            "new_today": new_today,
            "new_7_days": new_7_days,
            "active_24h": active_24h,
            "active_7d": active_7d,
        },
    )
@login_required
@require_POST
def profile_save(request):
    form = DashboardProfileForm(request.POST)
    if not form.is_valid():
        first_error = _first_form_error(form)
        return JsonResponse(
            {
                "ok": False,
                "message": f"{first_error['label']}: {first_error['message']}" if first_error["label"] else first_error["message"],
                "field": first_error["field"],
                "errors": form.errors,
            },
            status=400,
        )

    # Save the real logged-in account, not the legacy data-owner account.
    # ERPTenantMiddleware keeps the real login user in request.erp_actor and may
    # temporarily point request.user to the company admin so old owner=request.user
    # queries can still share company data.
    u = get_actor(request)
    if not u or not getattr(u, "is_authenticated", False):
        return JsonResponse({"ok": False, "message": "Login session expired. Please login again."}, status=403)

    u.first_name = form.cleaned_data["first_name"]
    u.last_name = form.cleaned_data["last_name"]
    u.email = form.cleaned_data["email"]
    u.save(update_fields=["first_name", "last_name", "email"])

    extra, _ = UserExtra.objects.get_or_create(user=u)
    extra.phone = form.cleaned_data["phone"]
    extra.address = form.cleaned_data["address"]
    extra.save(update_fields=["phone", "address"])

    return JsonResponse({"ok": True, "message": "Profile saved ✅"})


# ==========================
# JOBBERS (embed supported)
# ==========================
def _jobber_list_context(request):
    q = (request.GET.get("q") or "").strip()
    qs = Jobber.objects.filter(owner=request.user).select_related("jobber_type")

    if q:
        search_terms = [term for term in q.split() if term]

        for term in search_terms:
            phone_term = "".join(ch for ch in term if ch.isdigit())
            term_filter = (
                Q(name__icontains=term)
                | Q(phone__icontains=term)
                | Q(email__icontains=term)
                | Q(jobber_type__name__icontains=term)
                | Q(address__icontains=term)
            )

            if phone_term and phone_term != term:
                term_filter |= Q(phone__icontains=phone_term)

            qs = qs.filter(term_filter)

    qs = qs.order_by("name")

    all_jobbers = Jobber.objects.filter(owner=request.user)
    stats = {
        "total": all_jobbers.count(),
        "active": all_jobbers.filter(is_active=True).count(),
        "inactive": all_jobbers.filter(is_active=False).count(),
        "types": JobberType.objects.filter(owner=request.user).count(),
    }

    return {
        "jobbers": qs,
        "q": q,
        "stats": stats,
    }


def _jobbertype_list_context(request, form=None):
    if form is None:
        form = JobberTypeForm(user=request.user)

    types = (
        JobberType.objects
        .filter(owner=request.user)
        .annotate(jobber_count=Count("jobber"))
        .order_by("name")
    )

    return {
        "types": types,
        "form": form,
        "type_stats": {
            "total_types": JobberType.objects.filter(owner=request.user).count(),
            "linked_jobbers": Jobber.objects.filter(owner=request.user, jobber_type__isnull=False).count(),
        },
    }


def _jobber_usage_rows(jobber):
    rows = []

    usage_map = [
        ("BOM rows", jobber.bom_jobber_details.count()),
        ("Program rows", jobber.program_jobber_rows.count()),
        ("Program start rows", jobber.program_start_jobbers.count()),
        ("Challans", jobber.program_jobber_challans.count()),
    ]

    for label, count in usage_map:
        if count:
            rows.append({"label": label, "count": count})

    return rows


def _jobbertype_usage_rows(jobber_type):
    rows = []

    usage_map = [
        ("Linked jobbers", Jobber.objects.filter(jobber_type=jobber_type).count()),
        ("BOM process rows", jobber_type.bom_jobber_type_processes.count()),
        ("BOM detail rows", jobber_type.bom_jobber_detail_types.count()),
        ("Program rows", jobber_type.program_jobber_type_rows.count()),
        ("Program start rows", jobber_type.program_start_jobber_types.count()),
        ("Challans", jobber_type.program_jobber_challan_types.count()),
    ]

    for label, count in usage_map:
        if count:
            rows.append({"label": label, "count": count})

    return rows

@login_required
def jobber_list(request):
    template = "accounts/jobbers/embed_list.html" if _is_embed(request) else "accounts/jobbers/list.html"
    return render(request, template, _jobber_list_context(request))


@login_required
@require_http_methods(["GET", "POST"])
def jobber_create(request):
    form = JobberForm(request.POST or None, user=request.user)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.owner = request.user
        obj.save()

        url = reverse("accounts:jobber_list")
        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})
        return redirect(url)

    template = "accounts/jobbers/embed_form.html" if _is_embed(request) else "accounts/jobbers/form.html"
    return render(request, template, {"form": form, "mode": "add"})


@login_required
@require_http_methods(["GET", "POST"])
def jobber_update(request, pk):
    jobber = get_object_or_404(Jobber, pk=pk, owner=request.user)
    form = JobberForm(request.POST or None, instance=jobber)
    if "jobber_type" in form.fields:
        form.fields["jobber_type"].queryset = JobberType.objects.filter(owner=request.user)

    if request.method == "POST" and form.is_valid():
        form.save()

        url = reverse("accounts:jobber_list")
        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})
        return redirect(url)

    template = "accounts/jobbers/embed_form.html" if _is_embed(request) else "accounts/jobbers/form.html"
    return render(request, template, {"form": form, "mode": "edit", "jobber": jobber})


@login_required
@require_http_methods(["GET", "POST"])
def jobber_delete(request, pk):
    jobber = get_object_or_404(Jobber, pk=pk, owner=request.user)

    if request.method == "POST":
        usage_rows = _jobber_usage_rows(jobber)

        if usage_rows:
            error_message = (
                f'Cannot delete "{jobber.name}" because it is already used in linked records. '
                f'Mark it inactive instead.'
            )

            if _is_embed(request):
                context = _jobber_list_context(request)
                context["delete_error"] = error_message
                context["delete_label"] = jobber.name
                context["delete_usage"] = usage_rows
                return render(request, "accounts/jobbers/embed_list.html", context)

            messages.error(request, error_message)
            return redirect("accounts:jobber_list")

        jobber.delete()
        url = reverse("accounts:jobber_list")
        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})
        return redirect(url)

    template = "accounts/jobbers/embed_confirm_delete.html" if _is_embed(request) else "accounts/jobbers/confirm_delete.html"
    return render(request, template, {"jobber": jobber})


@login_required
@require_http_methods(["GET", "POST"])
def jobbertype_list_create(request):
    if request.method == "POST":
        form = JobberTypeForm(request.POST, user=request.user)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.owner = request.user
            obj.save()

            url = reverse("accounts:jobbertype_list")
            if _is_embed(request):
                return JsonResponse({"ok": True, "url": url})
            return redirect(url)
    else:
        form = JobberTypeForm(user=request.user)

    context = _jobbertype_list_context(request, form=form)

    template = _pick_template(
        "accounts/jobbers/embed_types.html" if _is_embed(request) else "accounts/jobbers/types.html",
        "accounts/jobbers/embed_types.html",
    )
    return render(request, template, context)


# ==========================
# MATERIALS (embed supported)
# ==========================

def _material_qs_for_user(user):
    return (
        Material.objects
        .filter(Q(owner=user) | Q(owner__isnull=True))
        .select_related(
            "material_type",
            "material_sub_type",
            "material_shade",
            "unit",
            "yarn",
            "greige",
            "finished",
            "trim",
        )
        .order_by("name")
    )


def _material_usage_rows(material):
    rows = []

    usage_map = [
        ("Yarn PO items", material.yarn_po_items.count()),
        ("Greige PO items", material.greige_po_items.count()),
        ("Dyeing finished items", material.dyeing_po_finished_items.count()),
        ("BOM rows", material.bom_material_items.count()),
        ("Program start rows", material.program_start_fabrics.count()),
        ("Inventory lots", material.inventory_lots.count()),
        ("Greige dyeing links", material.dyeing_material_links.count()),
        ("Finished dyeing links", material.dyeing_link_details.count()),
    ]

    for label, count in usage_map:
        if count:
            rows.append({"label": label, "count": count})

    return rows


def _material_list_context(request):
    q = (request.GET.get("q") or "").strip()
    selected_type = (request.GET.get("type") or "").strip()
    selected_kind = (request.GET.get("kind") or "").strip()
    status = (request.GET.get("status") or "").strip()

    qs = _material_qs_for_user(request.user)

    if selected_kind:
        qs = qs.filter(material_kind=selected_kind)

    if selected_type.isdigit():
        qs = qs.filter(material_type_id=int(selected_type))

    if status == "active":
        qs = qs.filter(is_active=True)
    elif status == "inactive":
        qs = qs.filter(is_active=False)

    if q:
        qs = qs.filter(
            Q(material_code__icontains=q)
            | Q(name__icontains=q)
            | Q(composition__icontains=q)
            | Q(hsn_code__icontains=q)
            | Q(remarks__icontains=q)
            | Q(material_type__name__icontains=q)
            | Q(material_sub_type__name__icontains=q)
            | Q(material_shade__name__icontains=q)
            | Q(unit__name__icontains=q)
        )

    type_choices = MaterialType.objects.filter(owner=request.user).order_by("name")
    if selected_kind:
        type_choices = type_choices.filter(material_kind=selected_kind)

    all_materials = _material_qs_for_user(request.user)

    page_data = _paginate_utility_queryset(request, qs)

    ctx = {
        "materials": page_data.pop("object_list"),
        "q": q,
        "selected_type": selected_type,
        "selected_kind": selected_kind,
        "status": status,
        "type_choices": type_choices.only("id", "name", "material_kind"),
        "kind_choices": Material.MATERIAL_KIND_CHOICES,
        "stats": {
            "total": all_materials.count(),
            "active": all_materials.filter(is_active=True).count(),
            "inactive": all_materials.filter(is_active=False).count(),
            "kinds": all_materials.values("material_kind").distinct().count(),
        },
    }
    ctx.update(page_data)
    return ctx

@login_required
def material_kind_picker(request):
    ctx = {
        "kind_choices": Material.MATERIAL_KIND_CHOICES,
    }
    tpl = _pick_template(
        "accounts/materials/kind_picker_embed.html" if _is_embed(request) else "accounts/materials/kind_picker_page.html",
        "accounts/materials/kind_picker_embed.html",
    )
    return render(request, tpl, ctx)


@login_required
@require_http_methods(["GET", "POST"])
def material_create(request):
    allowed_kinds = {value for value, _ in Material.MATERIAL_KIND_CHOICES}
    selected_kind = (
        request.GET.get("kind")
        or request.POST.get("material_kind")
        or ""
    ).strip()

    if request.method == "GET" and selected_kind not in allowed_kinds:
        ctx = {"kind_choices": Material.MATERIAL_KIND_CHOICES}
        tpl = (
            "accounts/materials/kind_picker_embed.html"
            if _is_embed(request)
            else "accounts/materials/kind_picker_page.html"
        )
        return render(request, tpl, ctx)

    if request.method == "POST":
        form = MaterialForm(
            request.POST,
            request.FILES,
            user=request.user,
            initial_kind=selected_kind,
        )

        if form.is_valid():
            form.save()
            url = reverse("accounts:material_list")
            if _is_embed(request):
                return JsonResponse({"ok": True, "url": url})
            return redirect(url)
    else:
        form = MaterialForm(
            user=request.user,
            initial_kind=selected_kind,
            initial={"material_kind": selected_kind},
        )

    ctx = {
        "form": form,
        "mode": "create",
        "selected_kind": selected_kind,
    }
    tpl = "accounts/materials/form_embed.html" if _is_embed(request) else "accounts/materials/form_page.html"
    return render(request, tpl, ctx)


@login_required
@require_http_methods(["GET", "POST"])
def material_edit(request, pk: int):
    material = get_object_or_404(_material_qs_for_user(request.user), pk=pk)

    if request.method == "POST":
        form = MaterialForm(
            request.POST,
            request.FILES,
            instance=material,
            user=request.user,
            initial_kind=material.material_kind,
        )

        if form.is_valid():
            form.save()
            url = reverse("accounts:material_list")
            if _is_embed(request):
                return JsonResponse({"ok": True, "url": url})
            return redirect(url)
    else:
        form = MaterialForm(
            instance=material,
            user=request.user,
            initial_kind=material.material_kind,
        )

    ctx = {
        "form": form,
        "mode": "edit",
        "material": material,
        "selected_kind": material.material_kind,
    }
    tpl = "accounts/materials/form_embed.html" if _is_embed(request) else "accounts/materials/form_page.html"
    return render(request, tpl, ctx)


@login_required
def material_list(request):
    ctx = _material_list_context(request)
    tpl = "accounts/materials/list_embed.html" if _is_embed(request) else "accounts/materials/list_page.html"
    return render(request, tpl, ctx)

@login_required
@require_POST
def material_delete(request, pk: int):
    material = get_object_or_404(_material_qs_for_user(request.user), pk=pk)
    usage_rows = _material_usage_rows(material)

    if usage_rows:
        error_message = (
            f'Cannot delete "{material.name}" because it is already used in linked records. '
            f'Mark it inactive instead.'
        )

        if _is_embed(request):
            ctx = _material_list_context(request)
            ctx["delete_error"] = error_message
            ctx["delete_label"] = material.name
            ctx["delete_usage"] = usage_rows
            return render(request, "accounts/materials/list_embed.html", ctx)

        messages.error(request, error_message)
        return redirect("accounts:material_list")

    material.delete()

    url = reverse("accounts:material_list")
    if _is_embed(request):
        return JsonResponse({"ok": True, "url": url})
    return redirect(url)

def _party_qs_for_user(user):
    return Party.objects.filter(Q(owner=user) | Q(owner__isnull=True)).order_by("party_name")


def _party_usage_rows(party):
    rows = []

    for rel in party._meta.related_objects:
        accessor_name = rel.get_accessor_name()
        if not accessor_name:
            continue

        try:
            related_manager = getattr(party, accessor_name)
            count = related_manager.count()
        except Exception:
            continue

        if count:
            label = rel.related_model._meta.verbose_name_plural.title()
            rows.append({"label": label, "count": count})

    return rows


def _party_list_context(request):
    q = (request.GET.get("q") or "").strip()
    status = (request.GET.get("status") or "").strip()
    category = (request.GET.get("category") or "").strip()

    qs = _party_qs_for_user(request.user)

    if q:
        qs = qs.filter(
            Q(party_code__icontains=q)
            | Q(party_name__icontains=q)
            | Q(contact_person__icontains=q)
            | Q(phone_number__icontains=q)
            | Q(alt_phone__icontains=q)
            | Q(email__icontains=q)
            | Q(gst_number__icontains=q)
            | Q(city__icontains=q)
        )

    if status == "active":
        qs = qs.filter(is_active=True)
    elif status == "inactive":
        qs = qs.filter(is_active=False)

    if category:
        qs = qs.filter(party_category=category)

    all_parties = _party_qs_for_user(request.user)

    return {
        "parties": qs,
        "q": q,
        "status": status,
        "category": category,
        "category_choices": Party.PARTY_CATEGORY_CHOICES,
        "stats": {
            "total": all_parties.count(),
            "active": all_parties.filter(is_active=True).count(),
            "inactive": all_parties.filter(is_active=False).count(),
            "categories": all_parties.exclude(party_category="").values("party_category").distinct().count(),
        },
    }


def _party_list_url(request=None):
    url = reverse("accounts:party_list")
    if request is not None and _is_embed(request):
        return f"{url}?embed=1"
    return url

@login_required
def party_list(request):
    template = "accounts/parties/list_embed.html" if _is_embed(request) else "accounts/parties/list.html"
    return render(request, template, _party_list_context(request))


@login_required
@require_http_methods(["GET", "POST"])
def party_create(request):
    form = PartyForm(request.POST or None, user=request.user)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.owner = request.user
        obj.save()

        url = _party_list_url(request)
        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})
        return redirect(url)

    template = "accounts/parties/form_embed.html" if _is_embed(request) else "accounts/parties/form.html"
    return render(request, template, {"form": form, "mode": "add"})


@login_required
@require_http_methods(["GET", "POST"])
def party_update(request, pk):
    party = get_object_or_404(_party_qs_for_user(request.user), pk=pk)
    form = PartyForm(request.POST or None, instance=party, user=request.user)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.owner = request.user
        obj.save()

        url = _party_list_url(request)
        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})
        return redirect(url)

    template = "accounts/parties/form_embed.html" if _is_embed(request) else "accounts/parties/form.html"
    return render(request, template, {"form": form, "mode": "edit", "party": party})


@login_required
@require_POST
def party_delete(request, pk):
    party = get_object_or_404(_party_qs_for_user(request.user), pk=pk)
    usage_rows = _party_usage_rows(party)

    if usage_rows:
        error_message = (
            f'Cannot delete "{party.party_name}" because it is already used in linked records. '
            f'Mark it inactive instead.'
        )

        if _is_embed(request):
            context = _party_list_context(request)
            context["delete_error"] = error_message
            context["delete_label"] = party.party_name
            context["delete_usage"] = usage_rows
            return render(request, "accounts/parties/list_embed.html", context)

        messages.error(request, error_message)
        return redirect("accounts:party_list")

    party.delete()

    url = _party_list_url(request)
    if _is_embed(request):
        return JsonResponse({"ok": True, "url": url})
    return redirect(url)

# ==========================
# LOCATIONS (embed supported)
# ==========================
def _location_list_url(request):
    url = reverse("accounts:location_list")
    if _is_embed(request):
        url += "?embed=1"
    return url


def _location_usage_rows(location):
    rows = []

    for rel in location._meta.related_objects:
        accessor_name = rel.get_accessor_name()
        if not accessor_name:
            continue

        try:
            related_manager = getattr(location, accessor_name)
            count = related_manager.count()
        except Exception:
            continue

        if count:
            label = rel.related_model._meta.verbose_name_plural.title()
            rows.append({"label": label, "count": count})

    return rows


def _location_list_context(request):
    q = (request.GET.get("q") or "").strip()

    qs = Location.objects.filter(owner=request.user)

    if q:
        qs = qs.filter(
            Q(name__icontains=q)
            | Q(address_line_1__icontains=q)
            | Q(address_line_2__icontains=q)
            | Q(landmark__icontains=q)
            | Q(city__icontains=q)
            | Q(state__icontains=q)
            | Q(pincode__icontains=q)
        )

    all_locations = Location.objects.filter(owner=request.user)

    return {
        "locations": qs.order_by("name"),
        "q": q,
        "stats": {
            "total": all_locations.count(),
            "cities": all_locations.exclude(city="").values("city").distinct().count(),
            "states": all_locations.exclude(state="").values("state").distinct().count(),
        },
    }


@login_required
def location_list(request):
    ctx = _location_list_context(request)
    tpl = "accounts/locations/list_embed.html" if _is_embed(request) else "accounts/locations/list.html"
    return render(request, tpl, ctx)


@login_required
@require_http_methods(["GET", "POST"])
def location_create(request):
    form = LocationForm(request.POST or None, user=request.user)

    if request.method == "POST":
        form.instance.owner = request.user

        if form.is_valid():
            obj = form.save(commit=False)
            obj.owner = request.user
            obj.save()

            url = _location_list_url(request)
            if _is_embed(request):
                return JsonResponse({"ok": True, "url": url})
            return redirect(url)

    tpl = "accounts/locations/form_embed.html" if _is_embed(request) else "accounts/locations/form.html"
    return render(request, tpl, {"form": form, "mode": "add"})


@login_required
@require_http_methods(["GET", "POST"])
def location_update(request, pk: int):
    loc = get_object_or_404(Location, pk=pk, owner=request.user)
    form = LocationForm(request.POST or None, instance=loc, user=request.user)

    if request.method == "POST" and form.is_valid():
        form.save()

        url = _location_list_url(request)
        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})
        return redirect(url)

    ctx = {"form": form, "mode": "edit", "location": loc}
    tpl = "accounts/locations/form_embed.html" if _is_embed(request) else "accounts/locations/form.html"
    return render(request, tpl, ctx)


@login_required
@require_POST
def location_delete(request, pk: int):
    loc = get_object_or_404(Location, pk=pk, owner=request.user)
    usage_rows = _location_usage_rows(loc)

    if usage_rows:
        error_message = (
            f'Cannot delete "{loc.name}" because it is already used in linked records.'
        )

        if _is_embed(request):
            context = _location_list_context(request)
            context["delete_error"] = error_message
            context["delete_label"] = loc.name
            context["delete_usage"] = usage_rows
            return render(request, "accounts/locations/list_embed.html", context)

        messages.error(request, error_message)
        return redirect("accounts:location_list")

    loc.delete()

    url = _location_list_url(request)
    if _is_embed(request):
        return JsonResponse({"ok": True, "url": url})
    return redirect(url)


# ==========================
# FIRM (single per user)
# ==========================
@login_required
def firm_list(request):
    firms = Firm.objects.filter(owner=request.user).order_by("firm_name", "id")

    tpl = _pick_template(
        "accounts/firms/list_embed.html" if _is_embed(request) else "accounts/firms/list.html",
        "accounts/firms/list_embed.html",
    )
    return render(request, tpl, {"firms": firms})


@login_required
@require_http_methods(["GET", "POST"])
def firm_create(request):
    form = FirmForm(request.POST or None)
    address_formset = FirmAddressFormSet(request.POST or None, prefix="addresses")

    if request.method == "POST" and form.is_valid() and address_formset.is_valid():
        with transaction.atomic():
            obj = form.save(commit=False)
            obj.owner = request.user
            obj.save()
            address_formset.instance = obj
            address_formset.save()

        if _is_embed(request):
            return JsonResponse({"ok": True, "url": reverse("accounts:firm_list")})
        return redirect("accounts:firm_list")

    tpl = "accounts/firms/form_embed.html" if _is_embed(request) else "accounts/firms/form.html"
    return render(request, tpl, {"form": form, "address_formset": address_formset, "mode": "add"})


@login_required
@require_http_methods(["GET", "POST"])
def firm_update(request, pk: int):
    firm = get_object_or_404(Firm.objects.prefetch_related("addresses"), pk=pk, owner=request.user)
    form = FirmForm(request.POST or None, instance=firm)
    address_formset = FirmAddressFormSet(request.POST or None, instance=firm, prefix="addresses")

    if request.method == "POST" and form.is_valid() and address_formset.is_valid():
        with transaction.atomic():
            form.save()
            address_formset.save()

        if _is_embed(request):
            return JsonResponse({"ok": True, "url": reverse("accounts:firm_list")})
        return redirect("accounts:firm_list")

    tpl = "accounts/firms/form_embed.html" if _is_embed(request) else "accounts/firms/form.html"
    return render(request, tpl, {"form": form, "address_formset": address_formset, "mode": "edit", "firm": firm})


@login_required
@require_POST
def firm_delete(request, pk: int):
    firm = get_object_or_404(Firm, pk=pk, owner=request.user)
    firm.delete()

    if _is_embed(request):
        return JsonResponse({"ok": True, "url": reverse("accounts:firm_list")})
    return redirect("accounts:firm_list")


@login_required
@require_http_methods(["GET", "POST"])
def firm_view(request):
    firm = Firm.objects.filter(owner=request.user).prefetch_related("addresses").first()
    if firm is None:
        firm = Firm(owner=request.user)

    form = FirmForm(request.POST or None, instance=firm)
    address_formset = FirmAddressFormSet(request.POST or None, instance=firm if firm.pk else None, prefix="addresses")

    if request.method == "POST" and form.is_valid() and address_formset.is_valid():
        with transaction.atomic():
            saved_firm = form.save(commit=False)
            saved_firm.owner = request.user
            saved_firm.save()
            address_formset.instance = saved_firm
            address_formset.save()

        if _is_embed(request):
            return JsonResponse({"ok": True, "url": reverse("accounts:firm")})
        return redirect("accounts:firm")

    template = _pick_template(
        "accounts/firms/form_embed.html" if _is_embed(request) else "accounts/firms/form.html",
        "accounts/firms/form_embed.html",
        "accounts/firms/form.html",
    )
    return render(request, template, {"form": form, "address_formset": address_formset})


# ==========================
# MATERIAL SHADES (Utilities)
# ==========================
def _shade_usage_rows(shade):
    rows = []

    for rel in shade._meta.related_objects:
        accessor_name = rel.get_accessor_name()
        if not accessor_name:
            continue

        try:
            related_manager = getattr(shade, accessor_name)
            count = related_manager.count()
        except Exception:
            continue

        if count:
            label = rel.related_model._meta.verbose_name_plural.title()
            rows.append({"label": label, "count": count})

    return rows


def _materialshade_list_context(request):
    q = (request.GET.get("q") or "").strip()
    selected_kind = (request.GET.get("kind") or "").strip()

    qs = MaterialShade.objects.filter(owner=request.user).order_by("name")

    if selected_kind:
        qs = qs.filter(material_kind=selected_kind)

    if q:
        qs = qs.filter(
            Q(name__icontains=q)
            | Q(code__icontains=q)
            | Q(notes__icontains=q)
            | Q(material_kind__icontains=q)
        )

    all_shades = MaterialShade.objects.filter(owner=request.user)
    filtered_base = all_shades
    if selected_kind:
        filtered_base = filtered_base.filter(material_kind=selected_kind)

    page_data = _paginate_utility_queryset(request, qs)

    ctx = {
        "shades": page_data.pop("object_list"),
        "q": q,
        "selected_kind": selected_kind,
        "kind_choices": Material.MATERIAL_KIND_CHOICES,
        "stats": {
            "total": filtered_base.count(),
            "with_code": filtered_base.exclude(code="").exclude(code__isnull=True).count(),
            "with_notes": filtered_base.exclude(notes="").exclude(notes__isnull=True).count(),
        },
    }
    ctx.update(page_data)
    return ctx


@login_required
def materialshade_list(request):
    ctx = _materialshade_list_context(request)
    tpl = "accounts/material_shades/list_embed.html" if _is_embed(request) else "accounts/material_shades/list.html"
    return render(request, tpl, ctx)


def _shade_list_url(request):
    url = reverse("accounts:materialshade_list")
    if _is_embed(request):
        url += "?embed=1"
    return url


@login_required
@require_http_methods(["GET", "POST"])
def materialshade_create(request):
    form = MaterialShadeForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.owner = request.user
        obj.save()

        url = _shade_list_url(request)

        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})

        return redirect(url)

    tpl = "accounts/material_shades/form_embed.html" if _is_embed(request) else "accounts/material_shades/form.html"
    return render(request, tpl, {"form": form, "mode": "add"})


@login_required
@require_http_methods(["GET", "POST"])
def materialshade_update(request, pk: int):
    shade = get_object_or_404(MaterialShade, pk=pk, owner=request.user)
    form = MaterialShadeForm(request.POST or None, instance=shade)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.owner = request.user
        obj.save()

        url = _shade_list_url(request)

        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})

        return redirect(url)

    tpl = "accounts/material_shades/form_embed.html" if _is_embed(request) else "accounts/material_shades/form.html"
    return render(request, tpl, {"form": form, "mode": "edit", "shade": shade})


@login_required
@require_POST
def materialshade_delete(request, pk: int):
    shade = get_object_or_404(MaterialShade, pk=pk, owner=request.user)
    usage_rows = _shade_usage_rows(shade)

    if usage_rows:
        error_message = (
            f'Cannot delete "{shade.name}" because it is already used in linked records.'
        )

        if _is_embed(request):
            ctx = _materialshade_list_context(request)
            ctx["delete_error"] = error_message
            ctx["delete_label"] = shade.name
            ctx["delete_usage"] = usage_rows
            return render(request, "accounts/material_shades/list_embed.html", ctx)

        messages.error(request, error_message)
        return redirect("accounts:materialshade_list")

    shade.delete()

    url = _shade_list_url(request)

    if _is_embed(request):
        return JsonResponse({"ok": True, "url": url})

    return redirect(url)


# ==========================
# MATERIAL TYPES (Utilities)
# ==========================
def _materialtype_list_url(request):
    url = reverse("accounts:materialtype_list")
    if _is_embed(request):
        url += "?embed=1"
    return url


def _materialtype_usage_rows(material_type):
    rows = []

    material_count = 0
    try:
        material_count = Material.objects.filter(
            owner=material_type.owner,
            material_type=material_type,
        ).count()
    except Exception:
        material_count = 0

    if material_count:
        rows.append({"label": "Materials", "count": material_count})

    return rows


def _materialtype_list_context(request):
    q = (request.GET.get("q") or "").strip()
    selected_kind = (request.GET.get("kind") or "").strip()

    qs = MaterialType.objects.filter(owner=request.user).order_by("name")

    if selected_kind:
        qs = qs.filter(material_kind=selected_kind)

    if q:
        qs = qs.filter(
            Q(name__icontains=q)
            | Q(description__icontains=q)
        )

    all_types = MaterialType.objects.filter(owner=request.user)

    filtered_base = all_types
    if selected_kind:
        filtered_base = filtered_base.filter(material_kind=selected_kind)

    page_data = _paginate_utility_queryset(request, qs)

    ctx = {
        "types": page_data.pop("object_list"),
        "q": q,
        "selected_kind": selected_kind,
        "kind_choices": Material.MATERIAL_KIND_CHOICES,
        "stats": {
            "total": filtered_base.count(),
            "with_description": filtered_base.exclude(description="").exclude(description__isnull=True).count(),
        },
    }
    ctx.update(page_data)
    return ctx


@login_required
def materialtype_list(request):
    ctx = _materialtype_list_context(request)
    tpl = "accounts/material_types/list_embed.html" if _is_embed(request) else "accounts/material_types/list.html"
    return render(request, tpl, ctx)


@login_required
@require_http_methods(["GET", "POST"])
def materialtype_create(request):
    form = MaterialTypeForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.owner = request.user
        obj.save()

        url = _materialtype_list_url(request)
        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})
        return redirect(url)

    tpl = "accounts/material_types/form_embed.html" if _is_embed(request) else "accounts/material_types/form.html"
    return render(request, tpl, {"form": form, "mode": "add"})


@login_required
@require_http_methods(["GET", "POST"])
def materialtype_update(request, pk: int):
    mt = get_object_or_404(MaterialType, pk=pk, owner=request.user)
    form = MaterialTypeForm(request.POST or None, instance=mt)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.owner = request.user
        obj.save()

        url = _materialtype_list_url(request)
        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})
        return redirect(url)

    tpl = "accounts/material_types/form_embed.html" if _is_embed(request) else "accounts/material_types/form.html"
    return render(request, tpl, {"form": form, "mode": "edit", "material_type": mt})


@login_required
@require_POST
def materialtype_delete(request, pk: int):
    mt = get_object_or_404(MaterialType, pk=pk, owner=request.user)
    usage_rows = _materialtype_usage_rows(mt)

    if usage_rows:
        error_message = (
            f'Cannot delete "{mt.name}" because it is already used in linked records.'
        )

        if _is_embed(request):
            ctx = _materialtype_list_context(request)
            ctx["delete_error"] = error_message
            ctx["delete_label"] = mt.name
            ctx["delete_usage"] = usage_rows
            return render(request, "accounts/material_types/list_embed.html", ctx)

        messages.error(request, error_message)
        return redirect("accounts:materialtype_list")

    mt.delete()

    url = _materialtype_list_url(request)
    if _is_embed(request):
        return JsonResponse({"ok": True, "url": url})
    return redirect(url)

def _materialsubtype_list_url(request):
    url = reverse("accounts:materialsubtype_list")
    if _is_embed(request):
        url += "?embed=1"
    return url


def _materialsubtype_usage_rows(sub_type):
    rows = []

    for rel in sub_type._meta.related_objects:
        accessor_name = rel.get_accessor_name()
        if not accessor_name:
            continue

        try:
            related_manager = getattr(sub_type, accessor_name)
            count = related_manager.count()
        except Exception:
            continue

        if count:
            label = rel.related_model._meta.verbose_name_plural.title()
            rows.append({"label": label, "count": count})

    return rows


@login_required
def materialsubtype_list(request):
    q = (request.GET.get("q") or "").strip()
    selected_kind = (request.GET.get("kind") or "").strip()

    qs = (
        MaterialSubType.objects.filter(owner=request.user)
        .select_related("material_type")
        .order_by("material_type__name", "name")
    )

    if selected_kind:
        qs = qs.filter(material_type__material_kind=selected_kind)

    if q:
        qs = qs.filter(
            Q(name__icontains=q)
            | Q(description__icontains=q)
            | Q(material_type__name__icontains=q)
        )

    all_sub_types = MaterialSubType.objects.filter(owner=request.user)
    filtered_base = all_sub_types
    if selected_kind:
        filtered_base = filtered_base.filter(material_type__material_kind=selected_kind)

    page_data = _paginate_utility_queryset(request, qs)

    ctx = {
        "sub_types": page_data.pop("object_list"),
        "q": q,
        "selected_kind": selected_kind,
        "kind_choices": Material.MATERIAL_KIND_CHOICES,
        "stats": {
            "total": filtered_base.count(),
            "with_description": filtered_base.exclude(description="").exclude(description__isnull=True).count(),
            "types": filtered_base.values("material_type").distinct().count(),
        },
    }
    ctx.update(page_data)

    tpl = "accounts/material_sub_types/list_embed.html" if _is_embed(request) else "accounts/material_sub_types/list.html"
    return render(request, tpl, ctx)


@login_required
@require_http_methods(["GET", "POST"])
def materialsubtype_create(request):
    form = MaterialSubTypeForm(request.POST or None, user=request.user)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.owner = request.user
        obj.save()

        url = _materialsubtype_list_url(request)
        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})
        return redirect(url)

    tpl = "accounts/material_sub_types/form_embed.html" if _is_embed(request) else "accounts/material_sub_types/form.html"
    return render(request, tpl, {"form": form, "mode": "add"})


@login_required
@require_http_methods(["GET", "POST"])
def materialsubtype_update(request, pk: int):
    sub_type = get_object_or_404(MaterialSubType, pk=pk, owner=request.user)
    form = MaterialSubTypeForm(request.POST or None, instance=sub_type, user=request.user)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.owner = request.user
        obj.save()

        url = _materialsubtype_list_url(request)
        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})
        return redirect(url)

    tpl = "accounts/material_sub_types/form_embed.html" if _is_embed(request) else "accounts/material_sub_types/form.html"
    return render(request, tpl, {"form": form, "mode": "edit", "material_sub_type": sub_type})


@login_required
@require_POST
def materialsubtype_delete(request, pk: int):
    sub_type = get_object_or_404(MaterialSubType, pk=pk, owner=request.user)
    usage_rows = _materialsubtype_usage_rows(sub_type)

    if usage_rows:
        error_message = (
            f'Cannot delete "{sub_type.name}" because it is already used in linked records.'
        )

        if _is_embed(request):
            q = (request.GET.get("q") or "").strip()
            selected_kind = (request.GET.get("kind") or "").strip()

            qs = (
                MaterialSubType.objects.filter(owner=request.user)
                .select_related("material_type")
                .order_by("material_type__name", "name")
            )

            if selected_kind:
                qs = qs.filter(material_type__material_kind=selected_kind)

            if q:
                qs = qs.filter(
                    Q(name__icontains=q)
                    | Q(description__icontains=q)
                    | Q(material_type__name__icontains=q)
                )

            all_sub_types = MaterialSubType.objects.filter(owner=request.user)
            filtered_base = all_sub_types
            if selected_kind:
                filtered_base = filtered_base.filter(material_type__material_kind=selected_kind)

            ctx = {
                "sub_types": qs,
                "q": q,
                "selected_kind": selected_kind,
                "kind_choices": Material.MATERIAL_KIND_CHOICES,
                "stats": {
                    "total": filtered_base.count(),
                    "with_description": filtered_base.exclude(description="").exclude(description__isnull=True).count(),
                    "types": filtered_base.values("material_type").distinct().count(),
                },
                "delete_error": error_message,
                "delete_label": sub_type.name,
                "delete_usage": usage_rows,
            }
            return render(request, "accounts/material_sub_types/list_embed.html", ctx)

        messages.error(request, error_message)
        return redirect("accounts:materialsubtype_list")

    sub_type.delete()

    url = _materialsubtype_list_url(request)
    if _is_embed(request):
        return JsonResponse({"ok": True, "url": url})
    return redirect(url)

# ==========================
# VENDORS (embed supported)
# ==========================

def _vendor_list_context(request):
    q = (request.GET.get("q") or "").strip()
    status = (request.GET.get("status") or "").strip()
    vendor_type = (request.GET.get("vendor_type") or "").strip()

    qs = Vendor.objects.filter(owner=request.user)

    if q:
        search_terms = [term for term in q.split() if term]
        for term in search_terms:
            digit_term = "".join(ch for ch in term if ch.isdigit())
            term_filter = (
                Q(vendor_code__icontains=term)
                | Q(name__icontains=term)
                | Q(contact_person__icontains=term)
                | Q(phone__icontains=term)
                | Q(alt_phone__icontains=term)
                | Q(email__icontains=term)
                | Q(gst_number__icontains=term)
                | Q(city__icontains=term)
                | Q(vendor_type__icontains=term)
            )
            if digit_term and digit_term != term:
                term_filter |= Q(phone__icontains=digit_term) | Q(alt_phone__icontains=digit_term)
            qs = qs.filter(term_filter)

    if status == "active":
        qs = qs.filter(is_active=True)
    elif status == "inactive":
        qs = qs.filter(is_active=False)

    if vendor_type:
        qs = qs.filter(vendor_type=vendor_type)

    all_vendors = Vendor.objects.filter(owner=request.user)

    return {
        "vendors": qs.order_by("name"),
        "q": q,
        "status": status,
        "vendor_type": vendor_type,
        "vendor_type_choices": Vendor.VENDOR_TYPE_CHOICES,
        "stats": {
            "total": all_vendors.count(),
            "active": all_vendors.filter(is_active=True).count(),
            "inactive": all_vendors.filter(is_active=False).count(),
            "types": all_vendors.exclude(vendor_type="").values("vendor_type").distinct().count(),
        },
    }


def _vendor_usage_rows(vendor):
    rows = []

    usage_map = [
        ("Yarn POs", vendor.yarn_purchase_orders.count()),
        ("Yarn Inwards", vendor.yarn_inwards.count()),
        ("Greige POs", vendor.greige_purchase_orders.count()),
        ("Greige Inwards", vendor.greige_inwards.count()),
        ("Dyeing POs", vendor.dyeing_purchase_orders.count()),
        ("Dyeing Inwards", vendor.dyeing_inwards.count()),
        ("Ready POs", vendor.ready_purchase_orders.count()),
        ("Dyeing Material Links", vendor.dyeing_material_links.count()),
    ]

    for label, count in usage_map:
        if count:
            rows.append({"label": label, "count": count})

    return rows

@login_required
def vendor_list(request):
    tpl = "accounts/vendors/list_embed.html" if _is_embed(request) else "accounts/vendors/list.html"
    return render(request, tpl, _vendor_list_context(request))


@login_required
@require_http_methods(["GET", "POST"])
def vendor_create(request):
    form = VendorForm(request.POST or None, user=request.user)

    if request.method == "POST":
        form.instance.owner = request.user

        if form.is_valid():
            obj = form.save(commit=False)
            obj.owner = request.user
            obj.save()

            url = reverse("accounts:vendor_list")
            if _is_embed(request):
                return JsonResponse({"ok": True, "url": url})
            return redirect(url)

    tpl = "accounts/vendors/form_embed.html" if _is_embed(request) else "accounts/vendors/form.html"
    return render(request, tpl, {"form": form, "mode": "add"})


@login_required
@require_http_methods(["GET", "POST"])
def vendor_update(request, pk: int):
    vendor = get_object_or_404(Vendor, pk=pk, owner=request.user)
    form = VendorForm(request.POST or None, instance=vendor, user=request.user)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.owner = request.user
        obj.save()

        url = reverse("accounts:vendor_list")
        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})
        return redirect(url)

    tpl = "accounts/vendors/form_embed.html" if _is_embed(request) else "accounts/vendors/form.html"
    return render(request, tpl, {"form": form, "mode": "edit", "vendor": vendor})


@login_required
@require_POST
def vendor_delete(request, pk: int):
    vendor = get_object_or_404(Vendor, pk=pk, owner=request.user)
    usage_rows = _vendor_usage_rows(vendor)

    if usage_rows:
        error_message = (
            f'Cannot delete "{vendor.name}" because it is already used in linked records. '
            f'Mark it inactive instead.'
        )

        if _is_embed(request):
            context = _vendor_list_context(request)
            context["delete_error"] = error_message
            context["delete_label"] = vendor.name
            context["delete_usage"] = usage_rows
            return render(request, "accounts/vendors/list_embed.html", context)

        messages.error(request, error_message)
        return redirect("accounts:vendor_list")

    vendor.delete()

    url = reverse("accounts:vendor_list")
    if _is_embed(request):
        return JsonResponse({"ok": True, "url": url})
    return redirect(url)


# ==========================
# YARN PURCHASE ORDERS
# ==========================
def _can_access_yarn_po(user, po):
    return bool(_can_review_yarn_po(user) or po.owner_id == user.id)


def _next_yarn_inward_number() -> str:
    last = YarnPOInward.objects.order_by("-id").first()
    next_id = (last.id + 1) if last else 1
    return f"YIN-{next_id:04d}"
def _attach_yarn_po_metrics(po):
    return po


def _build_yarn_po_pdf_response(po):
    try:
        import os
        from pathlib import Path
        from html import escape

        from django.conf import settings
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.lib.utils import ImageReader
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError:
        return HttpResponse(
            "ReportLab is required for PDF generation. Install it with: pip install reportlab",
            status=500,
        )

    brand_pink = colors.HexColor("#ED2F8C")
    brand_orange = colors.HexColor("#F6A33B")
    brand_blue = colors.HexColor("#1976F3")
    brand_navy = colors.HexColor("#0F172A")
    ink = colors.HexColor("#1F2937")
    muted = colors.HexColor("#667085")
    border = colors.HexColor("#D0D5DD")
    soft_bg = colors.HexColor("#F8FAFC")
    white = colors.white

    def text_or_dash(value):
        value = "" if value is None else str(value).strip()
        return value if value else "-"

    def fmt_money(value):
        try:
            return f"{float(value or 0):,.2f}"
        except Exception:
            return f"{value or '0.00'}"

    def fmt_qty(value):
        try:
            return f"{float(value or 0):,.2f}".rstrip("0").rstrip(".")
        except Exception:
            return text_or_dash(value)

    def line_if(label, value):
        value = "" if value is None else str(value).strip()
        if not value:
            return ""
        return f"<b>{escape(label)}:</b> {escape(value)}"

    def join_parts(*parts):
        clean = [str(p).strip() for p in parts if str(p).strip()]
        return ", ".join(clean)

    def resolve_logo_path():
        if po.firm and getattr(po.firm, "logo", None):
            try:
                logo_path = po.firm.logo.path
                if logo_path and os.path.exists(logo_path):
                    return logo_path
            except Exception:
                pass

        fallback = Path(settings.BASE_DIR) / "Logo.jpeg"
        if fallback.exists():
            return str(fallback)

        return None

    logo_path = resolve_logo_path()

    def draw_branding(canvas, doc):
        page_w, page_h = A4
        canvas.saveState()

        canvas.setStrokeColor(colors.HexColor("#E4E7EC"))
        canvas.setLineWidth(0.8)
        canvas.roundRect(8 * mm, 8 * mm, page_w - 16 * mm, page_h - 16 * mm, 4 * mm, stroke=1, fill=0)

        usable_w = page_w - 16 * mm
        stripe_w = usable_w / 3.0
        stripe_y = page_h - 13 * mm
        stripe_h = 4.5 * mm

        canvas.setFillColor(brand_pink)
        canvas.rect(8 * mm, stripe_y, stripe_w, stripe_h, fill=1, stroke=0)

        canvas.setFillColor(brand_orange)
        canvas.rect(8 * mm + stripe_w, stripe_y, stripe_w, stripe_h, fill=1, stroke=0)

        canvas.setFillColor(brand_blue)
        canvas.rect(8 * mm + 2 * stripe_w, stripe_y, stripe_w, stripe_h, fill=1, stroke=0)

        if logo_path:
            try:
                img = ImageReader(logo_path)
                iw, ih = img.getSize()
                draw_w = 26 * mm
                draw_h = draw_w * (ih / float(iw)) if iw and ih else 26 * mm
                x = (page_w - draw_w) / 2.0
                y = 10 * mm

                try:
                    canvas.setFillAlpha(0.10)
                    canvas.setStrokeAlpha(0.10)
                except Exception:
                    pass

                canvas.drawImage(
                    img,
                    x,
                    y,
                    width=draw_w,
                    height=draw_h,
                    preserveAspectRatio=True,
                    mask="auto",
                )
            except Exception:
                pass

        canvas.restoreState()

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
    )

    styles = getSampleStyleSheet()

    base_style = ParagraphStyle(
        "YPOBase",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8.5,
        leading=10.5,
        textColor=ink,
        spaceAfter=0,
    )

    header_left_style = ParagraphStyle(
        "YPOHeaderLeft",
        parent=base_style,
        fontName="Helvetica",
        fontSize=8.4,
        leading=10.4,
        textColor=white,
        alignment=TA_LEFT,
    )

    header_title_style = ParagraphStyle(
        "YPOHeaderTitle",
        parent=base_style,
        fontName="Helvetica-Bold",
        fontSize=14,
        leading=16,
        textColor=brand_navy,
        alignment=TA_RIGHT,
    )

    header_meta_style = ParagraphStyle(
        "YPOHeaderMeta",
        parent=base_style,
        fontName="Helvetica",
        fontSize=8.3,
        leading=10.2,
        textColor=ink,
        alignment=TA_RIGHT,
    )

    section_head_left = ParagraphStyle(
        "YPOSectionHeadLeft",
        parent=base_style,
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=10,
        textColor=white,
        alignment=TA_LEFT,
    )

    section_value_style = ParagraphStyle(
        "YPOSectionValue",
        parent=base_style,
        fontName="Helvetica",
        fontSize=8.2,
        leading=10.2,
        textColor=ink,
        alignment=TA_LEFT,
    )

    table_head_style = ParagraphStyle(
        "YPOTableHead",
        parent=base_style,
        fontName="Helvetica-Bold",
        fontSize=7.8,
        leading=9.5,
        textColor=white,
        alignment=TA_CENTER,
    )

    item_center_style = ParagraphStyle(
        "YPOItemCenter",
        parent=base_style,
        fontName="Helvetica",
        fontSize=8,
        leading=9.6,
        alignment=TA_CENTER,
    )

    item_desc_style = ParagraphStyle(
        "YPOItemDesc",
        parent=base_style,
        fontName="Helvetica",
        fontSize=8,
        leading=9.6,
        alignment=TA_LEFT,
    )

    money_style = ParagraphStyle(
        "YPOMoney",
        parent=base_style,
        fontName="Helvetica",
        fontSize=8,
        leading=9.6,
        alignment=TA_RIGHT,
    )

    block_title_style = ParagraphStyle(
        "YPOBlockTitle",
        parent=base_style,
        fontName="Helvetica-Bold",
        fontSize=8.4,
        leading=10.5,
        textColor=brand_navy,
        alignment=TA_LEFT,
    )

    block_text_style = ParagraphStyle(
        "YPOBlockText",
        parent=base_style,
        fontName="Helvetica",
        fontSize=8.1,
        leading=10.1,
        textColor=ink,
        alignment=TA_LEFT,
    )

    sign_style = ParagraphStyle(
        "YPOSign",
        parent=base_style,
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=10,
        textColor=ink,
        alignment=TA_LEFT,
    )

    total_label_style = ParagraphStyle(
        "YPOTotalLabel",
        parent=base_style,
        fontName="Helvetica-Bold",
        fontSize=8.2,
        leading=10.2,
        textColor=ink,
        alignment=TA_LEFT,
    )

    total_value_style = ParagraphStyle(
        "YPOTotalValue",
        parent=base_style,
        fontName="Helvetica-Bold",
        fontSize=8.2,
        leading=10.2,
        textColor=ink,
        alignment=TA_RIGHT,
    )

    footer_note_style = ParagraphStyle(
        "YPOFooterNote",
        parent=base_style,
        fontName="Helvetica-Bold",
        fontSize=7.6,
        leading=9.2,
        textColor=muted,
        alignment=TA_CENTER,
    )

    story = []

    firm = po.firm
    vendor = po.vendor
    po_items = list(po.items.all())

    firm_name = text_or_dash(firm.firm_name if firm else "InventTech")
    firm_type = ""
    if firm:
        try:
            firm_type = firm.get_firm_type_display()
        except Exception:
            firm_type = text_or_dash(getattr(firm, "firm_type", ""))

    firm_address = join_parts(
        getattr(firm, "address_line", ""),
        getattr(firm, "city", ""),
        getattr(firm, "state", ""),
        getattr(firm, "pincode", ""),
    )

    firm_contact_line = " | ".join([
        part for part in [
            line_if("Phone", getattr(firm, "phone", "")),
            line_if("Email", getattr(firm, "email", "")),
            line_if("GSTIN", getattr(firm, "gst_number", "")),
        ] if part
    ])

    firm_stat_line = " | ".join([
        part for part in [
            line_if("PAN", getattr(firm, "pan_number", "")),
            line_if("TAN", getattr(firm, "tan_number", "")),
            line_if("CIN", getattr(firm, "cin_number", "")),
        ] if part
    ])

    order_number = po.po_number or po.system_number or "-"
    po_date = po.po_date.strftime("%d-%m-%Y") if po.po_date else "-"
    cancel_date = po.cancel_date.strftime("%d-%m-%Y") if po.cancel_date else "-"
    approval_label = getattr(po, "get_approval_status_display", lambda: text_or_dash(po.approval_status))()

    firm_header_html = f"<font size='13'><b>{escape(firm_name)}</b></font>"
    if firm_type and firm_type != "-":
        firm_header_html += f"<br/>{escape(firm_type)}"
    if firm_address:
        firm_header_html += f"<br/>{escape(firm_address)}"
    if firm_contact_line:
        firm_header_html += f"<br/>{firm_contact_line}"
    if firm_stat_line:
        firm_header_html += f"<br/>{firm_stat_line}"

    po_meta_html = (
        f"<b>YARN PURCHASE ORDER</b><br/>"
        f"<b>PO No:</b> {escape(order_number)}<br/>"
        f"<b>PO Date:</b> {escape(po_date)}<br/>"
        f"<b>System No:</b> {escape(text_or_dash(po.system_number))}<br/>"
        f"<b>Status:</b> {escape(text_or_dash(approval_label))}"
    )
    if cancel_date != "-":
        po_meta_html += f"<br/><b>Cancel Date:</b> {escape(cancel_date)}"

    header_table = Table(
        [[
            Paragraph(firm_header_html, header_left_style),
            Table(
                [[
                    Paragraph("<b>YARN PURCHASE ORDER</b>", header_title_style),
                ], [
                    Paragraph(
                        f"<b>PO No:</b> {escape(order_number)}<br/>"
                        f"<b>PO Date:</b> {escape(po_date)}<br/>"
                        f"<b>System No:</b> {escape(text_or_dash(po.system_number))}<br/>"
                        f"<b>Status:</b> {escape(text_or_dash(approval_label))}"
                        + (f"<br/><b>Cancel Date:</b> {escape(cancel_date)}" if cancel_date != "-" else ""),
                        header_meta_style,
                    )
                ]],
                colWidths=[66 * mm],
            ),
        ]],
        colWidths=[124 * mm, 66 * mm],
    )
    header_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), brand_navy),
        ("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#F5F8FF")),
        ("BOX", (0, 0), (-1, -1), 0.9, border),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (0, 0), 9),
        ("RIGHTPADDING", (0, 0), (0, 0), 9),
        ("TOPPADDING", (0, 0), (0, 0), 9),
        ("BOTTOMPADDING", (0, 0), (0, 0), 9),
        ("LEFTPADDING", (1, 0), (1, 0), 0),
        ("RIGHTPADDING", (1, 0), (1, 0), 0),
        ("TOPPADDING", (1, 0), (1, 0), 0),
        ("BOTTOMPADDING", (1, 0), (1, 0), 0),
    ]))
    header_table._argW[1] = 66 * mm
    inner_header = header_table._cellvalues[0][1]
    inner_header.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F5F8FF")),
        ("BOX", (0, 0), (-1, -1), 0.9, brand_blue),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 6))

    vendor_html = f"<b>{escape(text_or_dash(vendor.name if vendor else ''))}</b>"
    vendor_lines = [
        line_if("Contact", vendor.contact_person if vendor else ""),
        line_if("Phone", vendor.phone if vendor else ""),
        line_if("Email", vendor.email if vendor else ""),
        line_if("GSTIN", vendor.gst_number if vendor else ""),
        line_if("Address", vendor.address if vendor else ""),
    ]
    vendor_body = "<br/>".join([line for line in vendor_lines if line])
    if vendor_body:
        vendor_html += "<br/>" + vendor_body

    bill_to_html = f"<b>{escape(text_or_dash(firm_name if firm else ''))}</b>"
    bill_lines = [
        line_if("Address", firm_address),
        line_if("Phone", getattr(firm, "phone", "")),
        line_if("Email", getattr(firm, "email", "")),
        line_if("GSTIN", getattr(firm, "gst_number", "")),
        line_if("Ship To", po.shipping_address),
    ]
    bill_body = "<br/>".join([line for line in bill_lines if line])
    if bill_body:
        bill_to_html += "<br/>" + bill_body

    party_table = Table(
        [
            [
                Paragraph("VENDOR", section_head_left),
                Paragraph("BILL TO", section_head_left),
            ],
            [
                Paragraph(vendor_html, section_value_style),
                Paragraph(bill_to_html, section_value_style),
            ],
        ],
        colWidths=[95 * mm, 95 * mm],
    )
    party_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), brand_orange),
        ("BACKGROUND", (1, 0), (1, 0), brand_blue),
        ("TEXTCOLOR", (0, 0), (-1, 0), white),
        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#FBFCFE")),
        ("BOX", (0, 0), (-1, -1), 0.9, border),
        ("INNERGRID", (0, 0), (-1, -1), 0.7, border),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.append(party_table)
    story.append(Spacer(1, 7))

    item_rows = [[
        Paragraph("Sr No", table_head_style),
        Paragraph("Description", table_head_style),
        Paragraph("Unit", table_head_style),
        Paragraph("Qty", table_head_style),
        Paragraph("Price (Rs.)", table_head_style),
        Paragraph("Amount (Rs.)", table_head_style),
    ]]

    for index, item in enumerate(po_items, start=1):
        material_name = "-"
        if item.material:
            material_name = item.material.name
        elif item.material_type:
            material_name = item.material_type.name

        details = []
        if item.count:
            details.append(f"Count: {item.count}")
        if item.dia:
            details.append(f"Dia: {item.dia}")
        if item.gauge:
            details.append(f"Gauge: {item.gauge}")
        if item.gsm:
            details.append(f"GSM: {item.gsm}")
        if item.sl:
            details.append(f"SL: {item.sl}")
        if item.rolls:
            details.append(f"Rolls: {item.rolls}")
        if item.hsn_code:
            details.append(f"HSN: {item.hsn_code}")
        if item.remark:
            details.append(f"Remark: {item.remark}")

        description_html = f"<b>{escape(text_or_dash(material_name))}</b>"
        if details:
            description_html += "<br/>" + escape(" | ".join(details))

        item_rows.append([
            Paragraph(str(index), item_center_style),
            Paragraph(description_html, item_desc_style),
            Paragraph(escape(text_or_dash(item.unit)), item_center_style),
            Paragraph(escape(fmt_qty(item.quantity)), item_center_style),
            Paragraph(escape(fmt_money(item.rate)), money_style),
            Paragraph(escape(fmt_money(item.final_amount)), money_style),
        ])

    min_visual_rows = 5
    for _ in range(max(0, min_visual_rows - len(po_items))):
        item_rows.append([
            Paragraph("", item_center_style),
            Paragraph("", item_desc_style),
            Paragraph("", item_center_style),
            Paragraph("", item_center_style),
            Paragraph("", money_style),
            Paragraph("", money_style),
        ])

    items_table = Table(
        item_rows,
        colWidths=[13 * mm, 86 * mm, 18 * mm, 18 * mm, 26 * mm, 29 * mm],
        repeatRows=1,
    )
    items_table_style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), brand_navy),
        ("TEXTCOLOR", (0, 0), (-1, 0), white),
        ("BOX", (0, 0), (-1, -1), 0.9, border),
        ("INNERGRID", (0, 0), (-1, -1), 0.55, border),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ])

    for row_index in range(1, len(item_rows)):
        if row_index % 2 == 0:
            items_table_style.add("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#F9FAFB"))

    items_table.setStyle(items_table_style)
    story.append(items_table)
    story.append(Spacer(1, 8))

    notes_parts = []
    if po.remarks:
        notes_parts.append(f"<font color='#0F172A'><b>Remarks</b></font><br/>{escape(po.remarks).replace(chr(10), '<br/>')}")
    if po.terms_conditions:
        notes_parts.append(f"<font color='#0F172A'><b>Terms &amp; Conditions</b></font><br/>{escape(po.terms_conditions).replace(chr(10), '<br/>')}")
    if po.shipping_address:
        notes_parts.append(f"<font color='#0F172A'><b>Shipping Address</b></font><br/>{escape(po.shipping_address).replace(chr(10), '<br/>')}")

    if not notes_parts:
        notes_parts.append("<font color='#0F172A'><b>Notes</b></font><br/>Standard terms and conditions apply.")

    notes_html = "<br/><br/>".join(notes_parts)

    signature_table = Table(
        [[
            Paragraph("<b>AUTHORISED SIGNATORY</b><br/><br/>_________________________", sign_style),
            Paragraph(f"<b>DATE</b><br/><br/>{escape(po_date)}", sign_style),
        ]],
        colWidths=[68 * mm, 30 * mm],
    )
    signature_table.setStyle(TableStyle([
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))

    left_block = Table(
        [
            [Paragraph("NOTES / TERMS", block_title_style)],
            [Paragraph(notes_html, block_text_style)],
            [signature_table],
        ],
        colWidths=[102 * mm],
    )
    left_block.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#FDF2F8")),
        ("BOX", (0, 0), (-1, -1), 0.9, border),
        ("INNERGRID", (0, 0), (-1, -1), 0.6, border),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))

    discount_amount = Decimal(po.subtotal or 0) - Decimal(po.after_discount_value or 0)
    if discount_amount < 0:
        discount_amount = Decimal("0")

    taxable_amount = Decimal(po.after_discount_value or 0) + Decimal(po.others or 0)
    cgst_amount = taxable_amount * Decimal(po.cgst_percent or 0) / Decimal("100")
    sgst_amount = taxable_amount * Decimal(po.sgst_percent or 0) / Decimal("100")

    totals_rows = [
        [Paragraph("Sub Total", total_label_style), Paragraph(escape(fmt_money(po.subtotal)), total_value_style)],
        [Paragraph("Discount Amount", total_label_style), Paragraph(escape(fmt_money(discount_amount)), total_value_style)],
        [Paragraph("After Discount Amount", total_label_style), Paragraph(escape(fmt_money(po.after_discount_value)), total_value_style)],
        [Paragraph("Other Charges", total_label_style), Paragraph(escape(fmt_money(po.others)), total_value_style)],
        [Paragraph(f"CGST ({fmt_qty(po.cgst_percent)}%)", total_label_style), Paragraph(escape(fmt_money(cgst_amount)), total_value_style)],
        [Paragraph(f"SGST ({fmt_qty(po.sgst_percent)}%)", total_label_style), Paragraph(escape(fmt_money(sgst_amount)), total_value_style)],
        [Paragraph("Total Amount", total_label_style), Paragraph(escape(fmt_money(po.grand_total)), total_value_style)],
    ]

    totals_table = Table(totals_rows, colWidths=[49 * mm, 39 * mm])
    totals_table_style = TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.9, border),
        ("INNERGRID", (0, 0), (-1, -1), 0.6, border),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EFF6FF")),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#FFF7ED")),
    ])
    for row_index in range(1, len(totals_rows) - 1):
        if row_index % 2 == 1:
            totals_table_style.add("BACKGROUND", (0, row_index), (-1, row_index), soft_bg)
    totals_table.setStyle(totals_table_style)

    lower_table = Table([[left_block, totals_table]], colWidths=[102 * mm, 88 * mm])
    lower_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    story.append(lower_table)
    story.append(Spacer(1, 8))

    footer_table = Table(
        [[Paragraph("THIS PO IS COMPUTER GENERATED, HENCE SIGNATURE IS NOT REQUIRED", footer_note_style)]],
        colWidths=[190 * mm],
    )
    footer_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
        ("BOX", (0, 0), (-1, -1), 0.9, border),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(footer_table)

    doc.build(story, onFirstPage=draw_branding, onLaterPages=draw_branding)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{po.system_number or "yarn_po"}.pdf"'
    return response

def _build_greige_po_pdf_response(po):
    try:
        import os
        from pathlib import Path
        from html import escape

        from django.conf import settings
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.lib.utils import ImageReader
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError:
        return HttpResponse(
            "ReportLab is required for PDF generation. Install it with: pip install reportlab",
            status=500,
        )

    brand_pink = colors.HexColor("#ED2F8C")
    brand_orange = colors.HexColor("#F6A33B")
    brand_blue = colors.HexColor("#1976F3")
    brand_navy = colors.HexColor("#0F172A")
    ink = colors.HexColor("#1F2937")
    muted = colors.HexColor("#667085")
    border = colors.HexColor("#D0D5DD")
    soft_bg = colors.HexColor("#F8FAFC")
    white = colors.white

    def text_or_dash(value):
        value = "" if value is None else str(value).strip()
        return value if value else "-"

    def fmt_money(value):
        try:
            return f"{float(value or 0):,.2f}"
        except Exception:
            return f"{value or '0.00'}"

    def fmt_qty(value):
        try:
            return f"{float(value or 0):,.2f}".rstrip("0").rstrip(".")
        except Exception:
            return text_or_dash(value)

    def line_if(label, value):
        value = "" if value is None else str(value).strip()
        if not value:
            return ""
        return f"<b>{escape(label)}:</b> {escape(value)}"

    def join_parts(*parts):
        clean = [str(p).strip() for p in parts if str(p).strip()]
        return ", ".join(clean)

    firm = getattr(po.source_yarn_po, "firm", None)
    vendor = po.vendor
    po_items = list(po.items.all())

    def resolve_logo_path():
        if firm and getattr(firm, "logo", None):
            try:
                logo_path = firm.logo.path
                if logo_path and os.path.exists(logo_path):
                    return logo_path
            except Exception:
                pass

        fallback = Path(settings.BASE_DIR) / "Logo.jpeg"
        if fallback.exists():
            return str(fallback)

        return None

    logo_path = resolve_logo_path()

    def draw_branding(canvas, doc):
        page_w, page_h = A4
        canvas.saveState()

        canvas.setStrokeColor(colors.HexColor("#E4E7EC"))
        canvas.setLineWidth(0.8)
        canvas.roundRect(8 * mm, 8 * mm, page_w - 16 * mm, page_h - 16 * mm, 4 * mm, stroke=1, fill=0)

        usable_w = page_w - 16 * mm
        stripe_w = usable_w / 3.0
        stripe_y = page_h - 13 * mm
        stripe_h = 4.5 * mm

        canvas.setFillColor(brand_pink)
        canvas.rect(8 * mm, stripe_y, stripe_w, stripe_h, fill=1, stroke=0)

        canvas.setFillColor(brand_orange)
        canvas.rect(8 * mm + stripe_w, stripe_y, stripe_w, stripe_h, fill=1, stroke=0)

        canvas.setFillColor(brand_blue)
        canvas.rect(8 * mm + 2 * stripe_w, stripe_y, stripe_w, stripe_h, fill=1, stroke=0)

        if logo_path:
            try:
                img = ImageReader(logo_path)
                iw, ih = img.getSize()
                draw_w = 26 * mm
                draw_h = draw_w * (ih / float(iw)) if iw and ih else 26 * mm
                x = (page_w - draw_w) / 2.0
                y = 10 * mm

                try:
                    canvas.setFillAlpha(0.10)
                    canvas.setStrokeAlpha(0.10)
                except Exception:
                    pass

                canvas.drawImage(
                    img,
                    x,
                    y,
                    width=draw_w,
                    height=draw_h,
                    preserveAspectRatio=True,
                    mask="auto",
                )
            except Exception:
                pass

        canvas.restoreState()

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
    )

    styles = getSampleStyleSheet()

    base_style = ParagraphStyle(
        "GPOBase",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8.5,
        leading=10.5,
        textColor=ink,
        spaceAfter=0,
    )

    header_left_style = ParagraphStyle(
        "GPOHeaderLeft",
        parent=base_style,
        fontName="Helvetica",
        fontSize=8.4,
        leading=10.4,
        textColor=white,
        alignment=TA_LEFT,
    )

    header_title_style = ParagraphStyle(
        "GPOHeaderTitle",
        parent=base_style,
        fontName="Helvetica-Bold",
        fontSize=14,
        leading=16,
        textColor=brand_navy,
        alignment=TA_RIGHT,
    )

    header_meta_style = ParagraphStyle(
        "GPOHeaderMeta",
        parent=base_style,
        fontName="Helvetica",
        fontSize=8.3,
        leading=10.2,
        textColor=ink,
        alignment=TA_RIGHT,
    )

    section_head_left = ParagraphStyle(
        "GPOSectionHeadLeft",
        parent=base_style,
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=10,
        textColor=white,
        alignment=TA_LEFT,
    )

    section_value_style = ParagraphStyle(
        "GPOSectionValue",
        parent=base_style,
        fontName="Helvetica",
        fontSize=8.2,
        leading=10.2,
        textColor=ink,
        alignment=TA_LEFT,
    )

    table_head_style = ParagraphStyle(
        "GPOTableHead",
        parent=base_style,
        fontName="Helvetica-Bold",
        fontSize=7.8,
        leading=9.5,
        textColor=white,
        alignment=TA_CENTER,
    )

    item_center_style = ParagraphStyle(
        "GPOItemCenter",
        parent=base_style,
        fontName="Helvetica",
        fontSize=8,
        leading=9.6,
        alignment=TA_CENTER,
    )

    item_desc_style = ParagraphStyle(
        "GPOItemDesc",
        parent=base_style,
        fontName="Helvetica",
        fontSize=8,
        leading=9.6,
        alignment=TA_LEFT,
    )

    money_style = ParagraphStyle(
        "GPOMoney",
        parent=base_style,
        fontName="Helvetica",
        fontSize=8,
        leading=9.6,
        alignment=TA_RIGHT,
    )

    block_title_style = ParagraphStyle(
        "GPOBlockTitle",
        parent=base_style,
        fontName="Helvetica-Bold",
        fontSize=8.4,
        leading=10.5,
        textColor=brand_navy,
        alignment=TA_LEFT,
    )

    block_text_style = ParagraphStyle(
        "GPOBlockText",
        parent=base_style,
        fontName="Helvetica",
        fontSize=8.1,
        leading=10.1,
        textColor=ink,
        alignment=TA_LEFT,
    )

    sign_style = ParagraphStyle(
        "GPOSign",
        parent=base_style,
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=10,
        textColor=ink,
        alignment=TA_LEFT,
    )

    total_label_style = ParagraphStyle(
        "GPOTotalLabel",
        parent=base_style,
        fontName="Helvetica-Bold",
        fontSize=8.2,
        leading=10.2,
        textColor=ink,
        alignment=TA_LEFT,
    )

    total_value_style = ParagraphStyle(
        "GPOTotalValue",
        parent=base_style,
        fontName="Helvetica-Bold",
        fontSize=8.2,
        leading=10.2,
        textColor=ink,
        alignment=TA_RIGHT,
    )

    footer_note_style = ParagraphStyle(
        "GPOFooterNote",
        parent=base_style,
        fontName="Helvetica-Bold",
        fontSize=7.6,
        leading=9.2,
        textColor=muted,
        alignment=TA_CENTER,
    )

    story = []

    firm_name = text_or_dash(firm.firm_name if firm else "InventTech")
    firm_type = ""
    if firm:
        try:
            firm_type = firm.get_firm_type_display()
        except Exception:
            firm_type = text_or_dash(getattr(firm, "firm_type", ""))

    firm_address = join_parts(
        getattr(firm, "address_line", ""),
        getattr(firm, "city", ""),
        getattr(firm, "state", ""),
        getattr(firm, "pincode", ""),
    )

    firm_contact_line = " | ".join([
        part for part in [
            line_if("Phone", getattr(firm, "phone", "")),
            line_if("Email", getattr(firm, "email", "")),
            line_if("GSTIN", getattr(firm, "gst_number", "")),
        ] if part
    ])

    firm_stat_line = " | ".join([
        part for part in [
            line_if("PAN", getattr(firm, "pan_number", "")),
            line_if("TAN", getattr(firm, "tan_number", "")),
            line_if("CIN", getattr(firm, "cin_number", "")),
        ] if part
    ])

    order_number = po.po_number or po.system_number or "-"
    po_date = po.po_date.strftime("%d-%m-%Y") if po.po_date else "-"
    cancel_date = po.cancel_date.strftime("%d-%m-%Y") if po.cancel_date else "-"
    approval_label = getattr(po, "get_approval_status_display", lambda: text_or_dash(po.approval_status))()

    firm_header_html = f"<font size='13'><b>{escape(firm_name)}</b></font>"
    if firm_type and firm_type != "-":
        firm_header_html += f"<br/>{escape(firm_type)}"
    if firm_address:
        firm_header_html += f"<br/>{escape(firm_address)}"
    if firm_contact_line:
        firm_header_html += f"<br/>{firm_contact_line}"
    if firm_stat_line:
        firm_header_html += f"<br/>{firm_stat_line}"

    source_yarn_po_no = text_or_dash(po.source_yarn_po.system_number if po.source_yarn_po else "")
    source_yarn_inward_no = text_or_dash(po.source_yarn_inward.inward_number if po.source_yarn_inward else "")

    header_table = Table(
        [[
            Paragraph(firm_header_html, header_left_style),
            Table(
                [[
                    Paragraph("<b>GREIGE PURCHASE ORDER</b>", header_title_style),
                ], [
                    Paragraph(
                        f"<b>PO No:</b> {escape(order_number)}<br/>"
                        f"<b>PO Date:</b> {escape(po_date)}<br/>"
                        f"<b>System No:</b> {escape(text_or_dash(po.system_number))}<br/>"
                        f"<b>Status:</b> {escape(text_or_dash(approval_label))}<br/>"
                        f"<b>Source Yarn PO:</b> {escape(source_yarn_po_no)}"
                        + (f"<br/><b>Source Inward:</b> {escape(source_yarn_inward_no)}" if source_yarn_inward_no != "-" else "")
                        + (f"<br/><b>Cancel Date:</b> {escape(cancel_date)}" if cancel_date != "-" else ""),
                        header_meta_style,
                    )
                ]],
                colWidths=[66 * mm],
            ),
        ]],
        colWidths=[124 * mm, 66 * mm],
    )
    header_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), brand_navy),
        ("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#F5F8FF")),
        ("BOX", (0, 0), (-1, -1), 0.9, border),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (0, 0), 9),
        ("RIGHTPADDING", (0, 0), (0, 0), 9),
        ("TOPPADDING", (0, 0), (0, 0), 9),
        ("BOTTOMPADDING", (0, 0), (0, 0), 9),
        ("LEFTPADDING", (1, 0), (1, 0), 0),
        ("RIGHTPADDING", (1, 0), (1, 0), 0),
        ("TOPPADDING", (1, 0), (1, 0), 0),
        ("BOTTOMPADDING", (1, 0), (1, 0), 0),
    ]))
    header_table._argW[1] = 66 * mm
    inner_header = header_table._cellvalues[0][1]
    inner_header.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F5F8FF")),
        ("BOX", (0, 0), (-1, -1), 0.9, brand_blue),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 6))

    vendor_html = f"<b>{escape(text_or_dash(vendor.name if vendor else ''))}</b>"
    vendor_lines = [
        line_if("Contact", vendor.contact_person if vendor else ""),
        line_if("Phone", vendor.phone if vendor else ""),
        line_if("Email", vendor.email if vendor else ""),
        line_if("GSTIN", vendor.gst_number if vendor else ""),
        line_if("Address", vendor.address if vendor else ""),
    ]
    vendor_body = "<br/>".join([line for line in vendor_lines if line])
    if vendor_body:
        vendor_html += "<br/>" + vendor_body

    bill_to_html = f"<b>{escape(text_or_dash(firm_name if firm else ''))}</b>"
    bill_lines = [
        line_if("Address", firm_address),
        line_if("Phone", getattr(firm, "phone", "")),
        line_if("Email", getattr(firm, "email", "")),
        line_if("GSTIN", getattr(firm, "gst_number", "")),
        line_if("Ship To", po.shipping_address),
    ]
    bill_body = "<br/>".join([line for line in bill_lines if line])
    if bill_body:
        bill_to_html += "<br/>" + bill_body

    party_table = Table(
        [
            [
                Paragraph("VENDOR", section_head_left),
                Paragraph("BILL TO", section_head_left),
            ],
            [
                Paragraph(vendor_html, section_value_style),
                Paragraph(bill_to_html, section_value_style),
            ],
        ],
        colWidths=[95 * mm, 95 * mm],
    )
    party_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), brand_orange),
        ("BACKGROUND", (1, 0), (1, 0), brand_blue),
        ("TEXTCOLOR", (0, 0), (-1, 0), white),
        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#FBFCFE")),
        ("BOX", (0, 0), (-1, -1), 0.9, border),
        ("INNERGRID", (0, 0), (-1, -1), 0.7, border),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.append(party_table)
    story.append(Spacer(1, 7))

    item_rows = [[
        Paragraph("Sr No", table_head_style),
        Paragraph("Description", table_head_style),
        Paragraph("Unit", table_head_style),
        Paragraph("Qty", table_head_style),
        Paragraph("Price (Rs.)", table_head_style),
        Paragraph("Amount (Rs.)", table_head_style),
    ]]

    total_amount = Decimal("0")
    total_qty = Decimal("0")
    total_inward = Decimal("0")
    remaining_qty = Decimal("0")

    for index, item in enumerate(po_items, start=1):
        label = item.fabric_name or (item.material.name if item.material else "Greige Item")

        details = []
        if item.yarn_name:
            details.append(f"Yarn: {item.yarn_name}")
        if item.count:
            details.append(f"Count: {item.count}")
        if item.dia:
            details.append(f"Dia: {item.dia}")
        if item.gauge:
            details.append(f"Gauge: {item.gauge}")
        if item.gsm:
            details.append(f"GSM: {item.gsm}")
        if item.sl:
            details.append(f"SL: {item.sl}")
        if item.rolls:
            details.append(f"Rolls: {item.rolls}")
        if item.hsn_code:
            details.append(f"HSN: {item.hsn_code}")
        if item.remark:
            details.append(f"Remark: {item.remark}")

        description_html = f"<b>{escape(text_or_dash(label))}</b>"
        if details:
            description_html += "<br/>" + escape(" | ".join(details))

        item_rows.append([
            Paragraph(str(index), item_center_style),
            Paragraph(description_html, item_desc_style),
            Paragraph(escape(text_or_dash(item.unit)), item_center_style),
            Paragraph(escape(fmt_qty(item.quantity)), item_center_style),
            Paragraph(escape(fmt_money(item.rate)), money_style),
            Paragraph(escape(fmt_money(item.final_amount)), money_style),
        ])

        total_amount += Decimal(item.final_amount or 0)
        total_qty += Decimal(item.quantity or 0)
        total_inward += Decimal(item.inward_qty_total or 0)
        remaining_qty += Decimal(item.remaining_qty_total or 0)

    min_visual_rows = 5
    for _ in range(max(0, min_visual_rows - len(po_items))):
        item_rows.append([
            Paragraph("", item_center_style),
            Paragraph("", item_desc_style),
            Paragraph("", item_center_style),
            Paragraph("", item_center_style),
            Paragraph("", money_style),
            Paragraph("", money_style),
        ])

    items_table = Table(
        item_rows,
        colWidths=[13 * mm, 86 * mm, 18 * mm, 18 * mm, 26 * mm, 29 * mm],
        repeatRows=1,
    )
    items_table_style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), brand_navy),
        ("TEXTCOLOR", (0, 0), (-1, 0), white),
        ("BOX", (0, 0), (-1, -1), 0.9, border),
        ("INNERGRID", (0, 0), (-1, -1), 0.55, border),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ])

    for row_index in range(1, len(item_rows)):
        if row_index % 2 == 0:
            items_table_style.add("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#F9FAFB"))

    items_table.setStyle(items_table_style)
    story.append(items_table)
    story.append(Spacer(1, 8))

    notes_parts = []
    if po.remarks:
        notes_parts.append(f"<font color='#0F172A'><b>Remarks</b></font><br/>{escape(po.remarks).replace(chr(10), '<br/>')}")
    if po.delivery_schedule:
        notes_parts.append(f"<font color='#0F172A'><b>Delivery Schedule</b></font><br/>{escape(po.delivery_schedule).replace(chr(10), '<br/>')}")
    if po.shipping_address:
        notes_parts.append(f"<font color='#0F172A'><b>Shipping Address</b></font><br/>{escape(po.shipping_address).replace(chr(10), '<br/>')}")
    if po.source_yarn_po:
        notes_parts.append(f"<font color='#0F172A'><b>Source Yarn PO</b></font><br/>{escape(text_or_dash(po.source_yarn_po.system_number))}")

    if not notes_parts:
        notes_parts.append("<font color='#0F172A'><b>Notes</b></font><br/>Standard terms and conditions apply.")

    notes_html = "<br/><br/>".join(notes_parts)

    signature_table = Table(
        [[
            Paragraph("<b>AUTHORISED SIGNATORY</b><br/><br/>_________________________", sign_style),
            Paragraph(f"<b>DATE</b><br/><br/>{escape(po_date)}", sign_style),
        ]],
        colWidths=[68 * mm, 30 * mm],
    )
    signature_table.setStyle(TableStyle([
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))

    left_block = Table(
        [
            [Paragraph("NOTES / TERMS", block_title_style)],
            [Paragraph(notes_html, block_text_style)],
            [signature_table],
        ],
        colWidths=[102 * mm],
    )
    left_block.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#FDF2F8")),
        ("BOX", (0, 0), (-1, -1), 0.9, border),
        ("INNERGRID", (0, 0), (-1, -1), 0.6, border),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))

    totals_rows = [
        [Paragraph("Total Qty", total_label_style), Paragraph(escape(fmt_qty(total_qty)), total_value_style)],
        [Paragraph("Total Inward", total_label_style), Paragraph(escape(fmt_qty(total_inward)), total_value_style)],
        [Paragraph("Remaining Qty", total_label_style), Paragraph(escape(fmt_qty(remaining_qty)), total_value_style)],
        [Paragraph("Total Amount", total_label_style), Paragraph(escape(fmt_money(total_amount)), total_value_style)],
    ]

    totals_table = Table(totals_rows, colWidths=[49 * mm, 39 * mm])
    totals_table_style = TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.9, border),
        ("INNERGRID", (0, 0), (-1, -1), 0.6, border),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EFF6FF")),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#FFF7ED")),
    ])
    for row_index in range(1, len(totals_rows) - 1):
        if row_index % 2 == 1:
            totals_table_style.add("BACKGROUND", (0, row_index), (-1, row_index), soft_bg)
    totals_table.setStyle(totals_table_style)

    lower_table = Table([[left_block, totals_table]], colWidths=[102 * mm, 88 * mm])
    lower_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    story.append(lower_table)
    story.append(Spacer(1, 8))

    footer_table = Table(
        [[Paragraph("THIS PO IS COMPUTER GENERATED, HENCE SIGNATURE IS NOT REQUIRED", footer_note_style)]],
        colWidths=[190 * mm],
    )
    footer_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
        ("BOX", (0, 0), (-1, -1), 0.9, border),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(footer_table)

    doc.build(story, onFirstPage=draw_branding, onLaterPages=draw_branding)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    return response


@login_required
def greigepo_pdf(request, pk: int):
    po = get_object_or_404(
        GreigePurchaseOrder.objects
        .select_related("vendor", "source_yarn_po", "source_yarn_inward", "owner")
        .prefetch_related(
            Prefetch(
                "items",
                queryset=GreigePurchaseOrderItem.objects.select_related("material", "source_yarn_po_item")
            )
        ),
        pk=pk,
    )

    if not _can_access_greige_po(request.user, po):
        raise PermissionDenied("You do not have access to this Greige PO.")

    try:
        response = _build_greige_po_pdf_response(po)
    except Exception:
        logger.exception(
            "Branded Greige PO PDF generation failed for PO id=%s system_no=%s",
            po.pk,
            po.system_number,
        )
        response = _build_simple_greige_po_pdf_response(po)

    if response.status_code == 200 and response.get("Content-Type", "").startswith("application/pdf"):
        filename = f'{po.system_number or "greige_po"}.pdf'
        disposition = "attachment" if request.GET.get("download") == "1" else "inline"
        response["Content-Disposition"] = f'{disposition}; filename="{filename}"'
        response["Cache-Control"] = "no-store"
        response["X-Content-Type-Options"] = "nosniff"
        try:
            response["Content-Length"] = str(len(response.content))
        except Exception:
            pass

    return response

@login_required
def yarnpo_list(request):
    q = (request.GET.get("q") or "").strip()

    qs = (
        YarnPurchaseOrder.objects
        .select_related("vendor", "firm", "reviewed_by", "owner")
        .prefetch_related(
            Prefetch(
                "items",
                queryset=YarnPurchaseOrderItem.objects.select_related("material", "material_type").prefetch_related("inward_items"),
            ),
            Prefetch(
                "inwards",
                queryset=YarnPOInward.objects.prefetch_related("items"),
            ),
        )
    )

    if not _can_review_yarn_po(request):
        qs = qs.filter(owner=request.user)

    if q:
        qs = qs.filter(
            Q(system_number__icontains=q)
            | Q(po_number__icontains=q)
            | Q(vendor__name__icontains=q)
            | Q(firm__firm_name__icontains=q)
            | Q(items__material__name__icontains=q)
            | Q(items__material__material_type__name__icontains=q)
        ).distinct()

    orders = [_attach_yarn_po_metrics(po) for po in qs.order_by("-id")]

    return render(
        request,
        "accounts/yarn_po/list.html",
        {
            "orders": orders,
            "q": q,
            "can_review_yarn_po": _can_review_yarn_po(request),
        },
    )


def _bind_yarnpo_item_formset(request, instance=None, user=None):
    effective_user = user or request.user

    kwargs = {
        "instance": instance,
        "prefix": "items",
        "form_kwargs": {"user": effective_user},
    }

    if request.method == "POST":
        formset = YarnPurchaseOrderItemFormSet(request.POST, **kwargs)
    else:
        formset = YarnPurchaseOrderItemFormSet(**kwargs)

    return formset
@login_required
@require_http_methods(["GET", "POST"])
def yarnpo_create(request):
    default_firm = Firm.objects.filter(owner=request.user).prefetch_related("addresses").first()
    if default_firm is None:
        messages.error(request, "Create your firm first before generating a Yarn PO.")
        return redirect("accounts:firm_add")

    po = YarnPurchaseOrder(owner=request.user, firm=default_firm, shipping_address=_firm_address(default_firm))

    form = YarnPurchaseOrderForm(request.POST or None, user=request.user, instance=po)
    formset = _bind_yarnpo_item_formset(request, instance=po)

    if request.method == "POST":
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                po = form.save(commit=False)
                po.owner = request.user
                po.firm = default_firm
                po.po_date = timezone.localdate()

                if not po.system_number:
                    po.system_number = _next_yarn_po_number()

                if not po.shipping_address:
                    po.shipping_address = _firm_address(po.firm)

                po.save()
                formset.instance = po
                formset.save()
                _recalculate_yarn_po(po)

            messages.success(request, f"Yarn PO {po.system_number} saved successfully.")
            return redirect("accounts:yarnpo_list")

        form.add_error(None, "Yarn PO was not saved. Please fix the highlighted fields and try again.")

    return render(request, "accounts/yarn_po/form.html", {
        "form": form,
        "formset": formset,
        "mode": "add",
        "po_obj": po,
        "system_number_preview": po.system_number or _next_yarn_po_number(),
        "auto_firm_name": default_firm.firm_name if default_firm else "",
        "terms_condition_map": {str(obj.pk): obj.content for obj in form.fields["terms_template"].queryset},
        "shipping_address_map": getattr(form, "shipping_address_map", {}),
    })


@login_required
@require_http_methods(["GET", "POST"])
def yarnpo_update(request, pk: int):
    po = get_object_or_404(
        YarnPurchaseOrder.objects.select_related("vendor", "firm", "owner").prefetch_related("inwards", "greige_pos"),
        pk=pk,
    )

    if not _can_access_yarn_po(request.user, po):
        raise PermissionDenied("You do not have access to this Yarn PO.")

    lock_reason = _yarn_po_lock_reason(po)
    if lock_reason:
        messages.error(request, lock_reason)
        return redirect("accounts:yarnpo_review", pk=po.pk)

    po_owner = po.owner
    default_firm = Firm.objects.filter(owner=po_owner).prefetch_related("addresses").first()
    display_firm = po.firm or default_firm

    if request.method == "GET" and display_firm and not po.firm:
        po.firm = display_firm
        if not po.shipping_address:
            po.shipping_address = _firm_address(display_firm)

    form = YarnPurchaseOrderForm(request.POST or None, user=po_owner, instance=po)
    formset = _bind_yarnpo_item_formset(request, instance=po, user=po_owner)

    if request.method == "POST":
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                po = form.save(commit=False)

                if default_firm:
                    po.firm = default_firm

                if not po.shipping_address and po.firm:
                    po.shipping_address = _firm_address(po.firm)

                po.save()
                formset.instance = po
                formset.save()
                _recalculate_yarn_po(po)

            messages.success(request, f"Yarn PO {po.system_number} updated successfully.")
            return redirect("accounts:yarnpo_list")

        form.add_error(None, "Yarn PO was not updated. Please fix the highlighted fields and try again.")

    return render(request, "accounts/yarn_po/form.html", {
        "form": form,
        "formset": formset,
        "mode": "edit",
        "po_obj": po,
        "system_number_preview": po.system_number,
        "auto_firm_name": display_firm.firm_name if display_firm else "",
        "terms_condition_map": {str(obj.pk): obj.content for obj in form.fields["terms_template"].queryset},
        "shipping_address_map": getattr(form, "shipping_address_map", {}),
    })
    
def _build_simple_yarn_po_pdf_response(po):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas
    except ImportError:
        return HttpResponse(
            "ReportLab is required for PDF generation. Install it with: pip install reportlab",
            status=500,
        )

    def safe_text(value):
        value = "" if value is None else str(value).strip()
        return value if value else "-"

    def qty_text(value):
        try:
            return f"{float(value or 0):,.2f}".rstrip("0").rstrip(".")
        except Exception:
            return safe_text(value)

    def money_text(value):
        try:
            return f"{float(value or 0):,.2f}"
        except Exception:
            return safe_text(value)

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    left = 15 * mm
    top = height - 18 * mm
    line_gap = 5.5 * mm

    def draw_line(label, value, y, bold=False):
        pdf.setFont("Helvetica-Bold" if bold else "Helvetica", 9)
        pdf.drawString(left, y, f"{label}: {safe_text(value)}")

    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(left, top, safe_text(po.firm.firm_name if po.firm else "InventTech"))

    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawRightString(width - 15 * mm, top, "YARN PURCHASE ORDER")

    y = top - 9 * mm
    draw_line("PO No", po.po_number or po.system_number, y)
    draw_line("PO Date", po.po_date.strftime("%d-%m-%Y") if po.po_date else "-", y - line_gap)
    draw_line("System No", po.system_number, y - (2 * line_gap))
    draw_line("Status", po.get_approval_status_display(), y - (3 * line_gap))

    y -= 28 * mm

    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(left, y, "Vendor")
    pdf.drawString(width / 2, y, "Bill To")

    y -= 6 * mm
    pdf.setFont("Helvetica", 9)
    vendor_lines = [
        safe_text(po.vendor.name if po.vendor else ""),
        f"Phone: {safe_text(po.vendor.phone if po.vendor else '')}",
        f"Email: {safe_text(po.vendor.email if po.vendor else '')}",
        f"GSTIN: {safe_text(po.vendor.gst_number if po.vendor else '')}",
        f"Address: {safe_text(po.vendor.address if po.vendor else '')}",
    ]
    bill_lines = [
        safe_text(po.firm.firm_name if po.firm else ""),
        f"Phone: {safe_text(getattr(po.firm, 'phone', ''))}",
        f"Email: {safe_text(getattr(po.firm, 'email', ''))}",
        f"GSTIN: {safe_text(getattr(po.firm, 'gst_number', ''))}",
        f"Ship To: {safe_text(po.shipping_address)}",
    ]

    for idx in range(max(len(vendor_lines), len(bill_lines))):
        vendor_text = vendor_lines[idx] if idx < len(vendor_lines) else ""
        bill_text = bill_lines[idx] if idx < len(bill_lines) else ""
        pdf.drawString(left, y, vendor_text)
        pdf.drawString(width / 2, y, bill_text)
        y -= 5 * mm

    y -= 3 * mm

    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawString(left, y, "Sr")
    pdf.drawString(left + 12 * mm, y, "Description")
    pdf.drawString(left + 98 * mm, y, "Unit")
    pdf.drawRightString(left + 128 * mm, y, "Qty")
    pdf.drawRightString(left + 158 * mm, y, "Price")
    pdf.drawRightString(width - 15 * mm, y, "Amount")

    y -= 4 * mm
    pdf.line(left, y, width - 15 * mm, y)
    y -= 6 * mm

    items = list(po.items.all())
    for index, item in enumerate(items, start=1):
        if y < 45 * mm:
            pdf.showPage()
            y = height - 20 * mm

        item_name = "-"
        if item.material:
            item_name = item.material.name
        elif item.material_type:
            item_name = item.material_type.name

        detail_bits = []
        if item.count:
            detail_bits.append(f"Count: {item.count}")
        if item.dia:
            detail_bits.append(f"Dia: {item.dia}")
        if item.gauge:
            detail_bits.append(f"Gauge: {item.gauge}")
        if item.gsm:
            detail_bits.append(f"GSM: {item.gsm}")
        if item.rolls:
            detail_bits.append(f"Rolls: {item.rolls}")
        if item.hsn_code:
            detail_bits.append(f"HSN: {item.hsn_code}")

        desc = safe_text(item_name)
        if detail_bits:
            desc += " | " + " | ".join(detail_bits)

        pdf.setFont("Helvetica", 8.5)
        pdf.drawString(left, y, str(index))
        pdf.drawString(left + 12 * mm, y, desc[:70])
        pdf.drawString(left + 98 * mm, y, safe_text(item.unit))
        pdf.drawRightString(left + 128 * mm, y, qty_text(item.quantity))
        pdf.drawRightString(left + 158 * mm, y, money_text(item.rate))
        pdf.drawRightString(width - 15 * mm, y, money_text(item.final_amount))
        y -= 6 * mm

    y -= 4 * mm
    pdf.line(left, y, width - 15 * mm, y)
    y -= 8 * mm

    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawRightString(left + 158 * mm, y, "Total Amount")
    pdf.drawRightString(width - 15 * mm, y, money_text(po.grand_total))

    y -= 12 * mm
    pdf.setFont("Helvetica", 8.5)
    pdf.drawString(left, y, "THIS PO IS COMPUTER GENERATED, HENCE SIGNATURE IS NOT REQUIRED")

    pdf.save()
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    return response


@login_required
def yarnpo_pdf(request, pk: int):
    po = get_object_or_404(
        YarnPurchaseOrder.objects
        .select_related("vendor", "firm", "owner")
        .prefetch_related(
            Prefetch("items", queryset=YarnPurchaseOrderItem.objects.select_related("material", "material_type"))
        ),
        pk=pk,
    )

    if not _can_access_yarn_po(request.user, po):
        raise PermissionDenied("You do not have access to this PO.")

    try:
        response = _build_yarn_po_pdf_response(po)
    except Exception:
        logger.exception("Branded Yarn PO PDF generation failed for PO id=%s system_no=%s", po.pk, po.system_number)
        response = _build_simple_yarn_po_pdf_response(po)

    if response.status_code == 200 and response.get("Content-Type", "").startswith("application/pdf"):
        filename = f'{po.system_number or "yarn_po"}.pdf'
        disposition = "attachment" if request.GET.get("download") == "1" else "inline"
        response["Content-Disposition"] = f'{disposition}; filename="{filename}"'
        response["Cache-Control"] = "no-store"
        response["X-Content-Type-Options"] = "nosniff"
        try:
            response["Content-Length"] = str(len(response.content))
        except Exception:
            pass

    return response

def _build_simple_greige_po_pdf_response(po):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas
    except ImportError:
        return HttpResponse(
            "ReportLab is required for PDF generation. Install it with: pip install reportlab",
            status=500,
        )

    def safe_text(value):
        value = "" if value is None else str(value).strip()
        return value if value else "-"

    def qty_text(value):
        try:
            return f"{float(value or 0):,.2f}".rstrip("0").rstrip(".")
        except Exception:
            return safe_text(value)

    def money_text(value):
        try:
            return f"{float(value or 0):,.2f}"
        except Exception:
            return safe_text(value)

    items = list(po.items.all())
    total_amount = sum((item.final_amount or Decimal("0")) for item in items)
    total_qty = sum((item.quantity or Decimal("0")) for item in items)

    firm = getattr(po.source_yarn_po, "firm", None)

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    left = 15 * mm
    top = height - 18 * mm
    line_gap = 5.5 * mm

    def draw_line(label, value, y, bold=False):
        pdf.setFont("Helvetica-Bold" if bold else "Helvetica", 9)
        pdf.drawString(left, y, f"{label}: {safe_text(value)}")

    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(left, top, safe_text(firm.firm_name if firm else "InventTech"))

    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawRightString(width - 15 * mm, top, "GREIGE PURCHASE ORDER")

    y = top - 9 * mm
    draw_line("PO No", po.po_number or po.system_number, y)
    draw_line("PO Date", po.po_date.strftime("%d-%m-%Y") if po.po_date else "-", y - line_gap)
    draw_line("System No", po.system_number, y - (2 * line_gap))
    draw_line("Status", po.get_approval_status_display(), y - (3 * line_gap))
    draw_line("Source Yarn PO", po.source_yarn_po.system_number if po.source_yarn_po else "-", y - (4 * line_gap))

    y -= 32 * mm

    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(left, y, "Vendor")
    pdf.drawString(width / 2, y, "Bill To")

    y -= 6 * mm
    pdf.setFont("Helvetica", 9)

    vendor_lines = [
        safe_text(po.vendor.name if po.vendor else ""),
        f"Phone: {safe_text(po.vendor.phone if po.vendor else '')}",
        f"Email: {safe_text(po.vendor.email if po.vendor else '')}",
        f"GSTIN: {safe_text(po.vendor.gst_number if po.vendor else '')}",
        f"Address: {safe_text(po.vendor.address if po.vendor else '')}",
    ]

    bill_lines = [
        safe_text(firm.firm_name if firm else ""),
        f"Phone: {safe_text(getattr(firm, 'phone', ''))}",
        f"Email: {safe_text(getattr(firm, 'email', ''))}",
        f"GSTIN: {safe_text(getattr(firm, 'gst_number', ''))}",
        f"Ship To: {safe_text(po.shipping_address)}",
    ]

    for idx in range(max(len(vendor_lines), len(bill_lines))):
        vendor_text = vendor_lines[idx] if idx < len(vendor_lines) else ""
        bill_text = bill_lines[idx] if idx < len(bill_lines) else ""
        pdf.drawString(left, y, vendor_text)
        pdf.drawString(width / 2, y, bill_text)
        y -= 5 * mm

    y -= 3 * mm

    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawString(left, y, "Sr")
    pdf.drawString(left + 12 * mm, y, "Fabric")
    pdf.drawString(left + 62 * mm, y, "Yarn")
    pdf.drawString(left + 108 * mm, y, "Unit")
    pdf.drawRightString(left + 135 * mm, y, "Qty")
    pdf.drawRightString(left + 162 * mm, y, "Rate")
    pdf.drawRightString(width - 15 * mm, y, "Amount")

    y -= 4 * mm
    pdf.line(left, y, width - 15 * mm, y)
    y -= 6 * mm

    for index, item in enumerate(items, start=1):
        if y < 45 * mm:
            pdf.showPage()
            y = height - 20 * mm

            pdf.setFont("Helvetica-Bold", 9)
            pdf.drawString(left, y, "Sr")
            pdf.drawString(left + 12 * mm, y, "Fabric")
            pdf.drawString(left + 62 * mm, y, "Yarn")
            pdf.drawString(left + 108 * mm, y, "Unit")
            pdf.drawRightString(left + 135 * mm, y, "Qty")
            pdf.drawRightString(left + 162 * mm, y, "Rate")
            pdf.drawRightString(width - 15 * mm, y, "Amount")
            y -= 4 * mm
            pdf.line(left, y, width - 15 * mm, y)
            y -= 6 * mm

        fabric_name = item.fabric_name or (item.material.name if item.material else "Greige Item")
        yarn_name = item.yarn_name or "-"
        unit = item.unit or "-"
        qty = item.quantity or Decimal("0")
        rate = item.rate or Decimal("0")
        amount = item.final_amount or Decimal("0")

        pdf.setFont("Helvetica", 8.5)
        pdf.drawString(left, y, str(index))
        pdf.drawString(left + 12 * mm, y, safe_text(fabric_name)[:28])
        pdf.drawString(left + 62 * mm, y, safe_text(yarn_name)[:24])
        pdf.drawString(left + 108 * mm, y, safe_text(unit)[:8])
        pdf.drawRightString(left + 135 * mm, y, qty_text(qty))
        pdf.drawRightString(left + 162 * mm, y, money_text(rate))
        pdf.drawRightString(width - 15 * mm, y, money_text(amount))
        y -= 6 * mm

    y -= 4 * mm
    pdf.line(left, y, width - 15 * mm, y)
    y -= 8 * mm

    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawString(left, y, f"Total Qty: {qty_text(total_qty)}")
    pdf.drawRightString(left + 162 * mm, y, "Total Amount")
    pdf.drawRightString(width - 15 * mm, y, money_text(total_amount))

    y -= 12 * mm
    pdf.setFont("Helvetica", 8.5)
    pdf.drawString(left, y, "THIS PO IS COMPUTER GENERATED, HENCE SIGNATURE IS NOT REQUIRED")

    pdf.save()
    buffer.seek(0)

    return HttpResponse(buffer.getvalue(), content_type="application/pdf")


def _build_yarn_inward_line_rows(po, line_inputs=None, item_errors=None):
    line_inputs = line_inputs or {}
    item_errors = item_errors or {}

    rows = []
    for item in po.items.all():
        row_input = line_inputs.get(item.id, {})
        rows.append({
            "item": item,
            "qty_value": row_input.get("qty", ""),
            "remark_value": row_input.get("remark", ""),
            "error": item_errors.get(item.id, ""),
        })
    return rows

@login_required
@require_http_methods(["GET", "POST"])
def yarnpo_inward(request, pk: int):
    po = get_object_or_404(
        YarnPurchaseOrder.objects
        .select_related("vendor", "firm", "owner")
        .prefetch_related(
            Prefetch(
                "items",
                queryset=YarnPurchaseOrderItem.objects.select_related("material", "material_type").prefetch_related("inward_items"),
            ),
            Prefetch(
                "inwards",
                queryset=YarnPOInward.objects.select_related("vendor", "inward_type").prefetch_related("items__po_item__material"),
            ),
        ),
        pk=pk,
    )

    if not _can_access_yarn_po(request.user, po):
        raise PermissionDenied("You do not have access to this PO.")
    
    if not _is_po_approved_for_inward(po):
        messages.error(request, "Yarn PO must be approved before inward can be generated.")
        return redirect("accounts:yarnpo_list")

    po = _attach_yarn_po_metrics(po)
    item_errors = {}
    line_inputs = {}

    inward_form = YarnPOInwardForm(request.POST or None, user=po.owner)

    has_remaining_qty = any(
        (item.remaining_qty_total or Decimal("0")) > 0
        for item in po.items.all()
    )

    if request.method == "POST":
        if not has_remaining_qty:
            inward_form.is_valid()
            inward_form.add_error(None, "No quantity is remaining for inward in this Yarn PO.")
        elif inward_form.is_valid():
            line_payload, line_inputs, item_errors = _collect_basic_po_inward_payload(request, po)

            if not line_payload:
                inward_form.add_error(None, "Enter at least one inward quantity.")

            if not inward_form.errors and not item_errors:
                inward = inward_form.save(commit=False)
                inward.owner = po.owner
                inward.po = po
                inward.inward_number = _next_yarn_inward_number()
                inward.save()

                YarnPOInwardItem.objects.bulk_create([
                    YarnPOInwardItem(
                        inward=inward,
                        po_item=row["item"],
                        quantity=row["quantity"],
                        received_qty=row["received_qty"],
                        accepted_qty=row["accepted_qty"],
                        rejected_qty=row["rejected_qty"],
                        hold_qty=row["hold_qty"],
                        actual_rolls=row["actual_rolls"],
                        actual_gsm=row["actual_gsm"],
                        actual_width=row["actual_width"],
                        dye_lot_no=row["dye_lot_no"],
                        batch_no=row["batch_no"],
                        shade_reference=row["shade_reference"],
                        qc_status=row["qc_status"],
                        remark=row["remark"],
                    )
                    for row in line_payload
                ])

                tracker_url = reverse("accounts:yarn_inward_tracker")
                return redirect(f"{tracker_url}?inward={inward.pk}")

    existing_inwards = po.inwards.all().order_by("-inward_date", "-id")
    line_rows = _build_dyeing_inward_line_rows(po, line_inputs=line_inputs, item_errors=item_errors)

    return render(
        request,
        "accounts/yarn_po/inward.html",
        {
            "po": po,
            "inward_form": inward_form,
            "item_errors": item_errors,
            "line_inputs": line_inputs,
            "line_rows": line_rows,
            "existing_inwards": existing_inwards,
            "next_inward_number_preview": _next_yarn_inward_number(),
            "has_remaining_qty": has_remaining_qty,
            "editing_inward": None,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def yarn_inward_update(request, pk: int):
    inward = get_object_or_404(
        YarnPOInward.objects
        .select_related("po__vendor", "po__firm", "po__owner", "vendor", "inward_type")
        .prefetch_related(
            Prefetch(
                "items",
                queryset=YarnPOInwardItem.objects.select_related("po_item__material", "po_item__material_type"),
            ),
            "generated_greige_pos",
            Prefetch(
                "po__items",
                queryset=YarnPurchaseOrderItem.objects.select_related("material", "material_type").prefetch_related("inward_items"),
            ),
            Prefetch(
                "po__inwards",
                queryset=YarnPOInward.objects.select_related("vendor", "inward_type").prefetch_related("items__po_item__material"),
            ),
        ),
        pk=pk,
    )

    po = inward.po

    if not _can_access_yarn_po(request.user, po):
        raise PermissionDenied("You do not have access to this inward.")
    
    if not _is_po_approved_for_inward(po):
        messages.error(request, "Yarn PO must be approved before inward can be updated.")
        return redirect("accounts:yarnpo_list")

    if inward.generated_greige_pos.exists():
        messages.error(request, "This inward cannot be edited because a Greige PO has already been generated from it.")
        tracker_url = reverse("accounts:yarn_inward_tracker")
        return redirect(f"{tracker_url}?inward={inward.pk}")

    po = _attach_yarn_po_metrics(po)
    item_errors = {}
    existing_item_map = {
        row.po_item_id: {
            "qty": str(row.received_qty or row.quantity or ""),
            "accepted_qty": str(row.accepted_qty or ""),
            "rejected_qty": str(row.rejected_qty or ""),
            "hold_qty": str(row.hold_qty or ""),
            "actual_rolls": str(row.actual_rolls or ""),
            "actual_gsm": str(row.actual_gsm or ""),
            "actual_width": str(row.actual_width or ""),
            "dye_lot_no": row.dye_lot_no or "",
            "batch_no": row.batch_no or "",
            "shade_reference": row.shade_reference or "",
            "remark": row.remark or "",
        }
        for row in inward.items.all()
    }
    line_inputs = dict(existing_item_map)

    inward_form = YarnPOInwardForm(request.POST or None, instance=inward, user=po.owner)

    if request.method == "POST" and inward_form.is_valid():
        line_payload, line_inputs, item_errors = _collect_basic_po_inward_payload(request, po, editing_inward=inward)

        if not line_payload:
            inward_form.add_error(None, "Enter at least one inward quantity.")

        if not inward_form.errors and not item_errors:
            inward = inward_form.save(commit=False)
            inward.owner = po.owner
            inward.po = po
            inward.save()

            inward.items.all().delete()

            YarnPOInwardItem.objects.bulk_create([
                YarnPOInwardItem(
                    inward=inward,
                    po_item=row["item"],
                    quantity=row["quantity"],
                    received_qty=row["received_qty"],
                    accepted_qty=row["accepted_qty"],
                    rejected_qty=row["rejected_qty"],
                    hold_qty=row["hold_qty"],
                    actual_rolls=row["actual_rolls"],
                    actual_gsm=row["actual_gsm"],
                    actual_width=row["actual_width"],
                    dye_lot_no=row["dye_lot_no"],
                    batch_no=row["batch_no"],
                    shade_reference=row["shade_reference"],
                    qc_status=row["qc_status"],
                    remark=row["remark"],
                )
                for row in line_payload
            ])

            messages.success(request, f"Inward {inward.inward_number} updated successfully.")
            tracker_url = reverse("accounts:yarn_inward_tracker")
            return redirect(f"{tracker_url}?inward={inward.pk}")

    line_rows = _build_yarn_inward_line_rows(po, line_inputs=line_inputs, item_errors=item_errors)

    return render(
        request,
        "accounts/yarn_po/inward.html",
        {
            "po": po,
            "inward_form": inward_form,
            "item_errors": item_errors,
            "line_inputs": line_inputs,
            "line_rows": line_rows,
            "existing_inwards": po.inwards.all().order_by("-inward_date", "-id"),
            "next_inward_number_preview": inward.inward_number,
            "has_remaining_qty": True,
            "editing_inward": inward,
        },
    )


def _po_tracker_qty(po, *names):
    for name in names:
        value = getattr(po, name, None)
        if value is not None:
            return value
    return Decimal("0.00")


def _po_tracker_item_payload(inward_item, fallback_name="Item"):
    po_item = getattr(inward_item, "po_item", None)

    fabric_name = fallback_name
    if po_item is not None:
        material = getattr(po_item, "material", None) or getattr(po_item, "finished_material", None)
        material_type = getattr(po_item, "material_type", None)
        fabric_name = (
            getattr(material, "name", "")
            or getattr(material_type, "name", "")
            or getattr(po_item, "fabric_name", "")
            or getattr(po_item, "yarn_name", "")
            or getattr(po_item, "dyeing_name", "")
            or fallback_name
        )

    ordered_qty = Decimal("0.00")
    unit = ""
    if po_item is not None:
        ordered_qty = (
            getattr(po_item, "expected_output_qty", None)
            or getattr(po_item, "quantity", None)
            or getattr(po_item, "qty", None)
            or Decimal("0.00")
        )
        unit = getattr(po_item, "unit", "") or ""

    received_qty = (
        getattr(inward_item, "received_qty", None)
        or getattr(inward_item, "quantity", None)
        or Decimal("0.00")
    )

    return {
        "inward_item": inward_item,
        "po_item": po_item,
        "fabric_name": fabric_name,
        "ordered_qty": ordered_qty,
        "received_qty": received_qty,
        "inward_qty": getattr(inward_item, "quantity", None) or received_qty,
        "accepted_qty": getattr(inward_item, "accepted_qty", None) or Decimal("0.00"),
        "rejected_qty": getattr(inward_item, "rejected_qty", None) or Decimal("0.00"),
        "hold_qty": getattr(inward_item, "hold_qty", None) or Decimal("0.00"),
        "actual_rolls": getattr(inward_item, "actual_rolls", None) or Decimal("0.00"),
        "actual_gsm": getattr(inward_item, "actual_gsm", None),
        "actual_width": getattr(inward_item, "actual_width", None),
        "dye_lot_no": getattr(inward_item, "dye_lot_no", "") or "",
        "batch_no": getattr(inward_item, "batch_no", "") or "",
        "shade_reference": getattr(inward_item, "shade_reference", "") or "",
        "qc_status": getattr(inward_item, "qc_status", "") or "pending",
        "unit": unit,
        "remark": getattr(inward_item, "remark", "") or "",
    }


def _po_tracker_progress_label(done_count, total_count):
    if total_count and done_count >= total_count:
        return "Done"
    if done_count:
        return "Partial"
    return "Pending"

@login_required
def yarn_inward_tracker(request):
    q = (request.GET.get("q") or "").strip()
    target_inward_id = (request.GET.get("inward") or "").strip()

    qs = (
        YarnPurchaseOrder.objects
        .select_related("vendor", "firm", "owner")
        .prefetch_related(
            Prefetch(
                "items",
                queryset=YarnPurchaseOrderItem.objects.select_related("material", "material_type"),
            ),
            Prefetch(
                "inwards",
                queryset=YarnPOInward.objects.select_related("vendor", "inward_type").prefetch_related(
                    "items__po_item__material",
                    "items__po_item__material_type",
                    "generated_greige_pos__items",
                ).order_by("-inward_date", "-id"),
            ),
        )
        .filter(inwards__isnull=False)
        .distinct()
        .order_by("-id")
    )

    if not _can_review_yarn_po(request):
        qs = qs.filter(owner=request.user)

    if q:
        qs = qs.filter(
            Q(system_number__icontains=q)
            | Q(po_number__icontains=q)
            | Q(vendor__name__icontains=q)
            | Q(firm__firm_name__icontains=q)
            | Q(inwards__inward_number__icontains=q)
            | Q(items__material__name__icontains=q)
            | Q(items__material_type__name__icontains=q)
        ).distinct()

    rows = []
    for po in qs:
        po = _attach_yarn_po_metrics(po)
        inward_entries = []
        generated_count = 0

        for inward in po.inwards.all():
            linked_po = inward.generated_greige_pos.order_by("-id").first()
            if linked_po:
                generated_count += 1

            items = [_po_tracker_item_payload(inward_item, fallback_name="Yarn Item") for inward_item in inward.items.all()]
            inward_entries.append({
                "inward": inward,
                "items": items,
                "is_target": str(inward.id) == target_inward_id,
                "linked_po": linked_po,
                "next_started": bool(linked_po),
                "next_view_url": reverse("accounts:greigepo_edit", args=[linked_po.id]) if linked_po else "",
                "next_generate_url": reverse("accounts:generate_greige_po_from_yarn", args=[po.id]),
                "next_generate_method": "post",
                "next_generate_label": "Generate Greige PO",
                "next_view_label": "View Greige PO",
                "edit_url": reverse("accounts:yarn_inward_edit", args=[inward.id]),
            })

        total_inwards = len(inward_entries)
        rows.append({
            "po": po,
            "inward_entries": inward_entries,
            "total_inwards": total_inwards,
            "generated_count": generated_count,
            "progress_label": _po_tracker_progress_label(generated_count, total_inwards),
            "progress_title": "Greige Progress",
            "next_list_label": "View Greige PO List",
            "next_list_url": reverse("accounts:greigepo_list"),
            "inward_url": reverse("accounts:yarnpo_inward", args=[po.id]),
            "total_qty": _po_tracker_qty(po, "total_weight", "total_qty", "total_quantity"),
            "inward_qty": _po_tracker_qty(po, "total_inward_qty", "inward_qty_total"),
            "remaining_qty": _po_tracker_qty(po, "remaining_qty_total", "pending_qty_total"),
        })

    return render(request, "accounts/yarn_po/inward_tracker.html", {
        "rows": rows,
        "q": q,
        "target_inward_id": target_inward_id,
        "tracker_title": "Yarn Inward Tracker",
        "tracker_subtitle": "Track inwarded yarn, accepted stock, rejected qty, hold qty, and Greige PO linkage",
        "tracker_reset_url": reverse("accounts:yarn_inward_tracker"),
        "tracker_list_label": "Yarn POs",
        "tracker_list_url": reverse("accounts:yarnpo_list"),
        "tracker_stock_url": reverse("accounts:stock_lot_wise"),
        "empty_message": "No inwarded Yarn POs found yet.",
        "anchor_prefix": "yarn-inward-",
    })
def yarn_inward_tracker(request):
    q = (request.GET.get("q") or "").strip()
    target_inward_id = (request.GET.get("inward") or "").strip()

    qs = (
        YarnPurchaseOrder.objects
        .select_related("vendor", "firm", "owner")
        .prefetch_related(
            Prefetch(
                "items",
                queryset=YarnPurchaseOrderItem.objects.select_related("material", "material_type").prefetch_related("inward_items"),
            ),
            Prefetch(
                "inwards",
                queryset=YarnPOInward.objects.select_related("vendor").prefetch_related(
                    "items__po_item__material",
                    "items__po_item__material_type",
                    "generated_greige_pos__items",
                ),
            ),
        )
        .filter(inwards__isnull=False)
        .distinct()
        .order_by("-id")
    )

    if not _can_review_yarn_po(request):
        qs = qs.filter(owner=request.user)

    if q:
        qs = qs.filter(
            Q(system_number__icontains=q)
            | Q(po_number__icontains=q)
            | Q(vendor__name__icontains=q)
            | Q(firm__firm_name__icontains=q)
        ).distinct()

    rows = []

    for po in qs:
        po = _attach_yarn_po_metrics(po)

        inward_entries = []
        greige_generated_count = 0

        for inward in po.inwards.all():
            linked_greige_po = inward.generated_greige_pos.order_by("-id").first()
            if linked_greige_po:
                greige_generated_count += 1

            inward_items = []
            for inward_item in inward.items.all():
                po_item = inward_item.po_item
                inward_items.append({
                    "inward_item": inward_item,
                    "po_item": po_item,
                    "fabric_name": (
                        po_item.material.name if po_item and po_item.material
                        else (po_item.material_type.name if po_item and po_item.material_type else "Yarn Item")
                    ),
                    "ordered_qty": po_item.quantity if po_item else 0,
                    "inward_qty": inward_item.quantity,
                    "unit": po_item.unit if po_item else "",
                })

            inward_entries.append({
                "inward": inward,
                "items": inward_items,
                "is_target": str(inward.id) == target_inward_id,
                "greige_po": linked_greige_po,
                "greige_started": bool(linked_greige_po),
                "greige_items_count": linked_greige_po.items.count() if linked_greige_po else 0,
            })

        rows.append({
            "po": po,
            "inward_entries": inward_entries,
            "greige_generated_count": greige_generated_count,
            "total_inwards": len(inward_entries),
        })

    return render(
        request,
        "accounts/yarn_po/inward_tracker.html",
        {
            "rows": rows,
            "q": q,
            "target_inward_id": target_inward_id,
        },
    )


@login_required
@require_POST
def generate_greige_po_from_yarn(request, pk: int):
    yarn_po = get_object_or_404(
        YarnPurchaseOrder.objects.select_related("owner"),
        pk=pk,
    )

    if not _can_access_yarn_po(request.user, yarn_po):
        raise PermissionDenied("You do not have access to this Yarn PO.")

    inward_id = (request.POST.get("inward_id") or "").strip()
    if not inward_id:
        return redirect("accounts:yarn_inward_tracker")

    yarn_inward = get_object_or_404(
        YarnPOInward.objects.select_related("po"),
        pk=inward_id,
        po=yarn_po,
    )

    return redirect(
        f"{reverse('accounts:greigepo_add_from_yarn', kwargs={'yarn_po_id': pk})}?inward={yarn_inward.pk}"
    )


@login_required
@require_POST
def yarnpo_delete(request, pk: int):
    po = get_object_or_404(YarnPurchaseOrder.objects.select_related("owner"), pk=pk, owner=request.user)
    lock_reason = _yarn_po_delete_lock_reason(po)
    if lock_reason:
        messages.error(request, lock_reason)
        return redirect("accounts:yarnpo_list")
    po.delete()
    messages.success(request, f"Yarn PO {po.system_number} deleted successfully.")
    return redirect("accounts:yarnpo_list")
@login_required
@require_http_methods(["GET", "POST"])
def yarnpo_review(request, pk: int):
    po = get_object_or_404(
        YarnPurchaseOrder.objects
        .select_related("vendor", "firm", "reviewed_by", "owner")
        .prefetch_related(
            Prefetch("items", queryset=YarnPurchaseOrderItem.objects.select_related("material"))
        ),
        pk=pk,
    )

    if not _can_access_yarn_po(request.user, po):
        raise PermissionDenied("You do not have access to this PO.")

    embed_mode = _is_embed(request)
    can_review = _can_review_yarn_po(request)

    review_form = YarnPOReviewForm(request.POST or None)

    review_checks = {
        "has_header": bool(po.vendor_id and po.po_date),
        "has_firm": bool(po.firm_id),
        "has_items": po.items.exists(),
        "has_shipping": bool((po.shipping_address or "").strip()),
    }
    review_ready_count = sum(1 for value in review_checks.values() if value)

    context = {
        "po": po,
        "review_form": review_form,
        "can_review_yarn_po": can_review,
        "embed_mode": embed_mode,
        "review_checks": review_checks,
        "review_ready_count": review_ready_count,
    }

    if request.method == "POST":
        if not can_review:
            return HttpResponseForbidden("You are not allowed to review this PO.")

        if review_form.is_valid():
            decision = review_form.cleaned_data["decision"]

            if decision == "approve":
                po.approval_status = "approved"
                po.rejection_reason = ""
            else:
                po.approval_status = "rejected"
                po.rejection_reason = review_form.cleaned_data["rejection_reason"].strip()

            po.reviewed_by = get_actor(request) or request.user
            po.reviewed_at = timezone.now()
            po.save(update_fields=[
                "approval_status",
                "rejection_reason",
                "reviewed_by",
                "reviewed_at",
            ])

            if embed_mode:
                return JsonResponse({
                    "ok": True,
                    "message": "Yarn PO reviewed successfully.",
                    "redirect_url": reverse("accounts:yarnpo_list"),
                })

            return redirect("accounts:yarnpo_list")

        if embed_mode:
            return render(
                request,
                "accounts/yarn_po/review_embed.html",
                context,
                status=400,
            )

    template_name = (
        "accounts/yarn_po/review_embed.html"
        if embed_mode
        else "accounts/yarn_po/review.html"
    )

    return render(request, template_name, context)

@login_required
@require_POST
def firm_save(request):
    firm = Firm.objects.filter(owner=request.user).prefetch_related("addresses").first()
    if firm is None:
        firm = Firm(owner=request.user)

    form = FirmForm(request.POST, request.FILES or None, instance=firm)
    if not form.is_valid():
        first_error = _first_form_error(form)
        return JsonResponse(
            {
                "ok": False,
                "message": f"{first_error['label']}: {first_error['message']}" if first_error["label"] else first_error["message"],
                "field": first_error["field"],
                "errors": form.errors,
            },
            status=400,
        )

    try:
        addresses_payload = json.loads(request.POST.get("addresses_payload") or "[]")
    except Exception:
        return JsonResponse(
            {
                "ok": False,
                "message": "Address data is invalid.",
            },
            status=400,
        )

    if not isinstance(addresses_payload, list):
        return JsonResponse(
            {
                "ok": False,
                "message": "Address data format is invalid.",
            },
            status=400,
        )

    cleaned_addresses = []
    has_default = False

    for index, row in enumerate(addresses_payload, start=1):
        if not isinstance(row, dict):
            continue

        label = (row.get("label") or "").strip()
        address_line = (row.get("address_line") or "").strip()
        city = (row.get("city") or "").strip()
        state = (row.get("state") or "").strip()
        pincode = (row.get("pincode") or "").strip()
        is_default = bool(row.get("is_default"))

        if not address_line and not city and not state and not pincode:
            continue

        if not address_line:
            return JsonResponse(
                {
                    "ok": False,
                    "message": f"Address Row {index}: Address line is required.",
                },
                status=400,
            )

        if not city:
            return JsonResponse(
                {
                    "ok": False,
                    "message": f"Address Row {index}: City is required.",
                },
                status=400,
            )

        if not state:
            return JsonResponse(
                {
                    "ok": False,
                    "message": f"Address Row {index}: State is required.",
                },
                status=400,
            )

        if pincode and (not pincode.isdigit() or len(pincode) != 6):
            return JsonResponse(
                {
                    "ok": False,
                    "message": f"Address Row {index}: Pincode must be 6 digits.",
                },
                status=400,
            )

        if is_default:
            has_default = True

        cleaned_addresses.append({
            "label": label,
            "address_line": address_line,
            "city": city,
            "state": state,
            "pincode": pincode,
            "is_default": is_default,
        })

    if cleaned_addresses and not has_default:
        cleaned_addresses[0]["is_default"] = True

    with transaction.atomic():
        saved_firm = form.save(commit=False)
        saved_firm.owner = request.user
        saved_firm.save()

        if hasattr(saved_firm, "addresses"):
            saved_firm.addresses.all().delete()

            for idx, row in enumerate(cleaned_addresses, start=1):
                addr = FirmAddress(firm=saved_firm)

                if hasattr(addr, "label"):
                    addr.label = row["label"]
                if hasattr(addr, "address_line"):
                    addr.address_line = row["address_line"]
                if hasattr(addr, "city"):
                    addr.city = row["city"]
                if hasattr(addr, "state"):
                    addr.state = row["state"]
                if hasattr(addr, "pincode"):
                    addr.pincode = row["pincode"]
                if hasattr(addr, "is_default"):
                    addr.is_default = row["is_default"]

                if hasattr(addr, "sort_order"):
                    addr.sort_order = idx

                addr.save()

    created_at_display = ""
    if hasattr(saved_firm, "created_at") and saved_firm.created_at:
        created_at_display = timezone.localtime(saved_firm.created_at).strftime("%d %b %Y, %H:%M")

    return JsonResponse({
        "ok": True,
        "message": "Firm saved ✅",
        "firm_name": saved_firm.firm_name,
        "created_at_display": created_at_display,
        "address_count": len(cleaned_addresses),
    })
    
# =================================================================
def _jobbertype_qs_for_user(request):
    qs = JobberType.objects.all()
    if hasattr(JobberType, "owner") and request.user.is_authenticated:
        qs = qs.filter(owner=request.user)
    return qs


@login_required
def jobbertype_edit(request, pk):
    jt = get_object_or_404(_jobbertype_qs_for_user(request), pk=pk)

    if request.method == "POST":
        form = JobberTypeForm(request.POST, instance=jt, user=request.user)
        if form.is_valid():
            obj = form.save(commit=False)
            if hasattr(obj, "owner_id") and not obj.owner_id:
                obj.owner = request.user
            obj.save()

            url = reverse("accounts:jobbertype_list")
            if _is_embed(request):
                return JsonResponse({"ok": True, "url": url})
            return redirect(url)
    else:
        form = JobberTypeForm(instance=jt, user=request.user)

    template = _pick_template(
        "accounts/jobbers/jobbertype_edit_embed.html" if _is_embed(request) else "accounts/jobbers/jobbertype_edit.html",
        "accounts/jobbers/jobbertype_edit_embed.html",
    )
    return render(request, template, {
        "form": form,
        "obj": jt,
    })


@login_required
@require_POST
def jobbertype_delete(request, pk):
    jt = get_object_or_404(_jobbertype_qs_for_user(request), pk=pk)
    usage_rows = _jobbertype_usage_rows(jt)

    if usage_rows:
        error_message = (
            f'Cannot delete "{jt.name}" because it is already used in linked records. '
            f'Remove the usage or change linked records first.'
        )

        if _is_embed(request):
            context = _jobbertype_list_context(request)
            context["delete_error"] = error_message
            context["delete_label"] = jt.name
            context["delete_usage"] = usage_rows
            return render(request, "accounts/jobbers/embed_types.html", context)

        messages.error(request, error_message)
        return redirect("accounts:jobbertype_list")

    jt.delete()

    url = reverse("accounts:jobbertype_list")
    if _is_embed(request):
        return JsonResponse({"ok": True, "url": url})
    return redirect(url)


def _can_access_greige_po(user, po):
    return bool(_can_review_yarn_po(user) or po.owner_id == user.id)


def _can_access_dyeing_po(user, po):
    return bool(_can_review_yarn_po(user) or po.owner_id == user.id)

def _can_access_ready_po(user, po):
    return bool(_can_review_yarn_po(user) or po.owner_id == user.id)

def _greige_source_queryset():
    return (
        YarnPurchaseOrder.objects
        .select_related("vendor", "firm", "owner")
        .prefetch_related(
            Prefetch(
                "items",
                queryset=YarnPurchaseOrderItem.objects.select_related("material", "material_type").prefetch_related("inward_items"),
            ),
            Prefetch(
                "inwards",
                queryset=YarnPOInward.objects.prefetch_related("items__po_item__material", "items__po_item__material_type"),
            ),
        )
    )
def _attach_greige_po_metrics(po):
    total_inward = Decimal("0")
    remaining_qty = Decimal("0")

    for item in po.items.all():
        total_inward += Decimal(item.inward_qty_total or 0)
        remaining_qty += Decimal(item.remaining_qty_total or 0)

    po.metric_total_inward_qty = total_inward
    po.metric_remaining_qty_total = remaining_qty
    return po

def _greige_po_queryset():
    return (
        GreigePurchaseOrder.objects
        .select_related("vendor", "source_yarn_po", "source_yarn_inward", "owner")
        .prefetch_related(
            Prefetch(
                "items",
                queryset=GreigePurchaseOrderItem.objects.select_related("source_yarn_po_item").prefetch_related("inward_items"),
            ),
            Prefetch(
            "inwards",
            queryset=GreigePOInward.objects.select_related("vendor").prefetch_related("items__po_item"),
            ),
            Prefetch(
                "dyeing_pos",
                queryset=DyeingPurchaseOrder.objects.prefetch_related("items").order_by("-id"),
            ),
        )
        .order_by("-id")
    )
def _selected_greige_inward_total(source_greige_inward):
    if source_greige_inward is None:
        return Decimal("0")
    return source_greige_inward.items.aggregate(total=Sum("quantity")).get("total") or Decimal("0") 


def _dyeing_po_queryset():
    return (
        DyeingPurchaseOrder.objects
        .select_related("vendor", "firm", "source_greige_po", "source_greige_inward", "reviewed_by", "owner")
        .prefetch_related(
            Prefetch(
                "items",
                queryset=DyeingPurchaseOrderItem.objects.select_related("source_greige_po_item").prefetch_related("inward_items"),
            ),
            Prefetch(
                "inwards",
                queryset=DyeingPOInward.objects.prefetch_related("items__po_item"),
            ),
            Prefetch(
            "ready_pos",
            queryset=ReadyPurchaseOrder.objects.prefetch_related("items").order_by("-id"),
            ),
        )
        .order_by("-id")
    )

def _ready_po_queryset():
    return (
        ReadyPurchaseOrder.objects
        .select_related("vendor", "firm", "source_dyeing_po", "owner")
        .prefetch_related(
            Prefetch(
                "items",
                queryset=ReadyPurchaseOrderItem.objects.select_related("source_dyeing_po_item").prefetch_related("inward_items"),
            ),
            Prefetch(
                "inwards",
                queryset=ReadyPOInward.objects.prefetch_related("items__po_item"),
            ),
        )
        .order_by("-id")
    )

def _sync_greige_po_items_from_source(
    greige_po,
    source_inward=None,
    greige_material_map=None,
    greige_unit_map=None,
):
    yarn_po = (
        _greige_source_queryset()
        .filter(pk=greige_po.source_yarn_po_id)
        .first()
    )
    if yarn_po is None:
        return Decimal("0")

    greige_material_map = greige_material_map or {}
    greige_unit_map = greige_unit_map or {}

    item_rows = []
    total_weight = Decimal("0")

    if source_inward is not None:
        inward_items = source_inward.items.select_related("po_item__material", "po_item__material_type")
        for inward_item in inward_items:
            yarn_item = inward_item.po_item
            inward_qty = inward_item.quantity or Decimal("0")
            if yarn_item is None or inward_qty <= 0:
                continue

            source_yarn_name = (
                yarn_item.material.name
                if yarn_item.material
                else (yarn_item.material_type.name if yarn_item.material_type else "Yarn Item")
            )

            selected_greige_material = greige_material_map.get(yarn_item.id)
            fabric_name = (
                selected_greige_material.name
                if selected_greige_material is not None
                else source_yarn_name
            )

            unit_name = greige_unit_map.get(yarn_item.id) or yarn_item.unit or ""

            item_rows.append(
                GreigePurchaseOrderItem(
                    po=greige_po,
                    source_yarn_po_item=yarn_item,
                    fabric_name=fabric_name,
                    yarn_name=source_yarn_name,
                    unit=unit_name,
                    quantity=inward_qty,
                    remark=f"Generated from Yarn inward {source_inward.inward_number}",
                )
            )
            total_weight += inward_qty
    else:
        for yarn_item in yarn_po.items.all():
            inward_qty = yarn_item.inward_qty_total or Decimal("0")
            if inward_qty <= 0:
                continue

            source_yarn_name = (
                yarn_item.material.name
                if yarn_item.material
                else (yarn_item.material_type.name if yarn_item.material_type else "Yarn Item")
            )

            item_rows.append(
                GreigePurchaseOrderItem(
                    po=greige_po,
                    source_yarn_po_item=yarn_item,
                    fabric_name=source_yarn_name,
                    yarn_name=source_yarn_name,
                    unit=yarn_item.unit or "",
                    quantity=inward_qty,
                    remark=f"Generated from Yarn inward of {yarn_po.system_number}",
                )
            )
            total_weight += inward_qty

    GreigePurchaseOrderItem.objects.filter(po=greige_po).delete()
    if item_rows:
        GreigePurchaseOrderItem.objects.bulk_create(item_rows)

    greige_po.available_qty = total_weight
    greige_po.save(update_fields=["available_qty", "updated_at"])
    return total_weight


def _selected_yarn_inward_total(source_yarn_inward):
    if source_yarn_inward is None:
        return Decimal("0")
    return source_yarn_inward.items.aggregate(total=Sum("quantity")).get("total") or Decimal("0")


def _build_greige_material_rows(source_yarn_inward, owner, submitted_data=None, existing_po=None, row_errors=None):
    greige_qs = Material.objects.filter(material_kind="greige")
    if owner is not None and hasattr(Material, "owner"):
        greige_qs = greige_qs.filter(owner=owner)

    greige_material_options = list(
        greige_qs.select_related("material_type").order_by("name")
    )

    unit_options = list(
        MaterialUnit.objects.filter(owner=owner).order_by("name")
    ) if owner else list(MaterialUnit.objects.order_by("name"))

    material_name_to_id = {m.name: str(m.id) for m in greige_material_options}
    unit_name_to_id = {u.name: str(u.id) for u in unit_options}

    existing_by_po_item = {}
    if existing_po is not None:
        for item in existing_po.items.all():
            if item.source_yarn_po_item_id:
                existing_by_po_item[item.source_yarn_po_item_id] = item

    row_errors = row_errors or {}
    rows = []
    if source_yarn_inward is None:
        return rows, greige_material_options, unit_options

    for inward_item in source_yarn_inward.items.all():
        po_item = inward_item.po_item
        if po_item is None:
            continue

        source_name = (
            po_item.material.name
            if po_item.material
            else (po_item.material_type.name if po_item.material_type else "Yarn Item")
        )

        existing_item = existing_by_po_item.get(po_item.id)

        selected_material_id = ""
        selected_unit_id = ""

        if submitted_data is not None:
            selected_material_id = (submitted_data.get(f"greige_material_{inward_item.id}") or "").strip()
            selected_unit_id = (submitted_data.get(f"greige_unit_{inward_item.id}") or "").strip()
        elif existing_item is not None:
            selected_material_id = material_name_to_id.get(existing_item.fabric_name, "")
            selected_unit_id = unit_name_to_id.get(existing_item.unit, "")
        else:
            selected_unit_id = unit_name_to_id.get(po_item.unit or "", "")

        rows.append({
            "inward_item_id": inward_item.id,
            "po_item_id": po_item.id,
            "source_name": source_name,
            "source_unit": po_item.unit or "",
            "quantity": inward_item.quantity,
            "selected_material_id": selected_material_id,
            "selected_unit_id": selected_unit_id,
            "error": row_errors.get(inward_item.id, ""),
        })

    return rows, greige_material_options, unit_options


def _extract_greige_material_selection(request, source_yarn_inward, owner):
    greige_qs = Material.objects.filter(material_kind="greige")
    if owner is not None and hasattr(Material, "owner"):
        greige_qs = greige_qs.filter(owner=owner)

    unit_qs = MaterialUnit.objects.filter(owner=owner) if owner else MaterialUnit.objects.all()

    greige_material_map = {}
    greige_unit_map = {}
    row_errors = {}

    if source_yarn_inward is None:
        return greige_material_map, greige_unit_map, row_errors

    for inward_item in source_yarn_inward.items.all():
        po_item = inward_item.po_item
        if po_item is None:
            continue

        raw_material_id = (request.POST.get(f"greige_material_{inward_item.id}") or "").strip()
        raw_unit_id = (request.POST.get(f"greige_unit_{inward_item.id}") or "").strip()

        if not raw_material_id:
            row_errors[inward_item.id] = "Select greige name."
            continue

        greige_material = greige_qs.filter(pk=raw_material_id).first()
        if greige_material is None:
            row_errors[inward_item.id] = "Select a valid greige name."
            continue

        greige_material_map[po_item.id] = greige_material

        if raw_unit_id:
            unit_obj = unit_qs.filter(pk=raw_unit_id).first()
            if unit_obj is None:
                row_errors[inward_item.id] = "Select a valid unit."
                continue
            greige_unit_map[po_item.id] = unit_obj.name
        else:
            greige_unit_map[po_item.id] = po_item.unit or ""

    return greige_material_map, greige_unit_map, row_errors
    
def _sync_dyeing_po_items_from_source(dyeing_po):
    greige_po = (
        _greige_po_queryset()
        .filter(pk=dyeing_po.source_greige_po_id)
        .first()
    )
    if greige_po is None:
        return Decimal("0")

    item_rows = []
    total_weight = Decimal("0")
    subtotal = Decimal("0")
    source_inward = getattr(dyeing_po, "source_greige_inward", None)

    if source_inward is not None:
        inward_items = source_inward.items.select_related("po_item")
        for inward_item in inward_items:
            greige_item = inward_item.po_item
            inward_qty = inward_item.quantity or Decimal("0")
            if greige_item is None or inward_qty <= 0:
                continue

            greige_name = greige_item.fabric_name or "Greige Item"

            item_rows.append(
                DyeingPurchaseOrderItem(
                    po=dyeing_po,
                    source_greige_po_item=greige_item,
                    fabric_name=greige_name,
                    greige_name=greige_name,
                    unit=greige_item.unit or "",
                    quantity=inward_qty,
                    total_qty=inward_qty,
                    source_input_qty=inward_qty,
                    expected_loss_percent=Decimal("0"),
                    expected_output_qty=inward_qty,
                    remaining_qty=inward_qty,
                    rate=Decimal("0"),
                    other_charge_amount=Decimal("0"),
                    job_work_charges=Decimal("0"),
                    remark=f"Generated from Greige inward of {greige_po.system_number}",
                    line_final_amount=Decimal("0"),
                )
            )
            total_weight += inward_qty
    else:
        for greige_item in greige_po.items.all():
            inward_qty = greige_item.inward_qty_total or Decimal("0")
            if inward_qty <= 0:
                continue

            greige_name = greige_item.fabric_name or "Greige Item"

            item_rows.append(
                DyeingPurchaseOrderItem(
                    po=dyeing_po,
                    source_greige_po_item=greige_item,
                    fabric_name=greige_name,
                    greige_name=greige_name,
                    unit=greige_item.unit or "",
                    quantity=inward_qty,
                    total_qty=inward_qty,
                    remaining_qty=inward_qty,
                    value=Decimal("0"),
                    rolls=Decimal("0"),
                    dyeing_type="",
                    dyeing_name="",
                    rate=Decimal("0"),
                    other_charge_amount=Decimal("0"),
                    job_work_charges=Decimal("0"),
                    description="",
                    remark=f"Generated from Greige inward of {greige_po.system_number}",
                    line_subtotal=Decimal("0"),
                    line_final_amount=Decimal("0"),
                )
            )
            total_weight += inward_qty

    DyeingPurchaseOrderItem.objects.filter(po=dyeing_po).delete()
    if item_rows:
        DyeingPurchaseOrderItem.objects.bulk_create(item_rows)

    update_fields = ["total_weight", "updated_at"]
    dyeing_po.total_weight = total_weight

    if hasattr(dyeing_po, "subtotal"):
        dyeing_po.subtotal = subtotal
        update_fields.append("subtotal")

    if hasattr(dyeing_po, "after_discount_value"):
        dyeing_po.after_discount_value = subtotal
        update_fields.append("after_discount_value")

    if hasattr(dyeing_po, "final_amount"):
        dyeing_po.final_amount = subtotal
        update_fields.append("final_amount")

    if hasattr(dyeing_po, "available_qty"):
        dyeing_po.available_qty = total_weight
        update_fields.append("available_qty")

    dyeing_po.save(update_fields=update_fields)
    return total_weight



def _sync_ready_po_items_from_source(po):
    existing_items = {item.source_dyeing_po_item_id: item for item in po.items.all()}
    seen_ids = set()

    for source_item in po.source_dyeing_po.items.all():
        qty = source_item.accepted_inward_qty_total or Decimal("0")
        if qty <= 0:
            continue

        finished_material = getattr(source_item, "finished_material", None)
        fabric_name = (
            finished_material.name
            if finished_material
            else (source_item.fabric_name or "")
        )

        defaults = {
            "finished_material": finished_material,
            "fabric_name": fabric_name,
            "dyeing_name": source_item.dyeing_name or "",
            "unit": source_item.unit or "",
            "quantity": qty,
            "remark": source_item.remark or "",
        }

        existing = existing_items.get(source_item.pk)
        if existing:
            existing.finished_material = defaults["finished_material"]
            existing.fabric_name = defaults["fabric_name"]
            existing.dyeing_name = defaults["dyeing_name"]
            existing.unit = defaults["unit"]
            existing.quantity = defaults["quantity"]
            existing.remark = defaults["remark"]
            existing.save()
            seen_ids.add(existing.pk)
        else:
            created = ReadyPurchaseOrderItem.objects.create(
                po=po,
                source_dyeing_po_item=source_item,
                finished_material=defaults["finished_material"],
                fabric_name=defaults["fabric_name"],
                dyeing_name=defaults["dyeing_name"],
                unit=defaults["unit"],
                quantity=defaults["quantity"],
                remark=defaults["remark"],
            )
            seen_ids.add(created.pk)

    po.items.exclude(pk__in=seen_ids, source_dyeing_po_item__isnull=False).delete()
    _recalculate_ready_po(po)

def _recalculate_ready_po(po):
    total_weight = po.items.aggregate(total=Sum("quantity")).get("total") or Decimal("0")
    po.total_weight = total_weight
    po.available_qty = total_weight
    po.save(update_fields=["total_weight", "available_qty", "updated_at"])
    return total_weight


@login_required
def po_home(request):
    return render(request, "accounts/po/index.html")


@login_required
def greige_inward_tracker(request):
    q = (request.GET.get("q") or "").strip()
    target_inward_id = (request.GET.get("inward") or "").strip()

    qs = _greige_po_queryset().filter(inwards__isnull=False).distinct()
    if not _can_review_yarn_po(request):
        qs = qs.filter(owner=request.user)

    if q:
        qs = qs.filter(
            Q(system_number__icontains=q)
            | Q(po_number__icontains=q)
            | Q(vendor__name__icontains=q)
            | Q(source_yarn_po__system_number__icontains=q)
            | Q(source_yarn_po__firm__firm_name__icontains=q)
            | Q(source_yarn_inward__inward_number__icontains=q)
            | Q(inwards__inward_number__icontains=q)
            | Q(items__material__name__icontains=q)
            | Q(items__fabric_name__icontains=q)
            | Q(items__yarn_name__icontains=q)
            | Q(firm__firm_name__icontains=q)
        ).distinct()

    rows = []
    for po in qs:
        inward_entries = []
        generated_count = 0

        for inward in po.inwards.all():
            linked_po = inward.generated_dyeing_pos.order_by("-id").first()
            if linked_po:
                generated_count += 1

            items = [_po_tracker_item_payload(inward_item, fallback_name="Greige Item") for inward_item in inward.items.all()]
            inward_entries.append({
                "inward": inward,
                "items": items,
                "is_target": str(inward.id) == target_inward_id,
                "linked_po": linked_po,
                "next_started": bool(linked_po),
                "next_view_url": reverse("accounts:dyeingpo_edit", args=[linked_po.id]) if linked_po else "",
                "next_generate_url": reverse("accounts:generate_dyeing_po_from_greige", args=[po.id]),
                "next_generate_method": "post",
                "next_generate_label": "Generate Dyeing PO",
                "next_view_label": "View Dyeing PO",
                "edit_url": reverse("accounts:greige_inward_edit", args=[inward.id]),
            })

        total_inwards = len(inward_entries)
        rows.append({
            "po": po,
            "inward_entries": inward_entries,
            "total_inwards": total_inwards,
            "generated_count": generated_count,
            "progress_label": _po_tracker_progress_label(generated_count, total_inwards),
            "progress_title": "Dyeing Progress",
            "next_list_label": "View Dyeing PO List",
            "next_list_url": reverse("accounts:dyeingpo_list"),
            "inward_url": reverse("accounts:greigepo_inward", args=[po.id]),
            "total_qty": _po_tracker_qty(po, "total_weight", "total_qty", "total_quantity"),
            "inward_qty": _po_tracker_qty(po, "total_inward_qty", "inward_qty_total"),
            "remaining_qty": _po_tracker_qty(po, "remaining_qty_total", "pending_qty_total"),
        })

    return render(request, "accounts/greige_po/inward_tracker.html", {
        "rows": rows,
        "q": q,
        "target_inward_id": target_inward_id,
        "tracker_title": "Greige Inward Tracker",
        "tracker_subtitle": "Track inwarded greige fabric, accepted stock, rejected qty, hold qty, and Dyeing PO linkage",
        "tracker_reset_url": reverse("accounts:greige_inward_tracker"),
        "tracker_list_label": "Greige POs",
        "tracker_list_url": reverse("accounts:greigepo_list"),
        "tracker_stock_url": reverse("accounts:stock_lot_wise"),
        "empty_message": "No inwarded Greige POs found yet.",
        "anchor_prefix": "greige-inward-",
    })
def greige_inward_tracker(request):
    q = (request.GET.get("q") or "").strip()
    target_inward_id = (request.GET.get("inward") or "").strip()

    qs = _greige_po_queryset().filter(inwards__isnull=False).distinct()

    if not _can_review_yarn_po(request):
        qs = qs.filter(owner=request.user)

    if q:
        qs = qs.filter(
            Q(system_number__icontains=q)
            | Q(po_number__icontains=q)
            | Q(vendor__name__icontains=q)
            | Q(source_yarn_po__system_number__icontains=q)
            | Q(source_yarn_po__firm__firm_name__icontains=q)
            | Q(source_yarn_inward__inward_number__icontains=q)
            | Q(inwards__inward_number__icontains=q)
            | Q(items__material__name__icontains=q)
            | Q(items__fabric_name__icontains=q)
            | Q(items__yarn_name__icontains=q)
        ).distinct()

    rows = []
    for po in qs:
        inward_entries = []
        generated_count = 0

        for inward in po.inwards.all():
            items = []
            dyeing_po = inward.generated_dyeing_pos.order_by("-id").first()
            if dyeing_po:
                generated_count += 1

            for inward_item in inward.items.select_related("po_item__material").all():
                po_item = inward_item.po_item
                if po_item is None:
                    continue

                items.append({
                    "fabric_name": (
                        po_item.material.name
                        if getattr(po_item, "material", None)
                        else (po_item.fabric_name or "Greige Item")
                    ),
                    "ordered_qty": po_item.quantity or Decimal("0"),
                    "inward_qty": inward_item.quantity or Decimal("0"),
                    "unit": po_item.unit or "",
                })

            inward_entries.append({
                "inward": inward,
                "items": items,
                "dyeing_po": dyeing_po,
                "is_target": str(inward.id) == target_inward_id,
            })

        rows.append({
            "po": po,
            "inward_entries": inward_entries,
            "total_inwards": len(inward_entries),
            "dyeing_generated_count": generated_count,
        })

    return render(
        request,
        "accounts/greige_po/inward_tracker.html",
        {
            "rows": rows,
            "q": q,
            "target_inward_id": target_inward_id,
        },
    )

@login_required
@require_http_methods(["GET", "POST"])
def greige_inward_edit(request, pk: int):
    inward = get_object_or_404(
        GreigePOInward.objects.select_related("po__vendor", "po__firm", "po__owner"),
        pk=pk,
    )
    po = inward.po

    if not _can_access_greige_po(request.user, po):
        raise PermissionDenied("You do not have access to this Greige inward.")

    if not _is_po_approved_for_inward(po):
        messages.error(request, "Greige PO must be approved before inward can be updated.")
        return redirect("accounts:greige_inward_tracker")

    if inward.generated_dyeing_pos.exists():
        messages.error(request, "This inward cannot be edited because a Dyeing PO has already been generated from it.")
        tracker_url = reverse("accounts:greige_inward_tracker")
        return redirect(f"{tracker_url}?inward={inward.pk}")

    item_errors = {}
    line_inputs = {
        row.po_item_id: {
            "qty": str(row.received_qty or row.quantity or ""),
            "accepted_qty": str(row.accepted_qty or ""),
            "rejected_qty": str(row.rejected_qty or ""),
            "hold_qty": str(row.hold_qty or ""),
            "actual_rolls": str(row.actual_rolls or ""),
            "actual_gsm": str(row.actual_gsm or ""),
            "actual_width": str(row.actual_width or ""),
            "dye_lot_no": row.dye_lot_no or "",
            "batch_no": row.batch_no or "",
            "shade_reference": row.shade_reference or "",
            "remark": row.remark or "",
        }
        for row in inward.items.all()
    }

    inward_form = GreigePOInwardForm(request.POST or None, instance=inward, user=request.user)

    if request.method == "POST" and inward_form.is_valid():
        line_payload, line_inputs, item_errors = _collect_basic_po_inward_payload(request, po, editing_inward=inward)

        if not line_payload:
            inward_form.add_error(None, "Enter at least one inward quantity.")

        if not inward_form.errors and not item_errors:
            inward = inward_form.save(commit=False)
            inward.owner = po.owner
            inward.po = po
            inward.save()

            inward.items.all().delete()

            GreigePOInwardItem.objects.bulk_create([
                GreigePOInwardItem(
                    inward=inward,
                    po_item=row["item"],
                    quantity=row["quantity"],
                    received_qty=row["received_qty"],
                    accepted_qty=row["accepted_qty"],
                    rejected_qty=row["rejected_qty"],
                    hold_qty=row["hold_qty"],
                    actual_rolls=row["actual_rolls"],
                    actual_gsm=row["actual_gsm"],
                    actual_width=row["actual_width"],
                    dye_lot_no=row["dye_lot_no"],
                    batch_no=row["batch_no"],
                    shade_reference=row["shade_reference"],
                    qc_status=row["qc_status"],
                    remark=row["remark"],
                )
                for row in line_payload
            ])

            messages.success(request, f"Inward {inward.inward_number} updated successfully.")
            tracker_url = reverse("accounts:greige_inward_tracker")
            return redirect(f"{tracker_url}?inward={inward.pk}")

    line_rows = _build_dyeing_inward_line_rows(po, line_inputs=line_inputs, item_errors=item_errors)


    return render(
        request,
        "accounts/greige_po/inward.html",
        {
            "po": po,
            "inward_form": inward_form,
            "line_rows": line_rows,
            "existing_inwards": po.inwards.all().order_by("-inward_date", "-id"),
            "editing_inward": inward,
            "next_inward_number_preview": inward.inward_number,
        },
    )

@login_required
def greigepo_list(request):
    q = (request.GET.get("q") or "").strip()

    qs = _greige_po_queryset()

    if not _can_review_yarn_po(request):
        qs = qs.filter(owner=request.user)

    if q:
        qs = qs.filter(
            Q(system_number__icontains=q)
            | Q(po_number__icontains=q)
            | Q(internal_po_number__icontains=q)
            | Q(vendor__name__icontains=q)
            | Q(firm__firm_name__icontains=q)
            | Q(source_yarn_po__system_number__icontains=q)
            | Q(source_yarn_inward__inward_number__icontains=q)
            | Q(items__material__name__icontains=q)
            | Q(items__fabric_name__icontains=q)
            | Q(items__yarn_name__icontains=q)
        ).distinct()

    orders = [_attach_greige_po_metrics(po) for po in qs.order_by("-id")]

    return render(
        request,
        "accounts/greige_po/list.html",
        {
            "orders": orders,
            "q": q,
            "can_review_greige_po": _can_review_yarn_po(request),
        },
    )

@login_required
@require_http_methods(["GET", "POST"])
def greigepo_create(request, yarn_po_id=None):
    source_yarn_po = None
    selected_source_inward = None

    if yarn_po_id is not None:
        source_yarn_po = get_object_or_404(_greige_source_queryset(), pk=yarn_po_id)
        if not _can_access_yarn_po(request.user, source_yarn_po):
            raise PermissionDenied("You do not have access to this Yarn PO.")

    inward_id = (request.POST.get("source_yarn_inward_id") or request.GET.get("inward") or "").strip()
    if source_yarn_po is not None and inward_id:
        selected_source_inward = get_object_or_404(
            source_yarn_po.inwards.prefetch_related("items__po_item__material", "items__po_item__material_type"),
            pk=inward_id,
        )

    owner_for_forms = source_yarn_po.owner if source_yarn_po is not None else request.user

    temp_po = GreigePurchaseOrder(
        source_yarn_po=source_yarn_po,
        source_yarn_inward=selected_source_inward,
    )

    initial = {}

    if request.method == "POST":
        form = GreigePurchaseOrderForm(
            request.POST,
            user=owner_for_forms,
            source_yarn_po=source_yarn_po,
            lock_source=bool(source_yarn_po),
        )
        formset = GreigePurchaseOrderItemFormSet(
            request.POST,
            instance=temp_po,
            prefix="items",
            form_kwargs={"user": owner_for_forms},
        )

        if form.is_valid() and formset.is_valid():
            selected_source = source_yarn_po or form.cleaned_data["source_yarn_po"]

            if not _can_access_yarn_po(request.user, selected_source):
                raise PermissionDenied("You do not have access to this Yarn PO.")

            if selected_source_inward is None:
                form.add_error(None, "Select a specific Yarn inward before generating Greige PO.")
            elif selected_source_inward.po_id != selected_source.id:
                form.add_error(None, "Selected Yarn inward does not belong to the chosen Yarn PO.")
            elif selected_source_inward.generated_greige_pos.exists():
                form.add_error(None, "Greige PO already exists for this Yarn inward.")
            else:
                with transaction.atomic():
                    greige_po = form.save(commit=False)
                    greige_po.owner = selected_source.owner
                    greige_po.system_number = _next_greige_po_number()
                    greige_po.source_yarn_po = selected_source
                    greige_po.source_yarn_inward = selected_source_inward

                    if not greige_po.firm_id and getattr(selected_source, "firm_id", None):
                        greige_po.firm = selected_source.firm

                    if not greige_po.shipping_address and greige_po.firm_id:
                        greige_po.shipping_address = _firm_address(greige_po.firm)

                    if not greige_po.terms_conditions:
                        greige_po.terms_conditions = getattr(selected_source, "terms_conditions", "") or ""

                    greige_po.save()

                    formset.instance = greige_po
                    items = formset.save(commit=False)

                    for obj in formset.deleted_objects:
                        obj.delete()

                    for item in items:
                        item.po = greige_po
                        item.save()

                    formset.save_m2m()

                    _apply_greige_source_links(greige_po)
                    _recalculate_greige_po(greige_po)

                messages.success(request, f"Greige PO {greige_po.system_number} saved successfully.")
                return redirect("accounts:greigepo_inward", pk=greige_po.pk)

        form.add_error(None, "Greige PO was not saved. Please fix the highlighted fields and try again.")

    else:
        if source_yarn_po is not None:
            source_total = (
                _selected_yarn_inward_total(selected_source_inward)
                if selected_source_inward is not None
                else sum((item.inward_qty_total or Decimal("0")) for item in source_yarn_po.items.all())
            )

            initial = {
                "po_number": source_yarn_po.po_number or "",
                "po_date": timezone.localdate(),
                "available_qty": source_total,
                "vendor": source_yarn_po.vendor_id,
                "firm": source_yarn_po.firm_id,
                "shipping_address": _firm_address(source_yarn_po.firm) if source_yarn_po.firm else "",
                "terms_conditions": getattr(source_yarn_po, "terms_conditions", "") or "",
            }

        form = GreigePurchaseOrderForm(
            initial=initial,
            user=owner_for_forms,
            source_yarn_po=source_yarn_po,
            lock_source=bool(source_yarn_po),
        )
        formset = GreigePurchaseOrderItemFormSet(
            instance=temp_po,
            prefix="items",
            form_kwargs={"user": owner_for_forms},
        )

    source_inwards = [selected_source_inward] if selected_source_inward else (list(source_yarn_po.inwards.all()) if source_yarn_po else [])
    effective_owner = source_yarn_po.owner if source_yarn_po else request.user

    terms_condition_options = _greige_terms_condition_options(effective_owner)

    selected_terms_condition_id = (request.POST.get("terms_condition_id") or "").strip() if request.method == "POST" else ""
    greige_description = (request.POST.get("greige_description") or "").strip() if request.method == "POST" else ""

    if request.method != "POST":
        greige_total_weight = str(initial.get("available_qty", "0.00"))
    else:
        greige_total_weight = request.POST.get("greige_total_weight") or "0.00"

    greige_subtotal = request.POST.get("greige_subtotal") or "0.00"
    greige_discount_percent = request.POST.get("greige_discount_percent") or "0"
    greige_after_discount = request.POST.get("greige_after_discount") or "0.00"
    greige_others = request.POST.get("greige_others") or "0"
    greige_cgst = request.POST.get("greige_cgst") or "2.5"
    greige_sgst = request.POST.get("greige_sgst") or "2.5"

    firm_shipping_map = {}
    for firm in form.fields["firm"].queryset:
        addresses = []

        if hasattr(firm, "addresses"):
            for address in firm.addresses.order_by("-is_default", "id"):
                addresses.append({
                    "id": str(address.pk),
                    "label": f"{address.label or f'Address {address.pk}'}{' (Default)' if address.is_default else ''}",
                    "value": address.full_address,
                })

        if not addresses and getattr(firm, "full_address", ""):
            addresses.append({
                "id": "firm-default",
                "label": f"{firm.firm_name} (Default)",
                "value": firm.full_address,
            })

        firm_shipping_map[str(firm.pk)] = addresses

    terms_condition_map = {
        str(obj.pk): obj.content
        for obj in form.fields["terms_template"].queryset
    }

    return render(
        request,
        "accounts/greige_po/form.html",
        {
            "form": form,
            "formset": formset,
            "mode": "add",
            "po_obj": None,
            "system_number_preview": _next_greige_po_number(),
            "source_yarn_po": source_yarn_po,
            "selected_source_inward": selected_source_inward,
            "source_inwards": source_inwards,
            "terms_condition_options": terms_condition_options,
            "terms_condition_map": terms_condition_map,
            "selected_terms_condition_id": selected_terms_condition_id,
            "greige_description": greige_description,
            "greige_total_weight": greige_total_weight,
            "greige_subtotal": greige_subtotal,
            "greige_discount_percent": greige_discount_percent,
            "greige_after_discount": greige_after_discount,
            "greige_others": greige_others,
            "greige_cgst": greige_cgst,
            "greige_sgst": greige_sgst,
            "shipping_address_map": getattr(form, "shipping_address_map", {}),
            "firm_shipping_map": firm_shipping_map,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def greigepo_update(request, pk: int):
    po = get_object_or_404(_greige_po_queryset(), pk=pk)
    if not _can_access_greige_po(request.user, po):
        raise PermissionDenied("You do not have access to this Greige PO.")

    lock_reason = _greige_po_lock_reason(po)
    if lock_reason:
        messages.error(request, lock_reason)
        return redirect("accounts:greigepo_list")

    owner_for_forms = po.owner

    form = GreigePurchaseOrderForm(
        request.POST or None,
        user=owner_for_forms,
        instance=po,
        source_yarn_po=po.source_yarn_po,
        lock_source=True,
    )

    formset = GreigePurchaseOrderItemFormSet(
        request.POST or None,
        instance=po,
        prefix="items",
        form_kwargs={"user": owner_for_forms},
    )

    if request.method == "POST":
        if form.is_valid() and formset.is_valid():
            with transaction.atomic():
                po = form.save(commit=False)

                if not po.firm_id and po.source_yarn_po and po.source_yarn_po.firm_id:
                    po.firm = po.source_yarn_po.firm

                if not po.shipping_address and po.firm_id:
                    po.shipping_address = _firm_address(po.firm)

                if not po.terms_conditions and po.source_yarn_po:
                    po.terms_conditions = getattr(po.source_yarn_po, "terms_conditions", "") or ""

                po.save()

                formset.instance = po
                items = formset.save(commit=False)

                for obj in formset.deleted_objects:
                    obj.delete()

                for item in items:
                    item.po = po
                    item.save()

                formset.save_m2m()

                _apply_greige_source_links(po)
                _recalculate_greige_po(po)

            messages.success(request, f"Greige PO {po.system_number} updated successfully.")
            return redirect("accounts:greigepo_inward", pk=po.pk)

        form.add_error(None, "Greige PO was not updated. Please fix the highlighted fields and try again.")

    source_inwards = [po.source_yarn_inward] if po.source_yarn_inward else (list(po.source_yarn_po.inwards.all()) if po.source_yarn_po else [])
    effective_owner = po.owner

    terms_condition_options = _greige_terms_condition_options(effective_owner)

    selected_terms_condition_id = (request.POST.get("terms_condition_id") or "").strip()
    greige_description = (request.POST.get("greige_description") or "").strip()
    greige_total_weight = request.POST.get("greige_total_weight") or str(po.available_qty or "0.00")
    greige_subtotal = request.POST.get("greige_subtotal") or "0.00"
    greige_discount_percent = request.POST.get("greige_discount_percent") or "0"
    greige_after_discount = request.POST.get("greige_after_discount") or "0.00"
    greige_others = request.POST.get("greige_others") or "0"
    greige_cgst = request.POST.get("greige_cgst") or "2.5"
    greige_sgst = request.POST.get("greige_sgst") or "2.5"

    firm_shipping_map = {}
    for firm in form.fields["firm"].queryset:
        addresses = []

        if hasattr(firm, "addresses"):
            for address in firm.addresses.order_by("-is_default", "id"):
                addresses.append({
                    "id": str(address.pk),
                    "label": f"{address.label or f'Address {address.pk}'}{' (Default)' if address.is_default else ''}",
                    "value": address.full_address,
                })

        if not addresses and getattr(firm, "full_address", ""):
            addresses.append({
                "id": "firm-default",
                "label": f"{firm.firm_name} (Default)",
                "value": firm.full_address,
            })

        firm_shipping_map[str(firm.pk)] = addresses

    terms_condition_map = {
        str(obj.pk): obj.content
        for obj in form.fields["terms_template"].queryset
    }

    return render(
        request,
        "accounts/greige_po/form.html",
        {
            "form": form,
            "formset": formset,
            "mode": "edit",
            "po_obj": po,
            "system_number_preview": po.system_number,
            "source_yarn_po": po.source_yarn_po,
            "selected_source_inward": po.source_yarn_inward,
            "source_inwards": source_inwards,
            "terms_condition_options": terms_condition_options,
            "terms_condition_map": terms_condition_map,
            "selected_terms_condition_id": selected_terms_condition_id,
            "greige_description": greige_description,
            "greige_total_weight": greige_total_weight,
            "greige_subtotal": greige_subtotal,
            "greige_discount_percent": greige_discount_percent,
            "greige_after_discount": greige_after_discount,
            "greige_others": greige_others,
            "greige_cgst": greige_cgst,
            "greige_sgst": greige_sgst,
            "shipping_address_map": getattr(form, "shipping_address_map", {}),
            "firm_shipping_map": firm_shipping_map,
        },
    )

@login_required
def greigepo_detail(request, pk: int):
    po = get_object_or_404(_greige_po_queryset(), pk=pk)
    if not _can_access_greige_po(request.user, po):
        raise PermissionDenied("You do not have access to this Greige PO.")

    return render(
        request,
        "accounts/greige_po/detail.html",
        {
            "po": po,
            "source_yarn_po": po.source_yarn_po,
            "source_inwards": list(po.source_yarn_po.inwards.all()) if po.source_yarn_po else [],
            "existing_dyeing_po": po.dyeing_pos.first(),
            "greige_inwards": list(po.inwards.all()),
        },
    )


@login_required
@require_POST
def greigepo_delete(request, pk: int):
    po = get_object_or_404(
        GreigePurchaseOrder.objects.select_related("owner").prefetch_related("inwards", "dyeing_pos"),
        pk=pk,
    )

    if not _can_access_greige_po(request.user, po):
        raise PermissionDenied("You do not have access to this Greige PO.")

    lock_reason = _greige_po_delete_lock_reason(po)
    if lock_reason:
        messages.error(request, lock_reason)
        return redirect("accounts:greigepo_list")

    po.delete()
    messages.success(request, f"Greige PO {po.system_number} deleted successfully.")
    return redirect("accounts:greigepo_list")


@login_required
@require_http_methods(["GET", "POST"])
def greigepo_inward(request, pk: int):
    po = get_object_or_404(_greige_po_queryset(), pk=pk)
    if not _can_access_greige_po(request.user, po):
        raise PermissionDenied("You do not have access to this Greige PO.")
    if not _is_po_approved_for_inward(po):
        messages.error(request, "Greige PO must be approved before inward can be generated.")
        return redirect("accounts:greigepo_list")

    item_errors = {}
    line_inputs = {}
    inward_form = GreigePOInwardForm(request.POST or None, user=po.owner)

    if request.method == "GET" and po.vendor_id and "vendor" in inward_form.fields:
        inward_form.fields["vendor"].initial = po.vendor_id

    if request.method == "POST" and inward_form.is_valid():
        line_payload, line_inputs, item_errors = _collect_basic_po_inward_payload(request, po)

        if not line_payload:
            inward_form.add_error(None, "Enter at least one inward quantity.")

        if not inward_form.errors and not item_errors:
            inward = inward_form.save(commit=False)
            inward.owner = po.owner
            inward.po = po
            inward.inward_number = _next_greige_inward_number()

            if not inward.vendor_id and po.vendor_id:
                inward.vendor = po.vendor

            inward.save()

            GreigePOInwardItem.objects.bulk_create([
                GreigePOInwardItem(
                    inward=inward,
                    po_item=row["item"],
                    quantity=row["quantity"],
                    received_qty=row["received_qty"],
                    accepted_qty=row["accepted_qty"],
                    rejected_qty=row["rejected_qty"],
                    hold_qty=row["hold_qty"],
                    actual_rolls=row["actual_rolls"],
                    actual_gsm=row["actual_gsm"],
                    actual_width=row["actual_width"],
                    dye_lot_no=row["dye_lot_no"],
                    batch_no=row["batch_no"],
                    shade_reference=row["shade_reference"],
                    qc_status=row["qc_status"],
                    remark=row["remark"],
                )
                for row in line_payload
            ])

            tracker_url = reverse("accounts:greige_inward_tracker")
            return redirect(f"{tracker_url}?inward={inward.pk}")

    line_rows = _build_dyeing_inward_line_rows(po, line_inputs=line_inputs, item_errors=item_errors)


    return render(
        request,
        "accounts/greige_po/inward.html",
        {
            "po": po,
            "inward_form": inward_form,
            "line_rows": line_rows,
            "existing_inwards": po.inwards.all().order_by("-inward_date", "-id"),
            "next_inward_number_preview": _next_greige_inward_number(),
        },
    )


@login_required
def greige_inward_tracker(request):
    q = (request.GET.get("q") or "").strip()
    target_inward_id = (request.GET.get("inward") or "").strip()

    qs = _greige_po_queryset().filter(inwards__isnull=False).distinct()
    if not _can_review_yarn_po(request):
        qs = qs.filter(owner=request.user)

    if q:
        qs = qs.filter(
            Q(system_number__icontains=q)
            | Q(po_number__icontains=q)
            | Q(vendor__name__icontains=q)
            | Q(source_yarn_po__system_number__icontains=q)
            | Q(source_yarn_po__firm__firm_name__icontains=q)
            | Q(source_yarn_inward__inward_number__icontains=q)
            | Q(inwards__inward_number__icontains=q)
            | Q(items__material__name__icontains=q)
            | Q(items__fabric_name__icontains=q)
            | Q(items__yarn_name__icontains=q)
            | Q(firm__firm_name__icontains=q)
        ).distinct()

    rows = []
    for po in qs:
        inward_entries = []
        generated_count = 0

        for inward in po.inwards.all():
            linked_po = inward.generated_dyeing_pos.order_by("-id").first()
            if linked_po:
                generated_count += 1

            items = [_po_tracker_item_payload(inward_item, fallback_name="Greige Item") for inward_item in inward.items.all()]
            inward_entries.append({
                "inward": inward,
                "items": items,
                "is_target": str(inward.id) == target_inward_id,
                "linked_po": linked_po,
                "next_started": bool(linked_po),
                "next_view_url": reverse("accounts:dyeingpo_edit", args=[linked_po.id]) if linked_po else "",
                "next_generate_url": reverse("accounts:generate_dyeing_po_from_greige", args=[po.id]),
                "next_generate_method": "post",
                "next_generate_label": "Generate Dyeing PO",
                "next_view_label": "View Dyeing PO",
                "edit_url": reverse("accounts:greige_inward_edit", args=[inward.id]),
            })

        total_inwards = len(inward_entries)
        rows.append({
            "po": po,
            "inward_entries": inward_entries,
            "total_inwards": total_inwards,
            "generated_count": generated_count,
            "progress_label": _po_tracker_progress_label(generated_count, total_inwards),
            "progress_title": "Dyeing Progress",
            "next_list_label": "View Dyeing PO List",
            "next_list_url": reverse("accounts:dyeingpo_list"),
            "inward_url": reverse("accounts:greigepo_inward", args=[po.id]),
            "total_qty": _po_tracker_qty(po, "total_weight", "total_qty", "total_quantity"),
            "inward_qty": _po_tracker_qty(po, "total_inward_qty", "inward_qty_total"),
            "remaining_qty": _po_tracker_qty(po, "remaining_qty_total", "pending_qty_total"),
        })

    return render(request, "accounts/greige_po/inward_tracker.html", {
        "rows": rows,
        "q": q,
        "target_inward_id": target_inward_id,
        "tracker_title": "Greige Inward Tracker",
        "tracker_subtitle": "Track inwarded greige fabric, accepted stock, rejected qty, hold qty, and Dyeing PO linkage",
        "tracker_reset_url": reverse("accounts:greige_inward_tracker"),
        "tracker_list_label": "Greige POs",
        "tracker_list_url": reverse("accounts:greigepo_list"),
        "tracker_stock_url": reverse("accounts:stock_lot_wise"),
        "empty_message": "No inwarded Greige POs found yet.",
        "anchor_prefix": "greige-inward-",
    })
def greige_inward_tracker(request):
    q = (request.GET.get("q") or "").strip()
    target_inward_id = (request.GET.get("inward") or "").strip()

    qs = _greige_po_queryset().filter(inwards__isnull=False).distinct()
    if not _can_review_yarn_po(request):
        qs = qs.filter(owner=request.user)

    if q:
        qs = qs.filter(
            Q(system_number__icontains=q)
            | Q(po_number__icontains=q)
            | Q(vendor__name__icontains=q)
            | Q(source_yarn_po__system_number__icontains=q)
            | Q(source_yarn_po__firm__firm_name__icontains=q)
        ).distinct()

    rows = []
    for po in qs:
        inward_entries = []
        dyeing_generated_count = 0

        for inward in po.inwards.all():
            linked_dyeing_po = inward.generated_dyeing_pos.order_by("-id").first()
            if linked_dyeing_po:
                dyeing_generated_count += 1

            inward_entries.append({
                "inward": inward,
                "items": [
                    {
                        "inward_item": inward_item,
                        "po_item": inward_item.po_item,
                        "fabric_name": inward_item.po_item.fabric_name if inward_item.po_item else "Greige Item",
                        "ordered_qty": inward_item.po_item.quantity if inward_item.po_item else 0,
                        "inward_qty": inward_item.quantity,
                        "unit": inward_item.po_item.unit if inward_item.po_item else "",
                    }
                    for inward_item in inward.items.all()
                ],
                "is_target": str(inward.id) == target_inward_id,
                "dyeing_po": linked_dyeing_po,
                "dyeing_started": bool(linked_dyeing_po),
                "dyeing_items_count": linked_dyeing_po.items.count() if linked_dyeing_po else 0,
            })

        rows.append({
            "po": po,
            "inward_entries": inward_entries,
            "dyeing_generated_count": dyeing_generated_count,
            "total_inwards": len(inward_entries),
        })

    return render(
        request,
        "accounts/greige_po/inward_tracker.html",
        {
            "rows": rows,
            "q": q,
            "target_inward_id": target_inward_id,
        },
    )


@login_required
@require_POST
def generate_dyeing_po_from_greige(request, pk: int):
    greige_po = get_object_or_404(_greige_po_queryset(), pk=pk)
    if not _can_access_greige_po(request.user, greige_po):
        raise PermissionDenied("You do not have access to this Greige PO.")

    inward_id = (request.POST.get("inward_id") or "").strip()
    if not inward_id:
        return redirect("accounts:greige_inward_tracker")

    greige_inward = get_object_or_404(
        GreigePOInward.objects.select_related("po"),
        pk=inward_id,
        po=greige_po,
    )

    return redirect(
        f"{reverse('accounts:dyeingpo_add_from_greige', kwargs={'greige_po_id': pk})}?inward={greige_inward.pk}"
    )

@login_required
def generate_ready_po_from_dyeing(request, pk: int):
    dyeing_po = get_object_or_404(_dyeing_po_queryset(), pk=pk)
    if not _can_access_dyeing_po(request.user, dyeing_po):
        raise PermissionDenied("You do not have access to this Dyeing PO.")
    if not _is_po_approved_for_inward(dyeing_po):
        messages.error(request, "Dyeing PO must be approved before Ready PO can be generated.")
        return redirect("accounts:dyeingpo_list")
    if not dyeing_po.inwards.exists():
        messages.error(request, "Create at least one Dyeing inward before generating Ready PO.")
        return redirect("accounts:dyeingpo_inward", pk=dyeing_po.pk)
    if not any((item.accepted_inward_qty_total or Decimal("0")) > 0 for item in dyeing_po.items.all()):
        messages.error(request, "Accepted dyed qty is required before Ready PO can be generated.")
        return redirect("accounts:dyeingpo_inward", pk=dyeing_po.pk)   
    
    return redirect("accounts:readypo_add_from_dyeing", dyeing_po_id=pk)


def _build_simple_dyeing_po_pdf_response(po):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas
    except ImportError:
        return HttpResponse(
            "ReportLab is required for PDF generation. Install it with: pip install reportlab",
            status=500,
        )

    def safe_text(value):
        value = "" if value is None else str(value).strip()
        return value if value else "-"

    def qty_text(value):
        try:
            return f"{float(value or 0):,.2f}".rstrip("0").rstrip(".")
        except Exception:
            return safe_text(value)

    def money_text(value):
        try:
            return f"{float(value or 0):,.2f}"
        except Exception:
            return safe_text(value)

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    left = 15 * mm
    top = height - 18 * mm
    y = top

    firm = getattr(po, "firm", None)
    vendor = getattr(po, "vendor", None)

    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(left, y, safe_text(firm.firm_name if firm else "InventTech"))
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawRightString(width - 15 * mm, y, "DYEING PURCHASE ORDER")

    y -= 9 * mm
    pdf.setFont("Helvetica", 9)
    pdf.drawString(left, y, f"PO No: {safe_text(po.po_number or po.system_number)}")
    pdf.drawString(left + 70 * mm, y, f"Date: {po.po_date.strftime('%d-%m-%Y') if po.po_date else '-'}")
    pdf.drawRightString(width - 15 * mm, y, f"Status: {safe_text(po.get_approval_status_display())}")
    y -= 6 * mm
    pdf.drawString(left, y, f"System No: {safe_text(po.system_number)}")
    if getattr(po, "source_greige_po", None):
        pdf.drawString(left + 70 * mm, y, f"Greige PO: {safe_text(po.source_greige_po.system_number)}")

    y -= 12 * mm
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(left, y, "Vendor")
    pdf.drawString(width / 2, y, "Bill To / Ship To")
    y -= 6 * mm
    pdf.setFont("Helvetica", 9)

    vendor_lines = [
        safe_text(vendor.name if vendor else ""),
        f"Phone: {safe_text(getattr(vendor, 'phone', ''))}",
        f"Email: {safe_text(getattr(vendor, 'email', ''))}",
        f"GSTIN: {safe_text(getattr(vendor, 'gst_number', ''))}",
        f"Address: {safe_text(getattr(vendor, 'address', ''))}",
    ]
    bill_lines = [
        safe_text(firm.firm_name if firm else ""),
        f"Phone: {safe_text(getattr(firm, 'phone', ''))}",
        f"Email: {safe_text(getattr(firm, 'email', ''))}",
        f"GSTIN: {safe_text(getattr(firm, 'gst_number', ''))}",
        f"Ship To: {safe_text(po.shipping_address)}",
    ]
    for idx in range(max(len(vendor_lines), len(bill_lines))):
        pdf.drawString(left, y, vendor_lines[idx] if idx < len(vendor_lines) else "")
        pdf.drawString(width / 2, y, bill_lines[idx] if idx < len(bill_lines) else "")
        y -= 5 * mm

    y -= 4 * mm
    pdf.setFont("Helvetica-Bold", 8.5)
    pdf.drawString(left, y, "Sr")
    pdf.drawString(left + 10 * mm, y, "Fabric / Greige")
    pdf.drawString(left + 62 * mm, y, "Dyeing")
    pdf.drawString(left + 100 * mm, y, "Unit")
    pdf.drawRightString(left + 126 * mm, y, "Qty")
    pdf.drawRightString(left + 153 * mm, y, "Rate")
    pdf.drawRightString(width - 15 * mm, y, "Amount")
    y -= 4 * mm
    pdf.line(left, y, width - 15 * mm, y)
    y -= 6 * mm

    total_qty = Decimal("0")
    total_amount = Decimal("0")
    for index, item in enumerate(po.items.all(), start=1):
        if y < 45 * mm:
            pdf.showPage()
            y = height - 20 * mm
        qty = getattr(item, "quantity", 0) or getattr(item, "total_qty", 0) or 0
        amount = getattr(item, "line_final_amount", 0) or getattr(item, "line_subtotal", 0) or 0
        total_qty += Decimal(qty or 0)
        total_amount += Decimal(amount or 0)
        fabric = item.fabric_name or getattr(getattr(item, "finished_material", None), "name", "") or "-"
        greige = item.greige_name or ""
        desc = fabric if not greige else f"{fabric} / {greige}"
        dyeing = item.dyeing_name or item.dyeing_type or "-"
        pdf.setFont("Helvetica", 8)
        pdf.drawString(left, y, str(index))
        pdf.drawString(left + 10 * mm, y, safe_text(desc)[:31])
        pdf.drawString(left + 62 * mm, y, safe_text(dyeing)[:23])
        pdf.drawString(left + 100 * mm, y, safe_text(item.unit)[:8])
        pdf.drawRightString(left + 126 * mm, y, qty_text(qty))
        pdf.drawRightString(left + 153 * mm, y, money_text(item.rate))
        pdf.drawRightString(width - 15 * mm, y, money_text(amount))
        y -= 6 * mm

    y -= 4 * mm
    pdf.line(left, y, width - 15 * mm, y)
    y -= 8 * mm
    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawString(left, y, f"Total Qty: {qty_text(po.total_weight or total_qty)}")
    pdf.drawRightString(left + 153 * mm, y, "Total Amount")
    pdf.drawRightString(width - 15 * mm, y, money_text(po.final_amount or total_amount))
    y -= 12 * mm
    pdf.setFont("Helvetica", 8.5)
    pdf.drawString(left, y, "THIS PO IS COMPUTER GENERATED, HENCE SIGNATURE IS NOT REQUIRED")
    pdf.save()
    buffer.seek(0)
    return HttpResponse(buffer.getvalue(), content_type="application/pdf")


@login_required
def dyeingpo_pdf(request, pk: int):
    po = get_object_or_404(
        DyeingPurchaseOrder.objects
        .select_related("vendor", "firm", "source_greige_po", "source_greige_inward", "owner")
        .prefetch_related(
            Prefetch(
                "items",
                queryset=DyeingPurchaseOrderItem.objects.select_related("finished_material", "source_greige_po_item"),
            )
        ),
        pk=pk,
    )

    if not _can_access_dyeing_po(request.user, po):
        raise PermissionDenied("You do not have access to this Dyeing PO.")

    response = _build_simple_dyeing_po_pdf_response(po)
    if response.status_code == 200 and response.get("Content-Type", "").startswith("application/pdf"):
        filename = f'{po.system_number or "dyeing_po"}.pdf'
        disposition = "attachment" if request.GET.get("download") == "1" else "inline"
        response["Content-Disposition"] = f'{disposition}; filename="{filename}"'
        response["Cache-Control"] = "no-store"
        response["X-Content-Type-Options"] = "nosniff"
        try:
            response["Content-Length"] = str(len(response.content))
        except Exception:
            pass
    return response


def _build_simple_ready_po_pdf_response(po):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas
    except ImportError:
        return HttpResponse(
            "ReportLab is required for PDF generation. Install it with: pip install reportlab",
            status=500,
        )

    def safe_text(value):
        value = "" if value is None else str(value).strip()
        return value if value else "-"

    def qty_text(value):
        try:
            return f"{float(value or 0):,.2f}".rstrip("0").rstrip(".")
        except Exception:
            return safe_text(value)

    buffer = BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    left = 15 * mm
    top = height - 18 * mm
    y = top
    firm = getattr(po, "firm", None)
    vendor = getattr(po, "vendor", None)

    pdf.setFont("Helvetica-Bold", 16)
    pdf.drawString(left, y, safe_text(firm.firm_name if firm else "InventTech"))
    pdf.setFont("Helvetica-Bold", 12)
    pdf.drawRightString(width - 15 * mm, y, "READY PURCHASE ORDER")

    y -= 9 * mm
    pdf.setFont("Helvetica", 9)
    pdf.drawString(left, y, f"PO No: {safe_text(po.po_number or po.system_number)}")
    pdf.drawString(left + 70 * mm, y, f"Date: {po.po_date.strftime('%d-%m-%Y') if po.po_date else '-'}")
    pdf.drawRightString(width - 15 * mm, y, f"Status: {safe_text(po.get_approval_status_display())}")
    y -= 6 * mm
    pdf.drawString(left, y, f"System No: {safe_text(po.system_number)}")
    if getattr(po, "source_dyeing_po", None):
        pdf.drawString(left + 70 * mm, y, f"Dyeing PO: {safe_text(po.source_dyeing_po.system_number)}")

    y -= 12 * mm
    pdf.setFont("Helvetica-Bold", 10)
    pdf.drawString(left, y, "Vendor")
    pdf.drawString(width / 2, y, "Bill To / Ship To")
    y -= 6 * mm
    pdf.setFont("Helvetica", 9)
    vendor_lines = [
        safe_text(vendor.name if vendor else ""),
        f"Phone: {safe_text(getattr(vendor, 'phone', ''))}",
        f"Email: {safe_text(getattr(vendor, 'email', ''))}",
        f"GSTIN: {safe_text(getattr(vendor, 'gst_number', ''))}",
        f"Address: {safe_text(getattr(vendor, 'address', ''))}",
    ]
    bill_lines = [
        safe_text(firm.firm_name if firm else ""),
        f"Phone: {safe_text(getattr(firm, 'phone', ''))}",
        f"Email: {safe_text(getattr(firm, 'email', ''))}",
        f"GSTIN: {safe_text(getattr(firm, 'gst_number', ''))}",
        f"Ship To: {safe_text(po.shipping_address)}",
    ]
    for idx in range(max(len(vendor_lines), len(bill_lines))):
        pdf.drawString(left, y, vendor_lines[idx] if idx < len(vendor_lines) else "")
        pdf.drawString(width / 2, y, bill_lines[idx] if idx < len(bill_lines) else "")
        y -= 5 * mm

    y -= 4 * mm
    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawString(left, y, "Sr")
    pdf.drawString(left + 12 * mm, y, "Fabric")
    pdf.drawString(left + 82 * mm, y, "Dyeing")
    pdf.drawString(left + 130 * mm, y, "Unit")
    pdf.drawRightString(width - 15 * mm, y, "Qty")
    y -= 4 * mm
    pdf.line(left, y, width - 15 * mm, y)
    y -= 6 * mm

    total_qty = Decimal("0")
    for index, item in enumerate(po.items.all(), start=1):
        if y < 45 * mm:
            pdf.showPage()
            y = height - 20 * mm
        qty = getattr(item, "quantity", 0) or 0
        total_qty += Decimal(qty or 0)
        fabric = item.fabric_name or getattr(getattr(item, "finished_material", None), "name", "") or "-"
        pdf.setFont("Helvetica", 8.5)
        pdf.drawString(left, y, str(index))
        pdf.drawString(left + 12 * mm, y, safe_text(fabric)[:42])
        pdf.drawString(left + 82 * mm, y, safe_text(item.dyeing_name)[:28])
        pdf.drawString(left + 130 * mm, y, safe_text(item.unit)[:10])
        pdf.drawRightString(width - 15 * mm, y, qty_text(qty))
        y -= 6 * mm

    y -= 4 * mm
    pdf.line(left, y, width - 15 * mm, y)
    y -= 8 * mm
    pdf.setFont("Helvetica-Bold", 9)
    pdf.drawString(left, y, f"Total Qty: {qty_text(po.total_weight or total_qty)}")
    y -= 12 * mm
    pdf.setFont("Helvetica", 8.5)
    pdf.drawString(left, y, "THIS PO IS COMPUTER GENERATED, HENCE SIGNATURE IS NOT REQUIRED")
    pdf.save()
    buffer.seek(0)
    return HttpResponse(buffer.getvalue(), content_type="application/pdf")


@login_required
def readypo_pdf(request, pk: int):
    po = get_object_or_404(
        ReadyPurchaseOrder.objects
        .select_related("vendor", "firm", "source_dyeing_po", "owner")
        .prefetch_related(
            Prefetch(
                "items",
                queryset=ReadyPurchaseOrderItem.objects.select_related("finished_material", "source_dyeing_po_item"),
            )
        ),
        pk=pk,
    )

    if not _can_access_ready_po(request.user, po):
        raise PermissionDenied("You do not have access to this Ready PO.")

    response = _build_simple_ready_po_pdf_response(po)
    if response.status_code == 200 and response.get("Content-Type", "").startswith("application/pdf"):
        filename = f'{po.system_number or "ready_po"}.pdf'
        disposition = "attachment" if request.GET.get("download") == "1" else "inline"
        response["Content-Disposition"] = f'{disposition}; filename="{filename}"'
        response["Cache-Control"] = "no-store"
        response["X-Content-Type-Options"] = "nosniff"
        try:
            response["Content-Length"] = str(len(response.content))
        except Exception:
            pass
    return response

@login_required
def dyeingpo_list(request):
    q = (request.GET.get("q") or "").strip()

    qs = _dyeing_po_queryset()
    if not _can_review_yarn_po(request):
        qs = qs.filter(owner=request.user)

    if q:
        qs = qs.filter(
            Q(system_number__icontains=q)
            | Q(po_number__icontains=q)
            | Q(internal_po_number__icontains=q)
            | Q(vendor__name__icontains=q)
            | Q(source_greige_po__system_number__icontains=q)
            | Q(source_greige_inward__inward_number__icontains=q)
            | Q(inwards__inward_number__icontains=q)
            | Q(firm__firm_name__icontains=q)
            | Q(items__fabric_name__icontains=q)
            | Q(items__greige_name__icontains=q)
            | Q(items__finished_material__name__icontains=q)
            | Q(items__dyeing_name__icontains=q)
        ).distinct()

    return render(
        request,
        "accounts/dyeing_po/list.html",
        {
            "orders": qs,
            "q": q,
            "can_review_dyeing_po": _can_review_yarn_po(request),
        },
    )

def _dyeing_link_details_for_source(owner, vendor_id=None, greige_material_id=None, greige_material_name="", finished_material_id=None):
    """Return active dyeing-link details for a source greige item.

    Match order:
    1. exact vendor + exact material id/name
    2. vendor + normalized material name fallback
    3. vendor-only fallback, so the PO form does not stay empty when old material IDs differ
    """
    qs = (
        DyeingMaterialLinkDetail.objects
        .select_related("link__vendor", "link__material", "finished_material__unit")
        .filter(link__owner=owner, link__is_active=True, is_active=True)
        .exclude(finished_material__isnull=True)
    )

    if vendor_id:
        qs = qs.filter(link__vendor_id=vendor_id)

    if finished_material_id:
        qs = qs.filter(finished_material_id=finished_material_id)

    base_qs = qs
    material_filter = Q()
    greige_material_name = _compact_spaces(greige_material_name)

    if greige_material_id:
        material_filter |= Q(link__material_id=greige_material_id)
    if greige_material_name:
        material_filter |= Q(link__material__name__iexact=greige_material_name)

    if material_filter:
        material_qs = base_qs.filter(material_filter)

        if not material_qs.exists() and greige_material_name:
            same_name_material_ids = [
                material.pk
                for material in Material.objects.filter(
                    Q(owner=owner) | Q(owner__isnull=True),
                    material_kind="greige",
                ).only("id", "name")
                if _compact_spaces(material.name).lower() == greige_material_name.lower()
            ]
            if same_name_material_ids:
                material_qs = base_qs.filter(link__material_id__in=same_name_material_ids)

        if material_qs.exists():
            return material_qs.order_by("sort_order", "dyeing_name", "pk")

    return base_qs.order_by("link__material__name", "sort_order", "dyeing_name", "pk")


def _first_dyeing_link_detail_for_source(owner, vendor_id=None, source_item=None):
    greige_material_id = getattr(source_item, "material_id", None) if source_item else None
    greige_material_name = ""

    if source_item is not None:
        greige_material_name = (
            getattr(source_item.material, "name", "")
            if getattr(source_item, "material_id", None)
            else ""
        ) or getattr(source_item, "fabric_name", "") or ""

    return _dyeing_link_details_for_source(
        owner=owner,
        vendor_id=vendor_id,
        greige_material_id=greige_material_id,
        greige_material_name=greige_material_name,
    ).first()


def _po_header_maps_from_form(form):
    firm_shipping_map = {}

    firm_field = form.fields.get("firm")
    if firm_field:
        for firm in firm_field.queryset:
            addresses = []

            if hasattr(firm, "addresses"):
                for address in firm.addresses.order_by("-is_default", "id"):
                    addresses.append({
                        "id": str(address.pk),
                        "label": f"{address.label or f'Address {address.pk}'}{' (Default)' if address.is_default else ''}",
                        "value": address.full_address,
                    })

            if not addresses and getattr(firm, "full_address", ""):
                addresses.append({
                    "id": "firm-default",
                    "label": f"{firm.firm_name} (Default)",
                    "value": firm.full_address,
                })

            firm_shipping_map[str(firm.pk)] = addresses

    terms_condition_map = {}
    terms_field = form.fields.get("terms_template")
    if terms_field:
        terms_condition_map = {
            str(obj.pk): obj.content
            for obj in terms_field.queryset
        }

    return {
        "terms_condition_map": terms_condition_map,
        "shipping_address_map": getattr(form, "shipping_address_map", {}),
        "firm_shipping_map": firm_shipping_map,
    }

@login_required
@require_http_methods(["GET", "POST"])
def dyeingpo_create(request, greige_po_id=None):
    source_greige_po = None
    selected_source_inward = None

    if greige_po_id is not None:
        source_greige_po = get_object_or_404(_greige_po_queryset(), pk=greige_po_id)
        if not _can_access_greige_po(request.user, source_greige_po):
            raise PermissionDenied("You do not have access to this Greige PO.")

    # Also support the normal add page where the user first selects a
    # Source Greige PO from the dropdown. This lets the item formset build
    # source-item, finished-material, and dyeing-master choices correctly
    # before validation runs.
    if source_greige_po is None:
        posted_source_id = (
            request.POST.get("source_greige_po")
            or request.GET.get("source_greige_po")
            or ""
        ).strip()
        if posted_source_id.isdigit():
            source_greige_po = get_object_or_404(_greige_po_queryset(), pk=int(posted_source_id))
            if not _can_access_greige_po(request.user, source_greige_po):
                raise PermissionDenied("You do not have access to this Greige PO.")

    inward_id = (request.POST.get("source_greige_inward") or request.GET.get("inward") or "").strip()
    if source_greige_po is not None and inward_id:
        selected_source_inward = get_object_or_404(
            source_greige_po.inwards.prefetch_related("items__po_item"),
            pk=inward_id,
        )

    temp_po = DyeingPurchaseOrder(
        source_greige_po=source_greige_po,
        source_greige_inward=selected_source_inward,
    )

    initial_items = []
    first_source_item = None

    if selected_source_inward is not None:
        first_inward_item = selected_source_inward.items.select_related("po_item__material").first()
        if first_inward_item and first_inward_item.po_item:
            first_source_item = first_inward_item.po_item
    elif source_greige_po is not None:
        first_source_item = source_greige_po.items.select_related("material").first()

    if first_source_item is not None:
        source_material_name = _compact_spaces(
            (first_source_item.material.name if first_source_item.material_id else "")
            or first_source_item.fabric_name
            or ""
        )

        linked_detail = _first_dyeing_link_detail_for_source(
            owner=request.user,
            vendor_id=(source_greige_po.vendor_id if source_greige_po else None),
            source_item=first_source_item,
        )

        initial_row = {
            "source_greige_po_item": first_source_item.pk,
            "greige_name": (
                first_source_item.fabric_name
                or (first_source_item.material.name if first_source_item.material_id else "")
            ),
            "total_qty": first_source_item.quantity or Decimal("0"),
            "source_input_qty": first_source_item.quantity or Decimal("0"),
            "remark": first_source_item.remark or "",
        }

        if linked_detail is not None:
            initial_row.update({
                "dyeing_master_detail": linked_detail.pk,
                "finished_material": linked_detail.finished_material_id,
                "fabric_name": linked_detail.finished_material.name if linked_detail.finished_material_id else "",
                "dyeing_type": linked_detail.dyeing_type or "",
                "dyeing_name": linked_detail.dyeing_name or "",
                "unit": (
                    linked_detail.finished_material.unit.name
                    if linked_detail.finished_material_id and getattr(linked_detail.finished_material, "unit", None)
                    else ""
                ),
                "rate": linked_detail.price or Decimal("0"),
                "expected_loss_percent": linked_detail.weight_loss or Decimal("0"),
            })

        initial_items = [initial_row]

    if request.method == "POST":
        form = DyeingPurchaseOrderForm(
            request.POST,
            user=request.user,
            source_greige_po=source_greige_po,
            lock_source=bool(source_greige_po),
        )

        formset = DyeingPurchaseOrderItemFormSet(
            request.POST,
            instance=temp_po,
            prefix="items",
            initial=initial_items,
            queryset=DyeingPurchaseOrderItem.objects.none(),
            form_kwargs={
                "user": request.user,
                "source_greige_po": source_greige_po,
                "selected_source_inward": selected_source_inward,
            },
        )

        if form.is_valid() and formset.is_valid():
            selected_source = source_greige_po or form.cleaned_data["source_greige_po"]

            if not _can_access_greige_po(request.user, selected_source):
                raise PermissionDenied("You do not have access to this Greige PO.")

            if selected_source_inward is None:
                form.add_error(None, "Select exact Greige inward before generating Dyeing PO.")
            elif selected_source_inward.po_id != selected_source.id:
                form.add_error(None, "Selected Greige inward does not belong to the chosen Greige PO.")
            elif selected_source_inward.generated_dyeing_pos.exists():
                form.add_error(None, "Dyeing PO already exists for this Greige inward.")
            else:
                with transaction.atomic():
                    dyeing_po = form.save(commit=False)
                    dyeing_po.owner = selected_source.owner
                    dyeing_po.system_number = _next_dyeing_po_number()
                    dyeing_po.source_greige_po = selected_source
                    dyeing_po.source_greige_inward = selected_source_inward

                    if not dyeing_po.firm_id and getattr(selected_source, "firm_id", None):
                        dyeing_po.firm = selected_source.firm

                    if not dyeing_po.shipping_address and dyeing_po.firm_id:
                        dyeing_po.shipping_address = _firm_address(dyeing_po.firm)

                    if not dyeing_po.terms_conditions:
                        dyeing_po.terms_conditions = getattr(selected_source, "terms_conditions", "") or ""

                    dyeing_po.save()

                    formset.instance = dyeing_po
                    items = formset.save(commit=False)

                    for obj in formset.deleted_objects:
                        obj.delete()

                    total_weight = Decimal("0")
                    subtotal = Decimal("0")

                    for item in items:
                        item.po = dyeing_po

                        if item.source_greige_po_item_id and item.source_greige_po_item:
                            if not item.greige_name:
                                item.greige_name = (
                                    item.source_greige_po_item.fabric_name
                                    or (
                                        item.source_greige_po_item.material.name
                                        if item.source_greige_po_item.material_id
                                        else ""
                                    )
                                )

                        if item.dyeing_master_detail_id:
                            detail = item.dyeing_master_detail

                            if not item.finished_material_id and detail.finished_material_id:
                                item.finished_material = detail.finished_material

                            if not item.rate:
                                item.rate = detail.price or Decimal("0")

                            if item.expected_loss_percent in (None, Decimal("0")):
                                item.expected_loss_percent = detail.weight_loss or Decimal("0")

                            item.dyeing_type = detail.dyeing_type or ""
                            item.dyeing_name = detail.dyeing_name or ""

                            if not item.unit and detail.finished_material_id and getattr(detail.finished_material, "unit", None):
                                item.unit = detail.finished_material.unit.name

                        if item.finished_material:
                            item.fabric_name = item.finished_material.name
                        else:
                            item.fabric_name = "Dyeing Item"

                        if not item.greige_name:
                            if item.source_greige_po_item_id and item.source_greige_po_item:
                                item.greige_name = (
                                    item.source_greige_po_item.fabric_name
                                    or (
                                        item.source_greige_po_item.material.name
                                        if item.source_greige_po_item.material_id
                                        else ""
                                    )
                                )
                            elif selected_source_inward and selected_source_inward.items.exists():
                                first_inward_item = selected_source_inward.items.select_related("po_item").first()
                                if first_inward_item and first_inward_item.po_item:
                                    item.greige_name = first_inward_item.po_item.fabric_name or ""
                            elif selected_source and selected_source.items.exists():
                                first_po_item = selected_source.items.first()
                                if first_po_item:
                                    item.greige_name = first_po_item.fabric_name or ""

                        item.quantity = item.total_qty or Decimal("0")
                        item.source_input_qty = item.total_qty or Decimal("0")

                        if item.expected_loss_percent is None:
                            item.expected_loss_percent = Decimal("0")

                        if not item.expected_output_qty:
                            loss = item.expected_loss_percent or Decimal("0")
                            item.expected_output_qty = item.total_qty - ((item.total_qty * loss) / Decimal("100"))

                        if not item.remaining_qty:
                            item.remaining_qty = item.expected_output_qty or item.total_qty or Decimal("0")

                        item.line_final_amount = item.line_final_amount or (
                            (item.total_qty or Decimal("0")) * (item.rate or Decimal("0"))
                        )

                        item.save()

                        total_weight += item.total_qty or Decimal("0")
                        subtotal += item.line_final_amount or Decimal("0")

                    formset.save_m2m()

                    discount = dyeing_po.discount_percent or Decimal("0")
                    others = dyeing_po.others or Decimal("0")
                    gst = dyeing_po.gst_percent or Decimal("0")
                    tcs = dyeing_po.tcs_percent or Decimal("0")

                    after_discount = subtotal - (subtotal * discount / Decimal("100"))
                    after_others = after_discount + others
                    gst_amount = after_others * gst / Decimal("100")
                    tcs_amount = after_others * tcs / Decimal("100")
                    final_amount = after_others + gst_amount + tcs_amount

                    dyeing_po.total_weight = total_weight
                    dyeing_po.subtotal = subtotal
                    dyeing_po.after_discount_value = after_discount
                    dyeing_po.final_amount = final_amount
                    dyeing_po.save(update_fields=[
                        "total_weight",
                        "subtotal",
                        "after_discount_value",
                        "final_amount",
                        "updated_at",
                    ])

                messages.success(request, f"Dyeing PO {dyeing_po.system_number} saved successfully.")
                tracker_url = reverse("accounts:dyeing_inward_tracker")
                return redirect(f"{tracker_url}?po={dyeing_po.pk}")

    else:
        initial = {}
        if source_greige_po is not None:
            initial = {
                "po_number": source_greige_po.po_number or "",
                "po_date": timezone.localdate(),
                "vendor": source_greige_po.vendor_id,
                "firm": source_greige_po.firm_id or (
                    source_greige_po.source_yarn_po.firm_id
                    if source_greige_po.source_yarn_po and source_greige_po.source_yarn_po.firm
                    else None
                ),
                "shipping_address": source_greige_po.shipping_address or (
                    _firm_address(source_greige_po.firm)
                    if source_greige_po.firm
                    else (
                        _firm_address(source_greige_po.source_yarn_po.firm)
                        if source_greige_po.source_yarn_po and source_greige_po.source_yarn_po.firm
                        else ""
                    )
                ),
                "terms_conditions": getattr(source_greige_po, "terms_conditions", "") or "",
                "remarks": (
                    f"Generated from Greige inward {selected_source_inward.inward_number}"
                    if selected_source_inward else
                    f"Generated from Greige PO {source_greige_po.system_number}"
                ),
            }

        form = DyeingPurchaseOrderForm(
            initial=initial,
            user=request.user,
            source_greige_po=source_greige_po,
            lock_source=bool(source_greige_po),
        )

        formset = DyeingPurchaseOrderItemFormSet(
            instance=temp_po,
            prefix="items",
            initial=initial_items,
            queryset=DyeingPurchaseOrderItem.objects.none(),
            form_kwargs={
                "user": request.user,
                "source_greige_po": source_greige_po,
                "selected_source_inward": selected_source_inward,
            },
        )

    dyeing_master_map = {}
    master_qs = (
        DyeingMaterialLinkDetail.objects
        .select_related(
            "link__vendor",
            "link__material",
            "finished_material__unit",
        )
        .filter(
            link__owner=request.user,
            link__is_active=True,
            is_active=True,
        )
        .exclude(finished_material__isnull=True)
        .order_by("link__vendor__name", "finished_material__name", "dyeing_name")
    )

    for detail in master_qs:
        dyeing_master_map[str(detail.pk)] = {
            "id": str(detail.pk),
            "vendor_id": str(detail.link.vendor_id) if detail.link_id and detail.link.vendor_id else "",
            "material_id": str(detail.link.material_id) if detail.link_id and detail.link.material_id else "",
            "finished_material_id": str(detail.finished_material_id) if detail.finished_material_id else "",
            "finished_material_name": detail.finished_material.name if detail.finished_material_id else "",
            "rate": str(detail.price or Decimal("0")),
            "expected_loss_percent": str(detail.weight_loss or Decimal("0")),
            "unit": (
                detail.finished_material.unit.name
                if detail.finished_material_id and getattr(detail.finished_material, "unit", None)
                else ""
            ),
            "dyeing_type": detail.get_dyeing_type_display() if detail.dyeing_type else "",
            "dyeing_type_value": detail.dyeing_type or "",
            "dyeing_name": detail.dyeing_name or "",
            "percentage_no_of_colors": str(detail.percentage_no_of_colors or Decimal("0")),
            "link_material_name": detail.link.material.name if detail.link_id and detail.link.material_id else "",
            "link_material_name_key": _compact_spaces(detail.link.material.name).lower() if detail.link_id and detail.link.material_id else "",
            "vendor_name": detail.link.vendor.name if detail.link_id and detail.link.vendor_id else "",
            "vendor_name_key": _compact_spaces(detail.link.vendor.name).lower() if detail.link_id and detail.link.vendor_id else "",
        }

    selected_source_id = ""
    source_item_map = {}

    if request.method == "POST":
        selected_source_value = request.POST.get("source_greige_po")
        if selected_source_value:
            selected_source_id = str(selected_source_value)
    elif source_greige_po is not None:
        selected_source_id = str(source_greige_po.pk)

    source_items_qs = GreigePurchaseOrderItem.objects.none()
    if selected_source_inward is not None:
        source_items_qs = GreigePurchaseOrderItem.objects.filter(
            pk__in=selected_source_inward.items.values_list("po_item_id", flat=True)
        ).select_related("material").order_by("id")
    elif source_greige_po is not None:
        source_items_qs = source_greige_po.items.select_related("material").order_by("id")

    for item in source_items_qs:
        source_item_map[str(item.pk)] = {
            "id": str(item.pk),
            "material_id": str(item.material_id) if item.material_id else "",
            "material_name": item.material.name if item.material_id else "",
            "material_name_key": _compact_spaces(item.material.name).lower() if item.material_id else "",
            "fabric_name": item.fabric_name or (item.material.name if item.material_id else ""),
            "fabric_name_key": _compact_spaces(item.fabric_name or (item.material.name if item.material_id else "")).lower(),
            "quantity": str(item.quantity or Decimal("0")),
            "remark": item.remark or "",
        }

    greige_item_material_map = {
        str(item.pk): str(item.material_id) if item.material_id else ""
        for item in source_items_qs
    }

    po_header_maps = _po_header_maps_from_form(form)

    context = {
        "form": form,
        "formset": formset,
        "page_title": "Generate Dyeing PO",
        "submit_label": "Save Dyeing PO",
        "is_create": True,
        "source_greige_po": source_greige_po,
        "selected_source_inward": selected_source_inward,
        "selected_source_id": selected_source_id,
        "source_vendor_id": str(source_greige_po.vendor_id) if source_greige_po and source_greige_po.vendor_id else "",
        "source_vendor_name": source_greige_po.vendor.name if source_greige_po and source_greige_po.vendor_id else "",
        "source_item_map_json": source_item_map,
        "dyeing_master_map_json": dyeing_master_map,
        "greige_item_material_map_json": greige_item_material_map,
        "mode": "add",
        "po_obj": None,
        "system_number_preview": _next_dyeing_po_number(),
        "source_inwards": list(source_greige_po.inwards.all()) if source_greige_po else [],
        "existing_po": None,
        **po_header_maps,
        "dyeing_master_options": [
            {
                "id": detail_id,
                "label": (
                    f"{values['dyeing_name']} · {values['finished_material_name']}"
                    if values["finished_material_name"]
                    else values["dyeing_name"]
                ),
            }
            for detail_id, values in dyeing_master_map.items()
        ],
    }

    return render(request, "accounts/dyeing_po/form.html", context)


@login_required
@require_http_methods(["GET", "POST"])
def dyeingpo_update(request, pk: int):
    po = get_object_or_404(_dyeing_po_queryset(), pk=pk)
    if not _can_access_dyeing_po(request.user, po):
        raise PermissionDenied("You do not have access to this Dyeing PO.")

    if po.ready_pos.exists():
        messages.error(request, "This Dyeing PO cannot be edited because Ready PO already exists.")
        return redirect("accounts:dyeingpo_list")

    if po.inwards.exists():
        messages.error(request, "This Dyeing PO cannot be edited because inward entries already exist.")
        return redirect("accounts:dyeingpo_list")

    if po.approval_status == "approved":
        messages.error(request, "Approved Dyeing PO cannot be edited.")
        return redirect("accounts:dyeingpo_list")
    
    form = DyeingPurchaseOrderForm(
        request.POST or None,
        user=request.user,
        instance=po,
        source_greige_po=po.source_greige_po,
        lock_source=True,
    )

    formset = DyeingPurchaseOrderItemFormSet(
        request.POST or None,
        instance=po,
        prefix="items",
        form_kwargs={
            "user": request.user,
            "source_greige_po": po.source_greige_po,
            "selected_source_inward": po.source_greige_inward,
        },
    )   

    if request.method == "POST" and form.is_valid() and formset.is_valid():
        with transaction.atomic():
            po = form.save(commit=False)
            if not po.firm_id and po.source_greige_po and po.source_greige_po.firm_id:
                po.firm = po.source_greige_po.firm

            if not po.shipping_address and po.firm_id:
                po.shipping_address = _firm_address(po.firm)

            if not po.terms_conditions and po.source_greige_po:
                po.terms_conditions = getattr(po.source_greige_po, "terms_conditions", "") or ""

            po.save()

            items = formset.save(commit=False)

            for obj in formset.deleted_objects:
                obj.delete()

            total_weight = Decimal("0")
            subtotal = Decimal("0")

            for item in items:
                item.po = po
                if item.source_greige_po_item_id and item.source_greige_po_item:
                    if not item.greige_name:
                        item.greige_name = (
                            item.source_greige_po_item.fabric_name
                            or (item.source_greige_po_item.material.name if item.source_greige_po_item.material_id else "")
                        )
                if item.dyeing_master_detail_id:
                    detail = item.dyeing_master_detail

                    if not item.finished_material_id and detail.finished_material_id:
                        item.finished_material = detail.finished_material

                    if not item.rate:
                        item.rate = detail.price or Decimal("0")

                    if item.expected_loss_percent in (None, Decimal("0")):
                        item.expected_loss_percent = detail.weight_loss or Decimal("0")

                    item.dyeing_type = detail.dyeing_type or ""
                    item.dyeing_name = detail.dyeing_name or ""

                    if not item.unit and detail.finished_material_id and getattr(detail.finished_material, "unit", None):
                        item.unit = detail.finished_material.unit.name

                if item.finished_material:
                    item.fabric_name = item.finished_material.name
                else:
                    item.fabric_name = "Dyeing Item"

                item.quantity = item.total_qty or Decimal("0")
                item.source_input_qty = item.total_qty or Decimal("0")

                if item.expected_loss_percent is None:
                    item.expected_loss_percent = Decimal("0")

                if not item.expected_output_qty:
                    item.expected_output_qty = item.total_qty or Decimal("0")

                if not item.remaining_qty:
                    item.remaining_qty = item.expected_output_qty or item.total_qty or Decimal("0")

                item.line_final_amount = item.line_final_amount or Decimal("0")

                item.save()

                total_weight += item.total_qty or Decimal("0")
                subtotal += item.line_final_amount or Decimal("0")

            formset.save_m2m()

            discount = po.discount_percent or Decimal("0")
            others = po.others or Decimal("0")
            gst = po.gst_percent or Decimal("0")
            tcs = po.tcs_percent or Decimal("0")

            after_discount = subtotal - (subtotal * discount / Decimal("100"))
            after_others = after_discount + others
            gst_amount = after_others * gst / Decimal("100")
            tcs_amount = after_others * tcs / Decimal("100")
            final_amount = after_others + gst_amount + tcs_amount

            po.total_weight = total_weight
            po.subtotal = subtotal
            po.after_discount_value = after_discount
            po.final_amount = final_amount
            po.save(update_fields=[
                "total_weight",
                "subtotal",
                "after_discount_value",
                "final_amount",
                "updated_at",
            ])

        messages.success(request, f"Dyeing PO {po.system_number} updated successfully.")
        tracker_url = reverse("accounts:dyeing_inward_tracker")
        return redirect(f"{tracker_url}?po={po.pk}")

    dyeing_master_map = {}
    master_qs = formset.empty_form.fields["dyeing_master_detail"].queryset.select_related(
        "link__vendor",
        "link__material",
        "finished_material__unit",
    )

    for detail in master_qs:
        dyeing_master_map[str(detail.pk)] = {
            "id": str(detail.pk),
            "vendor_id": str(detail.link.vendor_id) if detail.link_id and detail.link.vendor_id else "",
            "vendor_name": detail.link.vendor.name if detail.link_id and detail.link.vendor_id else "",
            "vendor_name_key": _compact_spaces(detail.link.vendor.name).lower() if detail.link_id and detail.link.vendor_id else "",
            "material_id": str(detail.link.material_id) if detail.link_id and detail.link.material_id else "",
            "link_material_name": detail.link.material.name if detail.link_id and detail.link.material_id else "",
            "link_material_name_key": _compact_spaces(detail.link.material.name).lower() if detail.link_id and detail.link.material_id else "",
            "finished_material_id": str(detail.finished_material_id) if detail.finished_material_id else "",
            "finished_material_name": detail.finished_material.name if detail.finished_material_id else "",
            "rate": str(detail.price or Decimal("0")),
            "expected_loss_percent": str(detail.weight_loss or Decimal("0")),
            "unit": (
                detail.finished_material.unit.name
                if detail.finished_material_id and getattr(detail.finished_material, "unit", None)
                else ""
            ),
            "dyeing_type": detail.get_dyeing_type_display() if detail.dyeing_type else "",
            "dyeing_name": detail.dyeing_name or "",
            "percentage_no_of_colors": str(detail.percentage_no_of_colors or Decimal("0")),
        }
    po_header_maps = _po_header_maps_from_form(form)
    return render(
        request,
        "accounts/dyeing_po/form.html",
        {
            "form": form,
            "formset": formset,
            "mode": "edit",
            "po_obj": po,
            "system_number_preview": po.system_number,
            "dyeing_master_map_json": dyeing_master_map,
            "source_greige_po": po.source_greige_po,
            "selected_source_inward": po.source_greige_inward,
            "source_inwards": list(po.source_greige_po.inwards.all()) if po.source_greige_po else [],
            "existing_po": None,
            **po_header_maps,   
        },
    )


@login_required
def dyeingpo_detail(request, pk: int):
    po = get_object_or_404(_dyeing_po_queryset(), pk=pk)
    if not _can_access_dyeing_po(request.user, po):
        raise PermissionDenied("You do not have access to this Dyeing PO.")

    return render(
        request,
        "accounts/dyeing_po/detail.html",
        {
            "po": po,
            "source_greige_po": po.source_greige_po,
            "source_inwards": list(po.source_greige_po.inwards.all()) if po.source_greige_po else [],
            "dyeing_inwards": list(po.inwards.all()),
            "existing_ready_po": po.ready_pos.order_by("-id").first(),
        },
    )

def _next_dyeing_inward_number() -> str:
    last = DyeingPOInward.objects.order_by("-id").first()
    next_id = (last.id + 1) if last else 1
    return f"DIN-{next_id:04d}"

@login_required
@require_http_methods(["GET", "POST"])
def ready_inward_edit(request, pk: int):
    inward = get_object_or_404(
        ReadyPOInward.objects.select_related("po__vendor", "po__firm", "po__owner"),
        pk=pk,
    )
    po = inward.po

    if not _can_access_ready_po(request.user, po):
        raise PermissionDenied("You do not have access to this Ready inward.")

    item_errors = {}
    line_inputs = {
        row.po_item_id: {
            "received_qty": row.received_qty if row.received_qty is not None else row.quantity,
            "accepted_qty": row.accepted_qty,
            "rejected_qty": row.rejected_qty,
            "hold_qty": row.hold_qty,
            "actual_rolls": row.actual_rolls,
            "actual_gsm": row.actual_gsm,
            "actual_width": row.actual_width,
            "remark": row.remark or "",
        }
        for row in inward.items.all()
    }

    inward_form = ReadyPOInwardForm(request.POST or None, instance=inward, user=request.user)

    if request.method == "POST" and inward_form.is_valid():
        line_payload, line_inputs, item_errors = _collect_ready_inward_lines(
            request,
            po,
            editing_inward=inward,
        )

        if not line_payload:
            inward_form.add_error(None, "Enter at least one inward row.")

        if not inward_form.errors and not item_errors:
            inward = inward_form.save(commit=False)
            inward.owner = po.owner
            inward.po = po
            inward.save()

            inward.items.all().delete()

            ReadyPOInwardItem.objects.bulk_create([
                ReadyPOInwardItem(
                    inward=inward,
                    po_item=row["item"],
                    quantity=row["quantity"],
                    received_qty=row["received_qty"],
                    accepted_qty=row["accepted_qty"],
                    rejected_qty=row["rejected_qty"],
                    hold_qty=row["hold_qty"],
                    actual_rolls=row["actual_rolls"],
                    actual_gsm=row["actual_gsm"],
                    actual_width=row["actual_width"],
                    dye_lot_no=row["dye_lot_no"],
                    batch_no=row["batch_no"],
                    shade_reference=row["shade_reference"],
                    qc_status=row["qc_status"],
                    remark=row["remark"],
                )
                for row in line_payload
            ])

            messages.success(request, f"Inward {inward.inward_number} updated successfully.")
            tracker_url = reverse("accounts:ready_inward_tracker")
            return redirect(f"{tracker_url}?inward={inward.pk}")

    line_rows = [
        {
            "item": item,
            "received_qty_value": line_inputs.get(item.id, {}).get("received_qty", ""),
            "accepted_qty_value": line_inputs.get(item.id, {}).get("accepted_qty", ""),
            "rejected_qty_value": line_inputs.get(item.id, {}).get("rejected_qty", ""),
            "hold_qty_value": line_inputs.get(item.id, {}).get("hold_qty", ""),
            "actual_rolls_value": line_inputs.get(item.id, {}).get("actual_rolls", ""),
            "actual_gsm_value": line_inputs.get(item.id, {}).get("actual_gsm", ""),
            "actual_width_value": line_inputs.get(item.id, {}).get("actual_width", ""),
            "remark_value": line_inputs.get(item.id, {}).get("remark", ""),
            "error": item_errors.get(item.id, ""),
        }
        for item in po.items.all()
    ]

    return render(
        request,
        "accounts/ready_po/inward.html",
        {
            "po": po,
            "inward_form": inward_form,
            "line_rows": line_rows,
            "existing_inwards": po.inwards.all().order_by("-inward_date", "-id"),
            "editing_inward": inward,
            "next_inward_number_preview": inward.inward_number,
        },
    )
    
    
def _next_ready_inward_number() -> str:
    last = ReadyPOInward.objects.order_by("-id").first()
    next_id = (last.id + 1) if last else 1
    return f"RIN-{next_id:04d}"


def _decimal_or_zero(value):
    raw = (value or "").strip()
    if raw == "":
        return Decimal("0")
    return Decimal(raw)

def _decimal_or_none(value):
    raw = (value or "").strip()
    if raw == "":
        return None
    return Decimal(raw)

def _derive_dyeing_qc_status(received_qty, accepted_qty, rejected_qty, hold_qty):
    if received_qty <= 0:
        return "pending"
    if accepted_qty == received_qty and rejected_qty == 0 and hold_qty == 0:
        return "approved"
    if rejected_qty == received_qty:
        return "rejected"
    if hold_qty == received_qty:
        return "hold"
    return "partial"


def _collect_basic_po_inward_payload(request, po, editing_inward=None):
    """Collect dyeing-style inward rows for Yarn/Greige/Ready PO lines."""
    line_payload = []
    line_inputs = {}
    item_errors = {}

    for item in po.items.all():
        raw_qty = (request.POST.get(f"qty_{item.id}") or request.POST.get(f"received_qty_{item.id}") or "").strip()
        raw_accepted = (request.POST.get(f"accepted_qty_{item.id}") or "").strip()
        raw_rejected = (request.POST.get(f"rejected_qty_{item.id}") or "").strip()
        raw_hold = (request.POST.get(f"hold_qty_{item.id}") or "").strip()
        raw_rolls = (request.POST.get(f"actual_rolls_{item.id}") or "").strip()
        raw_gsm = (request.POST.get(f"actual_gsm_{item.id}") or "").strip()
        raw_width = (request.POST.get(f"actual_width_{item.id}") or "").strip()
        raw_dye_lot = (request.POST.get(f"dye_lot_no_{item.id}") or "").strip()
        raw_batch = (request.POST.get(f"batch_no_{item.id}") or "").strip()
        raw_shade = (request.POST.get(f"shade_reference_{item.id}") or "").strip()
        raw_remark = (request.POST.get(f"remark_{item.id}") or "").strip()

        line_inputs[item.id] = {
            "qty": raw_qty,
            "accepted_qty": raw_accepted,
            "rejected_qty": raw_rejected,
            "hold_qty": raw_hold,
            "actual_rolls": raw_rolls,
            "actual_gsm": raw_gsm,
            "actual_width": raw_width,
            "dye_lot_no": raw_dye_lot,
            "batch_no": raw_batch,
            "shade_reference": raw_shade,
            "remark": raw_remark,
        }

        has_any_input = any([
            raw_qty, raw_accepted, raw_rejected, raw_hold, raw_rolls,
            raw_gsm, raw_width, raw_dye_lot, raw_batch, raw_shade, raw_remark,
        ])
        if not has_any_input:
            continue

        if not raw_qty:
            item_errors[item.id] = "Enter received qty."
            continue

        try:
            qty = _decimal_or_zero(raw_qty)
            accepted_qty = _decimal_or_zero(raw_accepted)
            rejected_qty = _decimal_or_zero(raw_rejected)
            hold_qty = _decimal_or_zero(raw_hold)
            actual_rolls = _decimal_or_zero(raw_rolls)
            actual_gsm = _decimal_or_none(raw_gsm)
            actual_width = _decimal_or_none(raw_width)
        except InvalidOperation:
            item_errors[item.id] = "Enter valid numeric values."
            continue

        if qty <= 0:
            item_errors[item.id] = "Received qty must be greater than 0."
            continue

        if raw_accepted == "" and raw_rejected == "" and raw_hold == "":
            accepted_qty = qty

        if accepted_qty < 0 or rejected_qty < 0 or hold_qty < 0 or actual_rolls < 0:
            item_errors[item.id] = "Negative values are not allowed."
            continue

        if accepted_qty > qty:
            item_errors[item.id] = "Accepted qty cannot be greater than received qty."
            continue

        if rejected_qty > qty:
            item_errors[item.id] = "Rejected qty cannot be greater than received qty."
            continue

        if hold_qty > qty:
            item_errors[item.id] = "Hold qty cannot be greater than received qty."
            continue

        if (accepted_qty + rejected_qty + hold_qty) != qty:
            item_errors[item.id] = "Accepted + Rejected + Hold must match Received Qty."
            continue

        if editing_inward is None:
            max_allowed_qty = item.remaining_qty_total or Decimal("0")
        else:
            other_inward_qty = (
                item.inward_items.exclude(inward=editing_inward).aggregate(total=Sum("quantity")).get("total")
                or Decimal("0")
            )
            max_allowed_qty = (item.quantity or Decimal("0")) - other_inward_qty

        if qty > max_allowed_qty:
            item_errors[item.id] = "Entered quantity is greater than remaining quantity."
            continue

        qc_status = _derive_dyeing_qc_status(qty, accepted_qty, rejected_qty, hold_qty)

        line_payload.append({
            "item": item,
            "quantity": qty,
            "received_qty": qty,
            "accepted_qty": accepted_qty,
            "rejected_qty": rejected_qty,
            "hold_qty": hold_qty,
            "actual_rolls": actual_rolls,
            "actual_gsm": actual_gsm,
            "actual_width": actual_width,
            "dye_lot_no": raw_dye_lot,
            "batch_no": raw_batch,
            "shade_reference": raw_shade,
            "qc_status": qc_status,
            "remark": raw_remark,
        })

    return line_payload, line_inputs, item_errors

def _collect_ready_inward_lines(request, po, editing_inward=None):
    return _collect_basic_po_inward_payload(request, po, editing_inward=editing_inward)

def _build_dyeing_inward_line_rows(po, line_inputs=None, item_errors=None):
    line_inputs = line_inputs or {}
    item_errors = item_errors or {}

    return [
        {
            "item": item,
            "qty_value": line_inputs.get(item.id, {}).get("qty", ""),
            "accepted_qty_value": line_inputs.get(item.id, {}).get("accepted_qty", ""),
            "rejected_qty_value": line_inputs.get(item.id, {}).get("rejected_qty", ""),
            "hold_qty_value": line_inputs.get(item.id, {}).get("hold_qty", ""),
            "actual_rolls_value": line_inputs.get(item.id, {}).get("actual_rolls", ""),
            "actual_gsm_value": line_inputs.get(item.id, {}).get("actual_gsm", ""),
            "actual_width_value": line_inputs.get(item.id, {}).get("actual_width", ""),
            "dye_lot_no_value": line_inputs.get(item.id, {}).get("dye_lot_no", ""),
            "batch_no_value": line_inputs.get(item.id, {}).get("batch_no", ""),
            "shade_reference_value": line_inputs.get(item.id, {}).get("shade_reference", ""),
            "remark_value": line_inputs.get(item.id, {}).get("remark", ""),
            "error": item_errors.get(item.id, ""),
        }
        for item in po.items.all()
    ]

def _collect_dyeing_inward_payload(request, po, editing_inward=None):
    line_payload = []
    line_inputs = {}
    item_errors = {}

    for item in po.items.all():
        raw_qty = (request.POST.get(f"qty_{item.id}") or "").strip()
        raw_accepted = (request.POST.get(f"accepted_qty_{item.id}") or "").strip()
        raw_rejected = (request.POST.get(f"rejected_qty_{item.id}") or "").strip()
        raw_hold = (request.POST.get(f"hold_qty_{item.id}") or "").strip()
        raw_rolls = (request.POST.get(f"actual_rolls_{item.id}") or "").strip()
        raw_gsm = (request.POST.get(f"actual_gsm_{item.id}") or "").strip()
        raw_width = (request.POST.get(f"actual_width_{item.id}") or "").strip()
        raw_dye_lot = (request.POST.get(f"dye_lot_no_{item.id}") or "").strip()
        raw_batch = (request.POST.get(f"batch_no_{item.id}") or "").strip()
        raw_shade = (request.POST.get(f"shade_reference_{item.id}") or "").strip()
        raw_remark = (request.POST.get(f"remark_{item.id}") or "").strip()

        line_inputs[item.id] = {
            "qty": raw_qty,
            "accepted_qty": raw_accepted,
            "rejected_qty": raw_rejected,
            "hold_qty": raw_hold,
            "actual_rolls": raw_rolls,
            "actual_gsm": raw_gsm,
            "actual_width": raw_width,
            "dye_lot_no": raw_dye_lot,
            "batch_no": raw_batch,
            "shade_reference": raw_shade,
            "remark": raw_remark,
        }

        has_any_input = any([
            raw_qty, raw_accepted, raw_rejected, raw_hold, raw_rolls,
            raw_gsm, raw_width, raw_dye_lot, raw_batch, raw_shade, raw_remark
        ])
        if not has_any_input:
            continue

        if not raw_qty:
            item_errors[item.id] = "Enter received qty."
            continue

        try:
            qty = _decimal_or_zero(raw_qty)
            accepted_qty = _decimal_or_zero(raw_accepted)
            rejected_qty = _decimal_or_zero(raw_rejected)
            hold_qty = _decimal_or_zero(raw_hold)
            actual_rolls = _decimal_or_zero(raw_rolls)
            actual_gsm = _decimal_or_none(raw_gsm)
            actual_width = _decimal_or_none(raw_width)
        except InvalidOperation:
            item_errors[item.id] = "Enter valid numeric values."
            continue

        if qty <= 0:
            item_errors[item.id] = "Received qty must be greater than 0."
            continue

        if accepted_qty < 0 or rejected_qty < 0 or hold_qty < 0 or actual_rolls < 0:
            item_errors[item.id] = "Negative values are not allowed."
            continue

        if accepted_qty > qty:
            item_errors[item.id] = "Accepted qty cannot be greater than received qty."
            continue

        if rejected_qty > qty:
            item_errors[item.id] = "Rejected qty cannot be greater than received qty."
            continue

        if hold_qty > qty:
            item_errors[item.id] = "Hold qty cannot be greater than received qty."
            continue

        if (accepted_qty + rejected_qty + hold_qty) != qty:
            item_errors[item.id] = "Accepted + Rejected + Hold must match Received Qty."
            continue

        if editing_inward is None:
            max_allowed_qty = item.remaining_qty_total or Decimal("0")
        else:
            other_inward_qty = (
                item.inward_items.exclude(inward=editing_inward).aggregate(total=Sum("quantity")).get("total")
                or Decimal("0")
            )
            max_allowed_qty = (item.quantity or Decimal("0")) - other_inward_qty

        if qty > max_allowed_qty:
            item_errors[item.id] = "Entered quantity is greater than remaining quantity."
            continue

        qc_status = _derive_dyeing_qc_status(qty, accepted_qty, rejected_qty, hold_qty)

        line_payload.append({
            "item": item,
            "quantity": qty,
            "received_qty": qty,
            "accepted_qty": accepted_qty,
            "rejected_qty": rejected_qty,
            "hold_qty": hold_qty,
            "actual_rolls": actual_rolls,
            "actual_gsm": actual_gsm,
            "actual_width": actual_width,
            "dye_lot_no": raw_dye_lot,
            "batch_no": raw_batch,
            "shade_reference": raw_shade,
            "qc_status": qc_status,
            "remark": raw_remark,
        })

    return line_payload, line_inputs, item_errors


@login_required
@require_POST
def dyeingpo_delete(request, pk: int):
    po = get_object_or_404(DyeingPurchaseOrder, pk=pk, owner=request.user)

    if po.ready_pos.exists():
        messages.error(request, "Cannot delete this Dyeing PO because Ready PO already exists.")
        return redirect("accounts:dyeingpo_list")

    if po.inwards.exists():
        messages.error(request, "Cannot delete this Dyeing PO because inward entries already exist.")
        return redirect("accounts:dyeingpo_list")

    if po.approval_status == "approved":
        messages.error(request, "Approved Dyeing PO cannot be deleted.")
        return redirect("accounts:dyeingpo_list")

    po.delete()
    messages.success(request, "Dyeing PO deleted successfully.")
    return redirect("accounts:dyeingpo_list")


@login_required
@require_http_methods(["GET", "POST"])
def dyeingpo_inward(request, pk: int):
    po = get_object_or_404(_dyeing_po_queryset(), pk=pk)

    if not _can_access_dyeing_po(request.user, po):
        raise PermissionDenied("You do not have access to this Dyeing PO.")

    if not _is_po_approved_for_inward(po):
        messages.error(request, "Dyeing PO must be approved before inward can be generated.")
        return redirect("accounts:dyeingpo_list")

    if po.ready_pos.exists():
        messages.error(request, "This Dyeing PO is locked because Ready PO already exists.")
        return redirect("accounts:dyeing_inward_tracker")

    inward_form = DyeingPOInwardForm(request.POST or None, user=request.user)

    if request.method == "GET":
        if "vendor" in inward_form.fields and po.vendor_id:
            inward_form.fields["vendor"].initial = po.vendor_id

        if "inward_type" in inward_form.fields:
            inward_type_qs = inward_form.fields["inward_type"].queryset
            if inward_type_qs.count() == 1:
                inward_form.fields["inward_type"].initial = inward_type_qs.first().pk

    item_errors = {}
    line_inputs = {}

    if request.method == "POST" and inward_form.is_valid():
        line_payload, line_inputs, item_errors = _collect_dyeing_inward_payload(request, po)

        if not line_payload:
            inward_form.add_error(None, "Enter at least one inward quantity.")

        if not inward_form.errors and not item_errors:
            with transaction.atomic():
                inward = inward_form.save(commit=False)
                inward.owner = po.owner
                inward.po = po
                inward.inward_number = _next_dyeing_inward_number()

                if not inward.vendor_id and po.vendor_id:
                    inward.vendor = po.vendor

                inward.save()

                DyeingPOInwardItem.objects.bulk_create([
                    DyeingPOInwardItem(
                        inward=inward,
                        po_item=row["item"],
                        quantity=row["quantity"],
                        received_qty=row["received_qty"],
                        accepted_qty=row["accepted_qty"],
                        rejected_qty=row["rejected_qty"],
                        hold_qty=row["hold_qty"],
                        actual_rolls=row["actual_rolls"],
                        actual_gsm=row["actual_gsm"],
                        actual_width=row["actual_width"],
                        dye_lot_no=row["dye_lot_no"],
                        batch_no=row["batch_no"],
                        shade_reference=row["shade_reference"],
                        qc_status=row["qc_status"],
                        remark=row["remark"],
                    )
                    for row in line_payload
                ])

            messages.success(request, f"Inward {inward.inward_number} saved successfully.")
            tracker_url = reverse("accounts:dyeing_inward_tracker")
            return redirect(f"{tracker_url}?inward={inward.pk}")

    line_rows = _build_dyeing_inward_line_rows(
        po,
        line_inputs=line_inputs,
        item_errors=item_errors,
    )

    return render(
        request,
        "accounts/dyeing_po/inward.html",
        {
            "po": po,
            "inward_form": inward_form,
            "line_rows": line_rows,
            "existing_inwards": po.inwards.all().order_by("-inward_date", "-id"),
            "existing_ready_po": po.ready_pos.order_by("-id").first(),
        },
    )


def _normalize_stock_lot_search_value(value: str) -> str:
    return (value or "").strip().lower()


def _build_stock_lot_rows_for_user(user):
    rows = []

    dyeing_inward_items = (
        DyeingPOInwardItem.objects
        .select_related(
            "inward",
            "po_item__po__vendor",
            "po_item__po__firm",
            "po_item__finished_material",
            "po_item__source_greige_po_item",
        )
        .filter(inward__owner=user)
        .order_by("-inward__inward_date", "-id")
    )

    for inward_item in dyeing_inward_items:
        inward = inward_item.inward
        po_item = inward_item.po_item
        po = po_item.po if po_item else None

        ready_material_name = "Ready Material"
        raw_material_name = "-"
        dyeing_name = "-"
        dyeing_type = "-"
        unit = "KG"

        if po_item:
            if po_item.finished_material:
                ready_material_name = po_item.finished_material.name
            elif po_item.fabric_name:
                ready_material_name = po_item.fabric_name
            else:
                ready_material_name = "Ready Material"

            raw_material_name = po_item.greige_name or "-"
            dyeing_name = po_item.dyeing_name or "-"
            dyeing_type = po_item.dyeing_type or "-"
            unit = po_item.unit or "KG"

        vendor_name = po.vendor.name if po and po.vendor else "-"
        firm_name = po.firm.firm_name if po and po.firm else "-"
        quantity = inward_item.accepted_qty or Decimal("0")

        rows.append({
            "stage": "dyeing",
            "stage_label": "Dyeing",
            "lot_number": inward.inward_number or f"DYEING-{inward.pk}",
            "lot_date": inward.inward_date,
            "material_name": ready_material_name,
            "material_key": _normalize_stock_lot_search_value(ready_material_name),
            "ready_material_name": ready_material_name,
            "raw_material_name": raw_material_name,
            "vendor_name": vendor_name,
            "firm_name": firm_name,
            "source_number": po.system_number if po and po.system_number else (po.po_number if po else "-"),
            "quantity": quantity,
            "used_quantity": Decimal("0"),
            "final_stock": quantity,
            "unit": unit,
            "remark": inward_item.remark or "",
            "dyeing_name": dyeing_name,
            "dyeing_type": dyeing_type,
            "detail_url": reverse("accounts:dyeingpo_inward", args=[po.pk]) if po else "",
            "detail_label": "Open Dyeing Inward",
            "pk": inward_item.pk,
        })

    rows.sort(
        key=lambda row: (
            row["lot_date"] or timezone.datetime.min.date(),
            row["pk"],
        ),
        reverse=True,
    )
    return rows


@login_required
def stock_lot_wise(request):
    _sync_phase2_lots_from_dyeing(request.user)

    q = (request.GET.get("q") or "").strip()
    selected_material = (request.GET.get("material") or "").strip()
    selected_stage = (request.GET.get("stage") or "").strip()
    selected_qc = (request.GET.get("qc") or "").strip()
    selected_location = (request.GET.get("location") or "").strip()
    selected_stock_type = (request.GET.get("stock_type") or "").strip()

    qs = (
        InventoryLot.objects.filter(owner=request.user)
        .select_related(
            "material",
            "dyeing_inward_item__inward__po",
            "ready_inward_item__inward__po",
        )
        .order_by("-updated_at", "-id")
    )

    if selected_material.isdigit():
        qs = qs.filter(material_id=int(selected_material))

    if selected_stage in {"yarn", "greige", "dyeing", "ready"}:
        qs = qs.filter(stage=selected_stage)

    if selected_qc in {"pending", "approved", "partial", "hold", "rejected"}:
        qs = qs.filter(qc_status=selected_qc)

    if selected_location:
        qs = qs.filter(location_name__iexact=selected_location)

    if selected_stock_type == "available":
        qs = qs.filter(available_qty__gt=0, is_closed=False)
    elif selected_stock_type == "zero":
        qs = qs.filter(available_qty__lte=0)
    elif selected_stock_type == "closed":
        qs = qs.filter(is_closed=True)
    elif selected_stock_type == "open":
        qs = qs.filter(is_closed=False)

    if q:
        qs = qs.filter(
            Q(lot_code__icontains=q)
            | Q(material__name__icontains=q)
            | Q(dye_lot_no__icontains=q)
            | Q(batch_no__icontains=q)
            | Q(shade_reference__icontains=q)
            | Q(location_name__icontains=q)
            | Q(dyeing_inward_item__inward__inward_number__icontains=q)
            | Q(ready_inward_item__inward__inward_number__icontains=q)
            | Q(dyeing_inward_item__inward__po__system_number__icontains=q)
            | Q(ready_inward_item__inward__po__system_number__icontains=q)
        )

    summary = qs.aggregate(
        total_lots=Count("id"),
        total_received=Sum("received_qty"),
        total_accepted=Sum("accepted_qty"),
        total_used=Sum("used_qty"),
        total_available=Sum("available_qty"),
        total_hold=Sum("hold_qty"),
        total_rejected=Sum("rejected_qty"),
    )
    summary["pending_qc"] = qs.filter(qc_status="pending").count()

    material_choices = (
        Material.objects.filter(inventory_lots__owner=request.user)
        .distinct()
        .order_by("name")
    )

    location_choices = sorted(
        {
            (value or "").strip()
            for value in InventoryLot.objects.filter(owner=request.user)
            .values_list("location_name", flat=True)
            if (value or "").strip()
        },
        key=lambda x: x.lower(),
    )

    paginator = Paginator(qs, 20)
    page_obj = paginator.get_page(request.GET.get("page"))

    context = {
        "rows": page_obj.object_list,
        "page_obj": page_obj,
        "q": q,
        "selected_material": selected_material,
        "selected_stage": selected_stage,
        "selected_qc": selected_qc,
        "selected_location": selected_location,
        "selected_stock_type": selected_stock_type,
        "material_choices": material_choices,
        "location_choices": location_choices,
        "summary": {
            "total_lots": summary.get("total_lots") or 0,
            "total_received": summary.get("total_received") or Decimal("0"),
            "total_accepted": summary.get("total_accepted") or Decimal("0"),
            "total_used": summary.get("total_used") or Decimal("0"),
            "total_available": summary.get("total_available") or Decimal("0"),
            "total_hold": summary.get("total_hold") or Decimal("0"),
            "total_rejected": summary.get("total_rejected") or Decimal("0"),
            "pending_qc": summary.get("pending_qc") or 0,
        },
        "stage_choices": [
            ("yarn", "Yarn"),
            ("greige", "Greige"),
            ("dyeing", "Dyeing"),
            ("ready", "Ready"),
        ],
        "qc_choices": [
            ("pending", "Pending"),
            ("approved", "Approved"),
            ("partial", "Partial"),
            ("hold", "Hold"),
            ("rejected", "Rejected"),
        ],
    }

    return render(request, "accounts/inventory/stock_lot_wise.html", context)


@login_required
def dyeing_inward_tracker(request):
    q = (request.GET.get("q") or "").strip()
    target_inward_id = (request.GET.get("inward") or "").strip()

    qs = _dyeing_po_queryset().filter(inwards__isnull=False).distinct()

    if not _can_review_yarn_po(request):
        qs = qs.filter(owner=request.user)

    if q:
        qs = qs.filter(
            Q(system_number__icontains=q)
            | Q(po_number__icontains=q)
            | Q(vendor__name__icontains=q)
            | Q(source_greige_po__system_number__icontains=q)
            | Q(source_greige_inward__inward_number__icontains=q)
            | Q(inwards__inward_number__icontains=q)
            | Q(items__fabric_name__icontains=q)
            | Q(items__finished_material__name__icontains=q)
            | Q(items__dyeing_name__icontains=q)
            | Q(firm__firm_name__icontains=q)
        ).distinct()

    rows = []
    for po in qs:
        inward_entries = []
        ready_po = po.ready_pos.order_by("-id").first()
        generated_count = 1 if ready_po else 0

        for inward in po.inwards.all():
            items = [_po_tracker_item_payload(inward_item, fallback_name="Dyeing Item") for inward_item in inward.items.all()]
            inward_entries.append({
                "inward": inward,
                "items": items,
                "is_target": str(inward.id) == target_inward_id,
                "linked_po": ready_po,
                "next_started": bool(ready_po),
                "next_view_url": reverse("accounts:readypo_detail", args=[ready_po.id]) if ready_po else "",
                "next_generate_url": f"{reverse('accounts:generate_ready_po_from_dyeing', args=[po.id])}?inward={inward.id}",
                "next_generate_method": "get",
                "next_generate_label": "Generate Ready PO",
                "next_view_label": "View Ready PO",
                "edit_url": reverse("accounts:dyeing_inward_edit", args=[inward.id]),
            })

        total_inwards = len(inward_entries)
        rows.append({
            "po": po,
            "inward_entries": inward_entries,
            "total_inwards": total_inwards,
            "generated_count": generated_count,
            "progress_label": "Created" if ready_po else "Pending",
            "progress_title": "Ready PO Progress",
            "next_list_label": "View Ready PO List",
            "next_list_url": reverse("accounts:readypo_list"),
            "inward_url": reverse("accounts:dyeingpo_inward", args=[po.id]),
            "total_qty": _po_tracker_qty(po, "total_weight", "total_qty", "total_quantity"),
            "inward_qty": _po_tracker_qty(po, "total_inward_qty", "inward_qty_total"),
            "remaining_qty": _po_tracker_qty(po, "remaining_qty_total", "pending_qty_total"),
        })

    return render(request, "accounts/dyeing_po/inward_tracker.html", {
        "rows": rows,
        "q": q,
        "target_inward_id": target_inward_id,
        "tracker_title": "Dyeing Inward Tracker",
        "tracker_subtitle": "Track inwarded dyed output, accepted stock, rejected qty, hold qty, and Ready PO linkage",
        "tracker_reset_url": reverse("accounts:dyeing_inward_tracker"),
        "tracker_list_label": "Dyeing POs",
        "tracker_list_url": reverse("accounts:dyeingpo_list"),
        "tracker_stock_url": reverse("accounts:stock_lot_wise"),
        "empty_message": "No inwarded Dyeing POs found yet.",
        "anchor_prefix": "dyeing-inward-",
    })
def dyeing_inward_tracker(request):
    q = (request.GET.get("q") or "").strip()
    target_inward_id = (request.GET.get("inward") or "").strip()

    qs = _dyeing_po_queryset().filter(inwards__isnull=False).distinct()

    if not _can_review_yarn_po(request):
        qs = qs.filter(owner=request.user)

    if q:
        qs = qs.filter(
            Q(system_number__icontains=q)
            | Q(po_number__icontains=q)
            | Q(vendor__name__icontains=q)
            | Q(source_greige_po__system_number__icontains=q)
            | Q(source_greige_inward__inward_number__icontains=q)
            | Q(inwards__inward_number__icontains=q)
            | Q(items__fabric_name__icontains=q)
            | Q(items__finished_material__name__icontains=q)
            | Q(items__dyeing_name__icontains=q)
            | Q(firm__firm_name__icontains=q)
        ).distinct()

    rows = []
    for po in qs:
        inward_entries = []

        for inward in po.inwards.all():
            items = []

            for inward_item in inward.items.select_related("po_item__finished_material").all():
                po_item = inward_item.po_item
                if po_item is None:
                    continue

                expected_qty = (
                    po_item.expected_output_qty
                    or po_item.quantity
                    or Decimal("0")
                )

                items.append({
                    "inward_item": inward_item,
                    "po_item": po_item,
                    "fabric_name": po_item.fabric_name or "Dyeing Item",
                    "ordered_qty": expected_qty,
                    "received_qty": inward_item.received_qty or inward_item.quantity or Decimal("0"),
                    "accepted_qty": inward_item.accepted_qty or Decimal("0"),
                    "rejected_qty": inward_item.rejected_qty or Decimal("0"),
                    "hold_qty": inward_item.hold_qty or Decimal("0"),
                    "actual_rolls": inward_item.actual_rolls or Decimal("0"),
                    "actual_gsm": inward_item.actual_gsm,
                    "actual_width": inward_item.actual_width,
                    "dye_lot_no": inward_item.dye_lot_no,
                    "batch_no": inward_item.batch_no,
                    "shade_reference": inward_item.shade_reference,
                    "qc_status": inward_item.qc_status,
                    "unit": po_item.unit or "",
                })

            inward_entries.append({
                "inward": inward,
                "items": items,
                "is_target": str(inward.id) == target_inward_id,
            })

        rows.append({
            "po": po,
            "inward_entries": inward_entries,
            "ready_po": po.ready_pos.order_by("-id").first(),
        })

    return render(
        request,
        "accounts/dyeing_po/inward_tracker.html",
        {
            "rows": rows,
            "q": q,
            "target_inward_id": target_inward_id,
        },
    )

@login_required
@require_http_methods(["GET", "POST"])
def dyeing_inward_edit(request, pk: int):
    inward = get_object_or_404(
        DyeingPOInward.objects.select_related("po__vendor", "po__firm", "po__owner"),
        pk=pk,
    )
    po = inward.po

    if not _is_po_approved_for_inward(po):
        messages.error(request, "Dyeing PO must be approved before inward can be updated.")
        return redirect("accounts:dyeing_inward_tracker")

    if po.ready_pos.exists():
        messages.error(
            request,
            "This inward cannot be edited because a Ready PO has already been generated from this Dyeing PO."
        )
        tracker_url = reverse("accounts:dyeing_inward_tracker")
        return redirect(f"{tracker_url}?inward={inward.pk}")

    if not _can_access_dyeing_po(request.user, po):
        raise PermissionDenied("You do not have access to this Dyeing inward.")

    line_inputs = {
        row.po_item_id: {
            "qty": row.received_qty or row.quantity,
            "accepted_qty": row.accepted_qty,
            "rejected_qty": row.rejected_qty,
            "hold_qty": row.hold_qty,
            "actual_rolls": row.actual_rolls,
            "actual_gsm": row.actual_gsm if row.actual_gsm is not None else "",
            "actual_width": row.actual_width if row.actual_width is not None else "",
            "dye_lot_no": row.dye_lot_no or "",
            "batch_no": row.batch_no or "",
            "shade_reference": row.shade_reference or "",
            "remark": row.remark or "",
        }
        for row in inward.items.all()
    }
    item_errors = {}

    inward_form = DyeingPOInwardForm(request.POST or None, instance=inward, user=request.user)

    if request.method == "GET" and not inward.vendor_id and po.vendor_id:
        inward_form.fields["vendor"].initial = po.vendor_id
    
    if request.method == "POST" and inward_form.is_valid():
        line_payload, posted_inputs, item_errors = _collect_dyeing_inward_payload(
            request, po, editing_inward=inward
        )
        line_inputs = posted_inputs

        if not line_payload:
            inward_form.add_error(None, "Enter at least one inward quantity.")

        if not inward_form.errors and not item_errors:
            with transaction.atomic():
                inward = inward_form.save(commit=False)
                inward.owner = po.owner
                inward.po = po

                if not inward.vendor_id and po.vendor_id:
                    inward.vendor = po.vendor

                inward.save()
                inward.items.all().delete()

                DyeingPOInwardItem.objects.bulk_create([
                    DyeingPOInwardItem(
                        inward=inward,
                        po_item=row["item"],
                        quantity=row["quantity"],
                        received_qty=row["received_qty"],
                        accepted_qty=row["accepted_qty"],
                        rejected_qty=row["rejected_qty"],
                        hold_qty=row["hold_qty"],
                        actual_rolls=row["actual_rolls"],
                        actual_gsm=row["actual_gsm"],
                        actual_width=row["actual_width"],
                        dye_lot_no=row["dye_lot_no"],
                        batch_no=row["batch_no"],
                        shade_reference=row["shade_reference"],
                        qc_status=row["qc_status"],
                        remark=row["remark"],
                    )
                    for row in line_payload
                ])

            messages.success(request, f"Inward {inward.inward_number} updated successfully.")
            tracker_url = reverse("accounts:dyeing_inward_tracker")
            return redirect(f"{tracker_url}?inward={inward.pk}")

    line_rows = _build_dyeing_inward_line_rows(po, line_inputs=line_inputs, item_errors=item_errors)

    return render(
        request,
        "accounts/dyeing_po/inward.html",
        {
            "po": po,
            "inward_form": inward_form,
            "line_rows": line_rows,
            "existing_inwards": po.inwards.all().order_by("-inward_date", "-id"),
            "existing_ready_po": po.ready_pos.order_by("-id").first(),
            "editing_inward": inward,
            "next_inward_number_preview": inward.inward_number,
        },
    )



@login_required
@require_POST
def dyeing_inward_delete(request, pk: int):
    inward = get_object_or_404(
        DyeingPOInward.objects.select_related("po__owner", "po__vendor", "po__firm"),
        pk=pk,
    )
    po = inward.po

    if not _can_access_dyeing_po(request.user, po):
        raise PermissionDenied("You do not have access to this Dyeing inward.")

    tracker_url = reverse("accounts:dyeing_inward_tracker")

    if po.ready_pos.exists():
        messages.error(
            request,
            "This inward cannot be deleted because a Ready PO has already been generated from this Dyeing PO."
        )
        return redirect(f"{tracker_url}?inward={inward.pk}")

    linked_lots = InventoryLot.objects.filter(dyeing_inward_item__inward=inward)
    used_or_closed_lots = linked_lots.filter(Q(used_qty__gt=0) | Q(is_closed=True))
    if used_or_closed_lots.exists():
        messages.error(
            request,
            "This inward cannot be deleted because its stock lot is already used or closed."
        )
        return redirect(f"{tracker_url}?inward={inward.pk}")

    inward_number = inward.inward_number

    with transaction.atomic():
        linked_lots.delete()
        inward.delete()

    messages.success(request, f"Inward {inward_number} deleted successfully.")
    return redirect(tracker_url)

@login_required
@require_http_methods(["GET", "POST"])
def readypo_create(request, dyeing_po_id=None):
    source_dyeing_po = None
    selected_source = None
    existing_po = None
    posted_source_id = ""

    if dyeing_po_id is not None:
        source_dyeing_po = get_object_or_404(_dyeing_po_queryset(), pk=dyeing_po_id)

        if not _can_access_dyeing_po(request.user, source_dyeing_po):
            raise PermissionDenied("You do not have access to this Dyeing PO.")

        if not _is_po_approved_for_inward(source_dyeing_po):
            messages.error(request, "Dyeing PO must be approved before Ready PO can be created.")
            return redirect("accounts:dyeingpo_list")

        if not source_dyeing_po.inwards.exists():
            messages.error(request, "Create at least one Dyeing inward before generating Ready PO.")
            return redirect("accounts:dyeingpo_inward", pk=source_dyeing_po.pk)

        existing_po = source_dyeing_po.ready_pos.order_by("-id").first()
        if existing_po and request.method == "GET":
            messages.info(request, "Ready PO already exists for this Dyeing PO.")
            return redirect("accounts:readypo_detail", pk=existing_po.pk)

    if request.method == "POST":
        form = ReadyPurchaseOrderForm(
            request.POST,
            user=request.user,
            source_dyeing_po=source_dyeing_po,
            lock_source=bool(source_dyeing_po),
        )

        posted_source_id = (
            str(source_dyeing_po.pk)
            if source_dyeing_po
            else (request.POST.get("source_dyeing_po") or "").strip()
        )

        if posted_source_id:
            selected_source = _dyeing_po_queryset().filter(pk=posted_source_id).first()

        temp_po = ReadyPurchaseOrder(source_dyeing_po=selected_source)

        item_formset = ReadyPurchaseOrderItemFormSet(
            request.POST,
            instance=temp_po,
            prefix="items",
            form_kwargs={"user": request.user},
        )

        if form.is_valid():
            selected_source = source_dyeing_po or form.cleaned_data.get("source_dyeing_po")

            if selected_source is not None:
                if isinstance(selected_source, str):
                    selected_source = _dyeing_po_queryset().filter(pk=selected_source).first()

                if not selected_source:
                    form.add_error("source_dyeing_po", "Select a valid Dyeing PO.")

                elif not _can_access_dyeing_po(request.user, selected_source):
                    raise PermissionDenied("You do not have access to this Dyeing PO.")

                elif not _is_po_approved_for_inward(selected_source):
                    form.add_error("source_dyeing_po", "Selected Dyeing PO must be approved before Ready PO can be created.")

                elif not selected_source.inwards.exists():
                    form.add_error("source_dyeing_po", "Selected Dyeing PO must have inward entries before Ready PO can be created.")

                elif not any((item.accepted_inward_qty_total or Decimal("0")) > 0 for item in selected_source.items.all()):
                    form.add_error("source_dyeing_po", "Accepted dyed qty is required before Ready PO can be created.")

                elif selected_source.ready_pos.exists():
                    existing_po = selected_source.ready_pos.order_by("-id").first()
                    form.add_error("source_dyeing_po", "Ready PO already exists for this Dyeing PO.")

                else:
                    with transaction.atomic():
                        po = form.save(commit=False)
                        po.owner = selected_source.owner
                        po.system_number = _next_ready_po_number()
                        po.source_dyeing_po = selected_source

                        if not po.firm_id and getattr(selected_source, "firm_id", None):
                            po.firm = selected_source.firm

                        if not po.vendor_id and getattr(selected_source, "vendor_id", None):
                            po.vendor = selected_source.vendor

                        if not po.shipping_address and po.firm:
                            po.shipping_address = _firm_address(po.firm)

                        if not po.terms_conditions:
                            po.terms_conditions = getattr(selected_source, "terms_conditions", "") or ""

                        po.save()
                        _sync_ready_po_items_from_source(po)

                    messages.success(request, f"Ready PO {po.system_number} created successfully.")
                    return redirect("accounts:readypo_list")

            else:
                if item_formset.is_valid():
                    with transaction.atomic():
                        po = form.save(commit=False)
                        po.owner = request.user
                        po.system_number = _next_ready_po_number()
                        po.source_dyeing_po = None

                        if po.firm and not po.shipping_address:
                            po.shipping_address = _firm_address(po.firm)

                        po.save()

                        item_formset.instance = po
                        items = item_formset.save(commit=False)

                        for obj in item_formset.deleted_objects:
                            obj.delete()

                        for item in items:
                            item.po = po
                            item.source_dyeing_po_item = None

                            if item.finished_material:
                                item.fabric_name = item.finished_material.name
                                material_unit = getattr(item.finished_material, "unit", None)

                                if not item.unit:
                                    if hasattr(material_unit, "name"):
                                        item.unit = material_unit.name
                                    elif material_unit:
                                        item.unit = str(material_unit)

                            item.save()

                        item_formset.save_m2m()
                        _recalculate_ready_po(po)

                    messages.success(request, f"Ready PO {po.system_number} created successfully.")
                    return redirect("accounts:readypo_list")

                form.add_error(None, "Ready PO was not saved. Please fix item rows and try again.")
        else:
            form.add_error(None, "Ready PO was not saved. Please fix the highlighted fields and try again.")

        selected_source = source_dyeing_po
        if selected_source is None and posted_source_id:
            selected_source = _dyeing_po_queryset().filter(pk=posted_source_id).first()

    else:
        form = ReadyPurchaseOrderForm(
            user=request.user,
            source_dyeing_po=source_dyeing_po,
            lock_source=bool(source_dyeing_po),
        )

        selected_source = source_dyeing_po
        temp_po = ReadyPurchaseOrder(source_dyeing_po=selected_source)

        item_formset = ReadyPurchaseOrderItemFormSet(
            instance=temp_po,
            prefix="items",
            form_kwargs={"user": request.user},
        )

    source_inwards = []
    effective_source = source_dyeing_po or selected_source

    if effective_source is not None and isinstance(effective_source, str):
        effective_source = _dyeing_po_queryset().filter(pk=effective_source).first()

    if effective_source is not None:
        source_inwards = effective_source.inwards.prefetch_related("items__po_item").all()

    is_linked_mode = effective_source is not None

    po_header_maps = _po_header_maps_from_form(form)

    return render(
        request,
        "accounts/ready_po/form.html",
        {
            "form": form,
            "item_formset": item_formset,
            "mode": "create",
            "po_obj": None,
            "system_number_preview": _next_ready_po_number(),
            "source_dyeing_po": effective_source,
            "source_inwards": source_inwards,
            "existing_po": existing_po,
            "is_linked_mode": is_linked_mode,
            **po_header_maps,
        },
    )

@login_required
def readypo_list(request):
    q = (request.GET.get("q") or "").strip()

    qs = _ready_po_queryset()
    if not _can_review_yarn_po(request):
        qs = qs.filter(owner=request.user)

    if q:
        qs = qs.filter(
            Q(system_number__icontains=q)
            | Q(po_number__icontains=q)
            | Q(vendor__name__icontains=q)
            | Q(source_dyeing_po__system_number__icontains=q)
            | Q(firm__firm_name__icontains=q)
        ).distinct()

    return render(
        request,
        "accounts/ready_po/list.html",
        {
            "orders": qs,
            "q": q,
            "can_review_ready_po": _can_review_yarn_po(request),
        },
    )

@login_required
def readypo_review(request, pk: int):
    po = get_object_or_404(
        _ready_po_queryset(),
        pk=pk,
    )

    if not _can_access_ready_po(request.user, po):
        raise PermissionDenied("You do not have access to this Ready PO.")

    review_field_names = ("approval_status", "rejection_reason", "reviewed_by", "reviewed_at")
    review_templates_available = (
        _template_exists("accounts/ready_po/review.html")
        and _template_exists("accounts/ready_po/review_embed.html")
    )

    if not _model_has_fields(ReadyPurchaseOrder, *review_field_names) or not review_templates_available:
        message = "Ready PO review needs model review fields and review templates before this page can work."
        if _is_embed(request):
            return JsonResponse({
                "ok": False,
                "message": message,
                "redirect_url": reverse("accounts:readypo_list"),
            }, status=400)
        messages.error(request, message)
        return redirect("accounts:readypo_list")

    embed_mode = _is_embed(request)
    can_review = _can_review_yarn_po(request)

    review_form = ReadyPOReviewForm(request.POST or None)

    review_checks = {
        "has_header": bool(po.vendor_id and po.po_date),
        "has_source": bool(po.source_dyeing_po_id),
        "has_items": po.items.exists(),
        "has_shipping": bool((po.shipping_address or "").strip()),
    }
    review_ready_count = sum(1 for value in review_checks.values() if value)

    context = {
        "po": po,
        "review_form": review_form,
        "can_review_ready_po": can_review,
        "embed_mode": embed_mode,
        "review_checks": review_checks,
        "review_ready_count": review_ready_count,
    }

    if request.method == "POST":
        if not can_review:
            return HttpResponseForbidden("You are not allowed to review this PO.")

        if review_form.is_valid():
            decision = review_form.cleaned_data["decision"]

            if decision == "approve":
                po.approval_status = "approved"
                po.rejection_reason = ""
            else:
                po.approval_status = "rejected"
                po.rejection_reason = review_form.cleaned_data["rejection_reason"].strip()

            po.reviewed_by = get_actor(request) or request.user
            po.reviewed_at = timezone.now()
            po.save(update_fields=[
                "approval_status",
                "rejection_reason",
                "reviewed_by",
                "reviewed_at",
            ])

            if embed_mode:
                return JsonResponse({
                    "ok": True,
                    "message": "Ready PO reviewed successfully.",
                    "redirect_url": reverse("accounts:readypo_list"),
                })

            return redirect("accounts:readypo_list")

        if embed_mode:
            return render(
                request,
                "accounts/ready_po/review_embed.html",
                context,
                status=400,
            )

    template_name = (
        "accounts/ready_po/review_embed.html"
        if embed_mode
        else "accounts/ready_po/review.html"
    )

    return render(request, template_name, context)

@login_required
@require_http_methods(["GET", "POST"])
def readypo_update(request, pk: int):
    po = get_object_or_404(_ready_po_queryset(), pk=pk)
    if not _can_access_ready_po(request.user, po):
        raise PermissionDenied("You do not have access to this Ready PO.")

    has_inwards = po.inwards.exists()

    if has_inwards:
        messages.error(request, "This Ready PO cannot be edited because inward entries already exist.")
        return redirect("accounts:readypo_list")

    if po.approval_status == "approved":
        messages.error(request, "Approved Ready PO cannot be edited.")
        return redirect("accounts:readypo_list")

    initial_source = po.source_dyeing_po
    lock_source = bool(initial_source) or has_inwards

    form = ReadyPurchaseOrderForm(
        request.POST or None,
        user=request.user,
        instance=po,
        source_dyeing_po=initial_source,
        lock_source=lock_source,
    )

    item_formset = ReadyPurchaseOrderItemFormSet(
        request.POST or None,
        instance=po,
        prefix="items",
        form_kwargs={"user": request.user},
    )

    selected_source = initial_source
    if request.method == "POST":
        posted_source_id = initial_source.pk if initial_source else (request.POST.get("source_dyeing_po") or "").strip()
        if posted_source_id:
            selected_source = _dyeing_po_queryset().filter(pk=posted_source_id).first()
        else:
            selected_source = None

        if form.is_valid():
            chosen_source = initial_source if lock_source else form.cleaned_data.get("source_dyeing_po")

            if has_inwards and ((po.source_dyeing_po_id and chosen_source is None) or ((not po.source_dyeing_po_id) and chosen_source is not None)):
                form.add_error("source_dyeing_po", "Cannot change Ready PO mode after inward entries exist.")
            elif chosen_source is not None:
                if not _can_access_dyeing_po(request.user, chosen_source):
                    raise PermissionDenied("You do not have access to this Dyeing PO.")
                elif not _is_po_approved_for_inward(chosen_source):
                    form.add_error("source_dyeing_po", "Selected Dyeing PO must be approved before Ready PO can be created.")
                elif not chosen_source.inwards.exists():
                    form.add_error("source_dyeing_po", "Selected Dyeing PO must have inward entries before Ready PO can be created.")
                elif not any((item.accepted_inward_qty_total or Decimal("0")) > 0 for item in chosen_source.items.all()):
                    form.add_error("source_dyeing_po", "Accepted dyed qty is required before Ready PO can be created.")
                elif chosen_source.ready_pos.exclude(pk=po.pk).exists():
                    form.add_error("source_dyeing_po", "Another Ready PO already exists for this Dyeing PO.")
                else:
                    with transaction.atomic():
                        po = form.save(commit=False)
                        po.source_dyeing_po = chosen_source

                        if po.firm and not po.shipping_address:
                            po.shipping_address = _firm_address(po.firm)

                        po.save()

                        if not has_inwards:
                            _sync_ready_po_items_from_source(po)

                    messages.success(request, f"Ready PO {po.system_number} updated successfully.")
                    return redirect("accounts:readypo_detail", pk=po.pk)

            else:
                if has_inwards:
                    form.add_error(None, "Manual Ready PO items cannot be changed after inward entries exist.")
                elif item_formset.is_valid():
                    with transaction.atomic():
                        po = form.save(commit=False)
                        po.source_dyeing_po = None

                        if po.firm and not po.shipping_address:
                            po.shipping_address = _firm_address(po.firm)

                        po.save()

                        item_formset.instance = po
                        items = item_formset.save(commit=False)

                        for obj in item_formset.deleted_objects:
                            obj.delete()

                        for item in items:
                            item.po = po
                            item.source_dyeing_po_item = None

                            if item.finished_material:
                                item.fabric_name = item.finished_material.name
                                material_unit = getattr(item.finished_material, "unit", None)
                                if not item.unit:
                                    if hasattr(material_unit, "name"):
                                        item.unit = material_unit.name
                                    elif material_unit:
                                        item.unit = str(material_unit)

                            item.save()

                        item_formset.save_m2m()
                        _recalculate_ready_po(po)

                    messages.success(request, f"Ready PO {po.system_number} updated successfully.")
                    return redirect("accounts:readypo_detail", pk=po.pk)

    effective_source = initial_source if lock_source else selected_source
    source_inwards = list(effective_source.inwards.all()) if effective_source else []
    is_linked_mode = effective_source is not None

    return render(
        request,
        "accounts/ready_po/form.html",
        {
            "form": form,
            "item_formset": item_formset,
            "mode": "edit",
            "po_obj": po,
            "source_dyeing_po": effective_source,
            "source_inwards": source_inwards,
            "existing_po": None,
            "is_linked_mode": is_linked_mode,
            "has_inwards": has_inwards,
        },
    )


@login_required
def readypo_detail(request, pk: int):
    po = get_object_or_404(_ready_po_queryset(), pk=pk)
    if not _can_access_ready_po(request.user, po):
        raise PermissionDenied("You do not have access to this Ready PO.")

    return render(
        request,
        "accounts/ready_po/detail.html",
        {
            "po": po,
            "source_dyeing_po": po.source_dyeing_po,
            "source_inwards": list(po.source_dyeing_po.inwards.all()) if po.source_dyeing_po else [],
            "ready_inwards": list(po.inwards.all()),
        },
    )


@login_required
@require_POST
def readypo_delete(request, pk: int):
    po = get_object_or_404(ReadyPurchaseOrder, pk=pk, owner=request.user)

    if po.inwards.exists():
        messages.error(request, "Cannot delete this Ready PO because inward entries already exist.")
        return redirect("accounts:readypo_list")

    if po.approval_status == "approved":
        messages.error(request, "Approved Ready PO cannot be deleted.")
        return redirect("accounts:readypo_list")

    po.delete()
    messages.success(request, "Ready PO deleted successfully.")
    return redirect("accounts:readypo_list")


@login_required
@require_http_methods(["GET", "POST"])
def readypo_inward(request, pk: int):
    po = get_object_or_404(_ready_po_queryset(), pk=pk)

    if not _can_access_ready_po(request.user, po):
        raise PermissionDenied("You do not have access to this Ready PO.")

    if not _is_po_approved_for_inward(po):
        messages.error(request, "Ready PO must be approved before inward can be generated.")
        return redirect("accounts:readypo_list")

    item_errors = {}
    line_inputs = {}
    inward_form = ReadyPOInwardForm(request.POST or None, user=request.user)

    if request.method == "GET":
        if "vendor" in inward_form.fields and po.vendor_id:
            inward_form.fields["vendor"].initial = po.vendor_id
        if "inward_type" in inward_form.fields:
            inward_type_qs = inward_form.fields["inward_type"].queryset
            if inward_type_qs.count() == 1:
                inward_form.fields["inward_type"].initial = inward_type_qs.first().pk

    if request.method == "POST" and inward_form.is_valid():
        line_payload, line_inputs, item_errors = _collect_ready_inward_lines(request, po)

        if not line_payload:
            inward_form.add_error(None, "Enter at least one inward row.")

        if not inward_form.errors and not item_errors:
            with transaction.atomic():
                inward = inward_form.save(commit=False)
                inward.owner = po.owner
                inward.po = po
                inward.inward_number = _next_ready_inward_number()
                if not inward.vendor_id and po.vendor_id:
                    inward.vendor = po.vendor
                inward.save()

                ReadyPOInwardItem.objects.bulk_create([
                    ReadyPOInwardItem(
                        inward=inward,
                        po_item=row["item"],
                        quantity=row["quantity"],
                        received_qty=row["received_qty"],
                        accepted_qty=row["accepted_qty"],
                        rejected_qty=row["rejected_qty"],
                        hold_qty=row["hold_qty"],
                        actual_rolls=row["actual_rolls"],
                        actual_gsm=row["actual_gsm"],
                        actual_width=row["actual_width"],
                        dye_lot_no=row["dye_lot_no"],
                        batch_no=row["batch_no"],
                        shade_reference=row["shade_reference"],
                        qc_status=row["qc_status"],
                        remark=row["remark"],
                    )
                    for row in line_payload
                ])

            messages.success(request, f"Inward {inward.inward_number} saved successfully.")
            tracker_url = reverse("accounts:ready_inward_tracker")
            return redirect(f"{tracker_url}?inward={inward.pk}")

    line_rows = _build_dyeing_inward_line_rows(po, line_inputs=line_inputs, item_errors=item_errors)


    return render(
        request,
        "accounts/ready_po/inward.html",
        {
            "po": po,
            "inward_form": inward_form,
            "line_rows": line_rows,
            "existing_inwards": po.inwards.all().order_by("-inward_date", "-id"),
            "next_inward_number_preview": _next_ready_inward_number(),
        },
    )


@login_required
def ready_inward_tracker(request):
    q = (request.GET.get("q") or "").strip()
    target_inward_id = (request.GET.get("inward") or "").strip()

    qs = _ready_po_queryset().filter(inwards__isnull=False).distinct()
    if not _can_review_yarn_po(request):
        qs = qs.filter(owner=request.user)

    if q:
        qs = qs.filter(
            Q(system_number__icontains=q)
            | Q(po_number__icontains=q)
            | Q(vendor__name__icontains=q)
            | Q(source_dyeing_po__system_number__icontains=q)
            | Q(source_dyeing_po__inwards__inward_number__icontains=q)
            | Q(inwards__inward_number__icontains=q)
            | Q(firm__firm_name__icontains=q)
            | Q(items__fabric_name__icontains=q)
        ).distinct()

    rows = []
    for po in qs:
        inward_entries = []
        for inward in po.inwards.all():
            items = [_po_tracker_item_payload(inward_item, fallback_name="Ready Item") for inward_item in inward.items.all()]
            inward_entries.append({
                "inward": inward,
                "items": items,
                "is_target": str(inward.id) == target_inward_id,
                "linked_po": None,
                "next_started": True,
                "next_view_url": "",
                "next_generate_url": "",
                "next_generate_method": "",
                "next_generate_label": "Completed",
                "next_view_label": "Completed",
                "edit_url": reverse("accounts:ready_inward_edit", args=[inward.id]),
            })

        total_inwards = len(inward_entries)
        rows.append({
            "po": po,
            "inward_entries": inward_entries,
            "total_inwards": total_inwards,
            "generated_count": total_inwards,
            "progress_label": "Completed" if total_inwards else "Pending",
            "progress_title": "Ready Stock Progress",
            "next_list_label": "Stock Lot Wise",
            "next_list_url": reverse("accounts:stock_lot_wise"),
            "inward_url": reverse("accounts:readypo_inward", args=[po.id]),
            "total_qty": _po_tracker_qty(po, "total_weight", "total_qty", "total_quantity"),
            "inward_qty": _po_tracker_qty(po, "total_inward_qty", "inward_qty_total"),
            "remaining_qty": _po_tracker_qty(po, "remaining_qty_total", "pending_qty_total"),
        })

    return render(request, "accounts/ready_po/inward_tracker.html", {
        "rows": rows,
        "q": q,
        "target_inward_id": target_inward_id,
        "tracker_title": "Ready Inward Tracker",
        "tracker_subtitle": "Track inwarded ready fabric, accepted stock, rejected qty, hold qty, and stock movement",
        "tracker_reset_url": reverse("accounts:ready_inward_tracker"),
        "tracker_list_label": "Ready POs",
        "tracker_list_url": reverse("accounts:readypo_list"),
        "tracker_stock_url": reverse("accounts:stock_lot_wise"),
        "empty_message": "No inwarded Ready POs found yet.",
        "anchor_prefix": "ready-inward-",
    })
def ready_inward_tracker(request):
    q = (request.GET.get("q") or "").strip()
    target_inward_id = (request.GET.get("inward") or "").strip()

    qs = _ready_po_queryset().filter(inwards__isnull=False).distinct()
    if not _can_review_yarn_po(request):
        qs = qs.filter(owner=request.user)

    if q:
        qs = qs.filter(
            Q(system_number__icontains=q)
            | Q(po_number__icontains=q)
            | Q(vendor__name__icontains=q)
            | Q(source_dyeing_po__system_number__icontains=q)
            | Q(firm__firm_name__icontains=q)
        ).distinct()

    rows = []
    for po in qs:
        inward_entries = []

        for inward in po.inwards.all():
            inward_entries.append({
                "inward": inward,
                "items": [
                    {
                        "inward_item": inward_item,
                        "po_item": inward_item.po_item,
                        "fabric_name": inward_item.po_item.fabric_name if inward_item.po_item else "Ready Item",
                        "ordered_qty": inward_item.po_item.quantity if inward_item.po_item else 0,
                        "inward_qty": inward_item.quantity,
                        "unit": inward_item.po_item.unit if inward_item.po_item else "",
                    }
                    for inward_item in inward.items.all()
                ],
                "is_target": str(inward.id) == target_inward_id,
            })

        rows.append({
            "po": po,
            "inward_entries": inward_entries,
        })

    return render(
        request,
        "accounts/ready_po/inward_tracker.html",
        {
            "rows": rows,
            "q": q,
            "target_inward_id": target_inward_id,
        },
    )
    
    
# ==========================
# BRANDS (embed supported)
# ==========================
def _brand_list_url(request):
    url = reverse("accounts:brand_list")
    if _is_embed(request):
        url += "?embed=1"
    return url


def _brand_usage_rows(brand):
    rows = []

    bom_count = 0
    try:
        bom_count = BOM.objects.filter(owner=brand.owner, brand=brand).count()
    except Exception:
        bom_count = 0

    if bom_count:
        rows.append({"label": "BOM Records", "count": bom_count})

    return rows


def _brand_list_context(request):
    q = (request.GET.get("q") or "").strip()

    qs = Brand.objects.filter(owner=request.user)
    if q:
        qs = qs.filter(
            Q(name__icontains=q)
            | Q(description__icontains=q)
        )

    all_brands = Brand.objects.filter(owner=request.user)

    return {
        "brands": qs.order_by("name"),
        "q": q,
        "stats": {
            "total": all_brands.count(),
            "with_description": all_brands.exclude(description="").exclude(description__isnull=True).count(),
        },
    }


@login_required
def brand_list(request):
    tpl = "accounts/brands/list_embed.html" if _is_embed(request) else "accounts/brands/list.html"
    return render(request, tpl, _brand_list_context(request))


@login_required
@require_http_methods(["GET", "POST"])
def brand_create(request):
    form = BrandForm(request.POST or None, user=request.user)

    if request.method == "POST":
        form.instance.owner = request.user

        if form.is_valid():
            obj = form.save(commit=False)
            obj.owner = request.user
            obj.save()

            url = _brand_list_url(request)
            if _is_embed(request):
                return JsonResponse({"ok": True, "url": url})
            return redirect(url)

    tpl = "accounts/brands/form_embed.html" if _is_embed(request) else "accounts/brands/form.html"
    return render(request, tpl, {"form": form, "mode": "add"})


@login_required
@require_http_methods(["GET", "POST"])
def brand_update(request, pk: int):
    brand = get_object_or_404(Brand, pk=pk, owner=request.user)
    form = BrandForm(request.POST or None, instance=brand, user=request.user)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.owner = request.user
        obj.save()

        url = _brand_list_url(request)
        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})
        return redirect(url)

    tpl = "accounts/brands/form_embed.html" if _is_embed(request) else "accounts/brands/form.html"
    return render(request, tpl, {"form": form, "mode": "edit", "brand": brand})


@login_required
@require_POST
def brand_delete(request, pk: int):
    brand = get_object_or_404(Brand, pk=pk, owner=request.user)
    usage_rows = _brand_usage_rows(brand)

    if usage_rows:
        error_message = (
            f'Cannot delete "{brand.name}" because it is already used in linked records.'
        )

        if _is_embed(request):
            context = _brand_list_context(request)
            context["delete_error"] = error_message
            context["delete_label"] = brand.name
            context["delete_usage"] = usage_rows
            return render(request, "accounts/brands/list_embed.html", context)

        messages.error(request, error_message)
        return redirect("accounts:brand_list")

    brand.delete()

    url = _brand_list_url(request)
    if _is_embed(request):
        return JsonResponse({"ok": True, "url": url})
    return redirect(url)

def _materialunit_list_url(request):
    url = reverse("accounts:materialunit_list")
    if _is_embed(request):
        url += "?embed=1"
    return url


def _materialunit_usage_rows(unit):
    rows = []

    for rel in unit._meta.related_objects:
        accessor_name = rel.get_accessor_name()
        if not accessor_name:
            continue

        try:
            related_manager = getattr(unit, accessor_name)
            count = related_manager.count()
        except Exception:
            continue

        if count:
            label = rel.related_model._meta.verbose_name_plural.title()
            rows.append({"label": label, "count": count})

    return rows


def _materialunit_list_context(request, form=None):
    q = (request.GET.get("q") or "").strip()

    units = MaterialUnit.objects.filter(owner=request.user).order_by("name")
    if q:
        units = units.filter(name__icontains=q)

    all_units = MaterialUnit.objects.filter(owner=request.user)

    page_data = _paginate_utility_queryset(request, units)

    ctx = {
        "units": page_data.pop("object_list"),
        "form": form if form is not None else MaterialUnitForm(user=request.user),
        "q": q,
        "stats": {
            "total": all_units.count(),
        },
    }
    ctx.update(page_data)
    return ctx


@login_required
@require_http_methods(["GET", "POST"])
def materialunit_list_create(request):
    form = MaterialUnitForm(request.POST or None, user=request.user)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.owner = request.user
        obj.save()

        url = _materialunit_list_url(request)
        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})
        return redirect(url)

    template = "accounts/material_units/list_embed.html" if _is_embed(request) else "accounts/material_units/list.html"
    return render(request, template, _materialunit_list_context(request, form=form))


@login_required
@require_http_methods(["GET", "POST"])
def materialunit_edit(request, pk: int):
    unit = get_object_or_404(MaterialUnit, pk=pk, owner=request.user)
    form = MaterialUnitForm(request.POST or None, instance=unit, user=request.user)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.owner = request.user
        obj.save()

        url = _materialunit_list_url(request)
        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})
        return redirect(url)

    template = "accounts/material_units/edit_embed.html" if _is_embed(request) else "accounts/material_units/edit.html"
    return render(
        request,
        template,
        {
            "form": form,
            "unit": unit,
        },
    )


@login_required
@require_POST
def materialunit_delete(request, pk: int):
    unit = get_object_or_404(MaterialUnit, pk=pk, owner=request.user)
    usage_rows = _materialunit_usage_rows(unit)

    if usage_rows:
        error_message = (
            f'Cannot delete "{unit.name}" because it is already used in linked records.'
        )

        if _is_embed(request):
            ctx = _materialunit_list_context(request)
            ctx["delete_error"] = error_message
            ctx["delete_label"] = unit.name
            ctx["delete_usage"] = usage_rows
            template = "accounts/material_units/list_embed.html"
            return render(request, template, ctx)

        messages.error(request, error_message)
        return redirect("accounts:materialunit_list")

    unit.delete()

    url = _materialunit_list_url(request)
    if _is_embed(request):
        return JsonResponse({"ok": True, "url": url})
    return redirect(url)


def _termscondition_list_url(request):
    url = reverse("accounts:termscondition_list")
    if _is_embed(request):
        url += "?embed=1"
    return url

def _termscondition_usage_rows(terms_condition):
    rows = []

    for rel in terms_condition._meta.related_objects:
        accessor_name = rel.get_accessor_name()
        if not accessor_name:
            continue

        try:
            related_manager = getattr(terms_condition, accessor_name)
            count = related_manager.count()
        except Exception:
            continue

        if count:
            label = rel.related_model._meta.verbose_name_plural.title()
            rows.append({"label": label, "count": count})

    return rows

def _expense_list_url(request):
    url = reverse("accounts:expense_list")
    if _is_embed(request):
        url += "?embed=1"
    return url


def _expense_usage_rows(expense):
    rows = []

    for rel in expense._meta.related_objects:
        accessor_name = rel.get_accessor_name()
        if not accessor_name:
            continue

        try:
            related_manager = getattr(expense, accessor_name)
            count = related_manager.count()
        except Exception:
            continue

        if count:
            label = rel.related_model._meta.verbose_name_plural.title()
            rows.append({"label": label, "count": count})

    return rows


def _expense_list_context(request):
    q = (request.GET.get("q") or "").strip()

    qs = Expense.objects.filter(owner=request.user).order_by("name")
    if q:
        qs = qs.filter(name__icontains=q)

    all_expenses = Expense.objects.filter(owner=request.user)

    page_data = _paginate_utility_queryset(request, qs)

    ctx = {
        "expenses": page_data.pop("object_list"),
        "q": q,
        "stats": {
            "total": all_expenses.count(),
        },
    }
    ctx.update(page_data)
    return ctx


@login_required
def expense_list(request):
    tpl = "accounts/expenses/list_embed.html" if _is_embed(request) else "accounts/expenses/list.html"
    return render(request, tpl, _expense_list_context(request))


@login_required
@require_http_methods(["GET", "POST"])
def expense_list_create(request):
    form = ExpenseForm(request.POST or None, user=request.user)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.owner = request.user
        obj.save()

        url = _expense_list_url(request)
        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})
        return redirect(url)

    tpl = "accounts/expenses/form_embed.html" if _is_embed(request) else "accounts/expenses/form.html"
    return render(request, tpl, {"form": form, "mode": "add"})



@login_required
@require_http_methods(["GET", "POST"])
def expense_edit(request, pk: int):
    expense = get_object_or_404(Expense, pk=pk, owner=request.user)
    form = ExpenseForm(request.POST or None, instance=expense, user=request.user)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.owner = request.user
        obj.save()

        url = _expense_list_url(request)
        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})
        return redirect(url)

    template = (
        "accounts/expenses/edit_embed.html"
        if _is_embed(request)
        else "accounts/expenses/edit.html"
    )
    return render(
        request,
        template,
        {
            "form": form,
            "mode": "edit",
            "expense": expense,
        },
    )


@login_required
@require_POST
def expense_delete(request, pk: int):
    expense = get_object_or_404(Expense, pk=pk, owner=request.user)
    usage_rows = _expense_usage_rows(expense)

    if usage_rows:
        error_message = (
            f'Cannot delete "{expense.name}" because it is already used in linked records.'
        )

        if _is_embed(request):
            context = _expense_list_context(request)
            context["delete_error"] = error_message
            context["delete_label"] = expense.name
            context["delete_usage"] = usage_rows
            return render(request, "accounts/expenses/list_embed.html", context)

        messages.error(request, error_message)
        return redirect("accounts:expense_list")

    expense.delete()

    url = _expense_list_url(request)
    if _is_embed(request):
        return JsonResponse({"ok": True, "url": url})
    return redirect(url)

def _accessory_list_url(request):
    url = reverse("accounts:accessory_list")
    if _is_embed(request):
        url += "?embed=1"
    return url


def _accessory_usage_rows(accessory):
    rows = []

    for rel in accessory._meta.related_objects:
        accessor_name = rel.get_accessor_name()
        if not accessor_name:
            continue

        try:
            related_manager = getattr(accessory, accessor_name)
            count = related_manager.count()
        except Exception:
            continue

        if count:
            label = rel.related_model._meta.verbose_name_plural.title()
            rows.append({"label": label, "count": count})

    return rows


def _accessory_list_context(request, form=None):
    q = (request.GET.get("q") or "").strip()

    accessories = Accessory.objects.filter(owner=request.user).select_related("default_unit").order_by("name")
    if q:
        accessories = accessories.filter(
            Q(name__icontains=q) | Q(description__icontains=q)
        )

    all_accessories = Accessory.objects.filter(owner=request.user)

    page_data = _paginate_utility_queryset(request, accessories)

    ctx = {
        "accessories": page_data.pop("object_list"),
        "form": form if form is not None else AccessoryForm(user=request.user),
        "q": q,
        "stats": {
            "total": all_accessories.count(),
            "with_description": all_accessories.exclude(description="").exclude(description__isnull=True).count(),
            "with_default_unit": all_accessories.exclude(default_unit__isnull=True).count(),
        },
    }
    ctx.update(page_data)
    return ctx


@login_required
@require_http_methods(["GET", "POST"])
def accessory_list_create(request):
    form = AccessoryForm(request.POST or None, user=request.user)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.owner = request.user
        obj.save()

        url = _accessory_list_url(request)
        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})
        return redirect(url)

    template = (
        "accounts/accessories/list_embed.html"
        if _is_embed(request)
        else "accounts/accessories/list.html"
    )
    return render(request, template, _accessory_list_context(request, form=form))


@login_required
@require_http_methods(["GET", "POST"])
def accessory_edit(request, pk: int):
    accessory = get_object_or_404(Accessory, pk=pk, owner=request.user)
    form = AccessoryForm(request.POST or None, instance=accessory, user=request.user)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.owner = request.user
        obj.save()

        url = _accessory_list_url(request)
        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})
        return redirect(url)

    template = (
        "accounts/accessories/edit_embed.html"
        if _is_embed(request)
        else "accounts/accessories/edit.html"
    )
    return render(
        request,
        template,
        {
            "form": form,
            "accessory": accessory,
        },
    )


@login_required
@require_POST
def accessory_delete(request, pk: int):
    accessory = get_object_or_404(Accessory, pk=pk, owner=request.user)
    usage_rows = _accessory_usage_rows(accessory)

    if usage_rows:
        error_message = (
            f'Cannot delete "{accessory.name}" because it is already used in linked records.'
        )

        if _is_embed(request):
            ctx = _accessory_list_context(request)
            ctx["delete_error"] = error_message
            ctx["delete_label"] = accessory.name
            ctx["delete_usage"] = usage_rows
            return render(request, "accounts/accessories/list_embed.html", ctx)

        messages.error(request, error_message)
        return redirect("accounts:accessory_list")

    accessory.delete()

    url = _accessory_list_url(request)
    if _is_embed(request):
        return JsonResponse({"ok": True, "url": url})
    return redirect(url)

def _dyeing_other_charge_list_url(request):
    url = reverse("accounts:dyeing_other_charge_list")
    if _is_embed(request):
        url += "?embed=1"
    return url

def _dyeing_other_charge_usage_rows(charge):
    rows = []

    for rel in charge._meta.related_objects:
        accessor_name = rel.get_accessor_name()
        if not accessor_name:
            continue

        try:
            related_manager = getattr(charge, accessor_name)
            count = related_manager.count()
        except Exception:
            continue

        if count:
            label = rel.related_model._meta.verbose_name_plural.title()
            rows.append({"label": label, "count": count})

    return rows


def _dyeing_other_charge_list_context(request):
    q = (request.GET.get("q") or "").strip()

    qs = DyeingOtherCharge.objects.filter(owner=request.user).order_by("name")
    if q:
        qs = qs.filter(name__icontains=q)

    all_charges = DyeingOtherCharge.objects.filter(owner=request.user)

    page_data = _paginate_utility_queryset(request, qs)

    ctx = {
        "charges": page_data.pop("object_list"),
        "q": q,
        "stats": {
            "total": all_charges.count(),
        },
    }
    ctx.update(page_data)
    return ctx


@login_required
@require_GET
def dyeing_other_charge_list(request):
    template = (
        "accounts/dyeing_other_charges/list_embed.html"
        if _is_embed(request)
        else "accounts/dyeing_other_charges/list.html"
    )
    return render(request, template, _dyeing_other_charge_list_context(request))


@login_required
@require_http_methods(["GET", "POST"])
def dyeing_other_charge_create(request):
    form = DyeingOtherChargeForm(request.POST or None, user=request.user)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.owner = request.user
        obj.save()

        url = _dyeing_other_charge_list_url(request)
        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})
        return redirect(url)

    template = (
        "accounts/dyeing_other_charges/form_embed.html"
        if _is_embed(request)
        else "accounts/dyeing_other_charges/form.html"
    )
    return render(
        request,
        template,
        {
            "form": form,
            "mode": "add",
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def dyeing_other_charge_edit(request, pk: int):
    charge = get_object_or_404(DyeingOtherCharge, pk=pk, owner=request.user)
    form = DyeingOtherChargeForm(request.POST or None, instance=charge, user=request.user)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.owner = request.user
        obj.save()

        url = _dyeing_other_charge_list_url(request)
        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})
        return redirect(url)

    template = (
        "accounts/dyeing_other_charges/form_embed.html"
        if _is_embed(request)
        else "accounts/dyeing_other_charges/form.html"
    )
    return render(
        request,
        template,
        {
            "form": form,
            "charge": charge,
            "mode": "edit",
        },
    )


@login_required
@require_POST
def dyeing_other_charge_delete(request, pk: int):
    charge = get_object_or_404(DyeingOtherCharge, pk=pk, owner=request.user)
    usage_rows = _dyeing_other_charge_usage_rows(charge)

    if usage_rows:
        error_message = (
            f'Cannot delete "{charge.name}" because it is already used in linked records.'
        )

        if _is_embed(request):
            context = _dyeing_other_charge_list_context(request)
            context["delete_error"] = error_message
            context["delete_label"] = charge.name
            context["delete_usage"] = usage_rows
            return render(request, "accounts/dyeing_other_charges/list_embed.html", context)

        messages.error(request, error_message)
        return redirect("accounts:dyeing_other_charge_list")

    charge.delete()

    url = _dyeing_other_charge_list_url(request)
    if _is_embed(request):
        return JsonResponse({"ok": True, "url": url})
    return redirect(url)


# ==========================
# CATEGORIES (embed supported)
# ==========================
def _category_list_url(request):
    url = reverse("accounts:category_list")
    if _is_embed(request):
        url += "?embed=1"
    return url


def _category_usage_rows(category):
    rows = []

    bom_count = 0
    try:
        bom_count = BOM.objects.filter(owner=category.owner, category=category).count()
    except Exception:
        bom_count = 0

    if bom_count:
        rows.append({"label": "BOM Records", "count": bom_count})

    return rows


def _category_list_context(request):
    q = (request.GET.get("q") or "").strip()

    qs = Category.objects.filter(owner=request.user).order_by("name")
    if q:
        qs = qs.filter(
            Q(name__icontains=q)
            | Q(description__icontains=q)
        )

    all_categories = Category.objects.filter(owner=request.user)

    return {
        "categories": qs,
        "q": q,
        "stats": {
            "total": all_categories.count(),
            "with_description": all_categories.exclude(description="").exclude(description__isnull=True).count(),
        },
    }


@login_required
def category_list(request):
    tpl = "accounts/categories/list_embed.html" if _is_embed(request) else "accounts/categories/list.html"
    return render(request, tpl, _category_list_context(request))


@login_required
@require_http_methods(["GET", "POST"])
def category_create(request):
    form = CategoryForm(request.POST or None, user=request.user)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.owner = request.user
        obj.save()

        url = _category_list_url(request)
        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})
        return redirect(url)

    tpl = "accounts/categories/form_embed.html" if _is_embed(request) else "accounts/categories/form.html"
    return render(request, tpl, {"form": form, "mode": "add"})


@login_required
@require_http_methods(["GET", "POST"])
def category_update(request, pk: int):
    category = get_object_or_404(Category, pk=pk, owner=request.user)
    form = CategoryForm(request.POST or None, instance=category, user=request.user)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.owner = request.user
        obj.save()

        url = _category_list_url(request)
        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})
        return redirect(url)

    ctx = {
        "form": form,
        "mode": "edit",
        "category": category,
    }
    tpl = "accounts/categories/form_embed.html" if _is_embed(request) else "accounts/categories/form.html"
    return render(request, tpl, ctx)


@login_required
@require_POST
def category_delete(request, pk: int):
    category = get_object_or_404(Category, pk=pk, owner=request.user)
    usage_rows = _category_usage_rows(category)

    if usage_rows:
        error_message = (
            f'Cannot delete "{category.name}" because it is already used in linked records.'
        )

        if _is_embed(request):
            context = _category_list_context(request)
            context["delete_error"] = error_message
            context["delete_label"] = category.name
            context["delete_usage"] = usage_rows
            return render(request, "accounts/categories/list_embed.html", context)

        messages.error(request, error_message)
        return redirect("accounts:category_list")

    category.delete()

    url = _category_list_url(request)
    if _is_embed(request):
        return JsonResponse({"ok": True, "url": url})
    return redirect(url)

def _maincategory_list_url(request):
    url = reverse("accounts:maincategory_list")
    if _is_embed(request):
        url += "?embed=1"
    return url


def _maincategory_usage_rows(main_category):
    rows = []

    bom_count = 0
    try:
        bom_count = BOM.objects.filter(owner=main_category.owner, main_category=main_category).count()
    except Exception:
        bom_count = 0

    if bom_count:
        rows.append({"label": "BOM Records", "count": bom_count})

    return rows


def _maincategory_list_context(request):
    q = (request.GET.get("q") or "").strip()

    qs = MainCategory.objects.filter(owner=request.user).order_by("name")
    if q:
        qs = qs.filter(
            Q(name__icontains=q)
            | Q(description__icontains=q)
        )

    all_main_categories = MainCategory.objects.filter(owner=request.user)

    return {
        "main_categories": qs,
        "q": q,
        "stats": {
            "total": all_main_categories.count(),
            "with_description": all_main_categories.exclude(description="").exclude(description__isnull=True).count(),
        },
    }


@login_required
def maincategory_list(request):
    tpl = "accounts/main_categories/list_embed.html" if _is_embed(request) else "accounts/main_categories/list.html"
    return render(request, tpl, _maincategory_list_context(request))


@login_required
@require_http_methods(["GET", "POST"])
def maincategory_create(request):
    form = MainCategoryForm(request.POST or None, user=request.user)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.owner = request.user
        obj.save()

        url = _maincategory_list_url(request)
        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})
        return redirect(url)

    tpl = "accounts/main_categories/form_embed.html" if _is_embed(request) else "accounts/main_categories/form.html"
    return render(request, tpl, {"form": form, "mode": "add"})


@login_required
@require_http_methods(["GET", "POST"])
def maincategory_edit(request, pk: int):
    main_category = get_object_or_404(MainCategory, pk=pk, owner=request.user)
    form = MainCategoryForm(request.POST or None, instance=main_category, user=request.user)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.owner = request.user
        obj.save()

        url = _maincategory_list_url(request)
        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})
        return redirect(url)

    ctx = {
        "form": form,
        "mode": "edit",
        "main_category": main_category,
    }
    tpl = "accounts/main_categories/form_embed.html" if _is_embed(request) else "accounts/main_categories/form.html"
    return render(request, tpl, ctx)


@login_required
@require_POST
def maincategory_delete(request, pk: int):
    main_category = get_object_or_404(MainCategory, pk=pk, owner=request.user)
    usage_rows = _maincategory_usage_rows(main_category)

    if usage_rows:
        error_message = (
            f'Cannot delete "{main_category.name}" because it is already used in linked records.'
        )

        if _is_embed(request):
            context = _maincategory_list_context(request)
            context["delete_error"] = error_message
            context["delete_label"] = main_category.name
            context["delete_usage"] = usage_rows
            return render(request, "accounts/main_categories/list_embed.html", context)

        messages.error(request, error_message)
        return redirect("accounts:maincategory_list")

    main_category.delete()

    url = _maincategory_list_url(request)
    if _is_embed(request):
        return JsonResponse({"ok": True, "url": url})
    return redirect(url)

def _patterntype_list_url(request):
    url = reverse("accounts:patterntype_list")
    if _is_embed(request):
        url += "?embed=1"
    return url

@login_required
@require_http_methods(["GET", "POST"])
def patterntype_list_create(request):
    q = (request.GET.get("q") or "").strip()

    pattern_types = PatternType.objects.filter(owner=request.user).order_by("name")
    if q:
        pattern_types = pattern_types.filter(
            Q(name__icontains=q) | Q(description__icontains=q)
        )

    if request.method == "POST":
        form = PatternTypeForm(request.POST, user=request.user)
        if form.is_valid():
            obj = form.save(commit=False)
            obj.owner = request.user
            obj.save()

            url = _patterntype_list_url(request)
            if _is_embed(request):
                return JsonResponse({"ok": True, "url": url})
            return redirect(url)
    else:
        form = PatternTypeForm(user=request.user)

    template = (
        "accounts/pattern_types/list_embed.html"
        if _is_embed(request)
        else "accounts/pattern_types/list.html"
    )
    return render(
        request,
        template,
        {
            "pattern_types": pattern_types,
            "form": form,
            "q": q,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def patterntype_edit(request, pk: int):
    pattern_type = get_object_or_404(PatternType, pk=pk, owner=request.user)
    form = PatternTypeForm(request.POST or None, instance=pattern_type, user=request.user)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.owner = request.user
        obj.save()

        url = _patterntype_list_url(request)
        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})
        return redirect(url)

    template = (
        "accounts/pattern_types/edit_embed.html"
        if _is_embed(request)
        else "accounts/pattern_types/edit.html"
    )
    return render(
        request,
        template,
        {
            "form": form,
            "pattern_type": pattern_type,
        },
    )


@login_required
@require_POST
def patterntype_delete(request, pk: int):
    pattern_type = get_object_or_404(PatternType, pk=pk, owner=request.user)
    pattern_type.delete()

    url = _patterntype_list_url(request)
    if _is_embed(request):
        return JsonResponse({"ok": True, "url": url})
    return redirect(url)

# ==========================
# CATALOGUES (embed supported)
# ==========================
@login_required
def catalogue_list(request):
    q = (request.GET.get("q") or "").strip()

    qs = Catalogue.objects.filter(owner=request.user)
    if q:
        qs = qs.filter(
            Q(name__icontains=q)
            | Q(wear_type__icontains=q)
            | Q(description__icontains=q)
        )

    ctx = {"catalogues": qs.order_by("name"), "q": q}
    tpl = "accounts/catalogues/list_embed.html" if _is_embed(request) else "accounts/catalogues/list.html"
    return render(request, tpl, ctx)


@login_required
@require_http_methods(["GET", "POST"])
def catalogue_create(request):
    form = CatalogueForm(request.POST or None, user=request.user)

    if request.method == "POST":
        form.instance.owner = request.user

        if form.is_valid():
            obj = form.save(commit=False)
            obj.owner = request.user
            obj.save()

            url = reverse("accounts:catalogue_list")
            if _is_embed(request):
                return JsonResponse({"ok": True, "url": url})
            return redirect(url)

    tpl = "accounts/catalogues/form_embed.html" if _is_embed(request) else "accounts/catalogues/form.html"
    return render(request, tpl, {"form": form, "mode": "add"})


@login_required
@require_http_methods(["GET", "POST"])
def catalogue_update(request, pk: int):
    catalogue = get_object_or_404(Catalogue, pk=pk, owner=request.user)
    form = CatalogueForm(request.POST or None, instance=catalogue, user=request.user)

    if request.method == "POST" and form.is_valid():
        form.save()

        url = reverse("accounts:catalogue_list")
        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})
        return redirect(url)

    tpl = "accounts/catalogues/form_embed.html" if _is_embed(request) else "accounts/catalogues/form.html"
    return render(request, tpl, {"form": form, "mode": "edit", "catalogue": catalogue})


@login_required
@require_POST
def catalogue_delete(request, pk: int):
    catalogue = get_object_or_404(Catalogue, pk=pk, owner=request.user)
    catalogue.delete()

    url = reverse("accounts:catalogue_list")
    if _is_embed(request):
        return JsonResponse({"ok": True, "url": url})
    return redirect(url)


def _greige_terms_condition_options(owner):

    qs = TermsCondition.objects.all()
    if owner is not None:
        qs = qs.filter(owner=owner)
    return qs.order_by("title")


def _termscondition_list_context(request):
    q = (request.GET.get("q") or "").strip()

    qs = TermsCondition.objects.filter(owner=request.user).order_by("title")

    if q:
        qs = qs.filter(
            Q(title__icontains=q) |
            Q(content__icontains=q)
        )

    all_terms = TermsCondition.objects.filter(owner=request.user)

    page_data = _paginate_utility_queryset(request, qs)

    ctx = {
        "terms_conditions": page_data.pop("object_list"),
        "q": q,
        "stats": {
            "total": all_terms.count(),
            "with_content": all_terms.exclude(content="").count(),
        },
    }
    ctx.update(page_data)
    return ctx


@login_required
@require_http_methods(["GET"])
def termscondition_list(request):
    ctx = _termscondition_list_context(request)

    tpl = (
        "accounts/terms_conditions/list_embed.html"
        if _is_embed(request)
        else "accounts/terms_conditions/list.html"
    )
    return render(request, tpl, ctx)


@login_required
@require_http_methods(["GET", "POST"])
def termscondition_create(request):
    form = TermsConditionForm(request.POST or None, user=request.user)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.owner = request.user
        obj.is_active = True
        obj.save()

        url = _termscondition_list_url(request)

        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})

        return redirect(url)

    tpl = (
        "accounts/terms_conditions/form_embed.html"
        if _is_embed(request)
        else "accounts/terms_conditions/form.html"
    )
    return render(request, tpl, {"form": form, "mode": "add"})


@login_required
@require_http_methods(["GET", "POST"])
def termscondition_update(request, pk: int):
    terms_condition = get_object_or_404(TermsCondition, pk=pk, owner=request.user)
    form = TermsConditionForm(request.POST or None, instance=terms_condition, user=request.user)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.owner = request.user
        obj.is_active = True
        obj.save()

        url = _termscondition_list_url(request)

        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})

        return redirect(url)

    tpl = (
        "accounts/terms_conditions/edit_embed.html"
        if _is_embed(request)
        else "accounts/terms_conditions/edit.html"
    )
    return render(
        request,
        tpl,
        {
            "form": form,
            "mode": "edit",
            "terms_condition": terms_condition,
        },
    )


@login_required
@require_POST
def termscondition_delete(request, pk: int):
    terms_condition = get_object_or_404(TermsCondition, pk=pk, owner=request.user)
    usage_rows = _termscondition_usage_rows(terms_condition)

    if usage_rows:
        error_message = (
            f'Cannot delete "{terms_condition.title}" because it is already used in linked records.'
        )

        if _is_embed(request):
            context = _termscondition_list_context(request)
            context["delete_error"] = error_message
            context["delete_label"] = terms_condition.title
            context["delete_usage"] = usage_rows
            return render(request, "accounts/terms_conditions/list_embed.html", context)

        messages.error(request, error_message)
        return redirect("accounts:termscondition_list")

    terms_condition.delete()

    url = _termscondition_list_url(request)

    if _is_embed(request):
        return JsonResponse({"ok": True, "url": url})

    return redirect(url)

def _inwardtype_list_url(request):
    url = reverse("accounts:inwardtype_list")
    if _is_embed(request):
        url += "?embed=1"
    return url


def _inwardtype_usage_rows(inward_type):
    rows = []

    for rel in inward_type._meta.related_objects:
        accessor_name = rel.get_accessor_name()
        if not accessor_name:
            continue

        try:
            related_manager = getattr(inward_type, accessor_name)
            count = related_manager.count()
        except Exception:
            continue

        if count:
            label = rel.related_model._meta.verbose_name_plural.title()
            rows.append({"label": label, "count": count})

    return rows


def _inwardtype_list_context(request):
    q = (request.GET.get("q") or "").strip()

    qs = InwardType.objects.filter(owner=request.user).order_by("name")
    if q:
        qs = qs.filter(
            Q(name__icontains=q) | Q(description__icontains=q)
        )

    all_inward_types = InwardType.objects.filter(owner=request.user)

    page_data = _paginate_utility_queryset(request, qs)

    ctx = {
        "inward_types": page_data.pop("object_list"),
        "q": q,
        "stats": {
            "total": all_inward_types.count(),
            "with_description": all_inward_types.exclude(description="").count(),
        },
    }
    ctx.update(page_data)
    return ctx


@login_required
@require_GET
def inwardtype_list(request):
    template = (
        "accounts/inward_types/list_embed.html"
        if _is_embed(request)
        else "accounts/inward_types/list.html"
    )
    return render(request, template, _inwardtype_list_context(request))


@login_required
@require_http_methods(["GET", "POST"])
def inwardtype_create(request):
    form = InwardTypeForm(request.POST or None, user=request.user)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.owner = request.user
        obj.save()

        url = _inwardtype_list_url(request)
        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})
        return redirect(url)

    template = (
        "accounts/inward_types/form_embed.html"
        if _is_embed(request)
        else "accounts/inward_types/form.html"
    )
    return render(
        request,
        template,
        {
            "form": form,
            "mode": "add",
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def inwardtype_edit(request, pk: int):
    inward_type = get_object_or_404(InwardType, pk=pk, owner=request.user)
    form = InwardTypeForm(request.POST or None, instance=inward_type, user=request.user)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.owner = request.user
        obj.save()

        url = _inwardtype_list_url(request)
        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})
        return redirect(url)

    template = (
        "accounts/inward_types/form_embed.html"
        if _is_embed(request)
        else "accounts/inward_types/form.html"
    )
    return render(
        request,
        template,
        {
            "form": form,
            "inward_type": inward_type,
            "mode": "edit",
        },
    )


@login_required
@require_POST
def inwardtype_delete(request, pk: int):
    inward_type = get_object_or_404(InwardType, pk=pk, owner=request.user)
    usage_rows = _inwardtype_usage_rows(inward_type)

    if usage_rows:
        error_message = (
            f'Cannot delete "{inward_type.name}" because it is already used in linked records.'
        )

        if _is_embed(request):
            context = _inwardtype_list_context(request)
            context["delete_error"] = error_message
            context["delete_label"] = inward_type.name
            context["delete_usage"] = usage_rows
            return render(request, "accounts/inward_types/list_embed.html", context)

        messages.error(request, error_message)
        return redirect("accounts:inwardtype_list")

    inward_type.delete()

    url = _inwardtype_list_url(request)
    if _is_embed(request):
        return JsonResponse({"ok": True, "url": url})
    return redirect(url)

def _subcategory_list_url(request):
    url = reverse("accounts:subcategory_list")
    if _is_embed(request):
        url += "?embed=1"
    return url


def _subcategory_usage_rows(sub_category):
    rows = []

    bom_count = 0
    try:
        bom_count = BOM.objects.filter(owner=sub_category.owner, sub_category=sub_category).count()
    except Exception:
        bom_count = 0

    if bom_count:
        rows.append({"label": "BOM Records", "count": bom_count})

    return rows


def _subcategory_list_context(request):
    q = (request.GET.get("q") or "").strip()

    qs = (
        SubCategory.objects.filter(owner=request.user)
        .select_related("main_category")
        .order_by("main_category__name", "name")
    )
    if q:
        qs = qs.filter(
            Q(name__icontains=q)
            | Q(description__icontains=q)
            | Q(main_category__name__icontains=q)
        )

    all_sub_categories = SubCategory.objects.filter(owner=request.user)

    return {
        "sub_categories": qs,
        "q": q,
        "stats": {
            "total": all_sub_categories.count(),
            "with_description": all_sub_categories.exclude(description="").exclude(description__isnull=True).count(),
        },
    }


@login_required
def subcategory_list(request):
    template = (
        "accounts/sub_categories/list_embed.html"
        if _is_embed(request)
        else "accounts/sub_categories/list.html"
    )
    return render(request, template, _subcategory_list_context(request))


@login_required
@require_http_methods(["GET", "POST"])
def subcategory_create(request):
    form = SubCategoryForm(request.POST or None, user=request.user)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.owner = request.user
        obj.save()

        url = _subcategory_list_url(request)
        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})
        return redirect(url)

    template = (
        "accounts/sub_categories/form_embed.html"
        if _is_embed(request)
        else "accounts/sub_categories/form.html"
    )
    return render(
        request,
        template,
        {
            "form": form,
            "mode": "add",
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def subcategory_edit(request, pk: int):
    sub_category = get_object_or_404(SubCategory, pk=pk, owner=request.user)
    form = SubCategoryForm(request.POST or None, instance=sub_category, user=request.user)

    if request.method == "POST" and form.is_valid():
        obj = form.save(commit=False)
        obj.owner = request.user
        obj.save()

        url = _subcategory_list_url(request)
        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})
        return redirect(url)

    template = (
        "accounts/sub_categories/form_embed.html"
        if _is_embed(request)
        else "accounts/sub_categories/form.html"
    )
    return render(
        request,
        template,
        {
            "form": form,
            "mode": "edit",
            "sub_category": sub_category,
        },
    )


@login_required
@require_POST
def subcategory_delete(request, pk: int):
    sub_category = get_object_or_404(SubCategory, pk=pk, owner=request.user)
    usage_rows = _subcategory_usage_rows(sub_category)

    if usage_rows:
        error_message = (
            f'Cannot delete "{sub_category.name}" because it is already used in linked records.'
        )

        if _is_embed(request):
            context = _subcategory_list_context(request)
            context["delete_error"] = error_message
            context["delete_label"] = sub_category.name
            context["delete_usage"] = usage_rows
            return render(request, "accounts/sub_categories/list_embed.html", context)

        messages.error(request, error_message)
        return redirect("accounts:subcategory_list")

    sub_category.delete()

    url = _subcategory_list_url(request)
    if _is_embed(request):
        return JsonResponse({"ok": True, "url": url})
    return redirect(url)

def _is_po_approved_for_inward(po) -> bool:
    current_status = str(getattr(po, "approval_status", "") or "").strip().lower()
    if current_status:
        return current_status == "approved"

    source_yarn_po = getattr(po, "source_yarn_po", None)
    source_greige_po = getattr(po, "source_greige_po", None)
    source_dyeing_po = getattr(po, "source_dyeing_po", None)

    for source in [source_yarn_po, source_greige_po, source_dyeing_po]:
        source_status = str(getattr(source, "approval_status", "") or "").strip().lower()
        if source_status:
            return source_status == "approved"

    return True

@login_required
@require_http_methods(["GET", "POST"])
def greigepo_review(request, pk: int):
    po = get_object_or_404(
        _greige_po_queryset(),
        pk=pk,
    )

    if not _can_access_greige_po(request.user, po):
        raise PermissionDenied("You do not have access to this Greige PO.")

    embed_mode = _is_embed(request)
    can_review = _can_review_yarn_po(request)

    review_form = GreigePOReviewForm(request.POST or None)

    review_checks = {
        "has_header": bool(po.vendor_id and po.po_date),
        "has_source": bool(po.source_yarn_po_id),
        "has_items": po.items.exists(),
        "has_shipping": bool((po.shipping_address or "").strip()),
    }
    review_ready_count = sum(1 for value in review_checks.values() if value)

    context = {
        "po": po,
        "review_form": review_form,
        "can_review_greige_po": can_review,
        "embed_mode": embed_mode,
        "review_checks": review_checks,
        "review_ready_count": review_ready_count,
    }

    if request.method == "POST":
        if not can_review:
            return HttpResponseForbidden("You are not allowed to review this PO.")

        if review_form.is_valid():
            decision = review_form.cleaned_data["decision"]

            if decision == "approve":
                po.approval_status = "approved"
                po.rejection_reason = ""
            else:
                po.approval_status = "rejected"
                po.rejection_reason = review_form.cleaned_data["rejection_reason"].strip()

            po.reviewed_by = get_actor(request) or request.user
            po.reviewed_at = timezone.now()
            po.save(update_fields=[
                "approval_status",
                "rejection_reason",
                "reviewed_by",
                "reviewed_at",
            ])

            if embed_mode:
                return JsonResponse({
                    "ok": True,
                    "message": "Greige PO reviewed successfully.",
                    "redirect_url": reverse("accounts:greigepo_list"),
                })

            return redirect("accounts:greigepo_list")

        if embed_mode:
            return render(
                request,
                "accounts/greige_po/review_embed.html",
                context,
                status=400,
            )

    template_name = (
        "accounts/greige_po/review_embed.html"
        if embed_mode
        else "accounts/greige_po/review.html"
    )

    return render(request, template_name, context)

@login_required
@require_http_methods(["GET", "POST"])
def dyeingpo_review(request, pk: int):
    po = get_object_or_404(
        _dyeing_po_queryset(),
        pk=pk,
    )

    if not _can_access_dyeing_po(request.user, po):
        raise PermissionDenied("You do not have access to this Dyeing PO.")

    review_field_names = ("approval_status", "rejection_reason", "reviewed_by", "reviewed_at")
    review_templates_available = (
        _template_exists("accounts/dyeing_po/review.html")
        and _template_exists("accounts/dyeing_po/review_embed.html")
    )

    if not _model_has_fields(DyeingPurchaseOrder, *review_field_names) or not review_templates_available:
        message = "Dyeing PO review needs model review fields and review templates before this page can work."
        if _is_embed(request):
            return JsonResponse({
                "ok": False,
                "message": message,
                "redirect_url": reverse("accounts:dyeingpo_list"),
            }, status=400)
        messages.error(request, message)
        return redirect("accounts:dyeingpo_list")

    embed_mode = _is_embed(request)
    can_review = _can_review_yarn_po(request)

    review_form = DyeingPOReviewForm(request.POST or None)

    review_checks = {
        "has_header": bool(po.vendor_id and po.po_date),
        "has_source": bool(po.source_greige_po_id),
        "has_items": po.items.exists(),
        "has_shipping": bool((po.shipping_address or "").strip()),
    }
    review_ready_count = sum(1 for value in review_checks.values() if value)

    context = {
        "po": po,
        "review_form": review_form,
        "can_review_dyeing_po": can_review,
        "embed_mode": embed_mode,
        "review_checks": review_checks,
        "review_ready_count": review_ready_count,
    }

    if request.method == "POST":
        if not can_review:
            return HttpResponseForbidden("You are not allowed to review this PO.")

        if review_form.is_valid():
            decision = review_form.cleaned_data["decision"]

            if decision == "approve":
                po.approval_status = "approved"
                po.rejection_reason = ""
            else:
                po.approval_status = "rejected"
                po.rejection_reason = review_form.cleaned_data["rejection_reason"].strip()

            po.reviewed_by = get_actor(request) or request.user
            po.reviewed_at = timezone.now()
            po.save(update_fields=[
                "approval_status",
                "rejection_reason",
                "reviewed_by",
                "reviewed_at",
            ])

            if embed_mode:
                return JsonResponse({
                    "ok": True,
                    "message": "Dyeing PO reviewed successfully.",
                    "redirect_url": reverse("accounts:dyeingpo_list"),
                })

            return redirect("accounts:dyeingpo_list")

        if embed_mode:
            return render(
                request,
                "accounts/dyeing_po/review_embed.html",
                context,
                status=400,
            )

    template_name = (
        "accounts/dyeing_po/review_embed.html"
        if embed_mode
        else "accounts/dyeing_po/review.html"
    )

    return render(request, template_name, context)

def _bom_usage_rows(bom):
    rows = []

    usage_map = [
        ("Programs", bom.programs.count()),
        ("Costing Snapshots", bom.costing_snapshots.count()),
    ]

    for label, count in usage_map:
        if count:
            rows.append({"label": label, "count": count})

    return rows


def _bom_list_context(request):
    q = (request.GET.get("q") or "").strip()

    qs = (
        BOM.objects.filter(owner=request.user)
        .select_related(
            "catalogue",
            "brand",
            "category",
            "main_category",
            "sub_category",
            "pattern_type",
        )
        .order_by("-id")
    )

    if q:
        qs = qs.filter(
            Q(bom_code__icontains=q)
            | Q(sku__icontains=q)
            | Q(product_name__icontains=q)
        )

    page_data = _paginate_utility_queryset(request, qs)

    ctx = {
        "boms": page_data.pop("object_list"),
        "q": q,
        "full_page": not _is_embed(request),
    }
    ctx.update(page_data)
    return ctx

def _bom_list_url(request):
    url = reverse("accounts:bom_list")
    if _is_embed(request):
        url += "?embed=1"
    return url


@login_required
def bom_list(request):
    template = "accounts/bom/list_embed.html" if _is_embed(request) else "accounts/bom/list.html"
    return render(request, template, _bom_list_context(request))

def _bom_truthy_flag(value):
    return str(value or "").lower() in {"on", "true", "1", "yes"}

def _save_bom_jobber_detail_formset(jobber_detail_formset, bom):
    process_price_map = {}

    data = getattr(jobber_detail_formset, "data", None)

    if data:
        total_forms = 0
        try:
            total_forms = int(data.get("jobber_processes-TOTAL_FORMS") or 0)
        except (TypeError, ValueError):
            total_forms = 0

        for index in range(total_forms):
            delete_flag = data.get(f"jobber_processes-{index}-DELETE")
            if _bom_truthy_flag(delete_flag):
                continue

            jobber_type_id = (data.get(f"jobber_processes-{index}-jobber_type") or "").strip()
            raw_price = (data.get(f"jobber_processes-{index}-price") or "").strip()

            if not jobber_type_id:
                continue

            try:
                process_price_map[int(jobber_type_id)] = Decimal(raw_price or "0")
            except InvalidOperation:
                process_price_map[int(jobber_type_id)] = Decimal("0")

    if not process_price_map:
        process_price_map = {
            row.jobber_type_id: (row.price or Decimal("0"))
            for row in bom.jobber_type_processes.select_related("jobber_type").all()
            if row.jobber_type_id
        }

    keep_ids = []

    for form in jobber_detail_formset.forms:
        if not getattr(form, "cleaned_data", None):
            continue

        if form.cleaned_data.get("DELETE"):
            continue

        jobber = form.cleaned_data.get("jobber")
        if not jobber:
            continue

        if not jobber.jobber_type_id:
            continue

        obj = form.instance
        obj.bom = bom
        obj.jobber = jobber
        obj.jobber_type = jobber.jobber_type
        obj.price = process_price_map.get(jobber.jobber_type_id, Decimal("0"))
        obj.sort_order = form.cleaned_data.get("sort_order") or 0
        obj.save()

        if obj.pk:
            keep_ids.append(obj.pk)

    if keep_ids:
        bom.jobber_details.exclude(pk__in=keep_ids).delete()
    else:
        bom.jobber_details.all().delete()



def _save_bom_image_formset(image_formset, bom):
    """
    Save BOM images reliably, including files posted from dynamically added rows.

    Django inline formsets handle this normally, but this helper also has a
    request.FILES fallback for dynamic rows so a selected BOM image is not lost.
    """
    image_formset.instance = bom

    saved_file_keys = set()

    for obj in getattr(image_formset, "deleted_objects", []):
        obj.delete()

    for form in image_formset.forms:
        cleaned = getattr(form, "cleaned_data", None)
        if not cleaned or cleaned.get("DELETE"):
            continue

        image = cleaned.get("image")
        has_existing = bool(getattr(form.instance, "pk", None))

        if not has_existing and not image:
            continue

        if has_existing and not image and not form.has_changed():
            continue

        obj = form.save(commit=False)
        obj.bom = bom
        obj.sort_order = cleaned.get("sort_order") or 0

        if obj.image:
            obj.save()
            saved_file_keys.add(form.add_prefix("image"))

    data = getattr(image_formset, "data", None)
    files = getattr(image_formset, "files", None)
    prefix = getattr(image_formset, "prefix", "images") or "images"

    if not data or not files:
        return

    try:
        total_forms = int(data.get(f"{prefix}-TOTAL_FORMS") or 0)
    except (TypeError, ValueError):
        total_forms = 0

    for index in range(total_forms):
        file_key = f"{prefix}-{index}-image"
        if file_key in saved_file_keys:
            continue

        if _bom_truthy_flag(data.get(f"{prefix}-{index}-DELETE")):
            continue

        uploaded = files.get(file_key)
        if not uploaded:
            continue

        raw_id = (data.get(f"{prefix}-{index}-id") or "").strip()
        image_obj = None

        if raw_id:
            try:
                image_obj = bom.images.get(pk=int(raw_id))
            except (TypeError, ValueError, BOMImage.DoesNotExist):
                image_obj = None

        if image_obj is None:
            image_obj = BOMImage(bom=bom)

        image_obj.image = uploaded

        raw_sort = (data.get(f"{prefix}-{index}-sort_order") or "0").strip()
        try:
            image_obj.sort_order = int(raw_sort or 0)
        except (TypeError, ValueError):
            image_obj.sort_order = 0

        image_obj.save()

@login_required
@require_http_methods(["GET", "POST"])
def bom_create(request):
    bom = BOM(owner=request.user)
    bom_debug_errors = []

    form = BOMForm(
        request.POST or None,
        request.FILES or None,
        instance=bom,
        user=request.user,
    )

    if request.method == "POST":
        material_formset = BOMMaterialItemFormSet(
            request.POST,
            instance=bom,
            prefix="materials",
            form_kwargs={"user": request.user},
        )
        accessory_formset = BOMAccessoryItemFormSet(
            request.POST,
            instance=bom,
            prefix="accessories",
            form_kwargs={"user": request.user},
        )
        image_formset = BOMImageFormSet(
            request.POST,
            request.FILES,
            instance=bom,
            prefix="images",
        )
        jobber_process_formset = BOMJobberTypeProcessFormSet(
            request.POST,
            instance=bom,
            prefix="jobber_processes",
            form_kwargs={"user": request.user},
        )
        jobber_detail_formset = BOMJobberDetailFormSet(
            request.POST,
            instance=bom,
            prefix="jobber_details",
            form_kwargs={"user": request.user},
        )
        expense_formset = BOMExpenseItemFormSet(
            request.POST,
            instance=bom,
            prefix="expenses",
            form_kwargs={"user": request.user},
        )

        form_valid = form.is_valid()
        material_formset_valid = material_formset.is_valid()
        accessory_formset_valid = accessory_formset.is_valid()
        image_formset_valid = image_formset.is_valid()
        jobber_process_formset_valid = jobber_process_formset.is_valid()
        jobber_detail_formset_valid = jobber_detail_formset.is_valid()
        expense_formset_valid = expense_formset.is_valid()

        is_valid = all(
            [
                form_valid,
                material_formset_valid,
                accessory_formset_valid,
                image_formset_valid,
                jobber_process_formset_valid,
                jobber_detail_formset_valid,
                expense_formset_valid,
            ]
        )

        if is_valid:
            with transaction.atomic():
                bom = form.save(commit=False)
                bom.owner = request.user
                bom.save()

                material_formset.instance = bom
                material_formset.save()

                accessory_formset.instance = bom
                accessory_formset.save()

                image_formset.instance = bom
                _save_bom_image_formset(image_formset, bom)

                jobber_process_formset.instance = bom
                jobber_process_formset.save()

                jobber_detail_formset.instance = bom
                _save_bom_jobber_detail_formset(jobber_detail_formset, bom)

                expense_formset.instance = bom
                expense_formset.save()

                if hasattr(bom, "recalculate_final_price"):
                    bom.recalculate_final_price(save=True)

            url = _bom_list_url(request)
            if _is_embed(request):
                return JsonResponse({"ok": True, "url": url})
            return redirect(url)

        bom_debug_errors = _collect_bom_debug_errors(
            form,
            material_formset,
            accessory_formset,
            image_formset,
            jobber_process_formset,
            jobber_detail_formset,
            expense_formset,
        )
        logger.warning("BOM create validation failed: %s", bom_debug_errors)

    else:
        material_formset = BOMMaterialItemFormSet(
            instance=bom,
            prefix="materials",
            form_kwargs={"user": request.user},
        )
        accessory_formset = BOMAccessoryItemFormSet(
            instance=bom,
            prefix="accessories",
            form_kwargs={"user": request.user},
        )
        image_formset = BOMImageFormSet(instance=bom, prefix="images")
        jobber_process_formset = BOMJobberTypeProcessFormSet(
            instance=bom,
            prefix="jobber_processes",
            form_kwargs={"user": request.user},
        )
        jobber_detail_formset = BOMJobberDetailFormSet(
            instance=bom,
            prefix="jobber_details",
            form_kwargs={"user": request.user},
        )
        expense_formset = BOMExpenseItemFormSet(
            instance=bom,
            prefix="expenses",
            form_kwargs={"user": request.user},
        )

    template = "accounts/bom/form_embed.html" if _is_embed(request) else "accounts/bom/form.html"
    return render(
        request,
        template,
        {
            "form": form,
            "material_formset": material_formset,
            "accessory_formset": accessory_formset,
            "image_formset": image_formset,
            "jobber_process_formset": jobber_process_formset,
            "jobber_detail_formset": jobber_detail_formset,
            "expense_formset": expense_formset,
            "bom_debug_errors": bom_debug_errors,
            "mode": "add",
            "jobber_defaults": _program_jobber_defaults(request.user),
            "full_page": not _is_embed(request),
            "action_url": reverse("accounts:bom_add"),
        },
    )
    

@login_required
@require_http_methods(["GET", "POST"])
def bom_update(request, pk: int):
    bom = get_object_or_404(BOM, pk=pk, owner=request.user)
    bom_debug_errors = []

    form = BOMForm(
        request.POST or None,
        request.FILES or None,
        instance=bom,
        user=request.user,
    )

    if request.method == "POST":
        material_formset = BOMMaterialItemFormSet(
            request.POST,
            instance=bom,
            prefix="materials",
            form_kwargs={"user": request.user},
        )
        accessory_formset = BOMAccessoryItemFormSet(
            request.POST,
            instance=bom,
            prefix="accessories",
            form_kwargs={"user": request.user},
        )
        image_formset = BOMImageFormSet(
            request.POST,
            request.FILES,
            instance=bom,
            prefix="images",
        )
        jobber_process_formset = BOMJobberTypeProcessFormSet(
            request.POST,
            instance=bom,
            prefix="jobber_processes",
            form_kwargs={"user": request.user},
        )
        jobber_detail_formset = BOMJobberDetailFormSet(
            request.POST,
            instance=bom,
            prefix="jobber_details",
            form_kwargs={"user": request.user},
        )
        expense_formset = BOMExpenseItemFormSet(
            request.POST,
            instance=bom,
            prefix="expenses",
            form_kwargs={"user": request.user},
        )

        form_valid = form.is_valid()
        material_formset_valid = material_formset.is_valid()
        accessory_formset_valid = accessory_formset.is_valid()
        image_formset_valid = image_formset.is_valid()
        jobber_process_formset_valid = jobber_process_formset.is_valid()
        jobber_detail_formset_valid = jobber_detail_formset.is_valid()
        expense_formset_valid = expense_formset.is_valid()

        is_valid = all(
            [
                form_valid,
                material_formset_valid,
                accessory_formset_valid,
                image_formset_valid,
                jobber_process_formset_valid,
                jobber_detail_formset_valid,
                expense_formset_valid,
            ]
        )

        if is_valid:
            with transaction.atomic():
                bom = form.save(commit=False)
                bom.owner = request.user
                bom.save()

                material_formset.instance = bom
                material_formset.save()

                accessory_formset.instance = bom
                accessory_formset.save()

                image_formset.instance = bom
                _save_bom_image_formset(image_formset, bom)

                jobber_process_formset.instance = bom
                jobber_process_formset.save()

                jobber_detail_formset.instance = bom
                _save_bom_jobber_detail_formset(jobber_detail_formset, bom)

                expense_formset.instance = bom
                expense_formset.save()

                if hasattr(bom, "recalculate_final_price"):
                    bom.recalculate_final_price(save=True)

            url = _bom_list_url(request)
            if _is_embed(request):
                return JsonResponse({"ok": True, "url": url})
            return redirect(url)

        bom_debug_errors = _collect_bom_debug_errors(
            form,
            material_formset,
            accessory_formset,
            image_formset,
            jobber_process_formset,
            jobber_detail_formset,
            expense_formset,
        )
        logger.warning("BOM update validation failed: %s", bom_debug_errors)

    else:
        material_formset = BOMMaterialItemFormSet(
            instance=bom,
            prefix="materials",
            form_kwargs={"user": request.user},
        )
        accessory_formset = BOMAccessoryItemFormSet(
            instance=bom,
            prefix="accessories",
            form_kwargs={"user": request.user},
        )
        image_formset = BOMImageFormSet(instance=bom, prefix="images")
        jobber_process_formset = BOMJobberTypeProcessFormSet(
            instance=bom,
            prefix="jobber_processes",
            form_kwargs={"user": request.user},
        )
        jobber_detail_formset = BOMJobberDetailFormSet(
            instance=bom,
            prefix="jobber_details",
            form_kwargs={"user": request.user},
        )
        expense_formset = BOMExpenseItemFormSet(
            instance=bom,
            prefix="expenses",
            form_kwargs={"user": request.user},
        )

    template = "accounts/bom/form_embed.html" if _is_embed(request) else "accounts/bom/form.html"
    return render(
        request,
        template,
        {
            "form": form,
            "material_formset": material_formset,
            "accessory_formset": accessory_formset,
            "image_formset": image_formset,
            "jobber_process_formset": jobber_process_formset,
            "jobber_detail_formset": jobber_detail_formset,
            "expense_formset": expense_formset,
            "bom_debug_errors": bom_debug_errors,
            "mode": "edit",
            "jobber_defaults": _program_jobber_defaults(request.user),
            "full_page": not _is_embed(request),
            "bom": bom,
            "action_url": reverse("accounts:bom_edit", args=[bom.pk]),
        },
    )
    
def _program_edit_url(request, program):
    url = reverse("accounts:program_edit", args=[program.pk])
    if _is_embed(request):
        url += "?embed=1"
    return url

def _bom_preview_image_url(bom):
    # 1) direct image-like fields on BOM itself
    for attr in ("image", "photo", "photo_update", "product_image"):
        field = getattr(bom, attr, None)
        if field and getattr(field, "name", None) and getattr(field, "url", None):
            return field.url

    # 2) child image rows under BOM
    images_manager = getattr(bom, "images", None)
    if images_manager is not None:
        try:
            for img in images_manager.all().order_by("sort_order", "id"):
                for attr in ("image", "photo", "file"):
                    field = getattr(img, attr, None)
                    if field and getattr(field, "name", None) and getattr(field, "url", None):
                        return field.url
        except Exception:
            pass

    return ""

def _program_sku_payloads(user):
    boms = (
        BOM.objects.filter(owner=user)
        .select_related(
            "brand",
            "category",
            "main_category",
            "sub_category",
            "pattern_type",
            "catalogue",
        )
        .prefetch_related(
            Prefetch(
                "material_items",
                queryset=BOMMaterialItem.objects.select_related("material").order_by("sort_order", "id"),
            ),
            Prefetch(
                "accessory_items",
                queryset=BOMAccessoryItem.objects.select_related("accessory").order_by("sort_order", "id"),
            ),
            Prefetch(
                "images",
                queryset=BOMImage.objects.order_by("sort_order", "id"),
            ),
            Prefetch(
                "jobber_type_processes",
                queryset=BOMJobberTypeProcess.objects.select_related("jobber_type").order_by("sort_order", "id"),
            ),
            Prefetch(
                "jobber_details",
                queryset=BOMJobberDetail.objects.select_related("jobber", "jobber_type").order_by("sort_order", "id"),
            ),
        )
        .order_by("sku")
    )

    payload = {}

    for bom in boms:
        linked_fabrics = [
            item.material.name
            for item in bom.material_items.all()
            if item.material_id and item.material
        ]
        linked_accessories = [
            item.accessory.name
            for item in bom.accessory_items.all()
            if item.accessory_id and item.accessory
        ]
        accessories_price = sum(
            ((item.cost or Decimal("0")) for item in bom.accessory_items.all()),
            Decimal("0"),
        )

        jobber_process_prices = {}
        linked_processes = []
        for row in bom.jobber_type_processes.all():
            if not row.jobber_type_id:
                continue
            process_key = str(row.jobber_type_id)
            jobber_process_prices.setdefault(process_key, str(row.price or Decimal("0")))
            linked_processes.append(
                f"{row.jobber_type.name} • {row.price or Decimal('0')}"
            )

        bom_jobber_details = {}
        linked_jobbers = []
        for row in bom.jobber_details.all():
            if not row.jobber_id or not row.jobber_type_id:
                continue

            jobber_key = str(row.jobber_id)
            process_key = str(row.jobber_type_id)

            entry = bom_jobber_details.setdefault(
                jobber_key,
                {
                    "jobber_name": row.jobber.name if row.jobber else "",
                    "default_jobber_type_id": "",
                    "process_prices": {},
                },
            )

            if not entry["default_jobber_type_id"]:
                entry["default_jobber_type_id"] = process_key

            entry["process_prices"][process_key] = str(row.price or Decimal("0"))

            linked_jobbers.append(
                f"{row.jobber.name} • {row.jobber_type.name} • {row.price or Decimal('0')}"
            )

        payload[str(bom.pk)] = {
            "bom_code": bom.bom_code or "",
            "sku": bom.sku or "",
            "product_name": bom.product_name or "",
            "catalogue": getattr(bom.catalogue, "name", "") or "",
            "brand": getattr(bom.brand, "name", "") or "",
            "gender": bom.gender or "",
            "main_category": getattr(bom.main_category, "name", "") or "",
            "category": getattr(bom.category, "name", "") or "",
            "sub_category": getattr(bom.sub_category, "name", "") or "",
            "pattern_type": getattr(bom.pattern_type, "name", "") or "",
            "character_name": bom.character_name or "",
            "license_name": bom.license_name or "",
            "size_type": bom.get_size_type_display() or "",
            "mrp": str(bom.mrp or Decimal("0")),
            "color": bom.color or "",
            "drawcord": bom.drawcord or "",
            "tie_dye_price": str(bom.tie_dye_price or Decimal("0")),
            "selling_price": str(bom.selling_price or Decimal("0")),
            "maintenance_price": str(bom.maintenance_price or Decimal("0")),
            "final_price": str(bom.final_price or Decimal("0")),
            "accessories_price": str(accessories_price),
            "notes": bom.notes or "",
            "status": bom.get_status_display() or "",
            "linked_fabrics": linked_fabrics,
            "linked_accessories": linked_accessories,
            "linked_processes": linked_processes,
            "linked_jobbers": linked_jobbers,
            "image_url": _bom_preview_image_url(bom),
            "jobber_process_prices": jobber_process_prices,
            "bom_jobber_details": bom_jobber_details,
        }

    return payload


def _program_jobber_defaults(user):
    rows = (
        Jobber.objects.filter(owner=user, is_active=True)
        .select_related("jobber_type")
        .order_by("name")
    )

    return {
        str(row.pk): {
            "jobber_type_id": str(row.jobber_type_id) if row.jobber_type_id else "",
            "jobber_type_name": row.jobber_type.name if row.jobber_type_id and row.jobber_type else "",
        }
        for row in rows
    }

def _program_bom_jobber_price(bom, jobber_id=None, jobber_type_id=None):
    if not bom or not jobber_type_id:
        return None

    if jobber_id:
        specific_row = (
            BOMJobberDetail.objects.filter(
                bom=bom,
                jobber_id=jobber_id,
                jobber_type_id=jobber_type_id,
            )
            .order_by("sort_order", "id")
            .first()
        )
        if specific_row is not None:
            return specific_row.price or Decimal("0")

    process_row = (
        BOMJobberTypeProcess.objects.filter(
            bom=bom,
            jobber_type_id=jobber_type_id,
        )
        .order_by("sort_order", "id")
        .first()
    )
    if process_row is not None:
        return process_row.price or Decimal("0")

    return None


def _program_start_jobber_price_payload(program):
    bom = getattr(program, "bom", None)
    payload = {
        "jobber_type_prices": {},
        "jobber_detail_prices": {},
    }

    if bom is None:
        return payload

    for row in bom.jobber_type_processes.select_related("jobber_type").all():
        if row.jobber_type_id:
            payload["jobber_type_prices"][str(row.jobber_type_id)] = str(row.price or Decimal("0"))

    for row in bom.jobber_details.select_related("jobber", "jobber_type").all():
        if row.jobber_id and row.jobber_type_id:
            payload["jobber_detail_prices"][f"{row.jobber_id}:{row.jobber_type_id}"] = str(row.price or Decimal("0"))

    return payload


def _apply_program_start_jobber_prices_from_bom(program, jobber_formset):
    bom = getattr(program, "bom", None)
    if bom is None:
        return

    for child_form in jobber_formset.forms:
        if not hasattr(child_form, "cleaned_data"):
            continue

        cleaned = child_form.cleaned_data
        if not cleaned or cleaned.get("DELETE"):
            continue

        jobber = cleaned.get("jobber")
        jobber_type = cleaned.get("jobber_type")
        jobber_price = cleaned.get("jobber_price")

        if jobber and not jobber_type and getattr(jobber, "jobber_type_id", None):
            try:
                jobber_type = jobber.jobber_type
            except Exception:
                jobber_type = None

            if jobber_type is not None:
                child_form.instance.jobber_type = jobber_type

        if not jobber_type:
            continue

        if jobber_price not in (None, "", 0, Decimal("0")):
            continue

        mapped_price = _program_bom_jobber_price(
            bom=bom,
            jobber_id=getattr(jobber, "id", None),
            jobber_type_id=getattr(jobber_type, "id", None),
        )
        if mapped_price is not None:
            child_form.instance.jobber_price = mapped_price


def _apply_program_jobber_prices_from_bom(program, jobber_formset):
    bom = getattr(program, "bom", None)
    if bom is None:
        return

    for child_form in jobber_formset.forms:
        if not hasattr(child_form, "cleaned_data"):
            continue

        cleaned = child_form.cleaned_data
        if not cleaned or cleaned.get("DELETE"):
            continue

        jobber = cleaned.get("jobber")
        jobber_type = cleaned.get("jobber_type")
        price = cleaned.get("price")

        if not jobber_type:
            continue

        if price not in (None, "", 0, Decimal("0")):
            continue

        mapped_price = _program_bom_jobber_price(
            bom=bom,
            jobber_id=getattr(jobber, "id", None),
            jobber_type_id=getattr(jobber_type, "id", None),
        )
        if mapped_price is not None:
            child_form.instance.price = mapped_price

def _ensure_program_size_rows(program):
    default_rows = [
        ("CQ", 1),
        ("FQ", 2),
        ("DQ", 3),
        ("FQ-DQ", 4),
        ("TP", 5),
    ]

    for line_name, order in default_rows:
        ProgramSizeDetail.objects.get_or_create(
            program=program,
            line_name=line_name,
            defaults={"sort_order": order},
        )

@login_required
@require_POST
def bom_delete(request, pk: int):
    bom = get_object_or_404(
        BOM.objects.prefetch_related("programs", "costing_snapshots"),
        pk=pk,
        owner=request.user,
    )

    usage_rows = _bom_usage_rows(bom)

    if usage_rows:
        error_message = (
            f'Cannot delete "{bom.bom_code} - {bom.product_name}" because it is already used in linked records. '
            f'Mark it inactive instead.'
        )

        if _is_embed(request):
            context = _bom_list_context(request)
            context["delete_error"] = error_message
            context["delete_label"] = f"{bom.bom_code} - {bom.product_name}"
            context["delete_usage"] = usage_rows
            return render(request, "accounts/bom/list_embed.html", context, status=400)

        messages.error(request, error_message)
        return redirect("accounts:bom_list")

    bom.delete()

    url = _bom_list_url(request)
    if _is_embed(request):
        return JsonResponse({"ok": True, "url": url})
    return redirect(url)

@login_required
@require_http_methods(["GET", "POST"])
def program_create(request):
    program = Program(owner=request.user)

    if request.method == "GET":
        program.program_no = Program.next_program_no(request.user)
        program.program_date = timezone.localdate()

        user_firm = Firm.objects.filter(owner=request.user).first()
        if user_firm:
            program.firm = user_firm

    form = ProgramForm(
        request.POST or None,
        instance=program,
        user=request.user,
    )

    if request.method == "POST":
        jobber_formset = ProgramJobberDetailFormSet(
            request.POST,
            instance=program,
            prefix="jobbers",
            form_kwargs={"user": request.user},
        )

        if form.is_valid() and jobber_formset.is_valid():
            with transaction.atomic():
                program = form.save(commit=False)
                program.owner = request.user
                program.save()
                _ensure_program_size_rows(program)
                _apply_program_jobber_prices_from_bom(program, jobber_formset)
                jobber_formset.instance = program
                jobber_formset.save()

                messages.success(request, "Program saved successfully.")
                list_url = reverse("accounts:program_list")
                if _is_embed(request):
                    return JsonResponse({"ok": True, "url": list_url})
                return redirect(list_url)
    else:
        jobber_formset = ProgramJobberDetailFormSet(
            instance=program,
            prefix="jobbers",
            form_kwargs={"user": request.user},
        )

    template = "accounts/programs/form_embed.html" if _is_embed(request) else "accounts/programs/form.html"
    return render(
        request,
        template,
        {
            "form": form,
            "jobber_formset": jobber_formset,
            "mode": "add",
            "full_page": not _is_embed(request),
            "action_url": reverse("accounts:program_add"),
            "sku_payloads": _program_sku_payloads(request.user),
            "jobber_defaults": _program_jobber_defaults(request.user),
        },
    )

@login_required
def program_list(request):
    q = (request.GET.get("q") or "").strip()

    qs = (
        Program.objects.filter(owner=request.user)
        .select_related("bom", "firm", "start_record")
        .prefetch_related(
            "size_rows",
            "jobber_rows__jobber",
            "jobber_rows__jobber_type",
            "start_record__size_rows",
            "start_record__jobber_rows__jobber",
            "start_record__jobber_rows__jobber_type",
            "start_record__jobber_rows__challans",
        )
        .order_by("-created_at", "-id")
    )

    if q:
        qs = qs.filter(
            Q(program_no__icontains=q)
            | Q(bom__sku__icontains=q)
            | Q(bom__product_name__icontains=q)
            | Q(firm__firm_name__icontains=q)
        ).distinct()

    programs = list(qs)

    for program in programs:
        start_record = getattr(program, "start_record", None)
        is_started = bool(start_record and getattr(start_record, "is_started", False))

        program.jobber_tracker_rows = _program_start_jobber_tracker_rows(program, start_record)
        program.size_display_rows = _program_list_size_rows(program)
        _size_keys = [
            ("xs_qty", "XS"),
            ("s_qty", "S"),
            ("m_qty", "M"),
            ("l_qty", "L"),
            ("xl_qty", "XL"),
            ("xxl_qty", "XXL"),
        ]
        program.visible_size_columns = [
            label
            for key, label in _size_keys
            if any((row.get(key) or Decimal("0")) != Decimal("0") for row in program.size_display_rows)
        ]
        if not program.visible_size_columns:
            program.visible_size_columns = ["XS"]
        _visible_keys = {"XS": "xs_qty", "S": "s_qty", "M": "m_qty", "L": "l_qty", "XL": "xl_qty", "XXL": "xxl_qty"}
        for _row in program.size_display_rows:
            _row["cells"] = [
                {"label": label, "value": _row.get(_visible_keys[label]) or Decimal("0")}
                for label in program.visible_size_columns
            ]
        if not program.size_display_rows:
            program.size_display_rows = [
                {"line_name": "CQ", "xs_qty": Decimal("0"), "s_qty": Decimal("0"), "m_qty": Decimal("0"), "l_qty": Decimal("0"), "xl_qty": Decimal("0"), "xxl_qty": Decimal("0"), "total_qty": Decimal("0")},
                {"line_name": "FQ", "xs_qty": Decimal("0"), "s_qty": Decimal("0"), "m_qty": Decimal("0"), "l_qty": Decimal("0"), "xl_qty": Decimal("0"), "xxl_qty": Decimal("0"), "total_qty": Decimal("0")},
                {"line_name": "DQ", "xs_qty": Decimal("0"), "s_qty": Decimal("0"), "m_qty": Decimal("0"), "l_qty": Decimal("0"), "xl_qty": Decimal("0"), "xxl_qty": Decimal("0"), "total_qty": Decimal("0")},
                {"line_name": "FQ-DQ", "xs_qty": Decimal("0"), "s_qty": Decimal("0"), "m_qty": Decimal("0"), "l_qty": Decimal("0"), "xl_qty": Decimal("0"), "xxl_qty": Decimal("0"), "total_qty": Decimal("0")},
                {"line_name": "TP", "xs_qty": Decimal("0"), "s_qty": Decimal("0"), "m_qty": Decimal("0"), "l_qty": Decimal("0"), "xl_qty": Decimal("0"), "xxl_qty": Decimal("0"), "total_qty": Decimal("0")},
            ]

        challans = ProgramJobberChallan.objects.filter(
            owner=request.user,
            program=program,
        )

        challan_count = challans.count()
        total_issued = challans.aggregate(total=Sum("total_issued_qty")).get("total") or Decimal("0")
        total_inward = challans.aggregate(total=Sum("inward_qty")).get("total") or Decimal("0")

        pending_qty = total_issued - total_inward
        if pending_qty < Decimal("0"):
            pending_qty = Decimal("0")

        if not is_started:
            production_status = "Not Started"
        elif challan_count == 0:
            production_status = "Started"
        elif total_issued > Decimal("0") and total_inward >= total_issued:
            production_status = "Completed"
        else:
            production_status = "In Progress"

        program.total_issued = total_issued
        program.total_inward = total_inward
        program.pending_qty = pending_qty
        program.production_status = production_status
        program.challan_count = challan_count

    verified_count = sum(1 for program in programs if getattr(program, "is_verified", False))
    unverified_count = len(programs) - verified_count

    return render(
        request,
        "accounts/programs/list.html",
        {
            "programs": programs,
            "q": q,
            "verified_count": verified_count,
            "unverified_count": unverified_count,
        },
    )

@login_required
@require_http_methods(["GET", "POST"])
def program_update(request, pk: int):
    program = get_object_or_404(Program, pk=pk, owner=request.user)

    form = ProgramForm(
        request.POST or None,
        instance=program,
        user=request.user,
    )

    if request.method == "POST":
        jobber_formset = ProgramJobberDetailFormSet(
            request.POST,
            instance=program,
            prefix="jobbers",
            form_kwargs={"user": request.user},
        )

        if form.is_valid() and jobber_formset.is_valid():
            with transaction.atomic():
                program = form.save()
                _ensure_program_size_rows(program)
                _apply_program_jobber_prices_from_bom(program, jobber_formset)
                jobber_formset.save()

            messages.success(request, "Program updated successfully.")
            list_url = reverse("accounts:program_list")
            if _is_embed(request):
                return JsonResponse({"ok": True, "url": list_url})
            return redirect(list_url)
    else:
        jobber_formset = ProgramJobberDetailFormSet(
            instance=program,
            prefix="jobbers",
            form_kwargs={"user": request.user},
        )

    template = "accounts/programs/form_embed.html" if _is_embed(request) else "accounts/programs/form.html"
    return render(
        request,
        template,
        {
            "form": form,
            "jobber_formset": jobber_formset,
            "mode": "edit",
            "program": program,
            "full_page": not _is_embed(request),
            "action_url": reverse("accounts:program_edit", args=[program.pk]),
            "sku_payloads": _program_sku_payloads(request.user),
            "jobber_defaults": _program_jobber_defaults(request.user),
        },
    )

def _program_list_size_rows(program):
    """
    Always return all Size Progress rows for the Program List:
    CQ / FQ / DQ / FQ-DQ / TP.
    The template keeps the full wide XS/S/M/L/XL/XXL/Total table.
    """
    size_order = ["XS", "S", "M", "L", "XL", "XXL"]

    def blank_map():
        return {size: Decimal("0") for size in size_order}

    tp_map = blank_map()
    cq_map = blank_map()
    fq_map = blank_map()
    dq_map = blank_map()

    for row in program.size_rows.all():
        line_name = (getattr(row, "line_name", "") or "").strip().upper()

        if line_name in size_order:
            value = getattr(row, f"{line_name.lower()}_qty", Decimal("0")) or Decimal("0")
            tp_map[line_name] += value
        else:
            tp_map["XS"] += row.xs_qty or Decimal("0")
            tp_map["S"] += row.s_qty or Decimal("0")
            tp_map["M"] += row.m_qty or Decimal("0")
            tp_map["L"] += row.l_qty or Decimal("0")
            tp_map["XL"] += row.xl_qty or Decimal("0")
            tp_map["XXL"] += row.xxl_qty or Decimal("0")

    challan_sizes = ProgramJobberChallanSize.objects.filter(
        challan__owner=program.owner,
        challan__program=program,
    )

    for row in challan_sizes:
        size_name = (row.size_name or "").strip().upper()
        if size_name not in cq_map:
            continue

        cq_map[size_name] += row.issued_qty or Decimal("0")
        fq_map[size_name] += row.inward_qty or Decimal("0")

    fq_dq_map = {
        size: (fq_map[size] - dq_map[size])
        for size in size_order
    }

    def build_row(label, source_map):
        return {
            "line_name": label,
            "xs_qty": source_map["XS"],
            "s_qty": source_map["S"],
            "m_qty": source_map["M"],
            "l_qty": source_map["L"],
            "xl_qty": source_map["XL"],
            "xxl_qty": source_map["XXL"],
            "total_qty": sum(source_map.values()),
        }

    return [
        build_row("CQ", cq_map),
        build_row("FQ", fq_map),
        build_row("DQ", dq_map),
        build_row("FQ-DQ", fq_dq_map),
        build_row("TP", tp_map),
    ]



def _can_verify_program(request_or_user):
    """Use ERP permissions so company admins and permitted staff can verify programs."""
    if hasattr(request_or_user, "erp_actor") or hasattr(request_or_user, "erp_is_company_admin"):
        return has_erp_permission(request_or_user, "program.verify")

    user = request_or_user
    if not user or not getattr(user, "is_authenticated", False):
        return False
    if getattr(user, "is_superuser", False) or getattr(user, "is_staff", False):
        return True

    try:
        profile = user.erp_profile
    except Exception:
        profile = None

    if not profile or not getattr(profile, "is_active", False):
        return False
    if not getattr(user, "is_active", True):
        return False
    if getattr(profile, "company", None) and not getattr(profile.company, "is_active_company", True):
        return False
    if getattr(profile, "is_company_admin", False):
        return True

    role = getattr(profile, "role", None)
    return bool(role and getattr(role, "is_active", False) and "program.verify" in (role.permissions or []))


def _program_verification_response(request, program, message, status=400):
    if _is_program_popup(request):
        return JsonResponse({"ok": False, "message": message}, status=status)
    messages.error(request, message)
    return redirect("accounts:program_list")


@login_required
@require_http_methods(["GET", "POST"])
def program_verify(request, pk: int):
    program = get_object_or_404(
        Program.objects.filter(owner=request.user).select_related("bom", "firm"),
        pk=pk,
    )

    popup_mode = _is_program_popup(request)

    if request.method == "POST":
        if not _can_verify_program(request):
            return _program_verification_response(
                request,
                program,
                "Only admin users can verify or unverify programs.",
                status=403,
            )

        decision = (request.POST.get("decision") or "").strip().lower()

        if decision == "approve":
            program.is_verified = True
            program.save(update_fields=["is_verified"])
            message = f"Program {program.program_no} verified successfully."
        elif decision in {"reject", "unverify"}:
            program.is_verified = False
            program.save(update_fields=["is_verified"])
            message = f"Program {program.program_no} marked unverified."
        else:
            return _program_verification_response(
                request,
                program,
                "Please choose Verify or Unverify.",
                status=400,
            )

        list_url = reverse("accounts:program_list")
        if popup_mode:
            return JsonResponse({"ok": True, "message": message, "redirect_url": list_url})

        messages.success(request, message)
        return redirect(list_url)

    return render(
        request,
        "accounts/programs/verify.html",
        {
            "program": program,
            "popup_mode": popup_mode,
            "can_verify_program": _can_verify_program(request),
        },
    )
@login_required
@require_POST
def program_toggle_verify(request, pk: int):
    program = get_object_or_404(Program, pk=pk, owner=request.user)
    program.is_verified = not program.is_verified
    program.save(update_fields=["is_verified"])
    messages.success(
        request,
        f"Program {'verified' if program.is_verified else 'marked unverified'} successfully."
    )
    return redirect("accounts:program_list")


@login_required
@require_POST
def program_toggle_status(request, pk: int):
    program = get_object_or_404(Program, pk=pk, owner=request.user)
    program.status = "closed" if program.status == "open" else "open"
    program.save(update_fields=["status"])
    messages.success(request, f"Program marked {program.status}.")
    return redirect("accounts:program_list")


@login_required
def program_print(request, pk: int):
    program = get_object_or_404(
        Program.objects.select_related("bom", "firm").prefetch_related(
            Prefetch(
                "jobber_rows",
                queryset=ProgramJobberDetail.objects.select_related("jobber", "jobber_type").order_by("sort_order", "id"),
            ),
            Prefetch(
                "size_rows",
                queryset=ProgramSizeDetail.objects.order_by("sort_order", "id"),
            ),
            Prefetch(
                "bom__images",
                queryset=BOMImage.objects.order_by("sort_order", "id"),
            ),
        ),
        pk=pk,
        owner=request.user,
    )
    return render(
        request,
        "accounts/programs/print.html",
        {
            "program": program,
        },
    )


def _dispatch_list_url(request):
    url = reverse("accounts:dispatch_list")
    if _is_embed(request):
        url += "?embed=1"
    return url


def _dispatch_feature_available() -> bool:
    return DispatchChallan is not None


def _dispatch_feature_unavailable_response(request):
    message = "Dispatch module is not available because DispatchChallan model/form is missing."
    if _is_embed(request):
        return JsonResponse({"ok": False, "message": message}, status=400)
    messages.error(request, message)
    return redirect("accounts:dashboard")


@login_required
def dispatch_list(request):
    if not _dispatch_feature_available():
        return _dispatch_feature_unavailable_response(request)

    q = (request.GET.get("q") or "").strip()

    challans = (
        DispatchChallan.objects.filter(owner=request.user)
        .select_related("program", "program__bom", "program__firm", "client", "firm")
        .order_by("-challan_date", "-id")
    )

    if q:
        challans = challans.filter(
            Q(challan_no__icontains=q)
            | Q(program__program_no__icontains=q)
            | Q(program__bom__sku__icontains=q)
            | Q(program__bom__product_name__icontains=q)
            | Q(client__name__icontains=q)
            | Q(lr_no__icontains=q)
            | Q(transport_name__icontains=q)
            | Q(vehicle_no__icontains=q)
            | Q(driver_name__icontains=q)
        )

    template = (
        "accounts/dispatch/list_embed.html"
        if _is_embed(request)
        else "accounts/dispatch/list.html"
    )
    return render(
        request,
        template,
        {
            "challans": challans,
            "q": q,
        },
    )


@login_required
def dispatch_program_picker(request):
    q = (request.GET.get("q") or "").strip()

    programs = (
        Program.objects.filter(owner=request.user)
        .select_related("bom", "firm")
        .annotate(challan_count=Count("dispatch_challans"))
        .order_by("-finishing_date", "-program_date", "-id")
    )

    if q:
        programs = programs.filter(
            Q(program_no__icontains=q)
            | Q(bom__sku__icontains=q)
            | Q(bom__product_name__icontains=q)
        )

    template = (
        "accounts/dispatch/program_picker_embed.html"
        if _is_embed(request)
        else "accounts/dispatch/program_picker.html"
    )
    return render(
        request,
        template,
        {
            "programs": programs,
            "q": q,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def dispatch_create(request, program_id: int):
    if not _dispatch_feature_available():
        return _dispatch_feature_unavailable_response(request)

    program = get_object_or_404(
        Program.objects.select_related("bom", "firm"),
        pk=program_id,
        owner=request.user,
    )

    challan = DispatchChallan(
        owner=request.user,
        program=program,
        firm=program.firm,
    )

    if request.method == "GET":
        challan.challan_no = DispatchChallan.next_challan_no(request.user)
        challan.challan_date = timezone.localdate()

    form = DispatchChallanForm(
        request.POST or None,
        instance=challan,
        user=request.user,
        program=program,
    )

    if request.method == "POST" and form.is_valid():
        with transaction.atomic():
            challan = form.save(commit=False)
            challan.owner = request.user
            challan.program = program
            challan.firm = program.firm
            challan.save()

        messages.success(request, f"Dispatch challan {challan.challan_no} created successfully.")

        url = _dispatch_list_url(request)
        if _is_embed(request):
            return JsonResponse({"ok": True, "url": url})
        return redirect(url)

    template = (
        "accounts/dispatch/form_embed.html"
        if _is_embed(request)
        else "accounts/dispatch/form.html"
    )
    return render(
        request,
        template,
        {
            "form": form,
            "mode": "add",
            "program": program,
            "challan_obj": None,
        },
    )


@login_required
def dispatch_detail(request, pk: int):
    if not _dispatch_feature_available():
        return _dispatch_feature_unavailable_response(request)

    challan = get_object_or_404(
        DispatchChallan.objects.select_related("program", "program__bom", "program__firm", "client", "firm"),
        pk=pk,
        owner=request.user,
    )

    template = (
        "accounts/dispatch/detail_embed.html"
        if _is_embed(request)
        else "accounts/dispatch/detail.html"
    )
    return render(
        request,
        template,
        {
            "challan": challan,
            "program": challan.program,
        },
    )
def _build_stitching_inwards_report_context(request):
    q = (request.GET.get("q") or "").strip()

    programs = (
        Program.objects.filter(owner=request.user)
        .select_related(
            "bom",
            "bom__brand",
            "bom__main_category",
            "bom__sub_category",
            "bom__pattern_type",
        )
        .prefetch_related(
            "jobber_challans__jobber",
            "jobber_challans__jobber_type",
        )
        .order_by("-program_date", "-id")
    )

    rows = []

    for program in programs:
        bom = program.bom

        challans = program.jobber_challans.all()

        for ch in challans:
            jobber_type_name = ""
            if ch.jobber_type:
                jobber_type_name = (ch.jobber_type.name or "").lower()

            # 🔥 Only stitching
            if "stitch" not in jobber_type_name:
                continue

            inward_qty = ch.inward_qty or 0

            if inward_qty <= 0:
                continue

            if q:
                if (
                    q.lower() not in (program.program_no or "").lower()
                    and q.lower() not in (bom.sku or "").lower()
                    and q.lower() not in (ch.jobber.name if ch.jobber else "").lower()
                ):
                    continue

            rows.append(
                {
                    "program_no": program.program_no,
                    "program_date": program.program_date,
                    "jobber": ch.jobber.name if ch.jobber else "-",
                    "category": getattr(bom, "category", "") or "-",
                    "main_category": bom.main_category.name if bom and bom.main_category else "-",
                    "sub_category": bom.sub_category.name if bom and bom.sub_category else "-",
                    "brand": bom.brand.name if bom and bom.brand else "-",
                    "gender": bom.gender or "-",
                    "pattern": bom.pattern_type.name if bom and bom.pattern_type else "-",
                    "sku": bom.sku or "-",
                    "total_qty": program.production_qty or program.qty or 0,
                    "inward_date": ch.updated_at if hasattr(ch, "updated_at") else None,
                }
            )

    return {
        "rows": rows,
        "q": q,
    }


@login_required
def report_stitching_inwards(request):
    context = _build_stitching_inwards_report_context(request)
    embed_mode = request.GET.get("embed") == "1"
    context["embed_mode"] = embed_mode

    template_name = (
        "accounts/reports/stitching_inwards_embed.html"
        if embed_mode
        else "accounts/reports/stitching_inwards.html"
    )

    return render(request, template_name, context)


@login_required
def report_stitching_inwards_excel(request):
    context = _build_stitching_inwards_report_context(request)
    rows = context["rows"]
    q = context.get("q", "")

    wb = Workbook()
    ws = wb.active
    ws.title = "Stitching Inwards"

    title_font = Font(bold=True, size=14, color="1F2937")
    header_font = Font(bold=True, color="111827")
    bold_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="E5E7EB")
    total_fill = PatternFill(fill_type="solid", fgColor="EEF2F7")

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    left_align = Alignment(horizontal="left", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    headers = [
        "Program No.",
        "Program Date",
        "Stitching Jobber",
        "Category",
        "Main Category",
        "Sub Category",
        "Brand",
        "Gender",
        "Pattern Name",
        "SKU Name",
        "Total",
        "Inward Date",
    ]

    total_columns = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    ws["A1"] = "Stitching Inwards Report"
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align

    ws["A2"] = "Search"
    ws["B2"] = q or "-"
    ws["A2"].font = bold_font

    header_row = 4
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align

    data_row = header_row + 1
    total_qty_sum = 0

    for row in rows:
        total_qty = row.get("total_qty") or 0
        total_qty_sum += total_qty

        program_date = row.get("program_date")
        inward_date = row.get("inward_date")

        values = [
            row.get("program_no") or "-",
            program_date.strftime("%d-%m-%Y") if program_date else "-",
            row.get("jobber") or "-",
            row.get("category") or "-",
            row.get("main_category") or "-",
            row.get("sub_category") or "-",
            row.get("brand") or "-",
            row.get("gender") or "-",
            row.get("pattern") or "-",
            row.get("sku") or "-",
            float(total_qty),
            inward_date.strftime("%d-%m-%Y") if inward_date else "-",
        ]

        for col_num, value in enumerate(values, start=1):
            cell = ws.cell(row=data_row, column=col_num, value=value)
            cell.border = border
            if col_num == 11:
                cell.alignment = right_align
            elif col_num in (2, 12):
                cell.alignment = center_align
            else:
                cell.alignment = left_align

        data_row += 1

    total_row = data_row
    for col_num in range(1, total_columns + 1):
        cell = ws.cell(row=total_row, column=col_num)
        cell.border = border
        cell.fill = total_fill
        cell.font = bold_font

    ws.cell(row=total_row, column=1, value="Total")
    ws.cell(row=total_row, column=11, value=float(total_qty_sum))
    ws.cell(row=total_row, column=11).alignment = right_align

    widths = {
        "A": 16,
        "B": 14,
        "C": 22,
        "D": 16,
        "E": 18,
        "F": 18,
        "G": 14,
        "H": 12,
        "I": 18,
        "J": 24,
        "K": 12,
        "L": 14,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="stitching_inwards_report.xlsx"'
    wb.save(response)
    return response

def _build_dispatch_challan_pdf_response(challan):
    try:
        import os
        from pathlib import Path
        from html import escape

        from django.conf import settings
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.lib.utils import ImageReader
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError:
        return HttpResponse(
            "ReportLab is required for PDF generation. Install it with: pip install reportlab",
            status=500,
        )

    brand_pink = colors.HexColor("#ED2F8C")
    brand_orange = colors.HexColor("#F6A33B")
    brand_blue = colors.HexColor("#1976F3")
    brand_navy = colors.HexColor("#0F172A")
    ink = colors.HexColor("#1F2937")
    muted = colors.HexColor("#667085")
    border = colors.HexColor("#D0D5DD")
    soft_bg = colors.HexColor("#F8FAFC")
    white = colors.white

    def text_or_dash(value):
        value = "" if value is None else str(value).strip()
        return value if value else "-"

    def fmt_qty(value):
        try:
            return f"{float(value or 0):,.2f}".rstrip("0").rstrip(".")
        except Exception:
            return text_or_dash(value)

    def join_parts(*parts):
        clean = [str(p).strip() for p in parts if str(p).strip()]
        return ", ".join(clean)

    def line_if(label, value):
        value = "" if value is None else str(value).strip()
        if not value:
            return ""
        return f"<b>{escape(label)}:</b> {escape(value)}"

    def resolve_logo_path():
        firm = challan.firm
        if firm and getattr(firm, "logo", None):
            try:
                logo_path = firm.logo.path
                if logo_path and os.path.exists(logo_path):
                    return logo_path
            except Exception:
                pass

        fallback = Path(settings.BASE_DIR) / "Logo.jpeg"
        if fallback.exists():
            return str(fallback)

        return None

    logo_path = resolve_logo_path()

    def draw_branding(canvas, doc):
        page_w, page_h = A4
        canvas.saveState()

        canvas.setStrokeColor(colors.HexColor("#E4E7EC"))
        canvas.setLineWidth(0.8)
        canvas.roundRect(8 * mm, 8 * mm, page_w - 16 * mm, page_h - 16 * mm, 4 * mm, stroke=1, fill=0)

        usable_w = page_w - 16 * mm
        stripe_w = usable_w / 3.0
        stripe_y = page_h - 13 * mm
        stripe_h = 4.5 * mm

        canvas.setFillColor(brand_pink)
        canvas.rect(8 * mm, stripe_y, stripe_w, stripe_h, fill=1, stroke=0)

        canvas.setFillColor(brand_orange)
        canvas.rect(8 * mm + stripe_w, stripe_y, stripe_w, stripe_h, fill=1, stroke=0)

        canvas.setFillColor(brand_blue)
        canvas.rect(8 * mm + 2 * stripe_w, stripe_y, stripe_w, stripe_h, fill=1, stroke=0)

        if logo_path:
            try:
                img = ImageReader(logo_path)
                iw, ih = img.getSize()
                draw_w = 26 * mm
                draw_h = draw_w * (ih / float(iw)) if iw and ih else 26 * mm
                x = (page_w - draw_w) / 2.0
                y = 10 * mm

                try:
                    canvas.setFillAlpha(0.10)
                    canvas.setStrokeAlpha(0.10)
                except Exception:
                    pass

                canvas.drawImage(
                    img,
                    x,
                    y,
                    width=draw_w,
                    height=draw_h,
                    preserveAspectRatio=True,
                    mask="auto",
                )
            except Exception:
                pass

        canvas.restoreState()

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
    )

    styles = getSampleStyleSheet()

    base_style = ParagraphStyle(
        "DispatchBase",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8.5,
        leading=10.5,
        textColor=ink,
        spaceAfter=0,
    )

    header_left_style = ParagraphStyle(
        "DispatchHeaderLeft",
        parent=base_style,
        fontName="Helvetica",
        fontSize=8.4,
        leading=10.4,
        textColor=white,
        alignment=TA_LEFT,
    )

    header_title_style = ParagraphStyle(
        "DispatchHeaderTitle",
        parent=base_style,
        fontName="Helvetica-Bold",
        fontSize=14,
        leading=16,
        textColor=brand_navy,
        alignment=TA_RIGHT,
    )

    header_meta_style = ParagraphStyle(
        "DispatchHeaderMeta",
        parent=base_style,
        fontName="Helvetica",
        fontSize=8.3,
        leading=10.2,
        textColor=ink,
        alignment=TA_RIGHT,
    )

    section_head_left = ParagraphStyle(
        "DispatchSectionHeadLeft",
        parent=base_style,
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=10,
        textColor=white,
        alignment=TA_LEFT,
    )

    section_value_style = ParagraphStyle(
        "DispatchSectionValue",
        parent=base_style,
        fontName="Helvetica",
        fontSize=8.2,
        leading=10.2,
        textColor=ink,
        alignment=TA_LEFT,
    )

    table_head_style = ParagraphStyle(
        "DispatchTableHead",
        parent=base_style,
        fontName="Helvetica-Bold",
        fontSize=7.8,
        leading=9.5,
        textColor=white,
        alignment=TA_CENTER,
    )

    item_style = ParagraphStyle(
        "DispatchItem",
        parent=base_style,
        fontName="Helvetica",
        fontSize=8,
        leading=9.6,
        alignment=TA_LEFT,
    )

    centered_style = ParagraphStyle(
        "DispatchCentered",
        parent=base_style,
        fontName="Helvetica",
        fontSize=8,
        leading=9.6,
        alignment=TA_CENTER,
    )

    sign_style = ParagraphStyle(
        "DispatchSign",
        parent=base_style,
        fontName="Helvetica-Bold",
        fontSize=8,
        leading=10,
        textColor=ink,
        alignment=TA_LEFT,
    )

    footer_note_style = ParagraphStyle(
        "DispatchFooterNote",
        parent=base_style,
        fontName="Helvetica-Bold",
        fontSize=7.6,
        leading=9.2,
        textColor=muted,
        alignment=TA_CENTER,
    )

    story = []

    program = challan.program
    bom = getattr(program, "bom", None)
    client = challan.client
    firm = challan.firm

    firm_name = text_or_dash(firm.firm_name if firm else "InventTech")
    firm_type = ""
    if firm:
        try:
            firm_type = firm.get_firm_type_display()
        except Exception:
            firm_type = text_or_dash(getattr(firm, "firm_type", ""))

    firm_address = join_parts(
        getattr(firm, "address_line", ""),
        getattr(firm, "city", ""),
        getattr(firm, "state", ""),
        getattr(firm, "pincode", ""),
    )

    firm_contact_line = " | ".join([
        part for part in [
            line_if("Phone", getattr(firm, "phone", "")),
            line_if("Email", getattr(firm, "email", "")),
            line_if("GSTIN", getattr(firm, "gst_number", "")),
        ] if part
    ])

    header_left_html = f"<font size='13'><b>{escape(firm_name)}</b></font>"
    if firm_type and firm_type != "-":
        header_left_html += f"<br/>{escape(firm_type)}"
    if firm_address:
        header_left_html += f"<br/>{escape(firm_address)}"
    if firm_contact_line:
        header_left_html += f"<br/>{firm_contact_line}"

    challan_date = challan.challan_date.strftime("%d-%m-%Y") if challan.challan_date else "-"
    finishing_date = program.finishing_date.strftime("%d-%m-%Y") if getattr(program, "finishing_date", None) else "-"
    program_date = program.program_date.strftime("%d-%m-%Y") if getattr(program, "program_date", None) else "-"

    right_meta_html = (
        f"<b>Challan No:</b> {escape(text_or_dash(challan.challan_no))}<br/>"
        f"<b>Challan Date:</b> {escape(challan_date)}<br/>"
        f"<b>Program No:</b> {escape(text_or_dash(getattr(program, 'program_no', '')))}<br/>"
        f"<b>Program Date:</b> {escape(program_date)}<br/>"
        f"<b>Finishing Date:</b> {escape(finishing_date)}"
    )

    header_table = Table(
        [[
            Paragraph(header_left_html, header_left_style),
            Table(
                [
                    [Paragraph("DISPATCH CHALLAN", header_title_style)],
                    [Paragraph(right_meta_html, header_meta_style)],
                ],
                colWidths=[68 * mm],
            ),
        ]],
        colWidths=[122 * mm, 68 * mm],
    )
    header_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), brand_navy),
        ("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#F4F8FF")),
        ("BOX", (0, 0), (-1, -1), 0.9, border),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (0, 0), 10),
        ("RIGHTPADDING", (0, 0), (0, 0), 10),
        ("TOPPADDING", (0, 0), (0, 0), 10),
        ("BOTTOMPADDING", (0, 0), (0, 0), 10),
        ("LEFTPADDING", (1, 0), (1, 0), 0),
        ("RIGHTPADDING", (1, 0), (1, 0), 0),
        ("TOPPADDING", (1, 0), (1, 0), 0),
        ("BOTTOMPADDING", (1, 0), (1, 0), 0),
    ]))
    inner_header = header_table._cellvalues[0][1]
    inner_header.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F4F8FF")),
        ("BOX", (0, 0), (-1, -1), 1.0, brand_blue),
        ("LEFTPADDING", (0, 0), (-1, -1), 9),
        ("RIGHTPADDING", (0, 0), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 7))

    client_html = f"<b>{escape(text_or_dash(client.name if client else ''))}</b>"
    client_lines = [
        line_if("Contact", getattr(client, "contact_person", "")),
        line_if("Phone", getattr(client, "phone", "")),
        line_if("Email", getattr(client, "email", "")),
        line_if("GSTIN", getattr(client, "gst_number", "")),
        line_if("Address", getattr(client, "address", "")),
    ]
    client_body = "<br/>".join([line for line in client_lines if line])
    if client_body:
        client_html += "<br/>" + client_body

    program_html = (
        f"<b>{escape(text_or_dash(getattr(program, 'program_no', '')))}</b><br/>"
        f"<b>SKU:</b> {escape(text_or_dash(getattr(bom, 'sku', '')))}<br/>"
        f"<b>Product:</b> {escape(text_or_dash(getattr(bom, 'product_name', '')))}<br/>"
        f"<b>Program Date:</b> {escape(program_date)}<br/>"
        f"<b>Finishing Date:</b> {escape(finishing_date)}<br/>"
        f"<b>Total Qty:</b> {escape(fmt_qty(getattr(program, 'total_qty', 0)))}"
    )

    party_table = Table(
        [
            [
                Paragraph("CLIENT DETAILS", section_head_left),
                Paragraph("PROGRAM DETAILS", section_head_left),
            ],
            [
                Paragraph(client_html, section_value_style),
                Paragraph(program_html, section_value_style),
            ],
        ],
        colWidths=[92 * mm, 98 * mm],
    )
    party_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), brand_orange),
        ("BACKGROUND", (1, 0), (1, 0), brand_blue),
        ("TEXTCOLOR", (0, 0), (-1, 0), white),
        ("BACKGROUND", (0, 1), (0, 1), colors.HexColor("#FFF9F3")),
        ("BACKGROUND", (1, 1), (1, 1), colors.HexColor("#F5F9FF")),
        ("BOX", (0, 0), (-1, -1), 0.9, border),
        ("INNERGRID", (0, 0), (-1, -1), 0.65, border),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.append(party_table)
    story.append(Spacer(1, 8))

    challan_rows = [
        [
            Paragraph("Driver Name", table_head_style),
            Paragraph("LR No", table_head_style),
            Paragraph("Transport", table_head_style),
            Paragraph("Vehicle No", table_head_style),
        ],
        [
            Paragraph(escape(text_or_dash(challan.driver_name)), item_style),
            Paragraph(escape(text_or_dash(challan.lr_no)), centered_style),
            Paragraph(escape(text_or_dash(challan.transport_name)), item_style),
            Paragraph(escape(text_or_dash(challan.vehicle_no)), centered_style),
        ],
    ]

    challan_table = Table(challan_rows, colWidths=[52 * mm, 32 * mm, 66 * mm, 40 * mm])
    challan_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), brand_navy),
        ("TEXTCOLOR", (0, 0), (-1, 0), white),
        ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#F8FAFC")),
        ("BOX", (0, 0), (-1, -1), 0.9, border),
        ("INNERGRID", (0, 0), (-1, -1), 0.55, border),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(challan_table)
    story.append(Spacer(1, 10))

    remarks_html = escape(text_or_dash(challan.remarks)).replace("\n", "<br/>")
    remarks_table = Table(
        [
            [Paragraph("REMARKS", section_head_left)],
            [Paragraph(remarks_html, section_value_style)],
        ],
        colWidths=[190 * mm],
    )
    remarks_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), brand_pink),
        ("TEXTCOLOR", (0, 0), (0, 0), white),
        ("BACKGROUND", (0, 1), (0, 1), soft_bg),
        ("BOX", (0, 0), (-1, -1), 0.9, border),
        ("INNERGRID", (0, 0), (-1, -1), 0.6, border),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.append(remarks_table)
    story.append(Spacer(1, 12))

    sign_table = Table(
        [[
            Paragraph("<b>RECEIVER SIGNATURE</b><br/><br/>______________________________", sign_style),
            Paragraph("<b>FOR " + escape(firm_name.upper()) + "</b><br/><br/>______________________________", sign_style),
        ]],
        colWidths=[95 * mm, 95 * mm],
    )
    sign_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    story.append(sign_table)
    story.append(Spacer(1, 10))

    footer_table = Table(
        [[Paragraph("THIS CHALLAN IS COMPUTER GENERATED, HENCE SIGNATURE IS NOT REQUIRED", footer_note_style)]],
        colWidths=[190 * mm],
    )
    footer_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")),
        ("BOX", (0, 0), (-1, -1), 0.9, border),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(footer_table)

    doc.build(story, onFirstPage=draw_branding, onLaterPages=draw_branding)
    buffer.seek(0)

    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    return response

@login_required
def dispatch_print(request, pk: int):
    if not _dispatch_feature_available():
        return _dispatch_feature_unavailable_response(request)

    challan = get_object_or_404(
        DispatchChallan.objects.select_related(
            "program",
            "program__bom",
            "program__firm",
            "client",
            "firm",
        ),
        pk=pk,
        owner=request.user,
    )

    response = _build_dispatch_challan_pdf_response(challan)

    if response.status_code == 200 and response.get("Content-Type", "").startswith("application/pdf"):
        filename = f'{challan.challan_no or "dispatch_challan"}.pdf'
        disposition = "attachment" if request.GET.get("download") == "1" else "inline"
        response["Content-Disposition"] = f'{disposition}; filename="{filename}"'
        response["Cache-Control"] = "no-store"
        response["X-Content-Type-Options"] = "nosniff"
        try:
            response["Content-Length"] = str(len(response.content))
        except Exception:
            pass

    return response


def _dyeing_material_link_list_url(request):
    url = reverse("accounts:dyeing_material_link_list")
    if _is_embed(request):
        url += "?embed=1"
    return url


def _dyeing_material_link_usage_rows(obj):
    rows = []

    for rel in obj._meta.related_objects:
        accessor_name = rel.get_accessor_name()
        if not accessor_name or accessor_name == "details":
            continue

        try:
            related_manager = getattr(obj, accessor_name)
            count = related_manager.count()
        except Exception:
            continue

        if count:
            label = rel.related_model._meta.verbose_name_plural.title()
            rows.append({"label": label, "count": count})

    return rows


def _dyeing_material_link_list_context(request):
    q = (request.GET.get("q") or "").strip()

    qs = (
        DyeingMaterialLink.objects
        .filter(owner=request.user)
        .select_related("vendor", "material_type", "material")
        .prefetch_related("details")
        .order_by("vendor__name", "material__name")
    )

    if q:
        qs = qs.filter(
            Q(vendor__name__icontains=q)
            | Q(material_type__name__icontains=q)
            | Q(material__name__icontains=q)
            | Q(details__dyeing_name__icontains=q)
            | Q(details__dyeing_type__icontains=q)
        ).distinct()

    all_links = DyeingMaterialLink.objects.filter(owner=request.user)
    detail_rows_count = DyeingMaterialLinkDetail.objects.filter(link__owner=request.user).count()

    page_data = _paginate_utility_queryset(request, qs)

    ctx = {
        "links": page_data.pop("object_list"),
        "q": q,
        "stats": {
            "total": all_links.count(),
            "vendors": all_links.values("vendor").distinct().count(),
            "materials": all_links.values("material").distinct().count(),
            "details": detail_rows_count,
        },
    }
    ctx.update(page_data)
    return ctx

def _dyeing_material_link_list_context(request):
    q = (request.GET.get("q") or "").strip()

    qs = (
        DyeingMaterialLink.objects
        .filter(owner=request.user)
        .select_related("vendor", "material_type", "material")
        .prefetch_related("details", "details__finished_material")
        .order_by("vendor__name", "material__name")
    )

    if q:
        qs = qs.filter(
            Q(vendor__name__icontains=q)
            | Q(material_type__name__icontains=q)
            | Q(material__name__icontains=q)
            | Q(details__dyeing_name__icontains=q)
            | Q(details__dyeing_type__icontains=q)
            | Q(details__finished_material__name__icontains=q)
        ).distinct()

    all_links = DyeingMaterialLink.objects.filter(owner=request.user)

    return {
        "links": qs,
        "q": q,
        "stats": {
            "total": all_links.count(),
        },
    }

@login_required
def dyeing_material_link_list(request):
    template = (
        "accounts/dyeing_material_links/list_embed.html"
        if _is_embed(request)
        else "accounts/dyeing_material_links/list.html"
    )
    return render(request, template, _dyeing_material_link_list_context(request))

def _dyeing_material_link_usage_rows(obj):
    rows = []

    detail_count = obj.details.count()
    if detail_count:
        rows.append({"label": "Detail Rows", "count": detail_count})

    try:
        used_in_po_items = DyeingPurchaseOrderItem.objects.filter(
            dyeing_master_detail__link=obj
        ).count()
    except Exception:
        used_in_po_items = 0

    if used_in_po_items:
        rows.append({"label": "Dyeing PO Items", "count": used_in_po_items})

    return rows

@login_required
@require_http_methods(["GET", "POST"])
def dyeing_material_link_create(request):
    link = DyeingMaterialLink(owner=request.user)

    form = DyeingMaterialLinkForm(
        request.POST or None,
        instance=link,
        user=request.user,
    )
    formset = DyeingMaterialLinkDetailFormSet(
        request.POST or None,
        instance=link,
        prefix="details",
        form_kwargs={"user": request.user},
    )

    if request.method == "POST":
        if form.is_valid() and formset.is_valid():
            non_deleted_forms = [
                f for f in formset.forms
                if f.cleaned_data and not f.cleaned_data.get("DELETE", False)
            ]
            if not non_deleted_forms:
                form.add_error(None, "Add at least one dyeing detail row.")
            else:
                obj = form.save(commit=False)
                obj.owner = request.user
                obj.save()

                formset.instance = obj
                formset.save()

                url = _dyeing_material_link_list_url(request)
                if _is_embed(request):
                    return JsonResponse({"ok": True, "url": url})
                return redirect(url)

    template = (
        "accounts/dyeing_material_links/form_embed.html"
        if _is_embed(request)
        else "accounts/dyeing_material_links/form.html"
    )
    return render(request, template, {
        "form": form,
        "formset": formset,
        "mode": "add",
    })


@login_required
@require_http_methods(["GET", "POST"])
def dyeing_material_link_update(request, pk: int):
    obj = get_object_or_404(
        DyeingMaterialLink.objects.prefetch_related("details"),
        pk=pk,
        owner=request.user,
    )

    form = DyeingMaterialLinkForm(
        request.POST or None,
        instance=obj,
        user=request.user,
    )
    formset = DyeingMaterialLinkDetailFormSet(
        request.POST or None,
        instance=obj,
        prefix="details",
        form_kwargs={"user": request.user},
    )

    if request.method == "POST":
        if form.is_valid() and formset.is_valid():
            non_deleted_forms = [
                f for f in formset.forms
                if f.cleaned_data and not f.cleaned_data.get("DELETE", False)
            ]
            if not non_deleted_forms:
                form.add_error(None, "Add at least one dyeing detail row.")
            else:
                form.save()
                formset.save()

                url = _dyeing_material_link_list_url(request)
                if _is_embed(request):
                    return JsonResponse({"ok": True, "url": url})
                return redirect(url)

    template = (
        "accounts/dyeing_material_links/form_embed.html"
        if _is_embed(request)
        else "accounts/dyeing_material_links/form.html"
    )
    return render(request, template, {
        "form": form,
        "formset": formset,
        "mode": "edit",
        "link_obj": obj,
    })


@login_required
@require_POST
def dyeing_material_link_delete(request, pk: int):
    obj = get_object_or_404(DyeingMaterialLink, pk=pk, owner=request.user)
    usage_rows = _dyeing_material_link_usage_rows(obj)

    used_in_po_items = any(row["label"] == "Dyeing PO Items" and row["count"] for row in usage_rows)

    if used_in_po_items:
        error_message = (
            f'Cannot delete dyeing master for "{obj.vendor} / {obj.material}" because it is already used in Dyeing PO items.'
        )

        if _is_embed(request):
            context = _dyeing_material_link_list_context(request)
            context["delete_error"] = error_message
            context["delete_label"] = f"{obj.vendor} / {obj.material}"
            context["delete_usage"] = usage_rows
            return render(request, "accounts/dyeing_material_links/list_embed.html", context)

        messages.error(request, error_message)
        return redirect("accounts:dyeing_material_link_list")

    obj.delete()

    url = reverse("accounts:dyeing_material_link_list")
    if _is_embed(request):
        return JsonResponse({"ok": True, "url": f"{url}?embed=1"})
    return redirect(url)

# ==========================================================
# PHASE 2 - PROGRAM EXECUTION / QC / LOT / QR / COSTING
# ==========================================================
def _sync_dyeing_lots_from_dyeing(owner):
    created = 0

    for item in DyeingPOInwardItem.objects.filter(inward__owner=owner).select_related(
        "inward",
        "po_item",
        "po_item__finished_material",
    ):
        accepted = item.accepted_qty or Decimal("0")
        if accepted <= 0:
            continue

        material = item.po_item.finished_material if item.po_item and item.po_item.finished_material_id else None
        if material is None and item.po_item and item.po_item.fabric_name:
            material = Material.objects.filter(name__iexact=item.po_item.fabric_name).first()
        if material is None:
            continue

        lot_code = item.dye_lot_no or item.batch_no or item.inward.inward_number

        lot, was_created = InventoryLot.objects.get_or_create(
            owner=owner,
            lot_code=lot_code,
            defaults={
                "stage": "dyeing",
                "material": material,
                "unit": item.po_item.unit if item.po_item else "",
                "dyeing_inward_item": item,
                "dye_lot_no": item.dye_lot_no or "",
                "batch_no": item.batch_no or "",
                "shade_reference": item.shade_reference or "",
                "received_qty": item.received_qty or item.quantity or Decimal("0"),
                "accepted_qty": accepted,
                "rejected_qty": item.rejected_qty or Decimal("0"),
                "hold_qty": item.hold_qty or Decimal("0"),
                "qc_status": item.qc_status or "pending",
            },
        )

        if not was_created:
            lot.owner = owner
            lot.stage = "dyeing"
            lot.material = material
            lot.unit = item.po_item.unit if item.po_item else lot.unit
            lot.dyeing_inward_item = item
            lot.ready_inward_item = None
            lot.dye_lot_no = item.dye_lot_no or lot.dye_lot_no
            lot.batch_no = item.batch_no or lot.batch_no
            lot.shade_reference = item.shade_reference or lot.shade_reference
            lot.received_qty = item.received_qty or item.quantity or Decimal("0")
            lot.accepted_qty = accepted
            lot.rejected_qty = item.rejected_qty or Decimal("0")
            lot.hold_qty = item.hold_qty or Decimal("0")
            lot.qc_status = item.qc_status or "pending"
            lot.save()

        if was_created:
            created += 1

    return created

def _sync_phase2_lots_from_dyeing(owner):
    return _sync_dyeing_lots_from_dyeing(owner)

@login_required
def inventory_lot_list(request):
    _sync_dyeing_lots_from_dyeing(request.user)
    q = (request.GET.get("q") or "").strip()
    qs = InventoryLot.objects.filter(owner=request.user).select_related("material").order_by("-id")
    if q:
        qs = qs.filter(Q(lot_code__icontains=q) | Q(material__name__icontains=q) | Q(dye_lot_no__icontains=q) | Q(batch_no__icontains=q))
    return render(request, "accounts/inventory/lot_list.html", {"lots": qs, "q": q})

@login_required
def inventory_lot_detail(request, pk):
    lot = get_object_or_404(
        InventoryLot.objects.filter(owner=request.user)
        .select_related(
            "material",
            "dyeing_inward_item__inward__po",
            "ready_inward_item__inward__po",
        )
        .prefetch_related("rolls", "qr_codes", "quality_checks", "movements"),
        pk=pk,
    )
    return render(
        request,
        "accounts/inventory/lot_detail.html",
        {
            "lot": lot,
            "full_page": not _is_embed(request),
        },
    )

@login_required
@require_http_methods(["GET", "POST"])
def quality_check_create(request):
    selected_lot = None
    lot_id = (request.GET.get("lot") or request.POST.get("lot") or "").strip()
    initial = {}

    if lot_id.isdigit():
        selected_lot = (
            InventoryLot.objects.filter(owner=request.user)
            .select_related("material")
            .filter(pk=int(lot_id))
            .first()
        )
        if selected_lot:
            initial["lot"] = selected_lot
            initial["stage"] = selected_lot.stage

    form = QualityCheckForm(request.POST or None, user=request.user, initial=initial)
    formset_params = QualityCheckParameterFormSet(request.POST or None, prefix="params")
    formset_defects = QualityCheckDefectFormSet(request.POST or None, prefix="defects")

    if request.method == "POST" and form.is_valid() and formset_params.is_valid() and formset_defects.is_valid():
        qc = form.save(commit=False)
        qc.owner = request.user
        if not qc.qc_number:
            qc.qc_number = next_quality_check_number()
        qc.inspected_by = request.user
        qc.save()

        formset_params.instance = qc
        formset_defects.instance = qc
        formset_params.save()
        formset_defects.save()

        if qc.lot_id:
            qc.lot.qc_status = qc.result
            qc.lot.save(update_fields=["qc_status", "available_qty"])

        messages.success(request, f"Quality Check {qc.qc_number} created successfully.")
        url = reverse("accounts:quality_check_detail", args=[qc.pk])
        if _is_embed(request):
            url += "?embed=1"
        return redirect(url)
    
    return render(
        request,
        "accounts/qc/form.html",
        {
            "form": form,
            "formset_params": formset_params,
            "formset_defects": formset_defects,
            "mode": "add",
            "selected_lot": selected_lot,
            "full_page": not _is_embed(request),
        },
    )

@login_required
def quality_check_list(request):
    qs = QualityCheck.objects.filter(owner=request.user).select_related("lot", "roll").order_by("-inspection_date", "-id")
    return render(request, "accounts/qc/list.html", {"checks": qs})

@login_required
def quality_check_detail(request, pk):
    qc = get_object_or_404(
        QualityCheck.objects.filter(owner=request.user).select_related("lot", "roll"),
        pk=pk,
    )
    
    return render(
        request,
        "accounts/qc/detail.html",
        {
            "qc": qc,
            "full_page": not _is_embed(request),
        },
    )

@login_required
@require_http_methods(["GET", "POST"])
def qr_code_create(request):
    selected_lot = None
    lot_id = (request.GET.get("lot") or request.POST.get("lot") or "").strip()
    initial = {}

    if lot_id.isdigit():
        selected_lot = (
            InventoryLot.objects.filter(owner=request.user)
            .select_related("material")
            .filter(pk=int(lot_id))
            .first()
        )
        if selected_lot:
            initial["lot"] = selected_lot
            initial["qr_type"] = "lot"

    form = QRCodeRecordForm(request.POST or None, user=request.user, initial=initial)

    if request.method == "POST" and form.is_valid():
        qr = form.save(commit=False)
        qr.owner = request.user

        if not qr.qr_code:
            qr.qr_code = next_qr_code_number()

        # Always make scan URL automatically
        qr.payload_url = request.build_absolute_uri(
            reverse("accounts:qr_code_scan", args=[qr.qr_code])
        )

        qr.save()

        # Generate PNG image
        qr.generate_qr_image(qr.payload_url)
        qr.save(update_fields=["payload_url", "qr_image"])

        url = reverse("accounts:qr_code_detail", args=[qr.pk])
        if _is_embed(request):
            url += "?embed=1"
        return redirect(url)

    return render(
        request,
        "accounts/qr/form.html",
        {
            "form": form,
            "selected_lot": selected_lot,
            "full_page": not _is_embed(request),
        },
    )

@login_required
def qr_code_scan(request, code):
    qr = get_object_or_404(
        QRCodeRecord.objects.select_related("lot", "roll"),
        qr_code=code,
        owner=request.user,
    )

    if hasattr(qr, "scan_count"):
        qr.scan_count = (qr.scan_count or 0) + 1
        qr.last_scanned_at = timezone.now()
        qr.save(update_fields=["scan_count", "last_scanned_at"])

    if qr.status != "active":
        messages.error(request, "This QR code is inactive.")
        return redirect("accounts:qr_code_detail", pk=qr.pk)

    if qr.roll_id and getattr(qr.roll, "lot_id", None):
        return redirect("accounts:inventory_lot_detail", pk=qr.roll.lot_id)

    if qr.lot_id:
        return redirect("accounts:inventory_lot_detail", pk=qr.lot_id)

    return redirect("accounts:qr_code_detail", pk=qr.pk)

@login_required
def qr_code_detail(request, pk):
    qr = get_object_or_404(
        QRCodeRecord.objects.filter(owner=request.user).select_related("lot", "roll"),
        pk=pk,
    )
    return render(
        request,
        "accounts/qr/detail.html",
        {
            "qr": qr,
            "full_page": not _is_embed(request),
        },
    )

@login_required
def costing_snapshot_list(request):
    qs = CostingSnapshot.objects.filter(owner=request.user).select_related("bom", "program").order_by("-id")
    form = CostingSnapshotForm(user=request.user)
    return render(
        request,
        "accounts/costing/list.html",
        {
            "snapshots": qs,
            "form": form,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def costing_snapshot_create(request):
    form = CostingSnapshotForm(request.POST or None, user=request.user)
    if request.method == "POST":
        if form.is_valid():
            obj = form.save(commit=False)
            obj.owner = request.user
            obj.save()
            messages.success(request, "Costing snapshot saved successfully.")
            return redirect("accounts:costing_snapshot_list")

        qs = CostingSnapshot.objects.filter(owner=request.user).select_related("bom", "program").order_by("-id")
        return render(
            request,
            "accounts/costing/list.html",
            {
                "snapshots": qs,
                "form": form,
                "open_costing_modal": True,
            },
        )

    return render(request, "accounts/costing/form.html", {"form": form})

def _program_start_standard_size_rows(program):
    source_row = program.size_rows.filter(line_name="TP").first()
    if source_row is None:
        source_row = program.size_rows.order_by("sort_order", "id").first()

    if source_row is None:
        return [
            ("XS", Decimal("0")),
            ("S", Decimal("0")),
            ("M", Decimal("0")),
            ("L", Decimal("0")),
            ("XL", Decimal("0")),
            ("XXL", Decimal("0")),
        ]

    return [
        ("XS", source_row.xs_qty or Decimal("0")),
        ("S", source_row.s_qty or Decimal("0")),
        ("M", source_row.m_qty or Decimal("0")),
        ("L", source_row.l_qty or Decimal("0")),
        ("XL", source_row.xl_qty or Decimal("0")),
        ("XXL", source_row.xxl_qty or Decimal("0")),
    ]

def _program_start_size_snapshot(program):
    size_keys = ["xs", "s", "m", "l", "xl", "xxl"]
    summary = {
        "xs": Decimal("0"),
        "s": Decimal("0"),
        "m": Decimal("0"),
        "l": Decimal("0"),
        "xl": Decimal("0"),
        "xxl": Decimal("0"),
        "3xl": Decimal("0"),
        "4xl": Decimal("0"),
        "5xl": Decimal("0"),
        "6xl": Decimal("0"),
        "7xl": Decimal("0"),
    }

    tp_row = program.size_rows.filter(line_name="TP").first()
    source_row = tp_row or program.size_rows.order_by("sort_order", "id").first()

    if source_row:
        summary["xs"] = source_row.xs_qty or Decimal("0")
        summary["s"] = source_row.s_qty or Decimal("0")
        summary["m"] = source_row.m_qty or Decimal("0")
        summary["l"] = source_row.l_qty or Decimal("0")
        summary["xl"] = source_row.xl_qty or Decimal("0")
        summary["xxl"] = source_row.xxl_qty or Decimal("0")

    total_qty = sum(summary[key] for key in ["xs", "s", "m", "l", "xl", "xxl"])
    avg_qty = (total_qty / Decimal("6")) if total_qty else Decimal("0")

    return {
        "sizes": summary,
        "total": total_qty,
        "avg": avg_qty.quantize(Decimal("0.01")),
        "a_avg": avg_qty.quantize(Decimal("0.01")),
        "ratio": program.ratio or "",
    }


def _program_start_fabric_lot_payload(user):
    rows = (
        InventoryLot.objects.filter(owner=user, is_closed=False, available_qty__gt=0)
        .values("material_id", "material__name")
        .annotate(
            lot_count=Count("id"),
            total_available=Sum("available_qty"),
        )
        .order_by("material__name")
    )

    payload = {}
    for row in rows:
        if not row["material_id"]:
            continue
        payload[str(row["material_id"])] = {
            "material_name": row["material__name"] or "",
            "lot_count": str(row["lot_count"] or 0),
            "available_qty": str(row["total_available"] or Decimal("0")),
        }
    return payload

def _sync_program_start_seed_rows(program, start_record):
    """
    Keep Program Start rows in sync with current Program/BOM
    until the program is actually started.

    Once started, preserve start snapshot.
    """
    if start_record.is_started:
        return

    # Rebuild fabric rows from BOM material rows
    start_record.fabric_rows.all().delete()
    fabric_seed_rows = []
    if program.bom_id and hasattr(program.bom, "material_items"):
        for idx, row in enumerate(
            program.bom.material_items.select_related("material", "unit").all(),
            start=1,
        ):
            bom_avg = row.avg or Decimal("0")
            fabric_seed_rows.append(
                ProgramStartFabric(
                    start_record=start_record,
                    material=row.material,
                    unit=row.unit,
                    used=bom_avg,
                    avg=bom_avg,
                    available_qty=Decimal("0"),
                    used_qty=Decimal("0"),
                    sort_order=idx,
                )
            )
    if fabric_seed_rows:
        ProgramStartFabric.objects.bulk_create(fabric_seed_rows)

       # Rebuild size rows from Program TP size breakup
    start_record.size_rows.all().delete()
    size_seed_rows = []

    for idx, (size_name, qty) in enumerate(_program_start_standard_size_rows(program), start=1):
        size_seed_rows.append(
            ProgramStartSize(
                start_record=start_record,
                size_name=size_name,
                qty=qty,
                sort_order=idx,
            )
        )

    if size_seed_rows:
        ProgramStartSize.objects.bulk_create(size_seed_rows)

    # Rebuild jobber rows from Program jobber rows
    start_record.jobber_rows.all().delete()
    jobber_seed_rows = []
    program_jobber_rows = list(program.jobber_rows.all().order_by("sort_order", "id"))

    if program_jobber_rows:
        for idx, row in enumerate(program_jobber_rows, start=1):
            jobber_seed_rows.append(
                ProgramStartJobber(
                    start_record=start_record,
                    jobber=row.jobber,
                    jobber_type=row.jobber_type,
                    jobber_price=row.price or Decimal("0"),
                    allocation_date=timezone.localdate(),
                    sort_order=idx,
                )
            )
    elif program.bom_id:
        # fallback: if program jobber rows are empty, seed from BOM mapping
        bom_rows = list(
            program.bom.jobber_details.select_related("jobber", "jobber_type").all().order_by("sort_order", "id")
        )
        for idx, row in enumerate(bom_rows, start=1):
            jobber_seed_rows.append(
                ProgramStartJobber(
                    start_record=start_record,
                    jobber=row.jobber,
                    jobber_type=row.jobber_type,
                    jobber_price=row.price or Decimal("0"),
                    allocation_date=timezone.localdate(),
                    sort_order=idx,
                )
            )

    if jobber_seed_rows:
        ProgramStartJobber.objects.bulk_create(jobber_seed_rows)

def _program_start_jobber_tracker_rows(program, start_record):
    rows = []

    if not start_record:
        return rows

    total_size_qty = (
        start_record.size_rows.aggregate(total=Sum("qty")).get("total")
        or Decimal("0")
    )

    start_jobber_rows = (
        start_record.jobber_rows.select_related("jobber", "jobber_type")
        .prefetch_related("challans")
        .order_by("sort_order", "id")
    )

    for row in start_jobber_rows:
        challans = ProgramJobberChallan.objects.filter(
            owner=program.owner,
            program=program,
            start_jobber=row,
        )
        active_challans = challans.exclude(status="rejected")

        challan_count = challans.count()
        active_challan_count = active_challans.count()
        last_challan = challans.order_by("-challan_date", "-id").first()
        last_challan_date = last_challan.challan_date if last_challan else None
        last_challan_no = last_challan.challan_no if last_challan else ""

        issued_qty = active_challans.aggregate(total=Sum("total_issued_qty")).get("total") or Decimal("0")
        inward_qty = active_challans.aggregate(total=Sum("inward_qty")).get("total") or Decimal("0")

        pending_qty = issued_qty - inward_qty
        if pending_qty < Decimal("0"):
            pending_qty = Decimal("0")

        assigned_qty = row.assigned_qty or Decimal("0")
        if assigned_qty <= 0:
            assigned_qty = total_size_qty

        assigned_balance = assigned_qty - issued_qty
        if assigned_balance < Decimal("0"):
            assigned_balance = Decimal("0")

        jobber_price = row.jobber_price or Decimal("0")
        total_value = assigned_qty * jobber_price

        latest_pending_challan = (
            challans.filter(status="pending")
            .order_by("-challan_date", "-id")
            .first()
        )

        latest_rejected_challan = (
            challans.filter(status="rejected")
            .order_by("-updated_at", "-challan_date", "-id")
            .first()
        )

        latest_approved_challan = (
            challans.filter(status__in=["approved", "partial"])
            .order_by("-challan_date", "-id")
            .first()
        )

        latest_open_approved_challan = None
        if latest_approved_challan:
            latest_balance = (
                (latest_approved_challan.total_issued_qty or Decimal("0"))
                - (latest_approved_challan.inward_qty or Decimal("0"))
            )
            if latest_balance > Decimal("0"):
                latest_open_approved_challan = latest_approved_challan

        can_generate = (
            start_record
            and getattr(start_record, "is_started", False)
            and not latest_pending_challan
            and not latest_open_approved_challan
            and assigned_balance > Decimal("0")
        )

        if active_challan_count and issued_qty > Decimal("0") and inward_qty >= issued_qty:
            status = "Completed"
            status_class = "success"
        elif latest_open_approved_challan:
            status = "Approved"
            status_class = "success"
        elif latest_pending_challan:
            status = "Pending Approval"
            status_class = "primary"
        elif active_challan_count and issued_qty > Decimal("0"):
            status = "In Progress"
            status_class = "primary"
        elif latest_rejected_challan:
            status = "Rejected - Generate Again"
            status_class = "danger"
        else:
            status = "Pending"
            status_class = "secondary"

        rows.append({
            "id": row.id,
            "jobber_name": row.jobber.name if row.jobber else "-",
            "jobber_type": row.jobber_type.name if row.jobber_type else "-",
            "assigned_qty": assigned_qty,
            "issued_qty": issued_qty,
            "inward_qty": inward_qty,
            "pending_qty": pending_qty,
            "assigned_balance": assigned_balance,
            "challan_count": challan_count,
            "active_challan_count": active_challan_count,
            "last_challan_date": last_challan_date,
            "last_challan_no": last_challan_no,
            "status": status,
            "status_class": status_class,
            "jobber_price": jobber_price,
            "total_value": total_value,
            "allocation_date": row.allocation_date,
            "create_url": (
                reverse("accounts:program_challan_create", args=[program.pk, row.pk])
                if can_generate
                else ""
            ),
            "latest_pending_challan": latest_pending_challan,
            "approve_url": (
                reverse("accounts:program_challan_approve", args=[latest_pending_challan.id])
                if latest_pending_challan
                else ""
            ),
            "latest_approved_challan": latest_open_approved_challan,
            "inward_url": (
                reverse("accounts:program_inward_form", args=[latest_open_approved_challan.id])
                if latest_open_approved_challan
                else ""
            ),
        })

    return rows

@login_required
@require_http_methods(["GET"])
def program_start_modal(request, pk):
    program = get_object_or_404(
        Program.objects.filter(owner=request.user).select_related("bom", "firm"),
        pk=pk,
    )

    start_record, _ = ProgramStart.objects.get_or_create(
        program=program,
        defaults={"owner": request.user},
    )

    if start_record.owner_id is None:
        start_record.owner = request.user
        start_record.save(update_fields=["owner"])

    # Seed rows only once for a fresh start record.
    if (
        not start_record.is_started
        and not start_record.fabric_rows.exists()
        and not start_record.size_rows.exists()
        and not start_record.jobber_rows.exists()
    ):
        _sync_program_start_seed_rows(program, start_record)

    form = ProgramStartForm(instance=start_record)
    fabric_formset = ProgramStartFabricFormSet(
        instance=start_record,
        prefix="fabrics",
        form_kwargs={"user": request.user},
    )
    size_formset = ProgramStartSizeFormSet(
        instance=start_record,
        prefix="sizes",
    )
    jobber_formset = ProgramStartJobberFormSet(
        instance=start_record,
        prefix="jobbers",
        form_kwargs={"user": request.user},
    )

    return render(
        request,
        "accounts/programs/start_program_modal.html",
        {
            "program": program,
            "start_record": start_record,
            "start_form": form,
            "fabric_formset": fabric_formset,
            "size_formset": size_formset,
            "jobber_formset": jobber_formset,
            "size_chart_image_url": program.preview_image_url,
            "size_snapshot": _program_start_size_snapshot(program),
            "fabric_lot_payload": _program_start_fabric_lot_payload(request.user),
            "bom_payload": _program_sku_payloads(request.user).get(str(program.bom_id), {}) if program.bom_id else {},
            "jobber_defaults": _program_jobber_defaults(request.user),
            "bom_jobber_price_map": _program_start_jobber_price_payload(program),
            "jobber_tracker_rows": _program_start_jobber_tracker_rows(program, start_record),
            "program_meta": {
                "program_no": program.program_no,
                "program_date": program.program_date,
                "style_code": program.bom.sku if program.bom else "",
                "style_name": program.bom.product_name if program.bom else "",
                "firm_name": program.firm.firm_name if program.firm else "",
            },
        },
    )

@login_required
@require_POST
def program_start_save(request, pk):
    program = get_object_or_404(
        Program.objects.filter(owner=request.user).select_related("bom", "firm"),
        pk=pk,
    )

    start_record, _ = ProgramStart.objects.get_or_create(
        program=program,
        defaults={"owner": request.user},
    )

    if start_record.owner_id is None:
        start_record.owner = request.user
        start_record.save(update_fields=["owner"])

    form = ProgramStartForm(request.POST, instance=start_record)
    fabric_formset = ProgramStartFabricFormSet(
        request.POST,
        instance=start_record,
        prefix="fabrics",
        form_kwargs={"user": request.user},
    )
    size_formset = ProgramStartSizeFormSet(
        request.POST,
        instance=start_record,
        prefix="sizes",
    )
    jobber_formset = ProgramStartJobberFormSet(
        request.POST,
        instance=start_record,
        prefix="jobbers",
        form_kwargs={"user": request.user},
    )

    is_ajax = request.headers.get("X-Requested-With") == "XMLHttpRequest"

    if form.is_valid() and fabric_formset.is_valid() and size_formset.is_valid() and jobber_formset.is_valid():
        with transaction.atomic():
            start_record = form.save(commit=False)
            start_record.owner = request.user
            start_record.program = program
            start_record.is_started = True
            start_record.started_at = timezone.now()
            start_record.save()

            program.status = "in_progress"
            program.save(update_fields=["status"])

            fabric_formset.instance = start_record
            size_formset.instance = start_record
            jobber_formset.instance = start_record

            fabric_formset.save()
            size_formset.save()
            jobber_formset.instance = start_record
            _apply_program_start_jobber_prices_from_bom(program, jobber_formset)
            jobber_formset.save()

        if is_ajax:
            return JsonResponse({
                "ok": True,
                "message": "Program started successfully.",
                "redirect_url": reverse("accounts:program_list"),
            })

        return redirect("accounts:program_list")

    if is_ajax:
        start_errors = _collect_program_start_errors(
            form,
            fabric_formset,
            size_formset,
            jobber_formset,
        )

        return render(
            request,
            "accounts/programs/start_program_modal.html",
            {
                "program": program,
                "start_record": start_record,
                "start_form": form,
                "fabric_formset": fabric_formset,
                "size_formset": size_formset,
                "jobber_formset": jobber_formset,
                "size_chart_image_url": program.preview_image_url,
                "size_snapshot": _program_start_size_snapshot(program),
                "fabric_lot_payload": _program_start_fabric_lot_payload(request.user),
                "bom_payload": _program_sku_payloads(request.user).get(str(program.bom_id), {}) if program.bom_id else {},
                "jobber_defaults": _program_jobber_defaults(request.user),
                "bom_jobber_price_map": _program_start_jobber_price_payload(program),
                "jobber_tracker_rows": _program_start_jobber_tracker_rows(program, start_record),
                "program_meta": {
                    "program_no": program.program_no,
                    "program_date": program.program_date,
                    "style_code": program.bom.sku if program.bom else "",
                    "style_name": program.bom.product_name if program.bom else "",
                    "firm_name": program.firm.firm_name if program.firm else "",
                },
                "start_errors": _collect_program_start_errors(
                    form,
                    fabric_formset,
                    size_formset,
                    jobber_formset,
                ),
            },
            status=400,
        )

    messages.error(request, "Program did not start. Please fix the errors in the Start Program form.")
    return redirect("accounts:program_edit", pk=program.pk)



def _collect_program_start_errors(form, fabric_formset, size_formset, jobber_formset):
    errors = []

    for field_name, field_errors in form.errors.items():
        label = "Form"
        if field_name != "__all__":
            try:
                label = form.fields[field_name].label or field_name.replace("_", " ").title()
            except Exception:
                label = field_name.replace("_", " ").title()

        for err in field_errors:
            errors.append({
                "section": "Program Start",
                "row": "",
                "field": label,
                "message": str(err),
            })

    for section_name, formset in [
        ("Fabric Rows", fabric_formset),
        ("Size Rows", size_formset),
        ("Jobber Rows", jobber_formset),
    ]:
        for err in formset.non_form_errors():
            errors.append({
                "section": section_name,
                "row": "",
                "field": "Formset",
                "message": str(err),
            })

        for index, row_form in enumerate(formset.forms, start=1):
            if not row_form.errors:
                continue

            for field_name, field_errors in row_form.errors.items():
                label = "Row"
                if field_name != "__all__":
                    try:
                        label = row_form.fields[field_name].label or field_name.replace("_", " ").title()
                    except Exception:
                        label = field_name.replace("_", " ").title()

                for err in field_errors:
                    errors.append({
                        "section": section_name,
                        "row": index,
                        "field": label,
                        "message": str(err),
                    })

    return errors


def _is_program_popup(request):
    return (
        request.GET.get("popup") == "1"
        or request.POST.get("popup") == "1"
        or request.headers.get("X-Requested-With") == "XMLHttpRequest"
    )


def _can_approve_program_jobber_challan(request_or_user):
    """Use ERP permissions so company admins and permitted staff can approve DCs."""
    if hasattr(request_or_user, "erp_actor") or hasattr(request_or_user, "erp_is_company_admin"):
        return has_erp_permission(request_or_user, "program.approve_challan")

    user = request_or_user
    if not user or not getattr(user, "is_authenticated", False):
        return False
    if getattr(user, "is_superuser", False) or getattr(user, "is_staff", False):
        return True

    try:
        profile = user.erp_profile
    except Exception:
        profile = None

    if not profile or not getattr(profile, "is_active", False):
        return False
    if not getattr(user, "is_active", True):
        return False
    if getattr(profile, "company", None) and not getattr(profile.company, "is_active_company", True):
        return False
    if getattr(profile, "is_company_admin", False):
        return True

    role = getattr(profile, "role", None)
    return bool(role and getattr(role, "is_active", False) and "program.approve_challan" in (role.permissions or []))


@login_required
@require_http_methods(["GET"])
def program_challan_manage(request, program_id):
    program = get_object_or_404(
        Program.objects.filter(owner=request.user)
        .select_related("bom", "firm", "start_record")
        .prefetch_related(
            "start_record__jobber_rows__jobber",
            "start_record__jobber_rows__jobber_type",
            "start_record__size_rows",
            "jobber_challans__jobber",
            "jobber_challans__jobber_type",
            "jobber_challans__size_rows",
        ),
        pk=program_id,
    )

    popup_mode = _is_program_popup(request)
    start_record = getattr(program, "start_record", None)
    assigned_jobbers = list(start_record.jobber_rows.all()) if start_record else []

    challans = program.jobber_challans.select_related(
        "jobber",
        "jobber_type",
        "owner",
        "approved_by",
        "start_jobber",
    ).prefetch_related("size_rows").order_by("-id")

    selected_start_jobber = None
    start_jobber_id = (request.GET.get("start_jobber") or "").strip()

    if start_jobber_id and start_record:
        selected_start_jobber = get_object_or_404(
            start_record.jobber_rows.select_related("jobber", "jobber_type"),
            pk=start_jobber_id,
        )
        challans = challans.filter(start_jobber=selected_start_jobber)

    total_issued_qty = challans.aggregate(total=Sum("total_issued_qty")).get("total") or Decimal("0")
    total_inward_qty = challans.aggregate(total=Sum("inward_qty")).get("total") or Decimal("0")
    total_pending_qty = total_issued_qty - total_inward_qty
    if total_pending_qty < Decimal("0"):
        total_pending_qty = Decimal("0")

    challans = list(challans)

    for ch in challans:
        issued = ch.total_issued_qty or Decimal("0")
        inward = ch.inward_qty or Decimal("0")

        if ch.status == "pending":
            ch.display_status = "Pending Approval"
            ch.next_action_label = "Approve DC"
            ch.next_action_url = reverse("accounts:program_challan_approve", args=[ch.id])
            ch.next_action_class = "pcp-link-btn--blue"
        elif ch.status == "rejected":
            ch.display_status = "Rejected"
            ch.next_action_label = ""
            ch.next_action_url = ""
            ch.next_action_class = ""
        elif ch.status in ["approved", "partial"] and inward < issued:
            ch.display_status = "Approved" if inward == Decimal("0") else "Partially Inwarded"
            ch.next_action_label = "Program Inward"
            ch.next_action_url = reverse("accounts:program_inward_form", args=[ch.id])
            ch.next_action_class = "pcp-link-btn--green"
        elif issued > Decimal("0") and inward >= issued:
            ch.display_status = "Completed"
            ch.next_action_label = ""
            ch.next_action_url = ""
            ch.next_action_class = ""
        else:
            ch.display_status = ch.get_status_display() if hasattr(ch, "get_status_display") else ch.status.title()
            ch.next_action_label = ""
            ch.next_action_url = ""
            ch.next_action_class = ""

    total_size_qty = Decimal("0")
    if start_record:
        total_size_qty = (
            start_record.size_rows.aggregate(total=Sum("qty")).get("total")
            or Decimal("0")
        )

    for row in assigned_jobbers:
        related_challans = ProgramJobberChallan.objects.filter(
            owner=request.user,
            program=program,
            start_jobber=row,
        ).order_by("-challan_date", "-id")
        active_challans = related_challans.exclude(status="rejected")

        row.latest_pending_challan = related_challans.filter(status="pending").first()
        row.latest_approved_challan = None
        row.latest_open_challan = row.latest_pending_challan or related_challans.first()

        issued_qty = active_challans.aggregate(total=Sum("total_issued_qty")).get("total") or Decimal("0")
        assigned_qty = row.assigned_qty or Decimal("0")
        if assigned_qty <= 0:
            assigned_qty = total_size_qty
        assigned_balance = assigned_qty - issued_qty
        if assigned_balance < Decimal("0"):
            assigned_balance = Decimal("0")

        latest_approved = related_challans.filter(status__in=["approved", "partial"]).first()
        if latest_approved:
            latest_balance = (
                (latest_approved.total_issued_qty or Decimal("0"))
                - (latest_approved.inward_qty or Decimal("0"))
            )
            if latest_balance > Decimal("0"):
                row.latest_approved_challan = latest_approved
                row.latest_open_challan = latest_approved

        if row.latest_pending_challan:
            row.next_action_label = "Approve DC"
            row.next_action_url = reverse("accounts:program_challan_approve", args=[row.latest_pending_challan.id])
            row.next_action_class = "pcm-link--blue"
        elif row.latest_approved_challan:
            row.next_action_label = "Program Inward"
            row.next_action_url = reverse("accounts:program_inward_form", args=[row.latest_approved_challan.id])
            row.next_action_class = "pcm-link--dark"
        elif start_record and getattr(start_record, "is_started", False) and assigned_balance > Decimal("0"):
            row.next_action_label = "Generate Challan"
            row.next_action_url = reverse("accounts:program_challan_create", args=[program.pk, row.pk])
            row.next_action_class = ""
        else:
            row.next_action_label = "Completed"
            row.next_action_url = ""
            row.next_action_class = "pcm-link--disabled"

    context = {
        "program": program,
        "start_record": start_record,
        "assigned_jobbers": assigned_jobbers,
        "challans": challans,
        "total_challans": len(challans),
        "total_issued_qty": total_issued_qty,
        "total_inward_qty": total_inward_qty,
        "total_pending_qty": total_pending_qty,
        "selected_start_jobber": selected_start_jobber,
        "is_program_level_popup": popup_mode and not selected_start_jobber,
    }

    if popup_mode:
        return render(
            request,
            "accounts/programs/challan_process_popup.html",
            context,
        )

    return render(
        request,
        "accounts/programs/challan_manage.html",
        context,
    )



@login_required
@require_http_methods(["GET", "POST"])
def program_challan_create(request, program_id, start_jobber_id):
    from django.forms import inlineformset_factory
    from .forms_legacy import ProgramJobberChallanSizeForm

    program = get_object_or_404(
        Program.objects.filter(owner=request.user).select_related("bom", "firm", "start_record"),
        pk=program_id,
    )

    popup_mode = _is_program_popup(request)
    start_record = getattr(program, "start_record", None)

    if not start_record or not start_record.is_started:
        message = "Start the program first before generating challan."
        if popup_mode:
            return JsonResponse({"ok": False, "message": message}, status=400)
        messages.error(request, message)
        return redirect("accounts:program_list")

    if not start_record.size_rows.exists():
        _sync_program_start_seed_rows(program, start_record)

    start_jobber = get_object_or_404(
        ProgramStartJobber.objects.select_related("start_record", "jobber", "jobber_type").filter(
            start_record=start_record,
            start_record__program=program,
        ),
        pk=start_jobber_id,
    )

    challan = ProgramJobberChallan(
        owner=request.user,
        program=program,
        start_record=start_record,
        start_jobber=start_jobber,
        jobber=start_jobber.jobber,
        jobber_type=start_jobber.jobber_type,
        firm=program.firm,
        production_sku=program.bom.sku if program.bom else "",
        product_name=program.bom.product_name if program.bom else "",
    )

    size_source_rows = list(start_record.size_rows.all().order_by("sort_order", "id"))

    size_initial = [
        {
            "size_name": row.size_name,
            "issued_qty": row.qty or Decimal("0"),
            "inward_qty": Decimal("0"),
            "sort_order": idx,
        }
        for idx, row in enumerate(size_source_rows, start=1)
    ]

    DynamicSizeFormSet = inlineformset_factory(
        ProgramJobberChallan,
        ProgramJobberChallanSize,
        form=ProgramJobberChallanSizeForm,
        extra=max(len(size_initial), 1),
        can_delete=False,
    )

    program_size_total = (
        start_record.size_rows.aggregate(total=Sum("qty")).get("total")
        or Decimal("0")
    )

    assigned_qty = start_jobber.assigned_qty or Decimal("0")
    if assigned_qty <= 0:
        assigned_qty = program_size_total

    already_issued_qty = (
        ProgramJobberChallan.objects.filter(
            owner=request.user,
            program=program,
            start_jobber=start_jobber,
        )
        .exclude(status="rejected")
        .aggregate(total=Sum("total_issued_qty"))
        .get("total")
        or Decimal("0")
    )

    remaining_qty = assigned_qty - already_issued_qty
    if remaining_qty < Decimal("0"):
        remaining_qty = Decimal("0")

    existing_challans = ProgramJobberChallan.objects.filter(
        owner=request.user,
        program=program,
        start_jobber=start_jobber,
    ).order_by("-challan_date", "-id")

    pending_challan = existing_challans.filter(status="pending").first()
    open_approved_challan = None
    for existing in existing_challans.filter(status__in=["approved", "partial"]):
        existing_balance = (
            (existing.total_issued_qty or Decimal("0"))
            - (existing.inward_qty or Decimal("0"))
        )
        if existing_balance > Decimal("0"):
            open_approved_challan = existing
            break

    blocked_message = ""
    if pending_challan:
        blocked_message = "This jobber already has a challan pending approval. Approve or reject it first."
    elif open_approved_challan:
        blocked_message = "This jobber already has an approved challan waiting for inward."
    elif remaining_qty <= Decimal("0"):
        blocked_message = "No remaining assigned quantity is available for this jobber."

    if blocked_message:
        if popup_mode:
            return JsonResponse({"ok": False, "message": blocked_message}, status=400)
        messages.error(request, blocked_message)
        return redirect("accounts:program_list")

    if request.method == "POST":
        post_data = request.POST.copy()

        if not post_data.get("challan_date"):
            post_data["challan_date"] = timezone.localdate().isoformat()

        form = ProgramJobberChallanForm(
            post_data,
            instance=challan,
            user=request.user,
            program=program,
            start_jobber=start_jobber,
        )

        size_formset = DynamicSizeFormSet(
            post_data,
            instance=challan,
            prefix="sizes",
        )

        if form.is_valid() and size_formset.is_valid():
            try:
                total_issued_qty = validate_program_jobber_challan_size_formset(size_formset)
            except forms.ValidationError as exc:
                size_formset._non_form_errors = size_formset.error_class(exc.messages)
            else:
                if total_issued_qty > remaining_qty:
                    size_formset._non_form_errors = size_formset.error_class([
                        f"Cannot issue more than remaining qty ({remaining_qty:.2f})."
                    ])
                else:
                    with transaction.atomic():
                        challan = form.save(commit=False)
                        challan.owner = request.user
                        challan.program = program
                        challan.start_record = start_record
                        challan.start_jobber = start_jobber
                        challan.jobber = start_jobber.jobber
                        challan.jobber_type = start_jobber.jobber_type
                        challan.firm = program.firm
                        challan.production_sku = program.bom.sku if program.bom else ""
                        challan.product_name = program.bom.product_name if program.bom else ""
                        challan.total_issued_qty = total_issued_qty
                        challan.status = "pending"
                        challan.save()

                        size_formset.instance = challan
                        size_formset.save()
                        challan.refresh_totals(save=True)
                        challan.status = "pending"
                        challan.save(update_fields=["status", "updated_at"])

                    list_url = reverse("accounts:program_list")

                    if popup_mode:
                        return JsonResponse({
                            "ok": True,
                            "message": "Program challan generated successfully. Approve it from the Program List.",
                            "redirect_url": list_url,
                        })

                    messages.success(request, "Program challan generated successfully. Approve it from the Program List.")
                    return redirect(list_url)

    else:
        form = ProgramJobberChallanForm(
            instance=challan,
            user=request.user,
            program=program,
            start_jobber=start_jobber,
        )

        size_formset = DynamicSizeFormSet(
            instance=challan,
            prefix="sizes",
            queryset=ProgramJobberChallanSize.objects.none(),
            initial=size_initial,
        )

    context = {
        "program": program,
        "start_record": start_record,
        "start_jobber": start_jobber,
        "challan": challan,
        "form": form,
        "size_formset": size_formset,
        "mode": "add",
        "popup_mode": popup_mode,
        "assigned_qty": assigned_qty,
        "already_issued_qty": already_issued_qty,
        "remaining_qty": remaining_qty,
        "program_meta": {
            "program_no": program.program_no,
            "program_date": program.program_date,
            "style_code": program.bom.sku if program.bom else "",
            "style_name": program.bom.product_name if program.bom else "",
            "firm_name": program.firm.firm_name if program.firm else "",
            "jobber_name": start_jobber.jobber.name if start_jobber.jobber else "-",
            "jobber_type": start_jobber.jobber_type.name if start_jobber.jobber_type else "-",
        },
    }

    return render(
        request,
        "accounts/programs/challan_form.html",
        context,
        status=400 if request.method == "POST" and popup_mode else 200,
    )


@login_required
@require_http_methods(["GET"])
def program_challan_detail(request, pk):
    challan = get_object_or_404(
        ProgramJobberChallan.objects.filter(owner=request.user)
        .select_related(
            "program",
            "program__bom",
            "firm",
            "jobber",
            "jobber_type",
            "owner",
            "approved_by",
        )
        .prefetch_related("size_rows"),
        pk=pk,
    )

    return render(
        request,
        "accounts/programs/challan_detail.html",
        {
            "challan": challan,
            "program": challan.program,
            "size_rows": challan.size_rows.all(),
        },
    )

@login_required
@require_http_methods(["GET", "POST"])
def program_challan_approve(request, pk):
    challan = get_object_or_404(
        ProgramJobberChallan.objects.select_related(
            "program",
            "program__bom",
            "jobber",
            "jobber_type",
            "start_record",
            "start_jobber",
            "approved_by",
        ).filter(owner=request.user),
        pk=pk,
    )

    popup_mode = _is_program_popup(request)
    program = challan.program

    if request.method == "POST":
        action = (
            request.POST.get("action")
            or request.POST.get("decision")
            or request.POST.get("approve")
            or ""
        ).strip().lower()

        rejection_reason = request.POST.get("rejection_reason", "").strip()

        if challan.status != "pending":
            message = "Only pending challans can be approved or rejected."
            if popup_mode:
                return JsonResponse({"ok": False, "message": message}, status=400)
            messages.error(request, message)
            return redirect("accounts:program_list")

        if action in ["approve", "approved"]:
            if (challan.total_issued_qty or Decimal("0")) <= Decimal("0"):
                message = "Cannot approve a challan with zero issued quantity."
                if popup_mode:
                    return JsonResponse({"ok": False, "message": message}, status=400)
                messages.error(request, message)
                return redirect("accounts:program_list")

            challan.status = "approved"
            challan.approved_by = get_actor(request) or request.user
            challan.approved_at = timezone.now()
            challan.rejection_reason = ""
            challan.save(update_fields=[
                "status",
                "approved_by",
                "approved_at",
                "rejection_reason",
                "updated_at",
            ])

            list_url = reverse("accounts:program_list")

            if popup_mode:
                return JsonResponse({
                    "ok": True,
                    "message": "Challan approved successfully. Program Inward is now available from the Program List.",
                    "redirect_url": list_url,
                })

            messages.success(request, "Challan approved successfully. Program Inward is now available from the Program List.")
            return redirect(list_url)

        if action in ["reject", "rejected"]:
            if not rejection_reason:
                return render(
                    request,
                    "accounts/programs/challan_approve.html",
                    {
                        "challan": challan,
                        "program": program,
                        "popup_mode": popup_mode,
                        "error": "Rejection reason is required.",
                    },
                    status=400,
                )

            challan.status = "rejected"
            challan.rejection_reason = rejection_reason
            challan.approved_by = None
            challan.approved_at = None
            challan.save(update_fields=[
                "status",
                "rejection_reason",
                "approved_by",
                "approved_at",
                "updated_at",
            ])

            list_url = reverse("accounts:program_list")

            if popup_mode:
                return JsonResponse({
                    "ok": True,
                    "message": "Challan rejected.",
                    "redirect_url": list_url,
                })

            messages.success(request, "Challan rejected.")
            return redirect(list_url)

        return render(
            request,
            "accounts/programs/challan_approve.html",
            {
                "challan": challan,
                "program": program,
                "popup_mode": popup_mode,
                "error": "Please choose approve or reject.",
            },
            status=400,
        )

    return render(
        request,
        "accounts/programs/challan_approve.html",
        {
            "challan": challan,
            "program": program,
            "popup_mode": popup_mode,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def program_inward_form(request, challan_id):
    challan = get_object_or_404(
        ProgramJobberChallan.objects.filter(owner=request.user)
        .select_related(
            "program",
            "program__bom",
            "firm",
            "jobber",
            "jobber_type",
            "start_record",
            "start_jobber",
            "last_inward_type",
        )
        .prefetch_related("size_rows"),
        pk=challan_id,
    )

    sibling_challans = ProgramJobberChallan.objects.filter(
        owner=request.user,
        start_jobber=challan.start_jobber,
    ).select_related("jobber", "jobber_type").order_by("-id")

    inward_type_options = InwardType.objects.filter(owner=request.user).order_by("name")
    size_rows = list(challan.size_rows.all())

    non_field_errors = []
    row_errors = {}
    entered_accepted_map = {}
    entered_rejected_map = {}

    selected_inward_type = str(challan.last_inward_type_id or "")
    inward_date_value = (
        challan.last_inward_date.isoformat()
        if challan.last_inward_date
        else timezone.localdate().isoformat()
    )
    inward_time_value = (
        challan.last_inward_time.strftime("%H:%M")
        if getattr(challan, "last_inward_time", None)
        else timezone.localtime().strftime("%H:%M")
    )
    inward_remarks = challan.last_inward_remarks or ""

    if request.method != "POST":
        if challan.status == "rejected":
            non_field_errors.append("Rejected challan cannot receive inward.")
        elif challan.status == "pending":
            non_field_errors.append("Approve the challan before inward.")
        elif challan.total_issued_qty > 0 and challan.inward_qty >= challan.total_issued_qty:
            non_field_errors.append("This challan is already fully inwarded.")

    if request.method == "POST":
        selected_inward_type = (request.POST.get("inward_type") or "").strip()
        inward_date_value = (request.POST.get("inward_date") or "").strip()
        inward_time_value = (request.POST.get("inward_time") or "").strip()
        inward_remarks = (request.POST.get("inward_remarks") or "").strip()

        if challan.status == "rejected":
            non_field_errors.append("Rejected challan cannot receive inward.")
        elif challan.status == "pending":
            non_field_errors.append("Approve the challan before inward.")
        elif challan.total_issued_qty > 0 and challan.inward_qty >= challan.total_issued_qty:
            non_field_errors.append("This challan is already fully inwarded.")

        inward_type = None
        if not selected_inward_type:
            non_field_errors.append("Select inward type.")
        else:
            inward_type = inward_type_options.filter(pk=selected_inward_type).first()
            if inward_type is None:
                non_field_errors.append("Select a valid inward type.")

        inward_date = None
        if not inward_date_value:
            non_field_errors.append("Select inward date.")
        else:
            try:
                inward_date = datetime.strptime(inward_date_value, "%Y-%m-%d").date()
            except ValueError:
                non_field_errors.append("Enter a valid inward date.")

        inward_time = None
        if not inward_time_value:
            non_field_errors.append("Select inward time.")
        else:
            try:
                inward_time = datetime.strptime(inward_time_value, "%H:%M").time()
            except ValueError:
                non_field_errors.append("Enter a valid inward time.")

        row_updates = []
        has_positive_qty = False

        for row in size_rows:
            accepted_field = f"row_{row.id}_accepted"
            rejected_field = f"row_{row.id}_rejected"

            raw_accepted = (request.POST.get(accepted_field) or "").strip()
            raw_rejected = (request.POST.get(rejected_field) or "").strip()

            entered_accepted_map[row.id] = raw_accepted
            entered_rejected_map[row.id] = raw_rejected

            try:
                accepted_qty = Decimal(raw_accepted or "0")
            except InvalidOperation:
                row_errors[row.id] = "Enter a valid accepted quantity."
                continue

            try:
                rejected_qty = Decimal(raw_rejected or "0")
            except InvalidOperation:
                row_errors[row.id] = "Enter a valid rejected quantity."
                continue

            if accepted_qty < 0:
                row_errors[row.id] = "Accepted quantity cannot be negative."
                continue

            if rejected_qty < 0:
                row_errors[row.id] = "Rejected quantity cannot be negative."
                continue

            issued_qty = row.issued_qty or Decimal("0")
            current_inward = row.inward_qty or Decimal("0")
            balance_qty = issued_qty - current_inward
            if balance_qty < 0:
                balance_qty = Decimal("0")

            add_qty = accepted_qty + rejected_qty

            if add_qty > balance_qty:
                row_errors[row.id] = f"Accepted + rejected cannot exceed balance {balance_qty:.2f}."
                continue

            if add_qty > 0:
                has_positive_qty = True

            row_updates.append((row, accepted_qty, rejected_qty, add_qty))

        if not has_positive_qty and not row_errors:
            non_field_errors.append("Enter inward quantity in at least one size.")

        total_add_qty = sum((add_qty for _, _, _, add_qty in row_updates), Decimal("0"))
        challan_balance = (challan.total_issued_qty or Decimal("0")) - (challan.inward_qty or Decimal("0"))
        if challan_balance < 0:
            challan_balance = Decimal("0")

        if total_add_qty > challan_balance:
            non_field_errors.append(
                f"Total inward cannot exceed challan balance {challan_balance:.2f}."
            )

        if not non_field_errors and not row_errors:
            with transaction.atomic():
                for row, accepted_qty, rejected_qty, add_qty in row_updates:
                    if add_qty > 0:
                        row.accepted_qty = (row.accepted_qty or Decimal("0")) + accepted_qty
                        row.rejected_qty = (row.rejected_qty or Decimal("0")) + rejected_qty
                        row.inward_qty = (row.inward_qty or Decimal("0")) + add_qty
                        row.save(update_fields=["accepted_qty", "rejected_qty", "inward_qty"])

                challan.last_inward_type = inward_type
                challan.last_inward_date = inward_date
                challan.last_inward_time = inward_time
                challan.last_inward_remarks = inward_remarks
                challan.save(
                    update_fields=[
                        "last_inward_type",
                        "last_inward_date",
                        "last_inward_time",
                        "last_inward_remarks",
                    ]
                )

                challan.refresh_totals(save=True)
                


            messages.success(request, "Program inward saved successfully.")
            url = reverse("accounts:program_list")
            if _is_program_popup(request):
                return JsonResponse({
                    "ok": True,
                    "message": "Program inward saved successfully.",
                    "redirect_url": url,
                })
            return redirect(url)

    row_payloads = []
    total_balance = Decimal("0")
    total_issued = Decimal("0")

    for row in size_rows:
        issued_qty = row.issued_qty or Decimal("0")
        current_inward = row.inward_qty or Decimal("0")
        balance_qty = issued_qty - current_inward
        if balance_qty < 0:
            balance_qty = Decimal("0")

        total_issued += issued_qty
        total_balance += balance_qty

        row_payloads.append(
            {
                "row": row,
                "balance_qty": balance_qty,
                "entered_accepted": entered_accepted_map.get(row.id, ""),
                "entered_rejected": entered_rejected_map.get(row.id, ""),
                "error": row_errors.get(row.id, ""),
            }
        )

    return render(
        request,
        "accounts/programs/inward_form.html",
        {
            "challan": challan,
            "program": challan.program,
            "row_payloads": row_payloads,
            "sibling_challans": sibling_challans,
            "selected_challan_id": str(challan.id),
            "inward_type_options": inward_type_options,
            "selected_inward_type": selected_inward_type,
            "inward_date_value": inward_date_value,
            "inward_time_value": inward_time_value,
            "inward_remarks": inward_remarks,
            "non_field_errors": non_field_errors,
            "total_balance": total_balance,
            "total_issued": total_issued,
        },
    )


@login_required
@require_http_methods(["GET"])
def program_challan_print(request, pk):
    challan = get_object_or_404(
        ProgramJobberChallan.objects.filter(owner=request.user)
        .select_related("program", "program__bom", "firm", "jobber", "jobber_type")
        .prefetch_related("size_rows"),
        pk=pk,
    )

    return render(
        request,
        "accounts/programs/challan_print.html",
        {
            "challan": challan,
            "program": challan.program,
            "size_rows": challan.size_rows.all(),
        },
    )

def _parse_report_date_value(value):
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def _report_size_sort_key(label):
    text = str(label or "").strip().upper()
    preferred = [
        "XS", "S", "M", "L", "XL", "XXL", "3XL", "4XL", "5XL", "6XL", "7XL",
        "2", "4", "6", "8", "10", "12", "14", "16", "18", "20", "22", "24", "26", "28", "30",
    ]
    if text in preferred:
        return (0, preferred.index(text), text)
    if text.isdigit():
        return (1, int(text), text)
    return (2, text, text)


def _build_jobber_type_wise_report_data(request):
    from_date_value = (request.GET.get("from_date") or "").strip()
    to_date_value = (request.GET.get("to_date") or "").strip()
    jobber_type_value = (request.GET.get("jobber_type") or "").strip()
    inward_type_value = (request.GET.get("inward_type") or "").strip()
    program_value = (request.GET.get("program_no") or "").strip()

    from_date = _parse_report_date_value(from_date_value)
    to_date = _parse_report_date_value(to_date_value)

    challans = (
        ProgramJobberChallan.objects.filter(owner=request.user)
        .select_related(
            "program",
            "jobber",
            "jobber_type",
            "last_inward_type",
        )
        .prefetch_related("size_rows")
        .filter(inward_qty__gt=0)
        .order_by("last_inward_date", "program__program_no", "jobber_type__name", "id")
    )

    if from_date:
        challans = challans.filter(last_inward_date__gte=from_date)
    if to_date:
        challans = challans.filter(last_inward_date__lte=to_date)
    if jobber_type_value:
        challans = challans.filter(jobber_type_id=jobber_type_value)
    if inward_type_value:
        challans = challans.filter(last_inward_type_id=inward_type_value)
    if program_value:
        challans = challans.filter(program__program_no__icontains=program_value)

    challans = list(challans)

    size_names = set()
    for challan in challans:
        for size_row in challan.size_rows.all():
            size_label = (size_row.size_name or "").strip()
            if size_label:
                size_names.add(size_label)

    size_headers = sorted(size_names, key=_report_size_sort_key)

    rows = []
    grand_total = Decimal("0")
    size_totals = {size_name: Decimal("0") for size_name in size_headers}

    for challan in challans:
        size_qty_map = {}
        for size_row in challan.size_rows.all():
            size_label = (size_row.size_name or "").strip()
            qty = size_row.inward_qty or Decimal("0")
            size_qty_map[size_label] = qty

        ordered_size_values = []
        for size_name in size_headers:
            qty = size_qty_map.get(size_name, Decimal("0"))
            ordered_size_values.append(qty)
            size_totals[size_name] += qty

        total_inward = challan.inward_qty or Decimal("0")
        grand_total += total_inward

        rows.append(
            {
                "program_no": challan.program.program_no if challan.program_id else "-",
                "inward_date": challan.last_inward_date,
                "jobber_type": challan.jobber_type.name if challan.jobber_type_id else "-",
                "total_inward": total_inward,
                "inward_type": challan.last_inward_type.name if challan.last_inward_type_id else "-",
                "size_values": ordered_size_values,
                "challan": challan,
            }
        )

    jobber_type_options = JobberType.objects.filter(owner=request.user).order_by("name")
    inward_type_options = InwardType.objects.filter(owner=request.user).order_by("name")

    return {
        "rows": rows,
        "size_headers": size_headers,
        "size_totals": [size_totals[size_name] for size_name in size_headers],
        "grand_total": grand_total,
        "jobber_type_options": jobber_type_options,
        "inward_type_options": inward_type_options,
        "filters": {
            "from_date": from_date_value,
            "to_date": to_date_value,
            "jobber_type": jobber_type_value,
            "inward_type": inward_type_value,
            "program_no": program_value,
        },
    }
def _build_jobber_type_wise_report_context(request):
    def _parse_report_date(value):
        if not value:
            return None
        try:
            return datetime.strptime(value, "%Y-%m-%d").date()
        except ValueError:
            return None

    def _report_size_sort_key(label):
        text = str(label or "").strip().upper()
        preferred = [
            "XS", "S", "M", "L", "XL", "XXL", "3XL", "4XL", "5XL", "6XL", "7XL",
            "2", "4", "6", "8", "10", "12", "14", "16", "18", "20", "22", "24", "26", "28", "30",
        ]
        if text in preferred:
            return (0, preferred.index(text), text)
        if text.isdigit():
            return (1, int(text), text)
        return (2, text, text)

    from_date_value = (request.GET.get("from_date") or "").strip()
    to_date_value = (request.GET.get("to_date") or "").strip()
    jobber_type_value = (request.GET.get("jobber_type") or "").strip()
    inward_type_value = (request.GET.get("inward_type") or "").strip()
    program_value = (request.GET.get("program_no") or "").strip()

    from_date = _parse_report_date(from_date_value)
    to_date = _parse_report_date(to_date_value)

    challans = (
        ProgramJobberChallan.objects.filter(owner=request.user)
        .select_related(
            "program",
            "jobber",
            "jobber_type",
            "last_inward_type",
        )
        .prefetch_related("size_rows")
        .filter(inward_qty__gt=0)
        .order_by("last_inward_date", "program__program_no", "jobber_type__name", "id")
    )

    if from_date:
        challans = challans.filter(last_inward_date__gte=from_date)
    if to_date:
        challans = challans.filter(last_inward_date__lte=to_date)
    if jobber_type_value:
        challans = challans.filter(jobber_type_id=jobber_type_value)
    if inward_type_value:
        challans = challans.filter(last_inward_type_id=inward_type_value)
    if program_value:
        challans = challans.filter(program__program_no__icontains=program_value)

    challans = list(challans)
    
    fixed_size_headers = [
        "XS", "S", "M", "L", "XL", "XXL", "3XL", "4XL", "5XL", "6XL", "7XL",
        "2", "4", "6", "8", "10", "12", "14", "16", "18", "20", "22", "24", "26", "28", "30",
        "31", "32", "33", "34", "35", "36",
    ]

    size_names = set(fixed_size_headers)
    for challan in challans:
        for size_row in challans and challan.size_rows.all():
            size_label = (size_row.size_name or "").strip()
            if size_label:
                size_names.add(size_label)

    size_headers = sorted(size_names, key=_report_size_sort_key)

    rows = []
    grand_total = Decimal("0")
    size_totals = {size_name: Decimal("0") for size_name in size_headers}

    for challan in challans:
        size_qty_map = {}
        for size_row in challan.size_rows.all():
            size_label = (size_row.size_name or "").strip()
            qty = size_row.inward_qty or Decimal("0")
            size_qty_map[size_label] = qty

        ordered_size_values = []
        for size_name in size_headers:
            qty = size_qty_map.get(size_name, Decimal("0"))
            ordered_size_values.append(qty)
            size_totals[size_name] += qty

        total_inward = challan.inward_qty or Decimal("0")
        grand_total += total_inward

        rows.append(
            {
                "program_no": challan.program.program_no if challan.program_id else "-",
                "inward_date": challan.last_inward_date,
                "jobber_type": challan.jobber_type.name if challan.jobber_type_id else "-",
                "total_inward": total_inward,
                "inward_type": challan.last_inward_type.name if challan.last_inward_type_id else "-",
                "size_values": ordered_size_values,
                "challan": challan,
            }
        )

    jobber_type_options = JobberType.objects.filter(owner=request.user).order_by("name")
    inward_type_options = InwardType.objects.filter(owner=request.user).order_by("name")

    selected_jobber_type_name = ""
    selected_inward_type_name = ""

    if jobber_type_value:
        selected_jobber_type = jobber_type_options.filter(pk=jobber_type_value).first()
        selected_jobber_type_name = selected_jobber_type.name if selected_jobber_type else ""

    if inward_type_value:
        selected_inward_type = inward_type_options.filter(pk=inward_type_value).first()
        selected_inward_type_name = selected_inward_type.name if selected_inward_type else ""

    return {
        "rows": rows,
        "size_headers": size_headers,
        "size_totals": [size_totals[size_name] for size_name in size_headers],
        "grand_total": grand_total,
        "jobber_type_options": jobber_type_options,
        "inward_type_options": inward_type_options,
        "selected_jobber_type_name": selected_jobber_type_name,
        "selected_inward_type_name": selected_inward_type_name,
        "filters": {
            "from_date": from_date_value,
            "to_date": to_date_value,
            "jobber_type": jobber_type_value,
            "inward_type": inward_type_value,
            "program_no": program_value,
        },
    }

@login_required
@require_http_methods(["GET"])
def report_jobber_type_wise(request):
    context = _build_jobber_type_wise_report_context(request)
    embed_mode = request.GET.get("embed") == "1"
    context["embed_mode"] = embed_mode

    template_name = (
        "accounts/reports/jobber_type_wise_embed.html"
        if embed_mode
        else "accounts/reports/jobber_type_wise.html"
    )

    return render(request, template_name, context)
    
@login_required
@require_http_methods(["GET"])
def reports_home(request):
    report_cards = [
        {
            "title": "Jobber Type Wise Report",
            "desc": "Size-wise inward report by jobber type.",
            "url": reverse("accounts:report_jobber_type_wise"),
            "excel_url": reverse("accounts:report_jobber_type_wise_excel"),
        },
        {
            "title": "Program Production Report",
            "desc": "Program-wise production summary with cutting, stitching and finishing quantities.",
            "url": reverse("accounts:report_program_production"),
            "excel_url": reverse("accounts:report_program_production_excel"),
        },
        {
            "title": "Ready Fabric PO Details Report",
            "desc": "Ready fabric PO item-wise details with vendor, material, sub type, rate and weight.",
            "url": reverse("accounts:report_ready_fabric_po_details"),
            "excel_url": reverse("accounts:report_ready_fabric_po_details_excel"),
        },
        {
            "title": "Stitching Inwards",
            "desc": "Program-wise stitching inward tracking",
            "url": reverse("accounts:report_stitching_inwards"),
        },
        {
            "title": "Used Lot Details Report",
            "desc": "Program-wise used lot tracking with inward, finished PO and dyeing PO linkage.",
            "url": reverse("accounts:report_used_lot_details"),
            "excel_url": reverse("accounts:report_used_lot_details_excel"),
        },
        {
            "title": "Program Stage Flow Report",
            "desc": "Program-wise cutting, stitching and finishing sent, received and pending quantities.",
            "url": reverse("accounts:report_program_stage_flow"),
            "excel_url": reverse("accounts:report_program_stage_flow_excel"),
        },
        {
            "title": "Ready PO Status Report",
            "desc": "Ready PO ordered, inwarded and pending quantities by PO and vendor.",
            "url": reverse("accounts:report_ready_po_status"),
            "excel_url": reverse("accounts:report_ready_po_status_excel"),
        },
        {
            "title": "Inventory Lot Stock Report",
            "desc": "Current lot stock with source stage, available quantity and QC status.",
            "url": reverse("accounts:report_inventory_lot_stock"),
            "excel_url": reverse("accounts:report_inventory_lot_stock_excel"),
        },
                {
            "title": "Dyeing PO Status Report",
            "desc": "Dyeing PO ordered, inwarded and pending quantities by PO and vendor.",
            "url": reverse("accounts:report_dyeing_po_status"),
            "excel_url": reverse("accounts:report_dyeing_po_status_excel"),
        },
        {
            "title": "Greige PO Status Report",
            "desc": "Greige PO ordered, inwarded and pending quantities by PO and vendor.",
            "url": reverse("accounts:report_greige_po_status"),
            "excel_url": reverse("accounts:report_greige_po_status_excel"),
        },
        {
            "title": "Yarn PO Status Report",
            "desc": "Yarn PO ordered, inwarded and pending quantities by PO and vendor.",
            "url": reverse("accounts:report_yarn_po_status"),
            "excel_url": reverse("accounts:report_yarn_po_status_excel"),
        },
                {
            "title": "Dispatch Pending Report",
            "desc": "Program-wise produced, dispatched and pending dispatch quantities.",
            "url": reverse("accounts:report_dispatch_pending"),
            "excel_url": reverse("accounts:report_dispatch_pending_excel"),
        },
        {
            "title": "Program Jobber Report",
            "desc": "Program-wise jobber challan movement with sent, inward and pending quantities.",
            "url": reverse("accounts:report_program_jobber"),
            "excel_url": reverse("accounts:report_program_jobber_excel"),
        },
        {
            "title": "Used Lot Summary Report",
            "desc": "Lot-wise used quantity summary across programs.",
            "url": reverse("accounts:report_used_lot_summary"),
            "excel_url": reverse("accounts:report_used_lot_summary_excel"),
        },
    ]

    return render(
        request,
        "accounts/reports/index.html",
        {
            "report_cards": report_cards,
        },
    )

def _build_dyeing_po_status_report_context(request):
    q = (request.GET.get("q") or "").strip()

    po_queryset = (
        DyeingPurchaseOrder.objects.filter(owner=request.user)
        .select_related(
            "vendor",
            "firm",
            "source_greige_po",
        )
        .prefetch_related(
            "items",
            "inwards__items",
        )
        .order_by("-po_date", "-id")
    )

    if q:
        po_queryset = po_queryset.filter(
            Q(po_number__icontains=q)
            | Q(system_number__icontains=q)
            | Q(vendor__name__icontains=q)
            | Q(items__fabric_name__icontains=q)
            | Q(items__finished_material__name__icontains=q)
            | Q(items__dyeing_name__icontains=q)
        ).distinct()

    rows = []
    totals = {
        "ordered_qty": Decimal("0"),
        "inward_qty": Decimal("0"),
        "pending_qty": Decimal("0"),
    }

    for po in po_queryset:
        ordered_qty = Decimal(po.total_weight or 0)
        inward_qty = Decimal(po.total_inward_qty or 0)
        pending_qty = ordered_qty - inward_qty if ordered_qty > inward_qty else Decimal("0")

        item_names = []
        for item in po.items.all():
            if getattr(item, "finished_material", None) and getattr(item.finished_material, "name", ""):
                item_names.append(item.finished_material.name)
            elif getattr(item, "fabric_name", ""):
                item_names.append(item.fabric_name)
            elif getattr(item, "dyeing_name", ""):
                item_names.append(item.dyeing_name)

        material_name = ", ".join(dict.fromkeys(item_names)) if item_names else "-"

        rows.append(
            {
                "po_no": po.po_number or po.system_number or "-",
                "po_date": po.po_date,
                "vendor_name": po.vendor.name if getattr(po, "vendor", None) else "-",
                "material_name": material_name,
                "ordered_qty": ordered_qty,
                "inward_qty": inward_qty,
                "pending_qty": pending_qty,
                "approval_status": po.approval_status or "pending",
            }
        )

        totals["ordered_qty"] += ordered_qty
        totals["inward_qty"] += inward_qty
        totals["pending_qty"] += pending_qty

    return {
        "rows": rows,
        "q": q,
        "filters": {"q": q},
        "totals": totals,
    }


@login_required
@require_http_methods(["GET"])
def report_dyeing_po_status(request):
    context = _build_dyeing_po_status_report_context(request)
    embed_mode = request.GET.get("embed") == "1"
    context["embed_mode"] = embed_mode

    template_name = (
        "accounts/reports/dyeing_po_status_embed.html"
        if embed_mode
        else "accounts/reports/dyeing_po_status.html"
    )

    return render(request, template_name, context)


@login_required
@require_http_methods(["GET"])
def report_dyeing_po_status_excel(request):
    context = _build_dyeing_po_status_report_context(request)
    rows = context["rows"]
    totals = context["totals"]
    filters = context["filters"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Dyeing PO Status"

    title_font = Font(bold=True, size=14, color="1F2937")
    header_font = Font(bold=True, color="111827")
    bold_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="E5E7EB")
    total_fill = PatternFill(fill_type="solid", fgColor="EEF2F7")

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    left_align = Alignment(horizontal="left", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    headers = [
        "PO No.",
        "PO Date",
        "Vendor",
        "Material",
        "Ordered Qty",
        "Inward Qty",
        "Pending Qty",
        "Approval Status",
    ]

    total_columns = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    ws["A1"] = "Dyeing PO Status Report"
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align

    ws["A2"] = "Search"
    ws["B2"] = filters.get("q") or "-"
    ws["A2"].font = bold_font

    header_row = 4
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align

    data_row = header_row + 1
    for row in rows:
        values = [
            row["po_no"],
            row["po_date"].strftime("%d-%m-%Y") if row["po_date"] else "-",
            row["vendor_name"],
            row["material_name"],
            float(row["ordered_qty"] or 0),
            float(row["inward_qty"] or 0),
            float(row["pending_qty"] or 0),
            row["approval_status"].title(),
        ]

        for col_num, value in enumerate(values, start=1):
            cell = ws.cell(row=data_row, column=col_num, value=value)
            cell.border = border
            if col_num in (5, 6, 7):
                cell.alignment = right_align
            elif col_num == 2:
                cell.alignment = center_align
            else:
                cell.alignment = left_align

        data_row += 1

    total_row = data_row
    total_values = [
        "Total", "", "", "",
        float(totals["ordered_qty"] or 0),
        float(totals["inward_qty"] or 0),
        float(totals["pending_qty"] or 0),
        "",
    ]

    for col_num, value in enumerate(total_values, start=1):
        cell = ws.cell(row=total_row, column=col_num, value=value)
        cell.font = bold_font
        cell.fill = total_fill
        cell.border = border
        if col_num in (5, 6, 7):
            cell.alignment = right_align
        else:
            cell.alignment = left_align

    widths = {
        "A": 18, "B": 14, "C": 24, "D": 34, "E": 14, "F": 14, "G": 14, "H": 16,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="dyeing_po_status_report.xlsx"'
    wb.save(response)
    return response


def _build_greige_po_status_report_context(request):
    q = (request.GET.get("q") or "").strip()

    po_queryset = (
        GreigePurchaseOrder.objects.filter(owner=request.user)
        .select_related(
            "vendor",
            "firm",
            "source_yarn_po",
        )
        .prefetch_related(
            "items",
            "inwards__items",
        )
        .order_by("-po_date", "-id")
    )

    if q:
        po_queryset = po_queryset.filter(
            Q(po_number__icontains=q)
            | Q(system_number__icontains=q)
            | Q(vendor__name__icontains=q)
            | Q(items__fabric_name__icontains=q)
            | Q(items__yarn_name__icontains=q)
            | Q(items__material__name__icontains=q)
        ).distinct()

    rows = []
    totals = {
        "ordered_qty": Decimal("0"),
        "inward_qty": Decimal("0"),
        "pending_qty": Decimal("0"),
    }

    for po in po_queryset:
        ordered_qty = Decimal(getattr(po, "total_weight", Decimal("0")) or Decimal("0"))
        inward_qty = Decimal(getattr(po, "total_inward_qty", Decimal("0")) or Decimal("0"))
        pending_qty = ordered_qty - inward_qty if ordered_qty > inward_qty else Decimal("0")

        item_names = []
        for item in po.items.all():
            if getattr(item, "material", None) and getattr(item.material, "name", ""):
                item_names.append(item.material.name)
            elif getattr(item, "fabric_name", ""):
                item_names.append(item.fabric_name)
            elif getattr(item, "yarn_name", ""):
                item_names.append(item.yarn_name)

        material_name = ", ".join(dict.fromkeys(item_names)) if item_names else "-"

        rows.append(
            {
                "po_no": po.po_number or po.system_number or "-",
                "po_date": po.po_date,
                "vendor_name": po.vendor.name if getattr(po, "vendor", None) else "-",
                "material_name": material_name,
                "ordered_qty": ordered_qty,
                "inward_qty": inward_qty,
                "pending_qty": pending_qty,
                "approval_status": po.approval_status or "pending",
            }
        )

        totals["ordered_qty"] += ordered_qty
        totals["inward_qty"] += inward_qty
        totals["pending_qty"] += pending_qty

    return {
        "rows": rows,
        "q": q,
        "filters": {"q": q},
        "totals": totals,
    }


@login_required
@require_http_methods(["GET"])
def report_greige_po_status(request):
    context = _build_greige_po_status_report_context(request)
    return render(
        request,
        "accounts/reports/greige_po_status.html",
        context,
    )


@login_required
@require_http_methods(["GET"])
def report_greige_po_status_excel(request):
    context = _build_greige_po_status_report_context(request)
    rows = context["rows"]
    totals = context["totals"]
    filters = context["filters"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Greige PO Status"

    title_font = Font(bold=True, size=14, color="1F2937")
    header_font = Font(bold=True, color="111827")
    bold_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="E5E7EB")
    total_fill = PatternFill(fill_type="solid", fgColor="EEF2F7")

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    left_align = Alignment(horizontal="left", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    headers = [
        "PO No.",
        "PO Date",
        "Vendor",
        "Material",
        "Ordered Qty",
        "Inward Qty",
        "Pending Qty",
        "Approval Status",
    ]

    total_columns = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    ws["A1"] = "Greige PO Status Report"
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align

    ws["A2"] = "Search"
    ws["B2"] = filters.get("q") or "-"
    ws["A2"].font = bold_font

    header_row = 4
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align

    data_row = header_row + 1
    for row in rows:
        values = [
            row["po_no"],
            row["po_date"].strftime("%d-%m-%Y") if row["po_date"] else "-",
            row["vendor_name"],
            row["material_name"],
            float(row["ordered_qty"] or 0),
            float(row["inward_qty"] or 0),
            float(row["pending_qty"] or 0),
            row["approval_status"].title(),
        ]

        for col_num, value in enumerate(values, start=1):
            cell = ws.cell(row=data_row, column=col_num, value=value)
            cell.border = border
            if col_num in (5, 6, 7):
                cell.alignment = right_align
            elif col_num == 2:
                cell.alignment = center_align
            else:
                cell.alignment = left_align

        data_row += 1

    total_row = data_row
    total_values = [
        "Total", "", "", "",
        float(totals["ordered_qty"] or 0),
        float(totals["inward_qty"] or 0),
        float(totals["pending_qty"] or 0),
        "",
    ]

    for col_num, value in enumerate(total_values, start=1):
        cell = ws.cell(row=total_row, column=col_num, value=value)
        cell.font = bold_font
        cell.fill = total_fill
        cell.border = border
        if col_num in (5, 6, 7):
            cell.alignment = right_align
        else:
            cell.alignment = left_align

    widths = {
        "A": 18, "B": 14, "C": 24, "D": 34, "E": 14, "F": 14, "G": 14, "H": 16,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="greige_po_status_report.xlsx"'
    wb.save(response)
    return response


def _build_yarn_po_status_report_context(request):
    q = (request.GET.get("q") or "").strip()

    po_queryset = (
        YarnPurchaseOrder.objects.filter(owner=request.user)
        .select_related(
            "vendor",
            "firm",
        )
        .prefetch_related(
            "items",
            "inwards__items",
        )
        .order_by("-po_date", "-id")
    )

    if q:
        po_queryset = po_queryset.filter(
            Q(po_number__icontains=q)
            | Q(system_number__icontains=q)
            | Q(vendor__name__icontains=q)
            | Q(items__material__name__icontains=q)
            | Q(items__material_name__icontains=q)
        ).distinct()

    rows = []
    totals = {
        "ordered_qty": Decimal("0"),
        "inward_qty": Decimal("0"),
        "pending_qty": Decimal("0"),
    }

    for po in po_queryset:
        ordered_qty = Decimal(getattr(po, "total_weight", None) or 0)
        if not ordered_qty:
            for item in po.items.all():
                ordered_qty += Decimal(getattr(item, "quantity", None) or 0)

        inward_qty = Decimal(getattr(po, "total_inward_qty", None) or 0)
        pending_qty = ordered_qty - inward_qty if ordered_qty > inward_qty else Decimal("0")

        item_names = []
        for item in po.items.all():
            if getattr(item, "material", None) and getattr(item.material, "name", ""):
                item_names.append(item.material.name)
            elif getattr(item, "material_type", None) and getattr(item.material_type, "name", ""):
                item_names.append(item.material_type.name)
            elif getattr(item, "material_name", ""):
                item_names.append(item.material_name)

        material_name = ", ".join(dict.fromkeys(item_names)) if item_names else "-"

        rows.append(
            {
                "po_no": po.po_number or po.system_number or "-",
                "po_date": po.po_date,
                "vendor_name": po.vendor.name if getattr(po, "vendor", None) else "-",
                "material_name": material_name,
                "ordered_qty": ordered_qty,
                "inward_qty": inward_qty,
                "pending_qty": pending_qty,
                "approval_status": po.approval_status or "pending",
            }
        )

        totals["ordered_qty"] += ordered_qty
        totals["inward_qty"] += inward_qty
        totals["pending_qty"] += pending_qty

    return {
        "rows": rows,
        "q": q,
        "filters": {"q": q},
        "totals": totals,
    }


@login_required
@require_http_methods(["GET"])
def report_yarn_po_status(request):
    context = _build_yarn_po_status_report_context(request)
    return render(
        request,
        "accounts/reports/yarn_po_status.html",
        context,
    )


@login_required
@require_http_methods(["GET"])
def report_yarn_po_status_excel(request):
    context = _build_yarn_po_status_report_context(request)
    rows = context["rows"]
    totals = context["totals"]
    filters = context["filters"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Yarn PO Status"

    title_font = Font(bold=True, size=14, color="1F2937")
    header_font = Font(bold=True, color="111827")
    bold_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="E5E7EB")
    total_fill = PatternFill(fill_type="solid", fgColor="EEF2F7")

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    left_align = Alignment(horizontal="left", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    headers = [
        "PO No.",
        "PO Date",
        "Vendor",
        "Material",
        "Ordered Qty",
        "Inward Qty",
        "Pending Qty",
        "Approval Status",
    ]

    total_columns = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    ws["A1"] = "Yarn PO Status Report"
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align

    ws["A2"] = "Search"
    ws["B2"] = filters.get("q") or "-"
    ws["A2"].font = bold_font

    header_row = 4
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align

    data_row = header_row + 1
    for row in rows:
        values = [
            row["po_no"],
            row["po_date"].strftime("%d-%m-%Y") if row["po_date"] else "-",
            row["vendor_name"],
            row["material_name"],
            float(row["ordered_qty"] or 0),
            float(row["inward_qty"] or 0),
            float(row["pending_qty"] or 0),
            row["approval_status"].title(),
        ]

        for col_num, value in enumerate(values, start=1):
            cell = ws.cell(row=data_row, column=col_num, value=value)
            cell.border = border
            if col_num in (5, 6, 7):
                cell.alignment = right_align
            elif col_num == 2:
                cell.alignment = center_align
            else:
                cell.alignment = left_align

        data_row += 1

    total_row = data_row
    total_values = [
        "Total", "", "", "",
        float(totals["ordered_qty"] or 0),
        float(totals["inward_qty"] or 0),
        float(totals["pending_qty"] or 0),
        "",
    ]

    for col_num, value in enumerate(total_values, start=1):
        cell = ws.cell(row=total_row, column=col_num, value=value)
        cell.font = bold_font
        cell.fill = total_fill
        cell.border = border
        if col_num in (5, 6, 7):
            cell.alignment = right_align
        else:
            cell.alignment = left_align

    widths = {
        "A": 18, "B": 14, "C": 24, "D": 34, "E": 14, "F": 14, "G": 14, "H": 16,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="yarn_po_status_report.xlsx"'
    wb.save(response)
    return response
    
    
def _build_used_lot_details_report_context(request):
    q = (request.GET.get("q") or "").strip()

    fabric_rows = (
        ProgramStartFabric.objects.filter(
            start_record__owner=request.user,
        )
        .select_related(
            "start_record",
            "start_record__program",
            "start_record__program__bom",
            "material",
        )
        .order_by("-start_record__started_at", "-id")
    )

    rows = []

    for row in fabric_rows:
        used_qty = row.used_qty or Decimal("0")
        lot_no = (row.lot_no or "").strip()

        if used_qty <= 0:
            continue
        if not lot_no:
            continue

        start_record = row.start_record
        program = getattr(start_record, "program", None)
        bom = getattr(program, "bom", None)

        inventory_lot = (
            InventoryLot.objects.select_related(
                "material",
                "ready_inward_item",
                "ready_inward_item__inward",
                "ready_inward_item__inward__po",
                "ready_inward_item__po_item",
                "ready_inward_item__po_item__po",
                "ready_inward_item__po_item__po__source_dyeing_po",
                "dyeing_inward_item",
                "dyeing_inward_item__inward",
                "dyeing_inward_item__inward__po",
            )
            .filter(owner=request.user, lot_code=lot_no)
            .first()
        )

        inward_date = None
        finished_po_no = "-"
        finished_po_date = None
        dyeing_po_no = "-"
        dyeing_po_date = None

        if inventory_lot:
            if inventory_lot.ready_inward_item_id:
                ready_inward_item = inventory_lot.ready_inward_item
                ready_inward = getattr(ready_inward_item, "inward", None)
                ready_po = getattr(ready_inward, "po", None) or getattr(getattr(ready_inward_item, "po_item", None), "po", None)

                inward_date = getattr(ready_inward, "inward_date", None)

                if ready_po:
                    finished_po_no = ready_po.po_number or ready_po.system_number or "-"
                    finished_po_date = ready_po.po_date

                    source_dyeing_po = getattr(ready_po, "source_dyeing_po", None)
                    if source_dyeing_po:
                        dyeing_po_no = source_dyeing_po.po_number or source_dyeing_po.system_number or "-"
                        dyeing_po_date = source_dyeing_po.po_date

            elif inventory_lot.dyeing_inward_item_id:
                dyeing_inward_item = inventory_lot.dyeing_inward_item
                dyeing_inward = getattr(dyeing_inward_item, "inward", None)
                dyeing_po = getattr(dyeing_inward, "po", None)

                inward_date = getattr(dyeing_inward, "inward_date", None)

                if dyeing_po:
                    dyeing_po_no = dyeing_po.po_number or dyeing_po.system_number or "-"
                    dyeing_po_date = dyeing_po.po_date

        material_name = "-"
        if getattr(row, "material", None) and getattr(row.material, "name", ""):
            material_name = row.material.name
        elif inventory_lot and getattr(inventory_lot, "material", None) and getattr(inventory_lot.material, "name", ""):
            material_name = inventory_lot.material.name

        data = {
            "program_no": program.program_no if program else "-",
            "used_qty": used_qty,
            "lot_used_at": getattr(start_record, "started_at", None),
            "product_sku": bom.sku if bom and getattr(bom, "sku", "") else "-",
            "program_date": getattr(program, "program_date", None),
            "lot_no": lot_no,
            "material_name": material_name,
            "inward_date": inward_date,
            "finished_po_no": finished_po_no,
            "finished_po_date": finished_po_date,
            "dyeing_po_no": dyeing_po_no,
            "dyeing_po_date": dyeing_po_date,
        }

        if q:
            ql = q.lower()
            haystack = " ".join(
                [
                    str(data["program_no"] or ""),
                    str(data["product_sku"] or ""),
                    str(data["lot_no"] or ""),
                    str(data["material_name"] or ""),
                    str(data["finished_po_no"] or ""),
                    str(data["dyeing_po_no"] or ""),
                ]
            ).lower()
            if ql not in haystack:
                continue

        rows.append(data)

    total_used_qty = sum((row["used_qty"] or Decimal("0")) for row in rows) if rows else Decimal("0")

    return {
        "rows": rows,
        "q": q,
        "filters": {"q": q},
        "totals": {
            "used_qty": total_used_qty,
        },
    }


@login_required
def report_used_lot_details(request):
    context = _build_used_lot_details_report_context(request)
    embed_mode = request.GET.get("embed") == "1"
    context["embed_mode"] = embed_mode

    template_name = (
        "accounts/reports/used_lot_details_embed.html"
        if embed_mode
        else "accounts/reports/used_lot_details.html"
    )
    return render(request, template_name, context)


@login_required
@require_http_methods(["GET"])
def report_used_lot_details_excel(request):
    context = _build_used_lot_details_report_context(request)
    rows = context["rows"]
    totals = context["totals"]
    filters = context["filters"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Used Lot Details"

    title_font = Font(bold=True, size=14, color="1F2937")
    header_font = Font(bold=True, color="111827")
    bold_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="E5E7EB")
    total_fill = PatternFill(fill_type="solid", fgColor="EEF2F7")

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    left_align = Alignment(horizontal="left", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    headers = [
        "Program No.",
        "Used Qty",
        "Lot Used At",
        "Product SKU",
        "Program Date",
        "Lot No.",
        "Material / Accessory Name",
        "Inward Date",
        "Finished PO No.",
        "Finished PO Date",
        "Dyeing PO No.",
        "Dyeing PO Date",
    ]

    total_columns = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    ws["A1"] = "Used Lot Details Report"
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align

    ws["A2"] = "Search"
    ws["B2"] = filters.get("q") or "-"
    ws["A2"].font = bold_font

    header_row = 4
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align

    data_row = header_row + 1
    for row in rows:
        values = [
            row["program_no"],
            float(row["used_qty"] or 0),
            row["lot_used_at"].strftime("%d-%m-%Y %H:%M") if row["lot_used_at"] else "-",
            row["product_sku"],
            row["program_date"].strftime("%d-%m-%Y") if row["program_date"] else "-",
            row["lot_no"],
            row["material_name"],
            row["inward_date"].strftime("%d-%m-%Y") if row["inward_date"] else "-",
            row["finished_po_no"],
            row["finished_po_date"].strftime("%d-%m-%Y") if row["finished_po_date"] else "-",
            row["dyeing_po_no"],
            row["dyeing_po_date"].strftime("%d-%m-%Y") if row["dyeing_po_date"] else "-",
        ]

        for col_num, value in enumerate(values, start=1):
            cell = ws.cell(row=data_row, column=col_num, value=value)
            cell.border = border
            if col_num == 2:
                cell.alignment = right_align
            elif col_num in (3, 5, 8, 10, 12):
                cell.alignment = center_align
            else:
                cell.alignment = left_align

        data_row += 1

    total_row = data_row
    for col_num in range(1, total_columns + 1):
        cell = ws.cell(row=total_row, column=col_num)
        cell.border = border
        cell.fill = total_fill
        cell.font = bold_font

    ws.cell(row=total_row, column=1, value="Total")
    ws.cell(row=total_row, column=2, value=float(totals["used_qty"] or 0))
    ws.cell(row=total_row, column=2).alignment = right_align

    widths = {
        "A": 16,
        "B": 12,
        "C": 20,
        "D": 24,
        "E": 14,
        "F": 16,
        "G": 28,
        "H": 14,
        "I": 16,
        "J": 16,
        "K": 16,
        "L": 16,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="used_lot_details_report.xlsx"'
    wb.save(response)
    return response

def _build_program_production_report_context(request):
    q = (request.GET.get("q") or "").strip()

    programs = (
        Program.objects.filter(owner=request.user)
        .select_related(
            "bom",
            "bom__brand",
            "bom__main_category",
            "bom__sub_category",
            "bom__pattern_type",
            "firm",
        )
        .prefetch_related(
            "jobber_challans__jobber_type",
        )
        .order_by("-program_date", "-id")
    )

    if q:
        programs = programs.filter(
            Q(program_no__icontains=q)
            | Q(bom__sku__icontains=q)
            | Q(bom__product_name__icontains=q)
            | Q(bom__brand__name__icontains=q)
            | Q(bom__main_category__name__icontains=q)
            | Q(bom__sub_category__name__icontains=q)
            | Q(bom__pattern_type__name__icontains=q)
        )

    rows = []
    total_production_qty = Decimal("0")
    total_cutting_sent = Decimal("0")
    total_cutting_received = Decimal("0")
    total_stitching_sent = Decimal("0")
    total_stitching_received = Decimal("0")
    total_finishing_sent = Decimal("0")
    total_finishing_received = Decimal("0")

    for program in programs:
        bom = getattr(program, "bom", None)

        cutting_sent = Decimal("0")
        cutting_received = Decimal("0")
        stitching_sent = Decimal("0")
        stitching_received = Decimal("0")
        finishing_sent = Decimal("0")
        finishing_received = Decimal("0")

        challans = program.jobber_challans.all()

        for challan in challans:
            jobber_type_name = ""
            if getattr(challan, "jobber_type", None):
                jobber_type_name = (challan.jobber_type.name or "").strip().lower()

            challan_qty = (
                Decimal(challan.total_qty or 0)
                if getattr(challan, "total_qty", None) is not None
                else Decimal("0")
            )
            inward_qty = (
                Decimal(challan.inward_qty or 0)
                if getattr(challan, "inward_qty", None) is not None
                else Decimal("0")
            )

            if "cut" in jobber_type_name:
                cutting_sent += challan_qty
                cutting_received += inward_qty
            elif "stitch" in jobber_type_name:
                stitching_sent += challan_qty
                stitching_received += inward_qty
            elif "finish" in jobber_type_name:
                finishing_sent += challan_qty
                finishing_received += inward_qty

        production_qty = Decimal("0")
        if getattr(program, "production_qty", None) is not None:
            production_qty = Decimal(program.production_qty or 0)
        elif getattr(program, "qty", None) is not None:
            production_qty = Decimal(program.qty or 0)
        elif getattr(program, "total_qty", None) is not None:
            production_qty = Decimal(program.total_qty or 0)
        elif getattr(program, "order_qty", None) is not None:
            production_qty = Decimal(program.order_qty or 0)

        total_production_qty += production_qty
        total_cutting_sent += cutting_sent
        total_cutting_received += cutting_received
        total_stitching_sent += stitching_sent
        total_stitching_received += stitching_received
        total_finishing_sent += finishing_sent
        total_finishing_received += finishing_received

        rows.append(
            {
                "program_no": program.program_no or "-",
                "program_date": program.program_date,
                "sku_name": bom.sku if bom and getattr(bom, "sku", "") else "-",
                "brand": bom.brand.name if bom and getattr(bom, "brand", None) else "-",
                "gender": bom.gender if bom and getattr(bom, "gender", "") else "-",
                "main_category": bom.main_category.name if bom and getattr(bom, "main_category", None) else "-",
                "sub_category": bom.sub_category.name if bom and getattr(bom, "sub_category", None) else "-",
                "pattern_name": bom.pattern_type.name if bom and getattr(bom, "pattern_type", None) else "-",
                "total_production_qty": production_qty,
                "cutting_sent": cutting_sent,
                "cutting_received": cutting_received,
                "stitching_sent": stitching_sent,
                "stitching_received": stitching_received,
                "finishing_sent": finishing_sent,
                "finishing_received": finishing_received,
            }
        )

    return {
        "rows": rows,
        "q": q,
        "filters": {
            "q": q,
        },
        "totals": {
            "total_production_qty": total_production_qty,
            "cutting_sent": total_cutting_sent,
            "cutting_received": total_cutting_received,
            "stitching_sent": total_stitching_sent,
            "stitching_received": total_stitching_received,
            "finishing_sent": total_finishing_sent,
            "finishing_received": total_finishing_received,
        },
    }


@login_required
@require_http_methods(["GET"])
def report_program_production(request):
    context = _build_program_production_report_context(request)
    embed_mode = request.GET.get("embed") == "1"
    context["embed_mode"] = embed_mode

    template_name = (
        "accounts/reports/program_production_embed.html"
        if embed_mode
        else "accounts/reports/program_production.html"
    )

    return render(request, template_name, context)


@login_required
@require_http_methods(["GET"])
def report_program_production_excel(request):
    context = _build_program_production_report_context(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Program Production"

    title_font = Font(bold=True, size=14, color="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="2563EB")
    total_fill = PatternFill(fill_type="solid", fgColor="EAF2FF")
    thin = Side(style="thin", color="D6DFEA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    left_align = Alignment(horizontal="left", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    rows = context["rows"]
    totals = context["totals"]
    filters = context["filters"]

    headers = [
        "Program No.",
        "Program Date",
        "SKU Name",
        "Brand",
        "Gender",
        "Main Category",
        "Sub Category",
        "Pattern Name",
        "Total Production Qty",
        "Cutting Sent",
        "Cutting Received",
        "Stitching Sent",
        "Stitching Received",
        "Finishing Sent",
        "Finishing Received",
    ]

    total_columns = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    ws["A1"] = "Program Production Report"
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align

    ws["A2"] = "Search"
    ws["B2"] = filters.get("q") or "-"
    ws["A2"].font = bold_font

    header_row = 4
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align

    row_num = header_row + 1
    for row in rows:
        values = [
            row["program_no"],
            row["program_date"].strftime("%d-%m-%Y %H:%M") if row["program_date"] else "",
            row["sku_name"],
            row["brand"],
            row["gender"],
            row["main_category"],
            row["sub_category"],
            row["pattern_name"],
            float(row["total_production_qty"] or 0),
            float(row["cutting_sent"] or 0),
            float(row["cutting_received"] or 0),
            float(row["stitching_sent"] or 0),
            float(row["stitching_received"] or 0),
            float(row["finishing_sent"] or 0),
            float(row["finishing_received"] or 0),
        ]

        for col_num, value in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            if col_num >= 9:
                cell.alignment = right_align
            else:
                cell.alignment = left_align
        row_num += 1

    total_row = row_num
    total_values = [
        "Total",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        float(totals["total_production_qty"] or 0),
        float(totals["cutting_sent"] or 0),
        float(totals["cutting_received"] or 0),
        float(totals["stitching_sent"] or 0),
        float(totals["stitching_received"] or 0),
        float(totals["finishing_sent"] or 0),
        float(totals["finishing_received"] or 0),
    ]

    for col_num, value in enumerate(total_values, start=1):
        cell = ws.cell(row=total_row, column=col_num, value=value)
        cell.font = bold_font
        cell.fill = total_fill
        cell.border = border
        if col_num >= 9:
            cell.alignment = right_align
        else:
            cell.alignment = left_align

    widths = {
        1: 14,
        2: 18,
        3: 20,
        4: 16,
        5: 10,
        6: 16,
        7: 16,
        8: 16,
        9: 18,
        10: 14,
        11: 16,
        12: 14,
        13: 16,
        14: 14,
        15: 16,
    }

    for col_num in range(1, total_columns + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = widths.get(col_num, 14)

    ws.freeze_panes = "A5"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="program_production_report.xlsx"'
    return response   
    

def _stage_bucket_from_name(name):
    value = (name or "").strip().lower()
    if "cut" in value:
        return "cutting"
    if "stitch" in value:
        return "stitching"
    if "finish" in value:
        return "finishing"
    return ""


def _build_program_stage_flow_report_context(request):
    q = (request.GET.get("q") or "").strip()

    programs = (
        Program.objects.filter(owner=request.user)
        .select_related(
            "bom",
            "bom__brand",
            "bom__main_category",
            "bom__sub_category",
            "bom__pattern_type",
            "firm",
        )
        .prefetch_related(
            "jobber_challans__jobber_type",
        )
        .order_by("-program_date", "-id")
    )

    if q:
        programs = programs.filter(
            Q(program_no__icontains=q)
            | Q(bom__sku__icontains=q)
            | Q(bom__product_name__icontains=q)
            | Q(bom__brand__name__icontains=q)
            | Q(bom__main_category__name__icontains=q)
            | Q(bom__sub_category__name__icontains=q)
            | Q(bom__pattern_type__name__icontains=q)
        )

    rows = []

    totals = {
        "program_qty": Decimal("0"),
        "cutting_sent": Decimal("0"),
        "cutting_received": Decimal("0"),
        "cutting_pending": Decimal("0"),
        "stitching_sent": Decimal("0"),
        "stitching_received": Decimal("0"),
        "stitching_pending": Decimal("0"),
        "finishing_sent": Decimal("0"),
        "finishing_received": Decimal("0"),
        "finishing_pending": Decimal("0"),
    }

    for program in programs:
        bom = getattr(program, "bom", None)

        program_qty = Decimal(program.total_qty or 0)

        cutting_sent = Decimal("0")
        cutting_received = Decimal("0")
        stitching_sent = Decimal("0")
        stitching_received = Decimal("0")
        finishing_sent = Decimal("0")
        finishing_received = Decimal("0")

        for challan in program.jobber_challans.all():
            stage = _stage_bucket_from_name(
                challan.jobber_type.name if getattr(challan, "jobber_type", None) else ""
            )
            if not stage:
                continue

            issued_qty = Decimal(challan.total_issued_qty or 0)
            inward_qty = Decimal(challan.inward_qty or 0)

            if stage == "cutting":
                cutting_sent += issued_qty
                cutting_received += inward_qty
            elif stage == "stitching":
                stitching_sent += issued_qty
                stitching_received += inward_qty
            elif stage == "finishing":
                finishing_sent += issued_qty
                finishing_received += inward_qty

        cutting_pending = cutting_sent - cutting_received if cutting_sent > cutting_received else Decimal("0")
        stitching_pending = stitching_sent - stitching_received if stitching_sent > stitching_received else Decimal("0")
        finishing_pending = finishing_sent - finishing_received if finishing_sent > finishing_received else Decimal("0")

        rows.append(
            {
                "program_no": program.program_no or "-",
                "program_date": program.program_date,
                "sku_name": bom.sku if bom and getattr(bom, "sku", "") else "-",
                "product_name": bom.product_name if bom and getattr(bom, "product_name", "") else "-",
                "brand": bom.brand.name if bom and getattr(bom, "brand", None) else "-",
                "main_category": bom.main_category.name if bom and getattr(bom, "main_category", None) else "-",
                "sub_category": bom.sub_category.name if bom and getattr(bom, "sub_category", None) else "-",
                "pattern_name": bom.pattern_type.name if bom and getattr(bom, "pattern_type", None) else "-",
                "program_qty": program_qty,
                "cutting_sent": cutting_sent,
                "cutting_received": cutting_received,
                "cutting_pending": cutting_pending,
                "stitching_sent": stitching_sent,
                "stitching_received": stitching_received,
                "stitching_pending": stitching_pending,
                "finishing_sent": finishing_sent,
                "finishing_received": finishing_received,
                "finishing_pending": finishing_pending,
            }
        )

        totals["program_qty"] += program_qty
        totals["cutting_sent"] += cutting_sent
        totals["cutting_received"] += cutting_received
        totals["cutting_pending"] += cutting_pending
        totals["stitching_sent"] += stitching_sent
        totals["stitching_received"] += stitching_received
        totals["stitching_pending"] += stitching_pending
        totals["finishing_sent"] += finishing_sent
        totals["finishing_received"] += finishing_received
        totals["finishing_pending"] += finishing_pending

    return {
        "rows": rows,
        "q": q,
        "filters": {"q": q},
        "totals": totals,
    }


@login_required
@require_http_methods(["GET"])
def report_program_stage_flow(request):
    context = _build_program_stage_flow_report_context(request)
    embed_mode = request.GET.get("embed") == "1"
    context["embed_mode"] = embed_mode

    template_name = (
        "accounts/reports/program_stage_flow_embed.html"
        if embed_mode
        else "accounts/reports/program_stage_flow.html"
    )

    return render(request, template_name, context)


@login_required
@require_http_methods(["GET"])
def report_program_stage_flow_excel(request):
    context = _build_program_stage_flow_report_context(request)
    rows = context["rows"]
    totals = context["totals"]
    filters = context["filters"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Program Stage Flow"

    title_font = Font(bold=True, size=14, color="1F2937")
    header_font = Font(bold=True, color="111827")
    bold_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="E5E7EB")
    total_fill = PatternFill(fill_type="solid", fgColor="EEF2F7")

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    left_align = Alignment(horizontal="left", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    headers = [
        "Program No.",
        "Program Date",
        "SKU Name",
        "Product Name",
        "Brand",
        "Main Category",
        "Sub Category",
        "Pattern Name",
        "Program Qty",
        "Cutting Sent",
        "Cutting Received",
        "Cutting Pending",
        "Stitching Sent",
        "Stitching Received",
        "Stitching Pending",
        "Finishing Sent",
        "Finishing Received",
        "Finishing Pending",
    ]

    total_columns = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    ws["A1"] = "Program Stage Flow Report"
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align

    ws["A2"] = "Search"
    ws["B2"] = filters.get("q") or "-"
    ws["A2"].font = bold_font

    header_row = 4
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align

    data_row = header_row + 1
    for row in rows:
        values = [
            row["program_no"],
            row["program_date"].strftime("%d-%m-%Y") if row["program_date"] else "-",
            row["sku_name"],
            row["product_name"],
            row["brand"],
            row["main_category"],
            row["sub_category"],
            row["pattern_name"],
            float(row["program_qty"] or 0),
            float(row["cutting_sent"] or 0),
            float(row["cutting_received"] or 0),
            float(row["cutting_pending"] or 0),
            float(row["stitching_sent"] or 0),
            float(row["stitching_received"] or 0),
            float(row["stitching_pending"] or 0),
            float(row["finishing_sent"] or 0),
            float(row["finishing_received"] or 0),
            float(row["finishing_pending"] or 0),
        ]

        for col_num, value in enumerate(values, start=1):
            cell = ws.cell(row=data_row, column=col_num, value=value)
            cell.border = border
            if col_num >= 9:
                cell.alignment = right_align
            elif col_num == 2:
                cell.alignment = center_align
            else:
                cell.alignment = left_align

        data_row += 1

    total_row = data_row
    total_values = [
        "Total", "", "", "", "", "", "", "",
        float(totals["program_qty"] or 0),
        float(totals["cutting_sent"] or 0),
        float(totals["cutting_received"] or 0),
        float(totals["cutting_pending"] or 0),
        float(totals["stitching_sent"] or 0),
        float(totals["stitching_received"] or 0),
        float(totals["stitching_pending"] or 0),
        float(totals["finishing_sent"] or 0),
        float(totals["finishing_received"] or 0),
        float(totals["finishing_pending"] or 0),
    ]

    for col_num, value in enumerate(total_values, start=1):
        cell = ws.cell(row=total_row, column=col_num, value=value)
        cell.font = bold_font
        cell.fill = total_fill
        cell.border = border
        if col_num >= 9:
            cell.alignment = right_align
        else:
            cell.alignment = left_align

    widths = {
        "A": 16, "B": 14, "C": 18, "D": 26, "E": 16, "F": 18, "G": 18, "H": 18,
        "I": 12, "J": 12, "K": 14, "L": 14, "M": 13, "N": 15, "O": 15, "P": 13, "Q": 15, "R": 15,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="program_stage_flow_report.xlsx"'
    wb.save(response)
    return response

def _build_dispatch_pending_report_context(request):
    q = (request.GET.get("q") or "").strip()

    programs = (
        Program.objects.filter(owner=request.user)
        .select_related(
            "bom",
            "bom__brand",
            "bom__main_category",
            "bom__sub_category",
            "firm",
        )
        .prefetch_related("dispatches")
        .order_by("-program_date", "-id")
    )

    if q:
        programs = programs.filter(
            Q(program_no__icontains=q)
            | Q(bom__sku__icontains=q)
            | Q(bom__product_name__icontains=q)
            | Q(firm__name__icontains=q)
        )

    rows = []
    totals = {
        "produced_qty": Decimal("0"),
        "dispatched_qty": Decimal("0"),
        "pending_qty": Decimal("0"),
    }

    for program in programs:
        bom = getattr(program, "bom", None)

        produced_qty = Decimal(getattr(program, "total_qty", None) or 0)

        dispatched_qty = Decimal("0")
        last_dispatch_date = None
        for dispatch in program.dispatches.all():
            dispatched_qty += Decimal(dispatch.total_qty or 0)
            if getattr(dispatch, "challan_date", None):
                if not last_dispatch_date or dispatch.challan_date > last_dispatch_date:
                    last_dispatch_date = dispatch.challan_date

        pending_qty = produced_qty - dispatched_qty if produced_qty > dispatched_qty else Decimal("0")

        rows.append(
            {
                "program_no": program.program_no or "-",
                "program_date": program.program_date,
                "sku_name": bom.sku if bom and getattr(bom, "sku", "") else "-",
                "product_name": bom.product_name if bom and getattr(bom, "product_name", "") else "-",
                "client_name": program.firm.name if getattr(program, "firm", None) else "-",
                "produced_qty": produced_qty,
                "dispatched_qty": dispatched_qty,
                "pending_qty": pending_qty,
                "last_dispatch_date": last_dispatch_date,
            }
        )

        totals["produced_qty"] += produced_qty
        totals["dispatched_qty"] += dispatched_qty
        totals["pending_qty"] += pending_qty

    return {
        "rows": rows,
        "q": q,
        "filters": {"q": q},
        "totals": totals,
    }


@login_required
@require_http_methods(["GET"])
def report_dispatch_pending(request):
    context = _build_dispatch_pending_report_context(request)
    return render(
        request,
        "accounts/reports/dispatch_pending.html",
        context,
    )


@login_required
@require_http_methods(["GET"])
def report_dispatch_pending_excel(request):
    context = _build_dispatch_pending_report_context(request)
    rows = context["rows"]
    totals = context["totals"]
    filters = context["filters"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Dispatch Pending"

    title_font = Font(bold=True, size=14, color="1F2937")
    header_font = Font(bold=True, color="111827")
    bold_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="E5E7EB")
    total_fill = PatternFill(fill_type="solid", fgColor="EEF2F7")

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    left_align = Alignment(horizontal="left", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    headers = [
        "Program No.",
        "Program Date",
        "SKU Name",
        "Product Name",
        "Client",
        "Produced Qty",
        "Dispatched Qty",
        "Pending Qty",
        "Last Dispatch Date",
    ]

    total_columns = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    ws["A1"] = "Dispatch Pending Report"
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align

    ws["A2"] = "Search"
    ws["B2"] = filters.get("q") or "-"
    ws["A2"].font = bold_font

    header_row = 4
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align

    data_row = header_row + 1
    for row in rows:
        values = [
            row["program_no"],
            row["program_date"].strftime("%d-%m-%Y") if row["program_date"] else "-",
            row["sku_name"],
            row["product_name"],
            row["client_name"],
            float(row["produced_qty"] or 0),
            float(row["dispatched_qty"] or 0),
            float(row["pending_qty"] or 0),
            row["last_dispatch_date"].strftime("%d-%m-%Y") if row["last_dispatch_date"] else "-",
        ]

        for col_num, value in enumerate(values, start=1):
            cell = ws.cell(row=data_row, column=col_num, value=value)
            cell.border = border
            if col_num in (6, 7, 8):
                cell.alignment = right_align
            elif col_num in (2, 9):
                cell.alignment = center_align
            else:
                cell.alignment = left_align

        data_row += 1

    total_row = data_row
    total_values = [
        "Total", "", "", "", "",
        float(totals["produced_qty"] or 0),
        float(totals["dispatched_qty"] or 0),
        float(totals["pending_qty"] or 0),
        "",
    ]

    for col_num, value in enumerate(total_values, start=1):
        cell = ws.cell(row=total_row, column=col_num, value=value)
        cell.font = bold_font
        cell.fill = total_fill
        cell.border = border
        if col_num in (6, 7, 8):
            cell.alignment = right_align
        else:
            cell.alignment = left_align

    widths = {
        "A": 16, "B": 14, "C": 18, "D": 26, "E": 22, "F": 14, "G": 14, "H": 14, "I": 16,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="dispatch_pending_report.xlsx"'
    wb.save(response)
    return response


def _build_program_jobber_report_context(request):
    q = (request.GET.get("q") or "").strip()

    challans = (
        ProgramJobberChallan.objects.filter(owner=request.user)
        .select_related(
            "program",
            "program__bom",
            "jobber",
            "jobber_type",
        )
        .order_by("-challan_date", "-id")
    )

    if q:
        challans = challans.filter(
            Q(program__program_no__icontains=q)
            | Q(program__bom__sku__icontains=q)
            | Q(jobber__name__icontains=q)
            | Q(jobber_type__name__icontains=q)
            | Q(challan_no__icontains=q)
        )

    rows = []
    totals = {
        "sent_qty": Decimal("0"),
        "inward_qty": Decimal("0"),
        "pending_qty": Decimal("0"),
    }

    for challan in challans:
        program = getattr(challan, "program", None)
        bom = getattr(program, "bom", None)

        sent_qty = Decimal(challan.total_issued_qty or 0)
        inward_qty = Decimal(challan.inward_qty or 0)
        pending_qty = sent_qty - inward_qty if sent_qty > inward_qty else Decimal("0")

        rows.append(
            {
                "challan_no": challan.challan_no or "-",
                "challan_date": challan.challan_date,
                "program_no": program.program_no if program else "-",
                "sku_name": bom.sku if bom and getattr(bom, "sku", "") else "-",
                "jobber_type": challan.jobber_type.name if getattr(challan, "jobber_type", None) else "-",
                "jobber_name": challan.jobber.name if getattr(challan, "jobber", None) else "-",
                "sent_qty": sent_qty,
                "inward_qty": inward_qty,
                "pending_qty": pending_qty,
            }
        )

        totals["sent_qty"] += sent_qty
        totals["inward_qty"] += inward_qty
        totals["pending_qty"] += pending_qty

    return {
        "rows": rows,
        "q": q,
        "filters": {"q": q},
        "totals": totals,
    }


@login_required
@require_http_methods(["GET"])
def report_program_jobber(request):
    context = _build_program_jobber_report_context(request)
    embed_mode = request.GET.get("embed") == "1"
    context["embed_mode"] = embed_mode

    template_name = (
        "accounts/reports/program_jobber_embed.html"
        if embed_mode
        else "accounts/reports/program_jobber.html"
    )

    return render(request, template_name, context)


@login_required
@require_http_methods(["GET"])
def report_program_jobber_excel(request):
    context = _build_program_jobber_report_context(request)
    rows = context["rows"]
    totals = context["totals"]
    filters = context["filters"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Program Jobber"

    title_font = Font(bold=True, size=14, color="1F2937")
    header_font = Font(bold=True, color="111827")
    bold_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="E5E7EB")
    total_fill = PatternFill(fill_type="solid", fgColor="EEF2F7")

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    left_align = Alignment(horizontal="left", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    headers = [
        "Challan No.",
        "Challan Date",
        "Program No.",
        "SKU Name",
        "Jobber Type",
        "Jobber Name",
        "Sent Qty",
        "Inward Qty",
        "Pending Qty",
    ]

    total_columns = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    ws["A1"] = "Program Jobber Report"
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align

    ws["A2"] = "Search"
    ws["B2"] = filters.get("q") or "-"
    ws["A2"].font = bold_font

    header_row = 4
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align

    data_row = header_row + 1
    for row in rows:
        values = [
            row["challan_no"],
            row["challan_date"].strftime("%d-%m-%Y") if row["challan_date"] else "-",
            row["program_no"],
            row["sku_name"],
            row["jobber_type"],
            row["jobber_name"],
            float(row["sent_qty"] or 0),
            float(row["inward_qty"] or 0),
            float(row["pending_qty"] or 0),
        ]

        for col_num, value in enumerate(values, start=1):
            cell = ws.cell(row=data_row, column=col_num, value=value)
            cell.border = border
            if col_num in (7, 8, 9):
                cell.alignment = right_align
            elif col_num == 2:
                cell.alignment = center_align
            else:
                cell.alignment = left_align

        data_row += 1

    total_row = data_row
    total_values = [
        "Total", "", "", "", "", "",
        float(totals["sent_qty"] or 0),
        float(totals["inward_qty"] or 0),
        float(totals["pending_qty"] or 0),
    ]

    for col_num, value in enumerate(total_values, start=1):
        cell = ws.cell(row=total_row, column=col_num, value=value)
        cell.font = bold_font
        cell.fill = total_fill
        cell.border = border
        if col_num in (7, 8, 9):
            cell.alignment = right_align
        else:
            cell.alignment = left_align

    widths = {
        "A": 16, "B": 14, "C": 16, "D": 18, "E": 18, "F": 22, "G": 12, "H": 12, "I": 12,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="program_jobber_report.xlsx"'
    wb.save(response)
    return response


def _build_used_lot_summary_report_context(request):
    q = (request.GET.get("q") or "").strip()

    fabric_rows = (
        ProgramStartFabric.objects.filter(start_record__owner=request.user)
        .select_related(
            "start_record",
            "start_record__program",
            "start_record__program__bom",
            "material",
        )
        .order_by("-id")
    )

    summary = {}

    for item in fabric_rows:
        lot_no = (item.lot_no or "").strip()
        used_qty = Decimal(item.used_qty or 0)

        if not lot_no or used_qty <= 0:
            continue

        program = getattr(getattr(item, "start_record", None), "program", None)
        bom = getattr(program, "bom", None)

        material_name = "-"
        if getattr(item, "material", None) and getattr(item.material, "name", ""):
            material_name = item.material.name

        key = lot_no
        if key not in summary:
            summary[key] = {
                "lot_no": lot_no,
                "material_name": material_name,
                "used_qty": Decimal("0"),
                "program_count": 0,
                "program_nos": [],
            }

        summary[key]["used_qty"] += used_qty

        program_no = program.program_no if program else "-"
        if program_no not in summary[key]["program_nos"]:
            summary[key]["program_nos"].append(program_no)
            summary[key]["program_count"] += 1

        if summary[key]["material_name"] == "-" and material_name != "-":
            summary[key]["material_name"] = material_name

    rows = []
    total_used_qty = Decimal("0")

    for lot_no, row in summary.items():
        row["program_nos_text"] = ", ".join(row["program_nos"])
        total_used_qty += row["used_qty"]

        if q:
            haystack = " ".join(
                [
                    row["lot_no"],
                    row["material_name"],
                    row["program_nos_text"],
                ]
            ).lower()
            if q.lower() not in haystack:
                continue

        rows.append(row)

    rows = sorted(rows, key=lambda x: x["lot_no"])

    return {
        "rows": rows,
        "q": q,
        "filters": {"q": q},
        "totals": {"used_qty": total_used_qty},
    }


@login_required
@require_http_methods(["GET"])
def report_used_lot_summary(request):
    context = _build_used_lot_summary_report_context(request)
    embed_mode = request.GET.get("embed") == "1"
    context["embed_mode"] = embed_mode

    template_name = (
        "accounts/reports/used_lot_summary_embed.html"
        if embed_mode
        else "accounts/reports/used_lot_summary.html"
    )

    return render(request, template_name, context)


@login_required
@require_http_methods(["GET"])
def report_used_lot_summary_excel(request):
    context = _build_used_lot_summary_report_context(request)
    rows = context["rows"]
    totals = context["totals"]
    filters = context["filters"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Used Lot Summary"

    title_font = Font(bold=True, size=14, color="1F2937")
    header_font = Font(bold=True, color="111827")
    bold_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="E5E7EB")
    total_fill = PatternFill(fill_type="solid", fgColor="EEF2F7")

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    headers = [
        "Lot No.",
        "Material",
        "Used Qty",
        "Program Count",
        "Programs",
    ]

    total_columns = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    ws["A1"] = "Used Lot Summary Report"
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align

    ws["A2"] = "Search"
    ws["B2"] = filters.get("q") or "-"
    ws["A2"].font = bold_font

    header_row = 4
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = left_align

    data_row = header_row + 1
    for row in rows:
        values = [
            row["lot_no"],
            row["material_name"],
            float(row["used_qty"] or 0),
            row["program_count"],
            row["program_nos_text"],
        ]

        for col_num, value in enumerate(values, start=1):
            cell = ws.cell(row=data_row, column=col_num, value=value)
            cell.border = border
            if col_num in (3, 4):
                cell.alignment = right_align
            else:
                cell.alignment = left_align

        data_row += 1

    total_row = data_row
    total_values = ["Total", "", float(totals["used_qty"] or 0), "", ""]

    for col_num, value in enumerate(total_values, start=1):
        cell = ws.cell(row=total_row, column=col_num, value=value)
        cell.font = bold_font
        cell.fill = total_fill
        cell.border = border
        if col_num == 3:
            cell.alignment = right_align
        else:
            cell.alignment = left_align

    widths = {
        "A": 18, "B": 28, "C": 12, "D": 14, "E": 40,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="used_lot_summary_report.xlsx"'
    wb.save(response)
    return response

def _build_ready_po_status_report_context(request):
    q = (request.GET.get("q") or "").strip()

    po_queryset = (
        ReadyPurchaseOrder.objects.filter(owner=request.user)
        .select_related(
            "vendor",
            "firm",
            "source_dyeing_po",
        )
        .prefetch_related(
            "items",
            "inwards__items",
        )
        .order_by("-po_date", "-id")
    )

    if q:
        po_queryset = po_queryset.filter(
            Q(po_number__icontains=q)
            | Q(system_number__icontains=q)
            | Q(vendor__name__icontains=q)
            | Q(items__fabric_name__icontains=q)
            | Q(items__finished_material__name__icontains=q)
        ).distinct()

    rows = []
    totals = {
        "ordered_qty": Decimal("0"),
        "inward_qty": Decimal("0"),
        "pending_qty": Decimal("0"),
    }

    for po in po_queryset:
        ordered_qty = Decimal(po.total_weight or 0)
        inward_qty = Decimal(po.total_inward_qty or 0)
        pending_qty = ordered_qty - inward_qty if ordered_qty > inward_qty else Decimal("0")

        item_names = []
        for item in po.items.all():
            if getattr(item, "finished_material", None) and getattr(item.finished_material, "name", ""):
                item_names.append(item.finished_material.name)
            elif item.fabric_name:
                item_names.append(item.fabric_name)

        material_name = ", ".join(dict.fromkeys(item_names)) if item_names else "-"

        rows.append(
            {
                "po_no": po.po_number or po.system_number or "-",
                "po_date": po.po_date,
                "vendor_name": po.vendor.name if getattr(po, "vendor", None) else "-",
                "material_name": material_name,
                "ordered_qty": ordered_qty,
                "inward_qty": inward_qty,
                "pending_qty": pending_qty,
                "approval_status": po.approval_status or "pending",
            }
        )

        totals["ordered_qty"] += ordered_qty
        totals["inward_qty"] += inward_qty
        totals["pending_qty"] += pending_qty

    return {
        "rows": rows,
        "q": q,
        "filters": {"q": q},
        "totals": totals,
    }


@login_required
@require_http_methods(["GET"])
def report_ready_po_status(request):
    context = _build_ready_po_status_report_context(request)
    embed_mode = request.GET.get("embed") == "1"
    context["embed_mode"] = embed_mode

    template_name = (
        "accounts/reports/ready_po_status_embed.html"
        if embed_mode
        else "accounts/reports/ready_po_status.html"
    )

    return render(request, template_name, context)

@login_required
@require_http_methods(["GET"])
def report_ready_po_status_excel(request):
    context = _build_ready_po_status_report_context(request)
    rows = context["rows"]
    totals = context["totals"]
    filters = context["filters"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Ready PO Status"

    title_font = Font(bold=True, size=14, color="1F2937")
    header_font = Font(bold=True, color="111827")
    bold_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="E5E7EB")
    total_fill = PatternFill(fill_type="solid", fgColor="EEF2F7")

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    left_align = Alignment(horizontal="left", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    headers = [
        "PO No.",
        "PO Date",
        "Vendor",
        "Material",
        "Ordered Qty",
        "Inward Qty",
        "Pending Qty",
        "Approval Status",
    ]

    total_columns = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    ws["A1"] = "Ready PO Status Report"
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align

    ws["A2"] = "Search"
    ws["B2"] = filters.get("q") or "-"
    ws["A2"].font = bold_font

    header_row = 4
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align

    data_row = header_row + 1
    for row in rows:
        values = [
            row["po_no"],
            row["po_date"].strftime("%d-%m-%Y") if row["po_date"] else "-",
            row["vendor_name"],
            row["material_name"],
            float(row["ordered_qty"] or 0),
            float(row["inward_qty"] or 0),
            float(row["pending_qty"] or 0),
            row["approval_status"].title(),
        ]

        for col_num, value in enumerate(values, start=1):
            cell = ws.cell(row=data_row, column=col_num, value=value)
            cell.border = border
            if col_num in (5, 6, 7):
                cell.alignment = right_align
            elif col_num == 2:
                cell.alignment = center_align
            else:
                cell.alignment = left_align

        data_row += 1

    total_row = data_row
    total_values = [
        "Total", "", "", "",
        float(totals["ordered_qty"] or 0),
        float(totals["inward_qty"] or 0),
        float(totals["pending_qty"] or 0),
        "",
    ]

    for col_num, value in enumerate(total_values, start=1):
        cell = ws.cell(row=total_row, column=col_num, value=value)
        cell.font = bold_font
        cell.fill = total_fill
        cell.border = border
        if col_num in (5, 6, 7):
            cell.alignment = right_align
        else:
            cell.alignment = left_align

    widths = {
        "A": 18, "B": 14, "C": 24, "D": 34, "E": 14, "F": 14, "G": 14, "H": 16,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="ready_po_status_report.xlsx"'
    wb.save(response)
    return response


def _build_inventory_lot_stock_report_context(request):
    q = (request.GET.get("q") or "").strip()

    lot_queryset = (
        InventoryLot.objects.filter(owner=request.user)
        .select_related("material")
        .order_by("-id")
    )

    if q:
        lot_queryset = lot_queryset.filter(
            Q(lot_code__icontains=q)
            | Q(material__name__icontains=q)
            | Q(stage__icontains=q)
            | Q(location_name__icontains=q)
            | Q(dye_lot_no__icontains=q)
            | Q(batch_no__icontains=q)
        )

    rows = []
    totals = {
        "received_qty": Decimal("0"),
        "accepted_qty": Decimal("0"),
        "used_qty": Decimal("0"),
        "available_qty": Decimal("0"),
    }

    for lot in lot_queryset:
        source_ref = "-"
        if lot.ready_inward_item_id:
            source_ref = "Ready"
        elif lot.dyeing_inward_item_id:
            source_ref = "Dyeing"
        elif lot.greige_inward_item_id:
            source_ref = "Greige"
        elif lot.yarn_inward_item_id:
            source_ref = "Yarn"
        elif lot.stage:
            source_ref = lot.stage.title()

        rows.append(
            {
                "lot_code": lot.lot_code or "-",
                "stage": lot.stage.title() if lot.stage else "-",
                "source_ref": source_ref,
                "material_name": lot.material.name if getattr(lot, "material", None) else "-",
                "unit": lot.unit or "-",
                "received_qty": Decimal(lot.received_qty or 0),
                "accepted_qty": Decimal(lot.accepted_qty or 0),
                "used_qty": Decimal(lot.used_qty or 0),
                "available_qty": Decimal(lot.available_qty or 0),
                "qc_status": lot.qc_status or "pending",
                "location_name": lot.location_name or "-",
                "dye_lot_no": lot.dye_lot_no or "-",
                "batch_no": lot.batch_no or "-",
            }
        )

        totals["received_qty"] += Decimal(lot.received_qty or 0)
        totals["accepted_qty"] += Decimal(lot.accepted_qty or 0)
        totals["used_qty"] += Decimal(lot.used_qty or 0)
        totals["available_qty"] += Decimal(lot.available_qty or 0)

    return {
        "rows": rows,
        "q": q,
        "filters": {"q": q},
        "totals": totals,
    }


@login_required
@require_http_methods(["GET"])
def report_inventory_lot_stock(request):
    context = _build_inventory_lot_stock_report_context(request)
    embed_mode = request.GET.get("embed") == "1"
    context["embed_mode"] = embed_mode

    template_name = (
        "accounts/reports/inventory_lot_stock_embed.html"
        if embed_mode
        else "accounts/reports/inventory_lot_stock.html"
    )

    return render(request, template_name, context)


@login_required
@require_http_methods(["GET"])
def report_inventory_lot_stock_excel(request):
    context = _build_inventory_lot_stock_report_context(request)
    rows = context["rows"]
    totals = context["totals"]
    filters = context["filters"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Lot Stock"

    title_font = Font(bold=True, size=14, color="1F2937")
    header_font = Font(bold=True, color="111827")
    bold_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="E5E7EB")
    total_fill = PatternFill(fill_type="solid", fgColor="EEF2F7")

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    left_align = Alignment(horizontal="left", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    headers = [
        "Lot Code",
        "Stage",
        "Source",
        "Material",
        "Unit",
        "Received Qty",
        "Accepted Qty",
        "Used Qty",
        "Available Qty",
        "QC Status",
        "Location",
        "Dye Lot No.",
        "Batch No.",
    ]

    total_columns = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    ws["A1"] = "Inventory Lot Stock Report"
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align

    ws["A2"] = "Search"
    ws["B2"] = filters.get("q") or "-"
    ws["A2"].font = bold_font

    header_row = 4
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align

    data_row = header_row + 1
    for row in rows:
        values = [
            row["lot_code"],
            row["stage"],
            row["source_ref"],
            row["material_name"],
            row["unit"],
            float(row["received_qty"] or 0),
            float(row["accepted_qty"] or 0),
            float(row["used_qty"] or 0),
            float(row["available_qty"] or 0),
            row["qc_status"].title(),
            row["location_name"],
            row["dye_lot_no"],
            row["batch_no"],
        ]

        for col_num, value in enumerate(values, start=1):
            cell = ws.cell(row=data_row, column=col_num, value=value)
            cell.border = border
            if col_num in (6, 7, 8, 9):
                cell.alignment = right_align
            else:
                cell.alignment = left_align

        data_row += 1

    total_row = data_row
    total_values = [
        "Total", "", "", "", "",
        float(totals["received_qty"] or 0),
        float(totals["accepted_qty"] or 0),
        float(totals["used_qty"] or 0),
        float(totals["available_qty"] or 0),
        "", "", "", "",
    ]

    for col_num, value in enumerate(total_values, start=1):
        cell = ws.cell(row=total_row, column=col_num, value=value)
        cell.font = bold_font
        cell.fill = total_fill
        cell.border = border
        if col_num in (6, 7, 8, 9):
            cell.alignment = right_align
        else:
            cell.alignment = left_align

    widths = {
        "A": 18, "B": 12, "C": 12, "D": 26, "E": 10, "F": 14, "G": 14, "H": 12, "I": 14,
        "J": 14, "K": 18, "L": 18, "M": 16,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="inventory_lot_stock_report.xlsx"'
    wb.save(response)
    return response


def _build_ready_fabric_po_details_report_context(request):
    q = (request.GET.get("q") or "").strip()

    ready_pos = (
        ReadyPurchaseOrder.objects.filter(owner=request.user)
        .select_related(
            "vendor",
            "source_dyeing_po",
            "firm",
        )
        .prefetch_related(
            "items__finished_material__material_sub_type",
            "items__source_dyeing_po_item__finished_material__material_sub_type",
        )
        .order_by("-po_date", "-id")
    )

    if q:
        ready_pos = ready_pos.filter(
            Q(po_number__icontains=q)
            | Q(system_number__icontains=q)
            | Q(vendor__name__icontains=q)
            | Q(items__fabric_name__icontains=q)
            | Q(items__finished_material__name__icontains=q)
            | Q(items__dyeing_name__icontains=q)
        ).distinct()

    rows = []
    total_weight = Decimal("0")

    for po in ready_pos:
        for item in po.items.all():
            source_item = getattr(item, "source_dyeing_po_item", None)

            material_name = "-"
            if getattr(item, "finished_material", None) and getattr(item.finished_material, "name", ""):
                material_name = item.finished_material.name
            elif item.fabric_name:
                material_name = item.fabric_name
            elif source_item and getattr(source_item, "finished_material", None) and getattr(source_item.finished_material, "name", ""):
                material_name = source_item.finished_material.name

            dye_po_sub_type = "-"
            if (
                source_item
                and getattr(source_item, "finished_material", None)
                and getattr(source_item.finished_material, "material_sub_type", None)
                and getattr(source_item.finished_material.material_sub_type, "name", "")
            ):
                dye_po_sub_type = source_item.finished_material.material_sub_type.name
            elif (
                getattr(item, "finished_material", None)
                and getattr(item.finished_material, "material_sub_type", None)
                and getattr(item.finished_material.material_sub_type, "name", "")
            ):
                dye_po_sub_type = item.finished_material.material_sub_type.name
            elif getattr(item, "dyeing_name", ""):
                dye_po_sub_type = item.dyeing_name

            rate = Decimal("0")
            if source_item and getattr(source_item, "rate", None) is not None:
                rate = Decimal(source_item.rate or 0)

            weight = Decimal(item.quantity or 0)
            total_weight += weight

            rows.append(
                {
                    "po_no": po.po_number or po.system_number or "-",
                    "vendor_name": po.vendor.name if getattr(po, "vendor", None) else "-",
                    "material_name": material_name,
                    "dye_po_sub_type": dye_po_sub_type,
                    "rate": rate,
                    "weight": weight,
                    "po_date": po.po_date,
                }
            )

    return {
        "rows": rows,
        "q": q,
        "filters": {
            "q": q,
        },
        "totals": {
            "weight": total_weight,
        },
    }


@login_required
@require_http_methods(["GET"])
def report_ready_fabric_po_details(request):
    context = _build_ready_fabric_po_details_report_context(request)
    embed_mode = request.GET.get("embed") == "1"
    context["embed_mode"] = embed_mode

    template_name = (
        "accounts/reports/ready_fabric_po_details_embed.html"
        if embed_mode
        else "accounts/reports/ready_fabric_po_details.html"
    )

    return render(request, template_name, context)


@login_required
@require_http_methods(["GET"])
def report_ready_fabric_po_details_excel(request):
    context = _build_ready_fabric_po_details_report_context(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ready Fabric PO Details"

    title_font = Font(bold=True, size=14, color="1F2937")
    header_font = Font(bold=True, color="111827")
    bold_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="E5E7EB")
    total_fill = PatternFill(fill_type="solid", fgColor="EEF2F7")

    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    left_align = Alignment(horizontal="left", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    rows = context["rows"]
    totals = context["totals"]
    filters = context["filters"]

    headers = [
        "PO No.",
        "Vendor Name",
        "Material Name",
        "Dye PO Sub Type",
        "Rate",
        "Weight",
        "PO Date",
    ]

    total_columns = len(headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    ws["A1"] = "Ready Fabric PO Details Report"
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align

    ws["A2"] = "Search"
    ws["B2"] = filters.get("q") or "-"
    ws["A2"].font = bold_font

    header_row = 4
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align

    data_row = header_row + 1
    for row in rows:
        values = [
            row["po_no"],
            row["vendor_name"],
            row["material_name"],
            row["dye_po_sub_type"],
            float(row["rate"] or 0),
            float(row["weight"] or 0),
            row["po_date"].strftime("%d-%m-%Y") if row["po_date"] else "-",
        ]

        for col_num, value in enumerate(values, start=1):
            cell = ws.cell(row=data_row, column=col_num, value=value)
            cell.border = border
            if col_num in (5, 6):
                cell.alignment = right_align
            elif col_num == 7:
                cell.alignment = center_align
            else:
                cell.alignment = left_align

        data_row += 1

    total_row = data_row
    for col_num in range(1, total_columns + 1):
        cell = ws.cell(row=total_row, column=col_num)
        cell.border = border
        cell.fill = total_fill
        cell.font = bold_font

    ws.cell(row=total_row, column=1, value="Total")
    ws.cell(row=total_row, column=6, value=float(totals["weight"] or 0))
    ws.cell(row=total_row, column=6).alignment = right_align

    widths = {
        "A": 16,
        "B": 28,
        "C": 24,
        "D": 22,
        "E": 12,
        "F": 12,
        "G": 14,
    }
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="ready_fabric_po_details_report.xlsx"'
    wb.save(response)
    return response

    
@login_required
@require_http_methods(["GET"])
def report_jobber_type_wise_excel(request):
    context = _build_jobber_type_wise_report_context(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Jobber Type Wise"

    title_font = Font(bold=True, size=14, color="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="2563EB")
    total_fill = PatternFill(fill_type="solid", fgColor="EAF2FF")
    thin = Side(style="thin", color="D6DFEA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    left_align = Alignment(horizontal="left", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    size_headers = context["size_headers"]
    rows = context["rows"]
    size_totals = context["size_totals"]
    grand_total = context["grand_total"]
    filters = context["filters"]

    total_columns = 5 + len(size_headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(total_columns, 5))
    ws["A1"] = "Jobber Type Wise Report"
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align

    ws["A2"] = "From Date"
    ws["B2"] = filters.get("from_date") or "-"
    ws["C2"] = "To Date"
    ws["D2"] = filters.get("to_date") or "-"
    ws["E2"] = "Program No"
    ws["F2"] = filters.get("program_no") or "-"

    ws["A3"] = "Jobber Type"
    ws["B3"] = context.get("selected_jobber_type_name") or "All"
    ws["C3"] = "Inward Type"
    ws["D3"] = context.get("selected_inward_type_name") or "All"

    for cell_ref in ["A2", "C2", "E2", "A3", "C3"]:
        ws[cell_ref].font = bold_font

    header_row = 5
    headers = ["Program No", "Inward Date", "Jobber Type", "Total Inward", "Inward Type"] + list(size_headers)

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align

    row_num = header_row + 1
    for row in rows:
        values = [
            row["program_no"],
            row["inward_date"].strftime("%d-%m-%Y") if row["inward_date"] else "",
            row["jobber_type"],
            float(row["total_inward"] or 0),
            row["inward_type"],
        ] + [float(value or 0) for value in row["size_values"]]

        for col_num, value in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            if col_num == 4 or col_num > 5:
                cell.alignment = right_align
            else:
                cell.alignment = left_align
        row_num += 1

    total_row = row_num
    total_values = ["Total", "", "", float(grand_total or 0), ""] + [float(value or 0) for value in size_totals]

    for col_num, value in enumerate(total_values, start=1):
        cell = ws.cell(row=total_row, column=col_num, value=value)
        cell.font = bold_font
        cell.fill = total_fill
        cell.border = border
        if col_num == 4 or col_num > 5:
            cell.alignment = right_align
        else:
            cell.alignment = left_align

    widths = {
        1: 18,
        2: 14,
        3: 18,
        4: 14,
        5: 16,
    }
    for col_num in range(1, total_columns + 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = widths.get(col_num, 10)

    ws.freeze_panes = "A6"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="jobber_type_wise_report.xlsx"'
    return response

@login_required
@require_http_methods(["GET"])
def report_jobber_type_wise_excel(request):
    context = _build_jobber_type_wise_report_context(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Jobber Type Wise"

    title_font = Font(bold=True, size=14, color="1F2937")
    header_font = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="2563EB")
    total_fill = PatternFill(fill_type="solid", fgColor="EAF2FF")
    thin = Side(style="thin", color="D6DFEA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    left_align = Alignment(horizontal="left", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    size_headers = context["size_headers"]
    rows = context["rows"]
    size_totals = context["size_totals"]
    grand_total = context["grand_total"]
    filters = context["filters"]

    total_columns = 5 + len(size_headers)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(total_columns, 5))
    ws["A1"] = "Jobber Type Wise Report"
    ws["A1"].font = title_font
    ws["A1"].alignment = left_align

    ws["A2"] = "From Date"
    ws["B2"] = filters.get("from_date") or "-"
    ws["C2"] = "To Date"
    ws["D2"] = filters.get("to_date") or "-"
    ws["E2"] = "Program No"
    ws["F2"] = filters.get("program_no") or "-"

    ws["A3"] = "Jobber Type"
    ws["B3"] = context.get("selected_jobber_type_name") or "All"
    ws["C3"] = "Inward Type"
    ws["D3"] = context.get("selected_inward_type_name") or "All"

    for cell_ref in ["A2", "C2", "E2", "A3", "C3"]:
        ws[cell_ref].font = bold_font

    header_row = 5
    headers = ["Program No", "Inward Date", "Jobber Type", "Total Inward", "Inward Type"] + list(size_headers)

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_align

    row_num = header_row + 1
    for row in rows:
        values = [
            row["program_no"],
            row["inward_date"].strftime("%d-%m-%Y") if row["inward_date"] else "",
            row["jobber_type"],
            float(row["total_inward"] or 0),
            row["inward_type"],
        ] + [float(value or 0) for value in row["size_values"]]

        for col_num, value in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            if col_num == 4 or col_num > 5:
                cell.alignment = right_align
            else:
                cell.alignment = left_align
        row_num += 1

    total_row = row_num
    total_values = ["Total", "", "", float(grand_total or 0), ""] + [float(value or 0) for value in size_totals]

    for col_num, value in enumerate(total_values, start=1):
        cell = ws.cell(row=total_row, column=col_num, value=value)
        cell.font = bold_font
        cell.fill = total_fill
        cell.border = border
        if col_num == 4 or col_num > 5:
            cell.alignment = right_align
        else:
            cell.alignment = left_align

    widths = {
        1: 18,
        2: 14,
        3: 18,
        4: 14,
        5: 16,
    }
    for col_num in range(1, total_columns + 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = widths.get(col_num, 10)

    ws.freeze_panes = "A6"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="jobber_type_wise_report.xlsx"'
    return response

# =========================================================
# Invoices
# =========================================================

def _invoice_month_range(month_key: str):
    year, month = month_key.split("-")
    year = int(year)
    month = int(month)
    start = datetime(year, month, 1).date()
    if month == 12:
        end = datetime(year + 1, 1, 1).date()
    else:
        end = datetime(year, month + 1, 1).date()
    return start, end


@login_required
def invoice_list(request):
    q = (request.GET.get("q") or "").strip()
    invoices = ProgramInvoice.objects.filter(owner=request.user).select_related("firm", "client", "program", "program__bom").order_by("-invoice_date", "-id")
    if q:
        invoices = invoices.filter(
            Q(invoice_no__icontains=q)
            | Q(program__program_no__icontains=q)
            | Q(program__bom__sku__icontains=q)
            | Q(client__name__icontains=q)
            | Q(firm__firm_name__icontains=q)
        )
    return render(request, "accounts/invoices/list.html", {"invoices": invoices, "q": q})


@login_required
@require_GET
def invoice_program_payload(request, program_id):
    program = get_object_or_404(Program.objects.filter(owner=request.user).select_related("bom", "firm"), pk=program_id)
    challans = list(program.dispatch_challans.all().order_by("-challan_date", "-id")) if hasattr(program, 'dispatch_challans') else []
    latest_snapshot = program.costing_snapshots.order_by("-id").first() if hasattr(program, 'costing_snapshots') else None
    price = Decimal("0")
    if latest_snapshot:
        price = latest_snapshot.target_selling_price or latest_snapshot.total_cost or Decimal("0")
    qty_default = program.total_qty or Decimal("0")
    items = []
    if challans:
        for idx, challan in enumerate(challans, start=1):
            amount = (qty_default * Decimal(price or 0)).quantize(Decimal("0.01"))
            items.append({
                "dispatch_challan_id": challan.id,
                "program": program.program_no or "",
                "sku": getattr(program.bom, "sku", "") or "",
                "challan_no": challan.challan_no or "",
                "quantity": str(qty_default),
                "hsn_code": "",
                "price": str(Decimal(price or 0).quantize(Decimal("0.01"))),
                "amount": str(amount),
                "vehicle_no": challan.vehicle_no or "",
                "remarks": challan.remarks or "",
            })
    else:
        amount = (qty_default * Decimal(price or 0)).quantize(Decimal("0.01"))
        items.append({
            "dispatch_challan_id": "",
            "program": program.program_no or "",
            "sku": getattr(program.bom, "sku", "") or "",
            "challan_no": "",
            "quantity": str(qty_default),
            "hsn_code": "",
            "price": str(Decimal(price or 0).quantize(Decimal("0.01"))),
            "amount": str(amount),
            "vehicle_no": "",
            "remarks": "",
        })
    sub_total = sum(Decimal(item["amount"]) for item in items)
    return JsonResponse({
        "program_no": program.program_no or "",
        "sku": getattr(program.bom, "sku", "") or "",
        "product_name": getattr(program.bom, "product_name", "") or "",
        "firm_id": program.firm_id or "",
        "vehicle_no": items[0]["vehicle_no"] if items else "",
        "remarks": items[0]["remarks"] if items else "",
        "items": items,
        "sub_total": str(sub_total.quantize(Decimal("0.01"))),
    })


@login_required
@require_http_methods(["GET", "POST"])
def invoice_create(request):
    invoice = ProgramInvoice(owner=request.user)
    if request.method == "GET":
        invoice.invoice_no = ProgramInvoice.next_invoice_no(request.user)
        invoice.invoice_date = timezone.localdate()

    form = ProgramInvoiceForm(request.POST or None, instance=invoice, user=request.user)

    if request.method == "POST" and form.is_valid():
        items_json = form.cleaned_data.get("items_json") or "[]"
        try:
            payload_items = json.loads(items_json)
        except Exception:
            payload_items = []

        cleaned_items = []
        for idx, item in enumerate(payload_items, start=1):
            challan_id = item.get("dispatch_challan_id") or None
            quantity = Decimal(str(item.get("quantity") or "0"))
            price = Decimal(str(item.get("price") or "0"))
            if quantity <= 0:
                continue
            amount = (quantity * price).quantize(Decimal("0.01"))
            cleaned_items.append({
                "dispatch_challan_id": challan_id,
                "program_label": item.get("program") or "",
                "sku": item.get("sku") or "",
                "challan_no": item.get("challan_no") or "",
                "quantity": quantity,
                "hsn_code": item.get("hsn_code") or "",
                "price": price,
                "amount": amount,
                "sort_order": idx,
            })

        if not cleaned_items:
            form.add_error(None, "Add at least one invoice row with quantity greater than zero.")
        else:
            with transaction.atomic():
                invoice = form.save(commit=False)
                invoice.owner = request.user
                if invoice.program_id and not invoice.firm_id and getattr(invoice.program, "firm_id", None):
                    invoice.firm = invoice.program.firm
                invoice.save()

                item_objs = []
                for item in cleaned_items:
                    dispatch_challan = None
                    if item["dispatch_challan_id"]:
                        dispatch_challan = DispatchChallan.objects.filter(owner=request.user, pk=item["dispatch_challan_id"]).first()
                    item_objs.append(ProgramInvoiceItem(invoice=invoice, dispatch_challan=dispatch_challan, **{k:v for k,v in item.items() if k != 'dispatch_challan_id'}))
                ProgramInvoiceItem.objects.bulk_create(item_objs)
                invoice.recompute_totals(save=True)

                messages.success(request, f"Invoice {invoice.invoice_no} created successfully.")
                return redirect("accounts:invoice_list")

    return render(request, "accounts/invoices/form.html", {"form": form, "mode": "add"})


@login_required
def invoice_detail(request, pk):
    invoice = get_object_or_404(ProgramInvoice.objects.filter(owner=request.user).select_related("firm", "client", "program", "program__bom").prefetch_related("items", "items__dispatch_challan"), pk=pk)
    return render(request, "accounts/invoices/detail.html", {"invoice": invoice})


def _build_program_invoice_pdf_response(invoice):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    except ImportError:
        return HttpResponse("ReportLab is required for PDF generation. Install it with: pip install reportlab", status=500)

    brand_pink = colors.HexColor("#ED2F8C")
    brand_orange = colors.HexColor("#F6A33B")
    brand_blue = colors.HexColor("#1976F3")
    brand_navy = colors.HexColor("#0F172A")
    ink = colors.HexColor("#1F2937")
    border = colors.HexColor("#D0D5DD")
    soft_bg = colors.HexColor("#F8FAFC")
    white = colors.white
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=10*mm, rightMargin=10*mm, topMargin=14*mm, bottomMargin=14*mm)
    styles = getSampleStyleSheet()
    base = ParagraphStyle('base', parent=styles['BodyText'], fontName='Helvetica', fontSize=8.5, leading=10, textColor=ink)
    head = ParagraphStyle('head', parent=base, fontName='Helvetica-Bold', fontSize=14, leading=16, textColor=brand_navy, alignment=TA_RIGHT)
    white_left = ParagraphStyle('wl', parent=base, fontName='Helvetica', fontSize=8.5, leading=10, textColor=white)
    meta = ParagraphStyle('meta', parent=base, alignment=TA_RIGHT)
    table_head = ParagraphStyle('thead', parent=base, fontName='Helvetica-Bold', textColor=white, fontSize=7.8, alignment=TA_CENTER)
    right = ParagraphStyle('right', parent=base, alignment=TA_RIGHT)
    center = ParagraphStyle('center', parent=base, alignment=TA_CENTER)
    story=[]
    firm = invoice.firm
    firm_html = f"<font size='13'><b>{firm.firm_name if firm else 'InventTech'}</b></font>"
    if firm and firm.full_address:
        firm_html += f"<br/>{firm.full_address}"
    header = Table([[Paragraph(firm_html, white_left), Table([[Paragraph('<b>PROGRAM INVOICE</b>', head)], [Paragraph(f'<b>Invoice No:</b> {invoice.invoice_no}<br/><b>Date:</b> {invoice.invoice_date.strftime('%d-%m-%Y')}<br/><b>Program:</b> {invoice.program.program_no}<br/><b>Client:</b> {invoice.client.name}', meta)]], colWidths=[66*mm])]], colWidths=[124*mm,66*mm])
    header.setStyle(TableStyle([('BACKGROUND',(0,0),(0,0),brand_navy),('BACKGROUND',(1,0),(1,0),colors.HexColor('#F5F8FF')),('BOX',(0,0),(-1,-1),0.9,border),('VALIGN',(0,0),(-1,-1),'TOP'),('LEFTPADDING',(0,0),(-1,-1),8),('RIGHTPADDING',(0,0),(-1,-1),8),('TOPPADDING',(0,0),(-1,-1),8),('BOTTOMPADDING',(0,0),(-1,-1),8)]))
    story += [header, Spacer(1,6)]
    rows = [[Paragraph('Program', table_head), Paragraph('SKU', table_head), Paragraph('Challan', table_head), Paragraph('Quantity', table_head), Paragraph('HSN', table_head), Paragraph('Price', table_head), Paragraph('Amount', table_head)]]
    for item in invoice.items.all().order_by('sort_order','id'):
        rows.append([Paragraph(item.program_label or '-', center), Paragraph(item.sku or '-', center), Paragraph(item.challan_no or '-', center), Paragraph(f'{item.quantity:.2f}', right), Paragraph(item.hsn_code or '-', center), Paragraph(f'{item.price:.2f}', right), Paragraph(f'{item.amount:.2f}', right)])
    tbl = Table(rows, colWidths=[22*mm,36*mm,22*mm,24*mm,22*mm,28*mm,30*mm], repeatRows=1)
    style = TableStyle([('BACKGROUND',(0,0),(-1,0),brand_navy),('TEXTCOLOR',(0,0),(-1,0),white),('BOX',(0,0),(-1,-1),0.9,border),('INNERGRID',(0,0),(-1,-1),0.5,border),('VALIGN',(0,0),(-1,-1),'MIDDLE'),('LEFTPADDING',(0,0),(-1,-1),4),('RIGHTPADDING',(0,0),(-1,-1),4),('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5)])
    for i in range(1,len(rows)):
        if i % 2 == 0: style.add('BACKGROUND',(0,i),(-1,i),soft_bg)
    tbl.setStyle(style)
    story += [tbl, Spacer(1,8)]
    totals = Table([[Paragraph('Sub Total', base), Paragraph(f'{invoice.sub_total:.2f}', right)], [Paragraph('Discount', base), Paragraph(f'{invoice.discount_amount:.2f}', right)], [Paragraph('After Discount', base), Paragraph(f'{invoice.after_discount_amount:.2f}', right)], [Paragraph('Others', base), Paragraph(f'{invoice.other_charges:.2f}', right)], [Paragraph('GST', base), Paragraph(f'{invoice.gst_amount:.2f}', right)], [Paragraph('IGST', base), Paragraph(f'{invoice.igst_amount:.2f}', right)], [Paragraph('<b>Final Amount</b>', base), Paragraph(f'<b>{invoice.final_amount:.2f}</b>', right)]], colWidths=[45*mm,35*mm])
    totals.setStyle(TableStyle([('BOX',(0,0),(-1,-1),0.9,border),('INNERGRID',(0,0),(-1,-1),0.5,border),('BACKGROUND',(0,0),(-1,0),colors.HexColor('#EFF6FF')),('BACKGROUND',(0,-1),(-1,-1),colors.HexColor('#FFF7ED')),('LEFTPADDING',(0,0),(-1,-1),6),('RIGHTPADDING',(0,0),(-1,-1),6),('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5)]))
    story += [totals, Spacer(1,8), Paragraph('This invoice is computer generated.', ParagraphStyle('foot', parent=base, alignment=TA_CENTER, fontName='Helvetica-Bold', textColor=colors.HexColor('#667085')))]
    doc.build(story)
    buf.seek(0)
    return HttpResponse(buf.getvalue(), content_type='application/pdf')


@login_required
def invoice_print(request, pk):
    invoice = get_object_or_404(ProgramInvoice.objects.filter(owner=request.user).select_related('firm','client','program','program__bom').prefetch_related('items'), pk=pk)
    response = _build_program_invoice_pdf_response(invoice)
    if response.status_code == 200 and response.get('Content-Type','').startswith('application/pdf'):
        filename = f'{invoice.invoice_no or "invoice"}.pdf'
        disposition = 'attachment' if request.GET.get('download') == '1' else 'inline'
        response['Content-Disposition'] = f'{disposition}; filename="{filename}"'
    return response

def _program_costing_decimal(value):
    if isinstance(value, Decimal):
        return value
    try:
        return Decimal(str(value or "0"))
    except Exception:
        return Decimal("0")


def _program_costing_money(value):
    return _program_costing_decimal(value).quantize(Decimal("0.01"))


def _program_costing_bucket(jobber_type_name):
    name = (jobber_type_name or "").strip().lower()
    if "cut" in name:
        return "cutting"
    if "stitch" in name:
        return "stitching"
    if "finish" in name:
        return "finishing"
    return "other"


def _build_program_costing_context(program):
    q0 = Decimal("0")
    bom = getattr(program, "bom", None)
    start_record = getattr(program, "start_record", None)

    # ---------------------------------------
    # Cutting Qty
    # ---------------------------------------
    cutting_qty = q0
    if start_record:
        for size_row in start_record.size_rows.all():
            cutting_qty += _program_costing_decimal(size_row.qty)

    if cutting_qty <= 0:
        for size_row in program.size_rows.all():
            cutting_qty += (
                _program_costing_decimal(size_row.xs_qty)
                + _program_costing_decimal(size_row.s_qty)
                + _program_costing_decimal(size_row.m_qty)
                + _program_costing_decimal(size_row.l_qty)
                + _program_costing_decimal(size_row.xl_qty)
                + _program_costing_decimal(size_row.xxl_qty)
            )

    if cutting_qty <= 0:
        cutting_qty = _program_costing_decimal(program.total_qty)

    cutting_qty = _program_costing_money(cutting_qty)

    # ---------------------------------------
    # Material costing breakup
    # Price from BOM material rows
    # Used qty from Program Start fabric rows
    # ---------------------------------------
    material_price_map = {}
    material_rows = []
    material_price_total = q0
    material_used_total = q0
    material_final_total = q0

    if bom:
        for item in bom.material_items.select_related("material").all():
            if item.material_id:
                material_price_map[item.material_id] = _program_costing_decimal(item.cost_per_unit)

    fabric_rows = []
    if start_record:
        fabric_rows = list(start_record.fabric_rows.select_related("material").all())

    if fabric_rows:
        for row in fabric_rows:
            if not row.material_id:
                continue

            used_qty = _program_costing_decimal(row.used_qty)
            if used_qty <= 0:
                used_qty = _program_costing_decimal(row.used)
            if used_qty <= 0:
                used_qty = _program_costing_decimal(row.avg)

            if used_qty <= 0:
                continue

            material_price = material_price_map.get(row.material_id, q0)
            line_total = _program_costing_money(material_price * used_qty)

            material_rows.append({
                "lot_no": (row.lot_no or "-").strip() or "-",
                "material_name": row.material.name if row.material_id and row.material else "-",
                "material_price": _program_costing_money(material_price),
                "used_qty": _program_costing_money(used_qty),
                "final_total": line_total,
            })

            material_price_total += _program_costing_decimal(material_price)
            material_used_total += used_qty
            material_final_total += line_total
    elif bom:
        # Fallback when program start is not filled yet
        for item in bom.material_items.select_related("material").all():
            used_qty = _program_costing_decimal(item.avg)
            material_price = _program_costing_decimal(item.cost_per_unit)
            line_total = _program_costing_money(material_price * used_qty)

            material_rows.append({
                "lot_no": "-",
                "material_name": item.material.name if item.material_id and item.material else "-",
                "material_price": _program_costing_money(material_price),
                "used_qty": _program_costing_money(used_qty),
                "final_total": line_total,
            })

            material_price_total += material_price
            material_used_total += used_qty
            material_final_total += line_total

    material_price_total = _program_costing_money(material_price_total)
    material_used_total = _program_costing_money(material_used_total)
    material_final_total = _program_costing_money(material_final_total)

    # ---------------------------------------
    # Process costing breakup
    # Use start jobber rows first, then program rows, then BOM rows
    # ---------------------------------------
    process_rows_source = []

    if start_record and start_record.jobber_rows.exists():
        process_rows_source = list(
            start_record.jobber_rows.select_related("jobber", "jobber_type").all()
        )
    elif program.jobber_rows.exists():
        process_rows_source = list(
            program.jobber_rows.select_related("jobber", "jobber_type").all()
        )
    elif bom and bom.jobber_details.exists():
        process_rows_source = list(
            bom.jobber_details.select_related("jobber", "jobber_type").all()
        )
    elif bom:
        process_rows_source = list(
            bom.jobber_type_processes.select_related("jobber_type").all()
        )

    process_map = {
        "cutting": {"unit_cost": q0, "names": []},
        "stitching": {"unit_cost": q0, "names": []},
        "finishing": {"unit_cost": q0, "names": []},
        "other": {"unit_cost": q0, "names": []},
    }

    for row in process_rows_source:
        if hasattr(row, "jobber_price"):
            unit_cost = _program_costing_decimal(row.jobber_price)
        else:
            unit_cost = _program_costing_decimal(getattr(row, "price", 0))

        if unit_cost <= 0:
            continue

        jobber_type_name = getattr(getattr(row, "jobber_type", None), "name", "") or ""
        jobber_name = getattr(getattr(row, "jobber", None), "name", "") or jobber_type_name

        bucket = _program_costing_bucket(jobber_type_name)
        process_map[bucket]["unit_cost"] += unit_cost

        if jobber_name and jobber_name not in process_map[bucket]["names"]:
            process_map[bucket]["names"].append(jobber_name)

    cutting_unit = _program_costing_money(process_map["cutting"]["unit_cost"])
    stitching_unit = _program_costing_money(process_map["stitching"]["unit_cost"])
    finishing_unit = _program_costing_money(process_map["finishing"]["unit_cost"])
    other_process_unit = _program_costing_money(process_map["other"]["unit_cost"])

    cutting_total = _program_costing_money(cutting_unit * cutting_qty)
    stitching_total = _program_costing_money(stitching_unit * cutting_qty)
    finishing_total = _program_costing_money(finishing_unit * cutting_qty)
    other_process_total = _program_costing_money(other_process_unit * cutting_qty)

    # ---------------------------------------
    # Accessory / other / maintenance
    # ---------------------------------------
    collar_drawcord_unit = q0
    accessories_unit = q0
    expense_unit = q0

    if bom:
        for item in bom.accessory_items.select_related("accessory").all():
            line_cost = _program_costing_decimal(item.cost)
            accessory_name = (getattr(getattr(item, "accessory", None), "name", "") or "").strip().lower()

            if "collar" in accessory_name or "drawcord" in accessory_name:
                collar_drawcord_unit += line_cost
            else:
                accessories_unit += line_cost

        expense_unit = _program_costing_decimal(bom.expense_total)
        maintenance_cost = _program_costing_money(bom.maintenance_price)
        damage_percent = _program_costing_decimal(program.damage or bom.damage_percent)
    else:
        maintenance_cost = q0
        damage_percent = _program_costing_decimal(program.damage)

    other_cost_unit = _program_costing_money(expense_unit + other_process_unit + _program_costing_decimal(getattr(bom, "price", 0)) + _program_costing_decimal(getattr(bom, "tie_dye_price", 0)))
    collar_drawcord_unit = _program_costing_money(collar_drawcord_unit)
    accessories_unit = _program_costing_money(accessories_unit)

    collar_drawcord_total = _program_costing_money(collar_drawcord_unit * cutting_qty)
    accessories_total = _program_costing_money(accessories_unit * cutting_qty)
    other_cost_total = _program_costing_money(other_cost_unit * cutting_qty)

    # ---------------------------------------
    # Final totals
    # Match the style shown in your reference:
    # maintenance is displayed separately
    # total cost is calculated without maintenance
    # ---------------------------------------
    total_cost = _program_costing_money(
        material_final_total
        + cutting_total
        + stitching_total
        + finishing_total
        + collar_drawcord_total
        + accessories_total
        + other_cost_total
    )

    program_cost_without_damage = q0
    if cutting_qty > 0:
        program_cost_without_damage = _program_costing_money(total_cost / cutting_qty)

    damage_cost = _program_costing_money(
        (program_cost_without_damage * damage_percent) / Decimal("100")
    )
    final_cost = _program_costing_money(program_cost_without_damage + damage_cost)

    breakdown_rows = [
        {
            "label": "Material Cost",
            "subtext": "",
            "amount": material_final_total,
        },
        {
            "label": f"Cutting Cost ({cutting_unit}) * Cutting Qty ({cutting_qty})",
            "subtext": f"({', '.join(process_map['cutting']['names'])})" if process_map["cutting"]["names"] else "",
            "amount": cutting_total,
        },
        {
            "label": f"Stitching Cost ({stitching_unit}) * Cutting Qty ({cutting_qty})",
            "subtext": f"({', '.join(process_map['stitching']['names'])})" if process_map["stitching"]["names"] else "",
            "amount": stitching_total,
        },
        {
            "label": f"Finishing Cost ({finishing_unit}) * Cutting Qty ({cutting_qty})",
            "subtext": f"({', '.join(process_map['finishing']['names'])})" if process_map["finishing"]["names"] else "",
            "amount": finishing_total,
        },
        {
            "label": f"Collar/Drawcord Cost ({collar_drawcord_unit}) * Cutting Qty ({cutting_qty})",
            "subtext": "",
            "amount": collar_drawcord_total,
        },
        {
            "label": f"Accessories Cost ({accessories_unit}) * Cutting Qty ({cutting_qty})",
            "subtext": "",
            "amount": accessories_total,
        },
        {
            "label": f"Other Cost ({other_cost_unit}) * Cutting Qty ({cutting_qty})",
            "subtext": "",
            "amount": other_cost_total,
        },
        {
            "label": "Production SKU Maintenance Cost",
            "subtext": "",
            "amount": maintenance_cost,
        },
        {
            "label": "Total Cost",
            "subtext": "",
            "amount": total_cost,
        },
        {
            "label": f"Program Cost Without Damage ({total_cost} / {cutting_qty})",
            "subtext": "",
            "amount": program_cost_without_damage,
        },
        {
            "label": f"Damage Cost ({_program_costing_money(damage_percent)}%)",
            "subtext": "",
            "amount": damage_cost,
        },
        {
            "label": "Final Cost",
            "subtext": "",
            "amount": final_cost,
        },
    ]

    return {
        "material_rows": material_rows,
        "material_price_total": material_price_total,
        "material_used_total": material_used_total,
        "material_final_total": material_final_total,
        "breakdown_rows": breakdown_rows,
        "cutting_qty": cutting_qty,
        "damage_percent": _program_costing_money(damage_percent),
        "final_cost": final_cost,
    }

@login_required
def program_costing_detail(request, program_id):
    program_qs = (
        Program.objects.filter(owner=request.user)
        .select_related("bom", "firm")
        .prefetch_related(
            "size_rows",
            "jobber_rows__jobber",
            "jobber_rows__jobber_type",
            "bom__material_items__material",
            "bom__accessory_items__accessory",
            "bom__expense_items__expense",
            "bom__jobber_details__jobber",
            "bom__jobber_details__jobber_type",
            "bom__jobber_type_processes__jobber_type",
            "start_record__fabric_rows__material",
            "start_record__size_rows",
            "start_record__jobber_rows__jobber",
            "start_record__jobber_rows__jobber_type",
        )
    )
    program = get_object_or_404(program_qs, pk=program_id)

    snapshots = program.costing_snapshots.all().order_by("-id")
    latest_snapshot = snapshots.first()
    live_context = _build_program_costing_context(program)

    costing_rows = [
        ("Material Cost", live_context.get("material_final_total", Decimal("0"))),
        ("Cutting Qty", live_context.get("cutting_qty", Decimal("0"))),
        ("Damage %", live_context.get("damage_percent", Decimal("0"))),
        ("Final Cost", live_context.get("final_cost", Decimal("0"))),
    ]

    context = {
        "program": program,
        "snapshots": snapshots,
        "latest_snapshot": latest_snapshot,
        "costing_rows": costing_rows,
        **live_context,
    }
    return render(request, "accounts/programs/costing_detail.html", context)


# =========================================================
# Maintenance
# =========================================================

@login_required
def maintenance_list(request):
    q = (request.GET.get('q') or '').strip()
    records = MaintenanceRecord.objects.filter(owner=request.user).order_by('-month_key', '-id')
    if q:
        records = records.filter(Q(month_key__icontains=q) | Q(remarks__icontains=q))
    return render(request, 'accounts/maintenance/list.html', {'records': records, 'q': q})


@login_required
@require_GET
def maintenance_month_payload(request):
    month_key = (request.GET.get("month_key") or "").strip()
    inward_total = 0
    inward_rows = []

    try:
        if not month_key:
            return JsonResponse({
                "inward_total": "0",
                "inward_rows": [],
            })

        start, end = _invoice_month_range(month_key)

        inward_sources = [
            ("Yarn", YarnPOInward.objects.filter(owner=request.user, inward_date__gte=start, inward_date__lt=end).select_related("po")),
            ("Greige", GreigePOInward.objects.filter(owner=request.user, inward_date__gte=start, inward_date__lt=end).select_related("po")),
            ("Dyeing", DyeingPOInward.objects.filter(owner=request.user, inward_date__gte=start, inward_date__lt=end).select_related("po")),
            ("Ready", ReadyPOInward.objects.filter(owner=request.user, inward_date__gte=start, inward_date__lt=end).select_related("po")),
        ]

        total_count = 0

        for inward_type, qs in inward_sources:
            for inward in qs.order_by("inward_date", "id"):
                total_count += 1
                po = getattr(inward, "po", None)

                inward_rows.append({
                    "inward_date": inward.inward_date.strftime("%d-%m-%Y") if inward.inward_date else "",
                    "inward_number": getattr(inward, "inward_number", "") or getattr(inward, "grn_no", "") or f"{inward_type}-{inward.id}",
                    "po_number": (
                        getattr(po, "po_number", "") or
                        getattr(po, "system_number", "") or
                        "-"
                    ),
                    "material_name": f"{inward_type} Inward",
                    "quantity": "1",
                })

        inward_total = total_count

        return JsonResponse({
            "inward_total": str(inward_total),
            "inward_rows": inward_rows,
        })

    except Exception:
        return JsonResponse({
            "inward_total": "0",
            "inward_rows": [],
        }, status=500)


@login_required
def maintenance_create(request):
    popup_mode = request.GET.get("popup") == "1" or request.POST.get("popup") == "1"

    if request.method == "POST":
        form = MaintenanceRecordForm(request.POST)
        formset = MaintenanceExpenseItemFormSet(
            request.POST,
            form_kwargs={"user": request.user}
        )

        if form.is_valid() and formset.is_valid():
            record = form.save(commit=False)
            record.owner = request.user

            # ✅ prevent duplicate month BEFORE saving
            if MaintenanceRecord.objects.filter(
                owner=request.user,
                month_key=record.month_key
            ).exists():
                form.add_error("month_key", "Maintenance for this month already exists.")

            else:
                # ✅ calculate inward count
                try:
                    start, end = _invoice_month_range(record.month_key)

                    total_count = 0
                    inward_sources = [
                        YarnPOInward.objects.filter(owner=request.user, inward_date__gte=start, inward_date__lt=end),
                        GreigePOInward.objects.filter(owner=request.user, inward_date__gte=start, inward_date__lt=end),
                        DyeingPOInward.objects.filter(owner=request.user, inward_date__gte=start, inward_date__lt=end),
                        ReadyPOInward.objects.filter(owner=request.user, inward_date__gte=start, inward_date__lt=end),
                    ]

                    for qs in inward_sources:
                        total_count += qs.count()

                    record.inward_total = Decimal(str(total_count)).quantize(Decimal("0.01"))

                except Exception:
                    record.inward_total = Decimal("0.00")

                # ✅ SAVE FIRST (VERY IMPORTANT)
                record.save()

                # ✅ now attach formset
                formset.instance = record
                items = formset.save(commit=False)

                # delete removed rows safely
                for obj in formset.deleted_objects:
                    if obj.pk:
                        obj.delete()

                # save items
                for index, obj in enumerate(items, start=1):
                    obj.maintenance = record
                    obj.sort_order = index
                    obj.save()

                # recalc totals
                record.recalculate_totals()
                record.save(update_fields=["expense_total", "cost_total"])

                messages.success(
                    request,
                    f"Maintenance record for {record.month_display} saved successfully."
                )

                # popup close
                if popup_mode:
                    return HttpResponse(
                        """
                        <html>
                          <body>
                            <script>
                              window.parent.location.reload();
                            </script>
                          </body>
                        </html>
                        """
                    )

                return redirect("accounts:maintenance_list")

    else:
        form = MaintenanceRecordForm()
        formset = MaintenanceExpenseItemFormSet(form_kwargs={"user": request.user})

    return render(
        request,
        "accounts/maintenance/form_popup.html" if popup_mode else "accounts/maintenance/form.html",
        {
            "form": form,
            "expense_formset": formset,
            "mode": "add",
            "popup_mode": popup_mode,
        },
    )

# -----------------------------------------------------------------------------
# FINAL INWARD TRACKER OVERRIDES
# These definitions intentionally live at the bottom so they override older
# duplicate tracker functions above and force Yarn / Greige / Dyeing / Ready to
# use the same premium normalized tracker context and templates.
# -----------------------------------------------------------------------------

@login_required
def yarn_inward_tracker(request):
    q = (request.GET.get("q") or "").strip()
    target_inward_id = (request.GET.get("inward") or "").strip()

    qs = (
        YarnPurchaseOrder.objects
        .select_related("vendor", "firm", "owner")
        .prefetch_related(
            Prefetch(
                "items",
                queryset=YarnPurchaseOrderItem.objects.select_related("material", "material_type"),
            ),
            Prefetch(
                "inwards",
                queryset=YarnPOInward.objects.select_related("vendor", "inward_type").prefetch_related(
                    "items__po_item__material",
                    "items__po_item__material_type",
                    "generated_greige_pos__items",
                ).order_by("-inward_date", "-id"),
            ),
        )
        .filter(inwards__isnull=False)
        .distinct()
        .order_by("-id")
    )

    if not _can_review_yarn_po(request):
        qs = qs.filter(owner=request.user)

    if q:
        qs = qs.filter(
            Q(system_number__icontains=q)
            | Q(po_number__icontains=q)
            | Q(vendor__name__icontains=q)
            | Q(firm__firm_name__icontains=q)
            | Q(inwards__inward_number__icontains=q)
            | Q(items__material__name__icontains=q)
            | Q(items__material_type__name__icontains=q)
        ).distinct()

    rows = []
    for po in qs:
        po = _attach_yarn_po_metrics(po)
        inward_entries = []
        generated_count = 0

        for inward in po.inwards.all():
            linked_po = inward.generated_greige_pos.order_by("-id").first()
            if linked_po:
                generated_count += 1

            items = [_po_tracker_item_payload(inward_item, fallback_name="Yarn Item") for inward_item in inward.items.all()]
            inward_entries.append({
                "inward": inward,
                "items": items,
                "is_target": str(inward.id) == target_inward_id,
                "linked_po": linked_po,
                "next_started": bool(linked_po),
                "next_view_url": reverse("accounts:greigepo_edit", args=[linked_po.id]) if linked_po else "",
                "next_generate_url": reverse("accounts:generate_greige_po_from_yarn", args=[po.id]),
                "next_generate_method": "post",
                "next_generate_label": "Generate Greige PO",
                "next_view_label": "View Greige PO",
                "edit_url": reverse("accounts:yarn_inward_edit", args=[inward.id]),
            })

        total_inwards = len(inward_entries)
        rows.append({
            "po": po,
            "inward_entries": inward_entries,
            "total_inwards": total_inwards,
            "generated_count": generated_count,
            "progress_label": _po_tracker_progress_label(generated_count, total_inwards),
            "progress_title": "Greige Progress",
            "next_list_label": "View Greige PO List",
            "next_list_url": reverse("accounts:greigepo_list"),
            "inward_url": reverse("accounts:yarnpo_inward", args=[po.id]),
            "total_qty": _po_tracker_qty(po, "total_weight", "total_qty", "total_quantity"),
            "inward_qty": _po_tracker_qty(po, "total_inward_qty", "inward_qty_total"),
            "remaining_qty": _po_tracker_qty(po, "remaining_qty_total", "pending_qty_total"),
        })

    return render(request, "accounts/yarn_po/inward_tracker.html", {
        "rows": rows,
        "q": q,
        "target_inward_id": target_inward_id,
        "tracker_title": "Yarn Inward Tracker",
        "tracker_subtitle": "Track inwarded yarn, accepted stock, rejected qty, hold qty, and Greige PO linkage",
        "tracker_reset_url": reverse("accounts:yarn_inward_tracker"),
        "tracker_list_label": "Yarn POs",
        "tracker_list_url": reverse("accounts:yarnpo_list"),
        "tracker_stock_url": reverse("accounts:stock_lot_wise"),
        "empty_message": "No inwarded Yarn POs found yet.",
        "anchor_prefix": "yarn-inward-",
    })


@login_required
def greige_inward_tracker(request):
    q = (request.GET.get("q") or "").strip()
    target_inward_id = (request.GET.get("inward") or "").strip()

    qs = _greige_po_queryset().filter(inwards__isnull=False).distinct()
    if not _can_review_yarn_po(request):
        qs = qs.filter(owner=request.user)

    if q:
        qs = qs.filter(
            Q(system_number__icontains=q)
            | Q(po_number__icontains=q)
            | Q(vendor__name__icontains=q)
            | Q(source_yarn_po__system_number__icontains=q)
            | Q(source_yarn_po__firm__firm_name__icontains=q)
            | Q(source_yarn_inward__inward_number__icontains=q)
            | Q(inwards__inward_number__icontains=q)
            | Q(items__material__name__icontains=q)
            | Q(items__fabric_name__icontains=q)
            | Q(items__yarn_name__icontains=q)
            | Q(firm__firm_name__icontains=q)
        ).distinct()

    rows = []
    for po in qs:
        inward_entries = []
        generated_count = 0

        for inward in po.inwards.all():
            linked_po = inward.generated_dyeing_pos.order_by("-id").first()
            if linked_po:
                generated_count += 1

            items = [_po_tracker_item_payload(inward_item, fallback_name="Greige Item") for inward_item in inward.items.all()]
            inward_entries.append({
                "inward": inward,
                "items": items,
                "is_target": str(inward.id) == target_inward_id,
                "linked_po": linked_po,
                "next_started": bool(linked_po),
                "next_view_url": reverse("accounts:dyeingpo_edit", args=[linked_po.id]) if linked_po else "",
                "next_generate_url": reverse("accounts:generate_dyeing_po_from_greige", args=[po.id]),
                "next_generate_method": "post",
                "next_generate_label": "Generate Dyeing PO",
                "next_view_label": "View Dyeing PO",
                "edit_url": reverse("accounts:greige_inward_edit", args=[inward.id]),
            })

        total_inwards = len(inward_entries)
        rows.append({
            "po": po,
            "inward_entries": inward_entries,
            "total_inwards": total_inwards,
            "generated_count": generated_count,
            "progress_label": _po_tracker_progress_label(generated_count, total_inwards),
            "progress_title": "Dyeing Progress",
            "next_list_label": "View Dyeing PO List",
            "next_list_url": reverse("accounts:dyeingpo_list"),
            "inward_url": reverse("accounts:greigepo_inward", args=[po.id]),
            "total_qty": _po_tracker_qty(po, "total_weight", "total_qty", "total_quantity"),
            "inward_qty": _po_tracker_qty(po, "total_inward_qty", "inward_qty_total"),
            "remaining_qty": _po_tracker_qty(po, "remaining_qty_total", "pending_qty_total"),
        })

    return render(request, "accounts/greige_po/inward_tracker.html", {
        "rows": rows,
        "q": q,
        "target_inward_id": target_inward_id,
        "tracker_title": "Greige Inward Tracker",
        "tracker_subtitle": "Track inwarded greige fabric, accepted stock, rejected qty, hold qty, and Dyeing PO linkage",
        "tracker_reset_url": reverse("accounts:greige_inward_tracker"),
        "tracker_list_label": "Greige POs",
        "tracker_list_url": reverse("accounts:greigepo_list"),
        "tracker_stock_url": reverse("accounts:stock_lot_wise"),
        "empty_message": "No inwarded Greige POs found yet.",
        "anchor_prefix": "greige-inward-",
    })


@login_required
def dyeing_inward_tracker(request):
    q = (request.GET.get("q") or "").strip()
    target_inward_id = (request.GET.get("inward") or "").strip()

    qs = _dyeing_po_queryset().filter(inwards__isnull=False).distinct()

    if not _can_review_yarn_po(request):
        qs = qs.filter(owner=request.user)

    if q:
        qs = qs.filter(
            Q(system_number__icontains=q)
            | Q(po_number__icontains=q)
            | Q(vendor__name__icontains=q)
            | Q(source_greige_po__system_number__icontains=q)
            | Q(source_greige_inward__inward_number__icontains=q)
            | Q(inwards__inward_number__icontains=q)
            | Q(items__fabric_name__icontains=q)
            | Q(items__finished_material__name__icontains=q)
            | Q(items__dyeing_name__icontains=q)
            | Q(firm__firm_name__icontains=q)
        ).distinct()

    rows = []
    for po in qs:
        inward_entries = []
        ready_po = po.ready_pos.order_by("-id").first()
        generated_count = 1 if ready_po else 0

        for inward in po.inwards.all():
            items = [_po_tracker_item_payload(inward_item, fallback_name="Dyeing Item") for inward_item in inward.items.all()]
            inward_entries.append({
                "inward": inward,
                "items": items,
                "is_target": str(inward.id) == target_inward_id,
                "linked_po": ready_po,
                "next_started": bool(ready_po),
                "next_view_url": reverse("accounts:readypo_detail", args=[ready_po.id]) if ready_po else "",
                "next_generate_url": f"{reverse('accounts:generate_ready_po_from_dyeing', args=[po.id])}?inward={inward.id}",
                "next_generate_method": "get",
                "next_generate_label": "Generate Ready PO",
                "next_view_label": "View Ready PO",
                "edit_url": reverse("accounts:dyeing_inward_edit", args=[inward.id]),
            })

        total_inwards = len(inward_entries)
        rows.append({
            "po": po,
            "inward_entries": inward_entries,
            "total_inwards": total_inwards,
            "generated_count": generated_count,
            "progress_label": "Created" if ready_po else "Pending",
            "progress_title": "Ready PO Progress",
            "next_list_label": "View Ready PO List",
            "next_list_url": reverse("accounts:readypo_list"),
            "inward_url": reverse("accounts:dyeingpo_inward", args=[po.id]),
            "total_qty": _po_tracker_qty(po, "total_weight", "total_qty", "total_quantity"),
            "inward_qty": _po_tracker_qty(po, "total_inward_qty", "inward_qty_total"),
            "remaining_qty": _po_tracker_qty(po, "remaining_qty_total", "pending_qty_total"),
        })

    return render(request, "accounts/dyeing_po/inward_tracker.html", {
        "rows": rows,
        "q": q,
        "target_inward_id": target_inward_id,
        "tracker_title": "Dyeing Inward Tracker",
        "tracker_subtitle": "Track inwarded dyed output, accepted stock, rejected qty, hold qty, and Ready PO linkage",
        "tracker_reset_url": reverse("accounts:dyeing_inward_tracker"),
        "tracker_list_label": "Dyeing POs",
        "tracker_list_url": reverse("accounts:dyeingpo_list"),
        "tracker_stock_url": reverse("accounts:stock_lot_wise"),
        "empty_message": "No inwarded Dyeing POs found yet.",
        "anchor_prefix": "dyeing-inward-",
    })


@login_required
def ready_inward_tracker(request):
    q = (request.GET.get("q") or "").strip()
    target_inward_id = (request.GET.get("inward") or "").strip()

    qs = _ready_po_queryset().filter(inwards__isnull=False).distinct()
    if not _can_review_yarn_po(request):
        qs = qs.filter(owner=request.user)

    if q:
        qs = qs.filter(
            Q(system_number__icontains=q)
            | Q(po_number__icontains=q)
            | Q(vendor__name__icontains=q)
            | Q(source_dyeing_po__system_number__icontains=q)
            | Q(source_dyeing_po__inwards__inward_number__icontains=q)
            | Q(inwards__inward_number__icontains=q)
            | Q(firm__firm_name__icontains=q)
            | Q(items__fabric_name__icontains=q)
        ).distinct()

    rows = []
    for po in qs:
        inward_entries = []
        for inward in po.inwards.all():
            items = [_po_tracker_item_payload(inward_item, fallback_name="Ready Item") for inward_item in inward.items.all()]
            inward_entries.append({
                "inward": inward,
                "items": items,
                "is_target": str(inward.id) == target_inward_id,
                "linked_po": None,
                "next_started": True,
                "next_view_url": "",
                "next_generate_url": "",
                "next_generate_method": "",
                "next_generate_label": "Completed",
                "next_view_label": "Completed",
                "edit_url": reverse("accounts:ready_inward_edit", args=[inward.id]),
            })

        total_inwards = len(inward_entries)
        rows.append({
            "po": po,
            "inward_entries": inward_entries,
            "total_inwards": total_inwards,
            "generated_count": total_inwards,
            "progress_label": "Completed" if total_inwards else "Pending",
            "progress_title": "Ready Stock Progress",
            "next_list_label": "Stock Lot Wise",
            "next_list_url": reverse("accounts:stock_lot_wise"),
            "inward_url": reverse("accounts:readypo_inward", args=[po.id]),
            "total_qty": _po_tracker_qty(po, "total_weight", "total_qty", "total_quantity"),
            "inward_qty": _po_tracker_qty(po, "total_inward_qty", "inward_qty_total"),
            "remaining_qty": _po_tracker_qty(po, "remaining_qty_total", "pending_qty_total"),
        })

    return render(request, "accounts/ready_po/inward_tracker.html", {
        "rows": rows,
        "q": q,
        "target_inward_id": target_inward_id,
        "tracker_title": "Ready Inward Tracker",
        "tracker_subtitle": "Track inwarded ready fabric, accepted stock, rejected qty, hold qty, and stock movement",
        "tracker_reset_url": reverse("accounts:ready_inward_tracker"),
        "tracker_list_label": "Ready POs",
        "tracker_list_url": reverse("accounts:readypo_list"),
        "tracker_stock_url": reverse("accounts:stock_lot_wise"),
        "empty_message": "No inwarded Ready POs found yet.",
        "anchor_prefix": "ready-inward-",
    })

# -----------------------------------------------------------------------------
# PDF DESIGN OVERRIDE - Dyeing/Ready PO styled like Yarn PO
# Kept at bottom intentionally so it overrides the older simple canvas builders
# without changing URL/view/import structure.
# -----------------------------------------------------------------------------
def _build_dye_ready_yarn_style_pdf(po, *, title, po_kind):
    try:
        import os
        from pathlib import Path
        from html import escape
        from django.conf import settings
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.lib.utils import ImageReader
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError:
        return HttpResponse(
            "ReportLab is required for PDF generation. Install it with: pip install reportlab",
            status=500,
        )

    brand_pink = colors.HexColor("#ED2F8C")
    brand_orange = colors.HexColor("#F6A33B")
    brand_blue = colors.HexColor("#1976F3")
    brand_navy = colors.HexColor("#0F172A")
    ink = colors.HexColor("#1F2937")
    muted = colors.HexColor("#667085")
    border = colors.HexColor("#D0D5DD")
    soft_bg = colors.HexColor("#F8FAFC")
    white = colors.white

    def safe(value):
        value = "" if value is None else str(value).strip()
        return value if value else "-"

    def money(value):
        try:
            return f"{float(value or 0):,.2f}"
        except Exception:
            return safe(value)

    def qty(value):
        try:
            return f"{float(value or 0):,.2f}".rstrip("0").rstrip(".")
        except Exception:
            return safe(value)

    def as_decimal(value):
        try:
            return Decimal(str(value or 0))
        except Exception:
            return Decimal("0")

    def line_if(label, value):
        value = "" if value is None else str(value).strip()
        if not value:
            return ""
        return f"<b>{escape(label)}:</b> {escape(value)}"

    def join_parts(*parts):
        clean = [str(p).strip() for p in parts if str(p).strip()]
        return ", ".join(clean)

    def date_value(value):
        try:
            return value.strftime("%d-%m-%Y") if value else "-"
        except Exception:
            return safe(value)

    def logo_path():
        firm_obj = getattr(po, "firm", None)
        if firm_obj and getattr(firm_obj, "logo", None):
            try:
                path = firm_obj.logo.path
                if path and os.path.exists(path):
                    return path
            except Exception:
                pass
        fallback = Path(settings.BASE_DIR) / "Logo.jpeg"
        return str(fallback) if fallback.exists() else None

    logo = logo_path()

    def draw_page_branding(canvas, doc):
        page_w, page_h = A4
        canvas.saveState()
        canvas.setStrokeColor(colors.HexColor("#E4E7EC"))
        canvas.setLineWidth(0.8)
        canvas.roundRect(8 * mm, 8 * mm, page_w - 16 * mm, page_h - 16 * mm, 4 * mm, stroke=1, fill=0)
        usable_w = page_w - 16 * mm
        stripe_w = usable_w / 3.0
        stripe_y = page_h - 13 * mm
        stripe_h = 4.5 * mm
        canvas.setFillColor(brand_pink)
        canvas.rect(8 * mm, stripe_y, stripe_w, stripe_h, fill=1, stroke=0)
        canvas.setFillColor(brand_orange)
        canvas.rect(8 * mm + stripe_w, stripe_y, stripe_w, stripe_h, fill=1, stroke=0)
        canvas.setFillColor(brand_blue)
        canvas.rect(8 * mm + 2 * stripe_w, stripe_y, stripe_w, stripe_h, fill=1, stroke=0)
        if logo:
            try:
                img = ImageReader(logo)
                iw, ih = img.getSize()
                draw_w = 26 * mm
                draw_h = draw_w * (ih / float(iw)) if iw and ih else 26 * mm
                try:
                    canvas.setFillAlpha(0.10)
                    canvas.setStrokeAlpha(0.10)
                except Exception:
                    pass
                canvas.drawImage(
                    img,
                    (page_w - draw_w) / 2.0,
                    10 * mm,
                    width=draw_w,
                    height=draw_h,
                    preserveAspectRatio=True,
                    mask="auto",
                )
            except Exception:
                pass
        canvas.restoreState()

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
    )

    styles = getSampleStyleSheet()
    base = ParagraphStyle("DYReadyBase", parent=styles["BodyText"], fontName="Helvetica", fontSize=8.4, leading=10.3, textColor=ink, spaceAfter=0)
    firm_style = ParagraphStyle("DYReadyFirm", parent=base, fontName="Helvetica", fontSize=8.4, leading=10.4, textColor=white, alignment=TA_LEFT)
    title_style = ParagraphStyle("DYReadyTitle", parent=base, fontName="Helvetica-Bold", fontSize=14, leading=16, textColor=brand_navy, alignment=TA_RIGHT)
    meta_style = ParagraphStyle("DYReadyMeta", parent=base, fontName="Helvetica", fontSize=8.3, leading=10.2, textColor=ink, alignment=TA_RIGHT)
    head_style = ParagraphStyle("DYReadyHead", parent=base, fontName="Helvetica-Bold", fontSize=8, leading=10, textColor=white, alignment=TA_LEFT)
    value_style = ParagraphStyle("DYReadyValue", parent=base, fontName="Helvetica", fontSize=8.2, leading=10.2, textColor=ink, alignment=TA_LEFT)
    th_style = ParagraphStyle("DYReadyTableHead", parent=base, fontName="Helvetica-Bold", fontSize=7.6, leading=9.3, textColor=white, alignment=TA_CENTER)
    center_style = ParagraphStyle("DYReadyCenter", parent=base, fontName="Helvetica", fontSize=7.9, leading=9.5, alignment=TA_CENTER)
    desc_style = ParagraphStyle("DYReadyDesc", parent=base, fontName="Helvetica", fontSize=7.9, leading=9.5, alignment=TA_LEFT)
    right_style = ParagraphStyle("DYReadyRight", parent=base, fontName="Helvetica", fontSize=7.9, leading=9.5, alignment=TA_RIGHT)
    block_title = ParagraphStyle("DYReadyBlockTitle", parent=base, fontName="Helvetica-Bold", fontSize=8.4, leading=10.5, textColor=brand_navy, alignment=TA_LEFT)
    block_text = ParagraphStyle("DYReadyBlockText", parent=base, fontName="Helvetica", fontSize=8.1, leading=10.1, textColor=ink, alignment=TA_LEFT)
    sign_style = ParagraphStyle("DYReadySign", parent=base, fontName="Helvetica-Bold", fontSize=8, leading=10, textColor=ink, alignment=TA_LEFT)
    total_label = ParagraphStyle("DYReadyTotalLabel", parent=base, fontName="Helvetica-Bold", fontSize=8.1, leading=10.1, textColor=ink, alignment=TA_LEFT)
    total_value = ParagraphStyle("DYReadyTotalValue", parent=base, fontName="Helvetica-Bold", fontSize=8.1, leading=10.1, textColor=ink, alignment=TA_RIGHT)
    footer_style = ParagraphStyle("DYReadyFooter", parent=base, fontName="Helvetica-Bold", fontSize=7.6, leading=9.2, textColor=muted, alignment=TA_CENTER)

    story = []
    firm = getattr(po, "firm", None)
    vendor = getattr(po, "vendor", None)
    items = list(po.items.all())

    firm_name = safe(getattr(firm, "firm_name", "InventTech") if firm else "InventTech")
    try:
        firm_type = firm.get_firm_type_display() if firm else ""
    except Exception:
        firm_type = safe(getattr(firm, "firm_type", ""))
    firm_address = join_parts(getattr(firm, "address_line", ""), getattr(firm, "city", ""), getattr(firm, "state", ""), getattr(firm, "pincode", ""))
    contact_line = " | ".join([p for p in [line_if("Phone", getattr(firm, "phone", "")), line_if("Email", getattr(firm, "email", "")), line_if("GSTIN", getattr(firm, "gst_number", ""))] if p])
    stat_line = " | ".join([p for p in [line_if("PAN", getattr(firm, "pan_number", "")), line_if("TAN", getattr(firm, "tan_number", "")), line_if("CIN", getattr(firm, "cin_number", ""))] if p])

    firm_html = f"<font size='13'><b>{escape(firm_name)}</b></font>"
    if firm_type and firm_type != "-":
        firm_html += f"<br/>{escape(firm_type)}"
    if firm_address:
        firm_html += f"<br/>{escape(firm_address)}"
    if contact_line:
        firm_html += f"<br/>{contact_line}"
    if stat_line:
        firm_html += f"<br/>{stat_line}"

    try:
        approval = po.get_approval_status_display()
    except Exception:
        approval = safe(getattr(po, "approval_status", ""))
    po_no = getattr(po, "po_number", "") or getattr(po, "system_number", "") or "-"
    source_line = ""
    if po_kind == "dyeing" and getattr(po, "source_greige_po", None):
        source_line = f"<br/><b>Greige PO:</b> {escape(safe(po.source_greige_po.system_number))}"
    if po_kind == "ready" and getattr(po, "source_dyeing_po", None):
        source_line = f"<br/><b>Dyeing PO:</b> {escape(safe(po.source_dyeing_po.system_number))}"
    cancel_date = date_value(getattr(po, "cancel_date", None))

    header_table = Table(
        [[
            Paragraph(firm_html, firm_style),
            Table(
                [[Paragraph(f"<b>{escape(title)}</b>", title_style)], [
                    Paragraph(
                        f"<b>PO No:</b> {escape(safe(po_no))}<br/>"
                        f"<b>PO Date:</b> {escape(date_value(getattr(po, 'po_date', None)))}<br/>"
                        f"<b>System No:</b> {escape(safe(getattr(po, 'system_number', '')))}<br/>"
                        f"<b>Status:</b> {escape(safe(approval))}"
                        + source_line
                        + (f"<br/><b>Cancel Date:</b> {escape(cancel_date)}" if cancel_date != "-" else ""),
                        meta_style,
                    )
                ]],
                colWidths=[66 * mm],
            ),
        ]],
        colWidths=[124 * mm, 66 * mm],
    )
    header_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), brand_navy),
        ("BACKGROUND", (1, 0), (1, 0), colors.HexColor("#F5F8FF")),
        ("BOX", (0, 0), (-1, -1), 0.9, border),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (0, 0), 9), ("RIGHTPADDING", (0, 0), (0, 0), 9),
        ("TOPPADDING", (0, 0), (0, 0), 9), ("BOTTOMPADDING", (0, 0), (0, 0), 9),
        ("LEFTPADDING", (1, 0), (1, 0), 0), ("RIGHTPADDING", (1, 0), (1, 0), 0),
        ("TOPPADDING", (1, 0), (1, 0), 0), ("BOTTOMPADDING", (1, 0), (1, 0), 0),
    ]))
    header_table._cellvalues[0][1].setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F5F8FF")),
        ("BOX", (0, 0), (-1, -1), 0.9, brand_blue),
        ("LEFTPADDING", (0, 0), (-1, -1), 8), ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 7), ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 6))

    vendor_html = f"<b>{escape(safe(getattr(vendor, 'name', '')))}</b>"
    vendor_body = "<br/>".join([line for line in [line_if("Contact", getattr(vendor, "contact_person", "")), line_if("Phone", getattr(vendor, "phone", "")), line_if("Email", getattr(vendor, "email", "")), line_if("GSTIN", getattr(vendor, "gst_number", "")), line_if("Address", getattr(vendor, "address", ""))] if line])
    if vendor_body:
        vendor_html += "<br/>" + vendor_body

    bill_html = f"<b>{escape(firm_name)}</b>"
    bill_body = "<br/>".join([line for line in [line_if("Address", firm_address), line_if("Phone", getattr(firm, "phone", "")), line_if("Email", getattr(firm, "email", "")), line_if("GSTIN", getattr(firm, "gst_number", "")), line_if("Ship To", getattr(po, "shipping_address", ""))] if line])
    if bill_body:
        bill_html += "<br/>" + bill_body

    party_table = Table([[Paragraph("VENDOR", head_style), Paragraph("BILL TO / SHIP TO", head_style)], [Paragraph(vendor_html, value_style), Paragraph(bill_html, value_style)]], colWidths=[95 * mm, 95 * mm])
    party_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), brand_orange), ("BACKGROUND", (1, 0), (1, 0), brand_blue),
        ("TEXTCOLOR", (0, 0), (-1, 0), white), ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#FBFCFE")),
        ("BOX", (0, 0), (-1, -1), 0.9, border), ("INNERGRID", (0, 0), (-1, -1), 0.7, border),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 7), ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.append(party_table)
    story.append(Spacer(1, 7))

    if po_kind == "dyeing":
        rows = [[Paragraph("Sr", th_style), Paragraph("Fabric / Greige", th_style), Paragraph("Dyeing", th_style), Paragraph("Unit", th_style), Paragraph("Qty", th_style), Paragraph("Rate", th_style), Paragraph("Amount", th_style)]]
        widths = [11 * mm, 62 * mm, 41 * mm, 16 * mm, 18 * mm, 20 * mm, 22 * mm]
        total_qty = Decimal("0")
        total_amount = Decimal("0")
        for index, item in enumerate(items, start=1):
            item_qty = getattr(item, "quantity", None) or getattr(item, "total_qty", None) or 0
            amount = getattr(item, "line_final_amount", None) or getattr(item, "line_subtotal", None) or 0
            total_qty += as_decimal(item_qty)
            total_amount += as_decimal(amount)
            fabric = getattr(item, "fabric_name", "") or getattr(getattr(item, "finished_material", None), "name", "") or "-"
            greige = getattr(item, "greige_name", "") or ""
            description = fabric if not greige else f"{fabric} / {greige}"
            detail = []
            if getattr(item, "source_input_qty", None):
                detail.append(f"Input: {qty(item.source_input_qty)}")
            if getattr(item, "expected_loss_percent", None):
                detail.append(f"Loss: {qty(item.expected_loss_percent)}%")
            if getattr(item, "expected_output_qty", None):
                detail.append(f"Output: {qty(item.expected_output_qty)}")
            if getattr(item, "rolls", None):
                detail.append(f"Rolls: {qty(item.rolls)}")
            if getattr(item, "remark", ""):
                detail.append(f"Remark: {item.remark}")
            desc_html = f"<b>{escape(safe(description))}</b>" + ("<br/>" + escape(" | ".join(detail)) if detail else "")
            dyeing = getattr(item, "dyeing_name", "") or getattr(item, "dyeing_type", "") or "-"
            rows.append([Paragraph(str(index), center_style), Paragraph(desc_html, desc_style), Paragraph(escape(safe(dyeing)), desc_style), Paragraph(escape(safe(getattr(item, "unit", ""))), center_style), Paragraph(escape(qty(item_qty)), center_style), Paragraph(escape(money(getattr(item, "rate", 0))), right_style), Paragraph(escape(money(amount)), right_style)])
        ordered_qty = getattr(po, "total_weight", None) or total_qty
        grand_total = getattr(po, "final_amount", None) or total_amount
    else:
        rows = [[Paragraph("Sr", th_style), Paragraph("Fabric", th_style), Paragraph("Dyeing", th_style), Paragraph("Unit", th_style), Paragraph("Qty", th_style), Paragraph("Remarks", th_style)]]
        widths = [12 * mm, 62 * mm, 42 * mm, 18 * mm, 20 * mm, 36 * mm]
        total_qty = Decimal("0")
        for index, item in enumerate(items, start=1):
            item_qty = getattr(item, "quantity", None) or 0
            total_qty += as_decimal(item_qty)
            fabric = getattr(item, "fabric_name", "") or getattr(getattr(item, "finished_material", None), "name", "") or "-"
            rows.append([Paragraph(str(index), center_style), Paragraph(f"<b>{escape(safe(fabric))}</b>", desc_style), Paragraph(escape(safe(getattr(item, "dyeing_name", ""))), desc_style), Paragraph(escape(safe(getattr(item, "unit", ""))), center_style), Paragraph(escape(qty(item_qty)), center_style), Paragraph(escape(safe(getattr(item, "remark", ""))), desc_style)])
        ordered_qty = getattr(po, "total_weight", None) or total_qty
        grand_total = None

    for _ in range(max(0, 5 - len(items))):
        rows.append([Paragraph("", center_style) for _ in range(len(widths))])

    item_table = Table(rows, colWidths=widths, repeatRows=1)
    item_style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), brand_navy), ("TEXTCOLOR", (0, 0), (-1, 0), white),
        ("BOX", (0, 0), (-1, -1), 0.9, border), ("INNERGRID", (0, 0), (-1, -1), 0.55, border),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ])
    for row_index in range(1, len(rows)):
        if row_index % 2 == 0:
            item_style.add("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#F9FAFB"))
    item_table.setStyle(item_style)
    story.append(item_table)
    story.append(Spacer(1, 8))

    notes = []
    if getattr(po, "remarks", ""):
        notes.append(f"<font color='#0F172A'><b>Remarks</b></font><br/>{escape(po.remarks).replace(chr(10), '<br/>')}")
    if getattr(po, "description", ""):
        notes.append(f"<font color='#0F172A'><b>Description</b></font><br/>{escape(po.description).replace(chr(10), '<br/>')}")
    if getattr(po, "terms_conditions", ""):
        notes.append(f"<font color='#0F172A'><b>Terms &amp; Conditions</b></font><br/>{escape(po.terms_conditions).replace(chr(10), '<br/>')}")
    if getattr(po, "shipping_address", ""):
        notes.append(f"<font color='#0F172A'><b>Shipping Address</b></font><br/>{escape(po.shipping_address).replace(chr(10), '<br/>')}")
    if not notes:
        notes.append("<font color='#0F172A'><b>Notes</b></font><br/>Standard terms and conditions apply.")

    signature_table = Table([[Paragraph("<b>AUTHORISED SIGNATORY</b><br/><br/>_________________________", sign_style), Paragraph(f"<b>DATE</b><br/><br/>{escape(date_value(getattr(po, 'po_date', None)))}", sign_style)]], colWidths=[68 * mm, 30 * mm])
    signature_table.setStyle(TableStyle([("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0), ("TOPPADDING", (0, 0), (-1, -1), 0), ("BOTTOMPADDING", (0, 0), (-1, -1), 0), ("VALIGN", (0, 0), (-1, -1), "TOP")]))

    left_block = Table([[Paragraph("NOTES / TERMS", block_title)], [Paragraph("<br/><br/>".join(notes), block_text)], [signature_table]], colWidths=[102 * mm])
    left_block.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), colors.HexColor("#FDF2F8")), ("BOX", (0, 0), (-1, -1), 0.9, border), ("INNERGRID", (0, 0), (-1, -1), 0.6, border),
        ("LEFTPADDING", (0, 0), (-1, -1), 7), ("RIGHTPADDING", (0, 0), (-1, -1), 7), ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 7), ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))

    totals = [[Paragraph("Total Qty", total_label), Paragraph(escape(qty(ordered_qty)), total_value)]]
    if po_kind == "dyeing":
        totals.append([Paragraph("Sub Total", total_label), Paragraph(escape(money(getattr(po, "subtotal", 0))), total_value)])
        totals.append([Paragraph("After Discount", total_label), Paragraph(escape(money(getattr(po, "after_discount_value", 0))), total_value)])
        totals.append([Paragraph("Other Charges", total_label), Paragraph(escape(money(getattr(po, "others", 0))), total_value)])
        totals.append([Paragraph(f"GST ({qty(getattr(po, 'gst_percent', 0))}%)", total_label), Paragraph("-", total_value)])
        totals.append([Paragraph(f"TCS ({qty(getattr(po, 'tcs_percent', 0))}%)", total_label), Paragraph("-", total_value)])
        totals.append([Paragraph("Total Amount", total_label), Paragraph(escape(money(grand_total)), total_value)])
    else:
        if getattr(po, "available_qty", None) is not None:
            totals.insert(0, [Paragraph("Available Qty", total_label), Paragraph(escape(qty(getattr(po, "available_qty", 0))), total_value)])
        for label, field in [("Delivery Period", "delivery_period"), ("Validity", "validity_period"), ("Director", "director")]:
            value = getattr(po, field, "")
            if value:
                totals.append([Paragraph(label, total_label), Paragraph(escape(safe(value)), total_value)])

    totals_table = Table(totals, colWidths=[49 * mm, 39 * mm])
    totals_style = TableStyle([
        ("BOX", (0, 0), (-1, -1), 0.9, border), ("INNERGRID", (0, 0), (-1, -1), 0.6, border),
        ("LEFTPADDING", (0, 0), (-1, -1), 7), ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EFF6FF")), ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#FFF7ED")),
    ])
    for row_index in range(1, len(totals) - 1):
        if row_index % 2 == 1:
            totals_style.add("BACKGROUND", (0, row_index), (-1, row_index), soft_bg)
    totals_table.setStyle(totals_style)

    lower_table = Table([[left_block, totals_table]], colWidths=[102 * mm, 88 * mm])
    lower_table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"), ("LEFTPADDING", (0, 0), (-1, -1), 0), ("RIGHTPADDING", (0, 0), (-1, -1), 0), ("TOPPADDING", (0, 0), (-1, -1), 0), ("BOTTOMPADDING", (0, 0), (-1, -1), 0)]))
    story.append(lower_table)
    story.append(Spacer(1, 8))

    footer = Table([[Paragraph("THIS PO IS COMPUTER GENERATED, HENCE SIGNATURE IS NOT REQUIRED", footer_style)]], colWidths=[190 * mm])
    footer.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FAFC")), ("BOX", (0, 0), (-1, -1), 0.9, border), ("LEFTPADDING", (0, 0), (-1, -1), 6), ("RIGHTPADDING", (0, 0), (-1, -1), 6), ("TOPPADDING", (0, 0), (-1, -1), 5), ("BOTTOMPADDING", (0, 0), (-1, -1), 5)]))
    story.append(footer)

    doc.build(story, onFirstPage=draw_page_branding, onLaterPages=draw_page_branding)
    buffer.seek(0)
    return HttpResponse(buffer.getvalue(), content_type="application/pdf")


def _build_simple_dyeing_po_pdf_response(po):
    return _build_dye_ready_yarn_style_pdf(po, title="DYEING PURCHASE ORDER", po_kind="dyeing")


def _build_simple_ready_po_pdf_response(po):
    return _build_dye_ready_yarn_style_pdf(po, title="READY PURCHASE ORDER", po_kind="ready")



from pathlib import Path

# ---------------------------------------------------------------------------
# DEEP SYSTEM HEALTH / ERROR LOG DASHBOARD
# ---------------------------------------------------------------------------


def _health_rel(path, base_dir):
    try:
        return str(Path(path).resolve().relative_to(Path(base_dir).resolve()))
    except Exception:
        try:
            return str(Path(path))
        except Exception:
            return str(path)


def _health_safe_tail(path, max_lines=2200):
    """Return a safe tail from a text log file without crashing on bad encoding."""
    try:
        path = Path(path)
    except Exception:
        return []

    if not path.exists() or not path.is_file():
        return []

    try:
        with path.open("r", encoding="utf-8", errors="replace") as handle:
            lines = handle.readlines()
    except Exception:
        try:
            with path.open("r", encoding="latin-1", errors="replace") as handle:
                lines = handle.readlines()
        except Exception:
            return []

    return [line.rstrip("\n") for line in lines[-max_lines:]]


def _health_project_files(base_dir, suffixes=None, include_migrations=True):
    """Yield project files while avoiding venv/git/cache/media folders."""
    base_dir = Path(base_dir)
    skip_parts = {
        ".git",
        ".idea",
        ".vscode",
        "__pycache__",
        "venv",
        "env",
        ".venv",
        "node_modules",
        "media",
        "staticfiles",
    }
    if not include_migrations:
        skip_parts.add("migrations")

    roots = [
        base_dir / "accounts",
        base_dir / "config",
        base_dir / "templates",
        base_dir / "static",
    ]

    for root in roots:
        if not root.exists():
            continue
        for path in root.rglob("*"):
            try:
                if not path.is_file():
                    continue
                if any(part in skip_parts for part in path.parts):
                    continue
                if suffixes and path.suffix.lower() not in suffixes:
                    continue
                yield path
            except Exception:
                continue


def _health_file_text(path):
    try:
        return Path(path).read_text(encoding="utf-8", errors="replace")
    except Exception:
        return ""


def _health_line_no(text, pos):
    try:
        return text[:pos].count("\n") + 1
    except Exception:
        return 1


def _health_collect_url_names():
    """Collect URL names including namespaced names such as accounts:dashboard and duplicate definitions."""
    from collections import Counter
    from django.urls import get_resolver

    names = set()
    counter = Counter()
    rows = []

    def walk(patterns, namespaces=None):
        namespaces = namespaces or []
        for pattern in patterns:
            child_patterns = getattr(pattern, "url_patterns", None)
            if child_patterns is not None:
                namespace = getattr(pattern, "namespace", None)
                next_namespaces = namespaces + ([namespace] if namespace else [])
                walk(child_patterns, next_namespaces)
                continue

            name = getattr(pattern, "name", None)
            if not name:
                continue
            full_name = ":".join(namespaces + [name]) if namespaces else name
            route = str(getattr(pattern, "pattern", ""))
            callback = getattr(pattern, "callback", None)
            callback_name = getattr(callback, "__name__", str(callback))
            names.add(name)
            names.add(full_name)
            counter[full_name] += 1
            rows.append({
                "name": full_name,
                "route": route,
                "view": callback_name,
            })

    try:
        walk(get_resolver().url_patterns)
    except Exception:
        pass

    duplicates = []
    for name, count in counter.items():
        if count <= 1:
            continue
        samples = [row for row in rows if row["name"] == name][:4]
        duplicates.append({
            "name": name,
            "count": count,
            "samples": "; ".join(f"{row['route']} → {row['view']}" for row in samples),
        })

    return names, duplicates[:80]


def _health_find_broken_url_references(base_dir):
    import re

    known_names, duplicate_url_names = _health_collect_url_names()
    rows = []
    seen = set()

    patterns = [
        ("template", re.compile(r"{%\s*url\s+(['\"])(?P<name>[^'\"]+)\1")),
        ("reverse", re.compile(r"\breverse(?:_lazy)?\s*\(\s*(['\"])(?P<name>[^'\"]+)\1")),
        ("redirect", re.compile(r"\bredirect\s*\(\s*(['\"])(?P<name>[^'\"]+)\1")),
    ]

    for path in _health_project_files(base_dir, {".html", ".py"}):
        text = _health_file_text(path)
        if not text:
            continue
        for source, regex in patterns:
            for match in regex.finditer(text):
                url_name = (match.group("name") or "").strip()
                if not url_name:
                    continue
                # redirect('/literal/path/') is valid and not a URL name.
                if source == "redirect" and (url_name.startswith("/") or url_name.startswith("http") or url_name.endswith(".html")):
                    continue
                # Skip variables accidentally captured from uncommon syntax.
                if "{" in url_name or "}" in url_name or " " in url_name:
                    continue
                if url_name in known_names:
                    continue
                key = (str(path), source, url_name)
                if key in seen:
                    continue
                seen.add(key)
                rows.append({
                    "file": _health_rel(path, base_dir),
                    "name": url_name,
                    "source": source,
                    "line": _health_line_no(text, match.start()),
                })

    return rows[:150], duplicate_url_names


def _health_template_name_from_path(path):
    parts = list(Path(path).parts)
    if "templates" in parts:
        index = parts.index("templates")
        return "/".join(parts[index + 1:])
    return Path(path).name


def _health_find_missing_templates(base_dir):
    import re
    from django.template import TemplateDoesNotExist
    from django.template.loader import get_template

    rows = []
    seen = set()

    template_patterns = [
        re.compile(r"render\s*\([^,]+,\s*['\"](?P<name>[^'\"]+\.html)['\"]"),
        re.compile(r"get_template\s*\(\s*['\"](?P<name>[^'\"]+\.html)['\"]"),
        re.compile(r"select_template\s*\(\s*\[\s*['\"](?P<name>[^'\"]+\.html)['\"]"),
        re.compile(r"{%\s*(?:include|extends)\s+['\"](?P<name>[^'\"]+\.html)['\"]"),
    ]

    for path in _health_project_files(base_dir, {".py", ".html"}):
        text = _health_file_text(path)
        if not text:
            continue

        for regex in template_patterns:
            for match in regex.finditer(text):
                template_name = (match.group("name") or "").strip()
                if not template_name:
                    continue
                key = (template_name, str(path))
                if key in seen:
                    continue
                seen.add(key)
                try:
                    get_template(template_name)
                except TemplateDoesNotExist as exc:
                    rows.append({
                        "file": _health_rel(path, base_dir),
                        "name": template_name,
                        "line": _health_line_no(text, match.start()),
                        "error": str(exc),
                    })
                except Exception as exc:
                    rows.append({
                        "file": _health_rel(path, base_dir),
                        "name": template_name,
                        "line": _health_line_no(text, match.start()),
                        "error": str(exc),
                    })

    return rows[:150]


def _health_compile_all_templates(base_dir):
    """Compile every project template to catch TemplateSyntaxError, bad tags, missing includes, etc."""
    from django.template import TemplateDoesNotExist, TemplateSyntaxError
    from django.template.loader import get_template

    rows = []
    checked = 0
    for path in _health_project_files(base_dir, {".html"}):
        if "templates" not in Path(path).parts:
            continue
        checked += 1
        template_name = _health_template_name_from_path(path)
        try:
            get_template(template_name)
        except TemplateSyntaxError as exc:
            rows.append({
                "file": _health_rel(path, base_dir),
                "template": template_name,
                "kind": "TemplateSyntaxError",
                "error": str(exc),
            })
        except TemplateDoesNotExist as exc:
            rows.append({
                "file": _health_rel(path, base_dir),
                "template": template_name,
                "kind": "TemplateDoesNotExist",
                "error": str(exc),
            })
        except Exception as exc:
            rows.append({
                "file": _health_rel(path, base_dir),
                "template": template_name,
                "kind": exc.__class__.__name__,
                "error": str(exc),
            })

    return {"checked": checked, "items": rows[:150]}


def _health_find_missing_static(base_dir):
    import re
    from django.contrib.staticfiles import finders

    rows = []
    seen = set()
    patterns = [
        ("template static tag", re.compile(r"{%\s*static\s+['\"](?P<path>[^'\"]+)['\"]")),
        ("hardcoded /static/", re.compile(r"(?:href|src)=[\"'](?P<path>/static/[^\"']+)[\"']")),
        ("css url", re.compile(r"url\([\"']?(?P<path>/static/[^\)\"']+)[\"']?\)")),
    ]

    for path in _health_project_files(base_dir, {".html", ".css", ".js"}):
        text = _health_file_text(path)
        if not text:
            continue
        for source, regex in patterns:
            for match in regex.finditer(text):
                static_path = (match.group("path") or "").strip()
                if not static_path or "{{" in static_path or "{%" in static_path:
                    continue
                if static_path.startswith("/static/"):
                    static_path = static_path[len("/static/"):]
                static_path = static_path.split("?")[0].split("#")[0].strip()
                if not static_path:
                    continue
                key = (static_path, str(path))
                if key in seen:
                    continue
                seen.add(key)
                try:
                    found = finders.find(static_path)
                except Exception:
                    found = None
                if not found:
                    rows.append({
                        "file": _health_rel(path, base_dir),
                        "path": static_path,
                        "source": source,
                        "line": _health_line_no(text, match.start()),
                    })

    return rows[:150]


def _health_collect_errors(base_dir):
    import re

    log_candidates = [
        Path(base_dir) / "logs" / "django-errors.log",
        Path(base_dir) / "logs" / "django.log",
        Path(base_dir) / "django-errors.log",
        Path(base_dir) / "error.log",
        Path(base_dir) / "server.log",
        Path(base_dir) / "debug.log",
    ]

    error_markers = (
        "ERROR", "CRITICAL", "WARNING", "Traceback", "Exception", "Internal Server Error",
        "OperationalError", "ProgrammingError", "IntegrityError", "DatabaseError",
        "TemplateDoesNotExist", "TemplateSyntaxError", "NoReverseMatch", "Resolver404",
        "AttributeError", "NameError", "ValueError", "TypeError", "ImportError", "ModuleNotFoundError",
        "FieldError", "ValidationError", "SuspiciousOperation", "DisallowedHost", "Forbidden",
        "Not Found:", "Bad Request:", " 500 ", " 404 ", " 403 ", " 400 ",
    )

    category_patterns = [
        ("Database", re.compile(r"OperationalError|ProgrammingError|IntegrityError|DatabaseError|no such table", re.I)),
        ("Template", re.compile(r"TemplateDoesNotExist|TemplateSyntaxError|Unclosed tag|Invalid block tag", re.I)),
        ("URL/Reverse", re.compile(r"NoReverseMatch|Resolver404|Reverse for", re.I)),
        ("Python", re.compile(r"AttributeError|NameError|TypeError|ValueError|ImportError|ModuleNotFoundError", re.I)),
        ("HTTP", re.compile(r"Internal Server Error|Not Found:|Forbidden|Bad Request| 500 | 404 | 403 | 400 ", re.I)),
        ("Security", re.compile(r"SuspiciousOperation|DisallowedHost|CSRF", re.I)),
    ]

    errors = []
    failed_pages = {}
    page_patterns = [
        re.compile(r"(?P<kind>Internal Server Error|Not Found|Forbidden|Bad Request):\s+(?P<path>\S+)"),
        re.compile(r"\"(?:GET|POST|PUT|PATCH|DELETE)\s+(?P<path>[^\"\s]+)[^\"]*\"\s+(?P<status>500|404|403|400)"),
    ]

    def classify(text):
        for label, regex in category_patterns:
            if regex.search(text):
                return label
        return "Runtime"

    for log_path in log_candidates:
        lines = _health_safe_tail(log_path, max_lines=2600)
        if not lines:
            continue

        source = log_path.name
        for idx, line in enumerate(lines):
            joined_context = "\n".join(lines[max(0, idx - 2): min(len(lines), idx + 10)])
            if any(marker in line for marker in error_markers):
                start = max(0, idx - 5)
                end = min(len(lines), idx + 18)
                snippet = "\n".join(lines[start:end]).strip()
                errors.append({
                    "source": source,
                    "line": idx + 1,
                    "title": line.strip()[:260],
                    "category": classify(joined_context),
                    "snippet": snippet[:5000],
                })

            for regex in page_patterns:
                match = regex.search(line)
                if match:
                    path = match.groupdict().get("path") or "-"
                    status = match.groupdict().get("status") or match.groupdict().get("kind") or "500"
                    key = (status, path)
                    failed_pages[key] = failed_pages.get(key, 0) + 1

    errors = errors[-60:][::-1]
    failed_rows = [
        {"status": status, "path": path, "count": count}
        for (status, path), count in sorted(failed_pages.items(), key=lambda item: item[1], reverse=True)
    ][:80]

    return errors, failed_rows, [_health_rel(p, base_dir) for p in log_candidates if p.exists()]


def _health_pending_migrations():
    from django.db import DEFAULT_DB_ALIAS, connections
    from django.db.migrations.executor import MigrationExecutor

    try:
        connection = connections[DEFAULT_DB_ALIAS]
        executor = MigrationExecutor(connection)
        conflicts = executor.loader.detect_conflicts()
        targets = executor.loader.graph.leaf_nodes()
        plan = executor.migration_plan(targets)
    except Exception as exc:
        return {
            "ok": False,
            "error": str(exc),
            "items": [],
            "conflicts": [],
        }

    rows = []
    for migration, backwards in plan:
        if backwards:
            continue
        rows.append({
            "app": migration.app_label,
            "name": migration.name,
        })

    conflict_rows = []
    for app_label, names in (conflicts or {}).items():
        conflict_rows.append({"app": app_label, "names": ", ".join(names)})

    return {
        "ok": True,
        "error": "",
        "items": rows,
        "conflicts": conflict_rows,
    }


def _health_database_status():
    from django.apps import apps
    from django.conf import settings
    from django.db import connection

    info = {
        "ok": True,
        "engine": settings.DATABASES.get("default", {}).get("ENGINE", ""),
        "name": str(settings.DATABASES.get("default", {}).get("NAME", "")),
        "vendor": "",
        "size": "",
        "table_count": 0,
        "missing_model_tables": [],
        "warnings": [],
        "error": "",
    }

    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT 1")
            cursor.fetchone()
        info["vendor"] = connection.vendor
    except Exception as exc:
        info["ok"] = False
        info["error"] = str(exc)
        return info

    try:
        existing_tables = set(connection.introspection.table_names())
        info["table_count"] = len(existing_tables)
        for model in apps.get_models():
            table = model._meta.db_table
            if table and table not in existing_tables:
                info["missing_model_tables"].append({
                    "model": f"{model._meta.app_label}.{model.__name__}",
                    "table": table,
                })
    except Exception as exc:
        info["warnings"].append(f"Could not inspect database tables: {exc}")

    try:
        db_name = settings.DATABASES.get("default", {}).get("NAME")
        db_path = Path(db_name)
        if db_path.exists() and db_path.is_file():
            size_mb = db_path.stat().st_size / (1024 * 1024)
            info["size"] = f"{size_mb:.2f} MB"
            if size_mb > 750:
                info["warnings"].append("SQLite database is large. Consider PostgreSQL before production use.")
    except Exception:
        pass

    if "sqlite3" in info["engine"]:
        info["warnings"].append("SQLite is fine for local development, but production ERP should use PostgreSQL or MySQL.")

    return info


def _health_settings_warnings():
    from django.conf import settings

    warnings = []

    if getattr(settings, "DEBUG", False):
        warnings.append("DEBUG=True is enabled. Keep it only for development.")
    if not getattr(settings, "ALLOWED_HOSTS", []):
        warnings.append("ALLOWED_HOSTS is empty. Configure it before deployment.")
    secret = getattr(settings, "SECRET_KEY", "")
    if secret.startswith("django-insecure-"):
        warnings.append("SECRET_KEY uses a development-style value. Use environment variables in production.")
    email_user = getattr(settings, "EMAIL_HOST_USER", "")
    if "yourgmail" in email_user or not email_user:
        warnings.append("Email SMTP credentials are placeholders.")
    if getattr(settings, "TIME_ZONE", "") == "UTC":
        warnings.append("TIME_ZONE is UTC. For India textile ERP usage, consider Asia/Kolkata if reports use local dates.")
    if not getattr(settings, "STATIC_ROOT", None):
        warnings.append("STATIC_ROOT is not configured. collectstatic deployment may be incomplete.")
    if str(getattr(settings, "DATABASES", {}).get("default", {}).get("NAME", "")).endswith("db.sqlite3"):
        warnings.append("Default db.sqlite3 is being used. Good for dev, weak for multi-company production.")

    return warnings


def _health_django_system_checks():
    from django.core.checks import ERROR, WARNING, run_checks

    rows = []
    try:
        messages = run_checks()
    except Exception as exc:
        return {"ok": False, "error": str(exc), "items": []}

    for msg in messages:
        if getattr(msg, "level", 0) < WARNING:
            continue
        rows.append({
            "level": "ERROR" if getattr(msg, "level", 0) >= ERROR else "WARNING",
            "id": getattr(msg, "id", ""),
            "message": str(getattr(msg, "msg", msg)),
            "hint": getattr(msg, "hint", "") or "",
            "object": str(getattr(msg, "obj", "") or ""),
        })

    return {"ok": True, "error": "", "items": rows[:120]}


def _health_python_compile_checks(base_dir):
    import py_compile

    rows = []
    checked = 0
    for path in _health_project_files(base_dir, {".py"}, include_migrations=True):
        checked += 1
        try:
            py_compile.compile(str(path), doraise=True)
        except py_compile.PyCompileError as exc:
            rows.append({
                "file": _health_rel(path, base_dir),
                "error": str(exc),
            })
        except Exception as exc:
            rows.append({
                "file": _health_rel(path, base_dir),
                "error": str(exc),
            })

    return {"checked": checked, "items": rows[:150]}


def _health_python_import_checks():
    import importlib

    modules = [
        "accounts.models",
        "accounts.forms",
        "accounts.forms_legacy",
        "accounts.permissions",
        "accounts.navigation",
        "accounts.urls",
        "accounts.views.core",
        "accounts.views.masters",
        "accounts.views.procurement",
        "accounts.views.production",
        "accounts.views.inventory",
        "accounts.views.sales",
    ]
    rows = []
    for module in modules:
        try:
            importlib.import_module(module)
        except Exception as exc:
            rows.append({
                "module": module,
                "error": f"{exc.__class__.__name__}: {exc}",
            })
    return rows


def _health_model_str_checks():
    """Call str() on one existing object per model to catch small bugs like __str__ using a missing field."""
    from django.apps import apps
    from django.db import DatabaseError, OperationalError, ProgrammingError

    rows = []
    checked = 0
    for model in apps.get_models():
        try:
            manager = getattr(model, "objects", None)
            if manager is None:
                continue
            obj = manager.all().only("pk").first()
            if obj is None:
                continue
            checked += 1
            str(obj)
        except (OperationalError, ProgrammingError) as exc:
            rows.append({
                "model": f"{model._meta.app_label}.{model.__name__}",
                "pk": "-",
                "error": f"DB table/query problem: {exc}",
            })
        except DatabaseError as exc:
            rows.append({
                "model": f"{model._meta.app_label}.{model.__name__}",
                "pk": "-",
                "error": f"DatabaseError: {exc}",
            })
        except Exception as exc:
            rows.append({
                "model": f"{model._meta.app_label}.{model.__name__}",
                "pk": getattr(locals().get("obj", None), "pk", "-"),
                "error": f"{exc.__class__.__name__}: {exc}",
            })

    return {"checked": checked, "items": rows[:150]}


def _health_duplicate_python_symbols(base_dir):
    import ast
    from collections import Counter, defaultdict

    rows = []
    parse_errors = []
    for path in _health_project_files(base_dir, {".py"}, include_migrations=False):
        text = _health_file_text(path)
        if not text:
            continue
        try:
            tree = ast.parse(text, filename=str(path))
        except SyntaxError as exc:
            parse_errors.append({"file": _health_rel(path, base_dir), "error": str(exc)})
            continue
        except Exception as exc:
            parse_errors.append({"file": _health_rel(path, base_dir), "error": str(exc)})
            continue

        names = defaultdict(list)
        for node in ast.iter_child_nodes(tree):
            if isinstance(node, (ast.FunctionDef, ast.AsyncFunctionDef, ast.ClassDef)):
                kind = "class" if isinstance(node, ast.ClassDef) else "function"
                names[(kind, node.name)].append(getattr(node, "lineno", 0))
        for (kind, name), lines in names.items():
            if len(lines) > 1:
                rows.append({
                    "file": _health_rel(path, base_dir),
                    "kind": kind,
                    "name": name,
                    "count": len(lines),
                    "lines": ", ".join(str(line) for line in lines[:8]),
                })

    return {"duplicates": rows[:150], "parse_errors": parse_errors[:80]}


def _health_project_file_warnings(base_dir):
    warnings = []
    base_dir = Path(base_dir)
    checks = [
        (base_dir / "venv", "venv folder is inside the project ZIP/folder. Do not ship it to clients or production."),
        (base_dir / ".git", ".git folder is inside the project folder. Do not include it in client ZIP/deploy builds."),
        (base_dir / "db.sqlite3", "db.sqlite3 exists in project root. Back it up before migrations and avoid shipping real customer data."),
    ]
    for path, message in checks:
        try:
            if path.exists():
                warnings.append(message)
        except Exception:
            pass
    return warnings


def _health_build_critical_issues(**groups):
    issues = []

    def add(title, detail, severity="bad", source=""):
        issues.append({
            "title": title,
            "detail": detail,
            "severity": severity,
            "source": source,
        })

    for row in groups.get("migration_conflicts", []):
        add("Migration conflict", f"{row.get('app')}: {row.get('names')}", "bad", "migrations")
    if groups.get("pending") and not groups["pending"].get("ok"):
        add("Migration check failed", groups["pending"].get("error", ""), "bad", "migrations")
    for row in groups.get("missing_tables", [])[:10]:
        add("Missing database table", f"{row.get('model')} expects {row.get('table')}", "bad", "database")
    for row in groups.get("python_errors", [])[:10]:
        add("Python compile error", f"{row.get('file')}: {row.get('error')}", "bad", "python")
    for row in groups.get("import_errors", [])[:10]:
        add("Import error", f"{row.get('module')}: {row.get('error')}", "bad", "imports")
    for row in groups.get("template_errors", [])[:10]:
        add("Template compile error", f"{row.get('template')}: {row.get('error')}", "bad", "templates")
    for row in groups.get("broken_links", [])[:10]:
        add("Broken URL reference", f"{row.get('name')} in {row.get('file')} line {row.get('line')}", "bad", "urls")
    for row in groups.get("model_str_errors", [])[:10]:
        add("Model __str__ error", f"{row.get('model')} pk={row.get('pk')}: {row.get('error')}", "bad", "models")
    for row in groups.get("system_check_errors", [])[:10]:
        add("Django system check", f"{row.get('id')} {row.get('message')}", "warn" if row.get("level") == "WARNING" else "bad", "checks")
    for warning in groups.get("file_warnings", [])[:6]:
        add("Project file warning", warning, "warn", "files")

    return issues[:40]


@login_required
def system_health_view(request):
    """
    Deep ERP internal System Health page.

    Shows runtime logs plus live diagnostic checks for Python files, templates,
    URL references, static files, database tables, migrations, model __str__,
    Django system checks, duplicate code symbols, and deployment warnings.
    No model/migration is needed because every check is calculated live.
    """
    from django.conf import settings

    base_dir = Path(settings.BASE_DIR)

    pending = _health_pending_migrations()
    database = _health_database_status()
    errors, failed_pages, log_files = _health_collect_errors(base_dir)
    broken_links, duplicate_url_names = _health_find_broken_url_references(base_dir)
    missing_templates = _health_find_missing_templates(base_dir)
    template_compile = _health_compile_all_templates(base_dir)
    missing_static = _health_find_missing_static(base_dir)
    settings_warnings = _health_settings_warnings()
    system_checks = _health_django_system_checks()
    python_compile = _health_python_compile_checks(base_dir)
    import_errors = _health_python_import_checks()
    model_str_checks = _health_model_str_checks()
    duplicate_symbols = _health_duplicate_python_symbols(base_dir)
    file_warnings = _health_project_file_warnings(base_dir)

    migration_conflicts = pending.get("conflicts", []) if pending.get("ok") else []
    missing_tables = database.get("missing_model_tables", [])
    system_check_items = system_checks.get("items", [])

    critical_issues = _health_build_critical_issues(
        pending=pending,
        migration_conflicts=migration_conflicts,
        missing_tables=missing_tables,
        python_errors=python_compile.get("items", []),
        import_errors=import_errors,
        template_errors=template_compile.get("items", []),
        broken_links=broken_links,
        model_str_errors=model_str_checks.get("items", []),
        system_check_errors=system_check_items,
        file_warnings=file_warnings,
    )

    total_issues = (
        len(errors)
        + len(failed_pages)
        + len(broken_links)
        + len(duplicate_url_names)
        + len(missing_templates)
        + len(template_compile.get("items", []))
        + len(missing_static)
        + len(pending.get("items", []))
        + len(migration_conflicts)
        + len(missing_tables)
        + len(settings_warnings)
        + len(database.get("warnings", []))
        + len(system_check_items)
        + len(python_compile.get("items", []))
        + len(import_errors)
        + len(model_str_checks.get("items", []))
        + len(duplicate_symbols.get("duplicates", []))
        + len(duplicate_symbols.get("parse_errors", []))
        + len(file_warnings)
    )

    cards = [
        {
            "label": "Total issues",
            "value": total_issues,
            "state": "bad" if critical_issues else ("warn" if total_issues else "good"),
            "note": "all deep checks combined",
        },
        {
            "label": "Critical issues",
            "value": len(critical_issues),
            "state": "bad" if critical_issues else "good",
            "note": "fix these first",
        },
        {
            "label": "Recent log errors",
            "value": len(errors),
            "state": "bad" if errors else "good",
            "note": "runtime tracebacks/warnings",
        },
        {
            "label": "Failed pages",
            "value": len(failed_pages),
            "state": "bad" if failed_pages else "good",
            "note": "500/404/403/400 in logs",
        },
        {
            "label": "Python compile",
            "value": len(python_compile.get("items", [])),
            "state": "bad" if python_compile.get("items") else "good",
            "note": f"{python_compile.get('checked', 0)} files checked",
        },
        {
            "label": "Template compile",
            "value": len(template_compile.get("items", [])),
            "state": "bad" if template_compile.get("items") else "good",
            "note": f"{template_compile.get('checked', 0)} templates checked",
        },
        {
            "label": "Broken URL names",
            "value": len(broken_links),
            "state": "bad" if broken_links else "good",
            "note": "template/reverse/redirect refs",
        },
        {
            "label": "Missing static",
            "value": len(missing_static),
            "state": "warn" if missing_static else "good",
            "note": "CSS/JS/image references",
        },
        {
            "label": "Pending migrations",
            "value": len(pending.get("items", [])) if pending.get("ok") else "!",
            "state": "bad" if (not pending.get("ok") or pending.get("items") or migration_conflicts) else "good",
            "note": "database schema status",
        },
        {
            "label": "DB missing tables",
            "value": len(missing_tables),
            "state": "bad" if missing_tables else "good",
            "note": "model table existence",
        },
        {
            "label": "Model __str__ bugs",
            "value": len(model_str_checks.get("items", [])),
            "state": "bad" if model_str_checks.get("items") else "good",
            "note": f"{model_str_checks.get('checked', 0)} models sampled",
        },
        {
            "label": "Settings warnings",
            "value": len(settings_warnings),
            "state": "warn" if settings_warnings else "good",
            "note": "deployment readiness",
        },
    ]

    context = {
        "cards": cards,
        "critical_issues": critical_issues,
        "recent_errors": errors,
        "failed_pages": failed_pages,
        "broken_links": broken_links,
        "duplicate_url_names": duplicate_url_names,
        "missing_templates": missing_templates,
        "template_compile": template_compile,
        "missing_static": missing_static,
        "pending": pending,
        "migration_conflicts": migration_conflicts,
        "database": database,
        "settings_warnings": settings_warnings,
        "system_checks": system_checks,
        "python_compile": python_compile,
        "import_errors": import_errors,
        "model_str_checks": model_str_checks,
        "duplicate_symbols": duplicate_symbols,
        "file_warnings": file_warnings,
        "log_files": log_files,
        "generated_at": timezone.localtime(),
        "base_dir": str(base_dir),
    }
    return render(request, "accounts/system_health/index.html", context)



@login_required
def activity_log_view(request):
    """Scoped audit trail page.

    Company admins see their own company activity. Platform super admin sees only
    platform/system-scope logs by default, not daily company ERP activity.
    """
    from .models import AuditLog

    actor = get_actor(request)
    profile = getattr(actor, "erp_profile", None) if actor and getattr(actor, "is_authenticated", False) else None
    is_legacy_admin = bool(actor and getattr(actor, "is_authenticated", False) and not actor.is_superuser and not profile)

    if not (actor and actor.is_authenticated and (actor.is_superuser or is_company_admin(request) or is_legacy_admin)):
        return render(
            request,
            "accounts/permissions/forbidden.html",
            {"required_permission": "Admin only: activity log", "erp_actor": actor},
            status=403,
        )

    base_qs = AuditLog.objects.select_related("company", "actor", "owner").all()
    scope_label = "Company Activity"

    if actor.is_superuser:
        # Platform owner privacy rule: do not show company activity by default.
        base_qs = base_qs.filter(company__isnull=True)
        scope_label = "Platform Activity"
    else:
        company = get_company(request)
        if company:
            base_qs = base_qs.filter(company=company)
            scope_label = f"{company.name} Activity"
        else:
            owner = getattr(request, "erp_owner", None) or actor
            base_qs = base_qs.filter(Q(owner=owner) | Q(actor=actor))
            scope_label = "Own Activity"

    qs = base_qs

    q = (request.GET.get("q") or "").strip()
    action = (request.GET.get("action") or "").strip()
    severity = (request.GET.get("severity") or "").strip()
    module = (request.GET.get("module") or "").strip()
    user = (request.GET.get("user") or "").strip()
    date_from = (request.GET.get("date_from") or "").strip()
    date_to = (request.GET.get("date_to") or "").strip()

    if q:
        qs = qs.filter(
            Q(message__icontains=q)
            | Q(actor_username__icontains=q)
            | Q(actor_display__icontains=q)
            | Q(actor_ip__icontains=q)
            | Q(module__icontains=q)
            | Q(object_model__icontains=q)
            | Q(object_repr__icontains=q)
            | Q(object_pk__icontains=q)
            | Q(path__icontains=q)
        )
    if action:
        qs = qs.filter(action=action)
    if severity:
        qs = qs.filter(severity=severity)
    if module:
        qs = qs.filter(module=module)
    if user:
        qs = qs.filter(actor_username=user)
    if date_from:
        try:
            qs = qs.filter(created_at__date__gte=date_from)
        except Exception:
            pass
    if date_to:
        try:
            qs = qs.filter(created_at__date__lte=date_to)
        except Exception:
            pass

    total_count = qs.count()
    critical_count = qs.filter(severity__in=[AuditLog.SEVERITY_ERROR, AuditLog.SEVERITY_SECURITY]).count()
    warning_count = qs.filter(severity=AuditLog.SEVERITY_WARNING).count()
    change_count = qs.filter(action__in=[AuditLog.ACTION_CREATE, AuditLog.ACTION_UPDATE, AuditLog.ACTION_DELETE]).count()
    login_count = qs.filter(action__in=[AuditLog.ACTION_LOGIN, AuditLog.ACTION_LOGIN_FAILED, AuditLog.ACTION_LOGOUT]).count()

    action_breakdown = list(qs.values("action").annotate(total=Count("id")).order_by("-total")[:10])
    severity_breakdown = list(qs.values("severity").annotate(total=Count("id")).order_by("severity"))
    module_breakdown = list(qs.exclude(module="").values("module").annotate(total=Count("id")).order_by("-total")[:10])
    top_users = list(qs.exclude(actor_username="").values("actor_username").annotate(total=Count("id")).order_by("-total")[:10])

    action_choices = AuditLog.ACTION_CHOICES
    severity_choices = AuditLog.SEVERITY_CHOICES
    # Filter dropdowns must use the scoped queryset, never all companies.
    module_choices = list(
        base_qs.exclude(module="")
        .values_list("module", flat=True)
        .distinct()
        .order_by("module")[:120]
    )
    user_choices = list(
        base_qs.exclude(actor_username="")
        .values_list("actor_username", flat=True)
        .distinct()
        .order_by("actor_username")[:120]
    )

    paginator = Paginator(qs, 50)
    page_obj = paginator.get_page(request.GET.get("page"))

    cards = [
        {"label": "Total Events", "value": total_count, "state": "good", "note": "matching current filters"},
        {"label": "Critical", "value": critical_count, "state": "bad" if critical_count else "good", "note": "errors/security events"},
        {"label": "Warnings", "value": warning_count, "state": "warn" if warning_count else "good", "note": "HTTP 4xx/login failed"},
        {"label": "Data Changes", "value": change_count, "state": "good", "note": "create/update/delete"},
        {"label": "Login Events", "value": login_count, "state": "good", "note": "login/logout/failed"},
    ]

    context = {
        "page_obj": page_obj,
        "logs": page_obj.object_list,
        "cards": cards,
        "action_breakdown": action_breakdown,
        "severity_breakdown": severity_breakdown,
        "module_breakdown": module_breakdown,
        "top_users": top_users,
        "action_choices": action_choices,
        "severity_choices": severity_choices,
        "module_choices": module_choices,
        "user_choices": user_choices,
        "filters": {
            "q": q,
            "action": action,
            "severity": severity,
            "module": module,
            "user": user,
            "date_from": date_from,
            "date_to": date_to,
        },
        "generated_at": timezone.localtime(),
        "is_platform_admin": bool(actor and actor.is_superuser),
        "scope_label": scope_label,
    }
    return render(request, "accounts/system_activity/index.html", context)
